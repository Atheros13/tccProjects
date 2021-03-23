#-----------------------------------------------------------------------------#

### PRIME IMPORTS ###

from __future__ import division

### KIVY IMPORTS ###

from kivy.app import App
from kivy.clock import Clock
from kivy.core.window import Window
from kivy.graphics import Color, Rectangle
from kivy.graphics.vertex_instructions import Line
from kivy.lang import Builder
from kivy.properties import StringProperty, ObjectProperty, NumericProperty
from kivy.properties import ListProperty, DictProperty, BooleanProperty

from kivy.uix.button import Button
from kivy.uix.carousel import Carousel
from kivy.uix.checkbox import CheckBox
from kivy.uix.colorpicker import ColorPicker, ColorWheel
from kivy.uix.dropdown import DropDown
from kivy.uix.filechooser import FileChooserIconView, FileChooserListView
from kivy.uix.image import Image, AsyncImage
from kivy.uix.label import Label 
from kivy.uix.popup import Popup
from kivy.uix.textinput import TextInput

from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.tabbedpanel import TabbedPanel, TabbedPanelItem

### PYTHON IMPORTS ###

from datetime import date, datetime, timedelta

from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
from email.MIMEBase import MIMEBase
from email import encoders

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as Image_pyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side 

from operator import itemgetter

from shutil import copyfile

import os
import os.path
import platform
import sqlite3
import smtplib
import subprocess
import time

#-----------------------------------------------------------------------------#

### CONSTRUCTION CONCEPT DEFINITIONS ###

'''
OUTPUT:		
			These are methods which are called to extract information from
			their class and return their value (in set format)
			i.e. Time24(); calling its .output() method calculates the
			current text input value, converts it to seconds and returns it
			as an INT value

PROCESS : 	
			These are methods which collate information from their class and 
			self/auto update the database.
			i.e. FrameAddMinus(style=gender ...); calling its .process_gender() 
			method makes it directly update the template_groups table in the
			tournament database
'''

#-----------------------------------------------------------------------------#

### BUILDER ###

Builder.load_string('''



''')

#-----------------------------------------------------------------------------#

### GLOBAL VARIABLES ###

global zohome
global zologin
global zotournament

### GLOBAL DATABASES ### 

zosportsdb = {

	'Athletics':[
	('100m','time', '', 0, 1), ('100m Hurdles', 'time', '', 0, 1), 
	('200m', 'time', '', 0, 1), ('400m', 'time', '', 0, 1),
	('800m', 'time', '', 0, 1), ('1500m', 'time', '', 0, 1),
	('3000m', 'time', '', 0, 1), ('Steeple Chase', 'time', '', 0, 1),
	('4x100m Relay', 'time', '', 1, 1),
	('4x400m Relay', 'time', '', 1, 1),
	('Long Jump', 'distance', 'length', 0, 1), 
	('Triple Jump', 'distance', 'length', 0, 1),
	('High Jump', 'distance', 'height', 0, 1),
	('Pole Vault', 'distance', 'height', 0, 1),
	('Discus', 'distance', 'length', 0, 1),
	('Shot Put', 'distance', 'length', 0, 1),
	('Javelin', 'distance', 'length', 0, 1),
	('Hammer Throw', 'distance', 'length', 0, 1)], 
	
	'Swimming':[
	('50m Freestyle','time', '', 0, 1), ('50m Breaststroke', 'time', '', 0, 1), 
	('50m Backstroke', 'time', '', 0, 1), ('50m Butterfly', 'time', '', 0, 1),
	('100m Freestyle','time', '', 0, 1), ('100m Breaststroke', 'time', '', 0, 1), 
	('100m Backstroke', 'time', '', 0, 1), ('100m Butterfly', 'time', '', 0, 1),
	('200m Freestyle','time', '', 0, 1), ('200m Breaststroke', 'time', '', 0, 1), 
	('200m Backstroke', 'time', '', 0, 1), ('200m Butterfly', 'time', '', 0, 1),
	('200m Individual Medley','time', '', 0, 1), 
	('400m Individual Medley', 'time', '', 0, 1), 
	('4x50m Freestyle Relay', 'time', '', 1, 1), 
	('4x100m Freestyle Relay', 'time', '', 1, 1), 
	('4x50m Medley Relay', 'time', '', 1, 1)]

	}
	
### GLOBAL FUNCTIONS ###

def calculate_age_filters(tournament, **kwargs):

	db = Database(tournament)
	age_db = db.cursor.execute('''SELECT title, under, year, date
				FROM template_groups WHERE category="age"''').fetchall()

	output = {}

	filter_under = {}
	for tup in age_db:
		if tup[0] == 'Open':
			continue
				
		title = tup[0]
		than = tup[1].title()

		date_split = tup[3].split('/')
		year = str(int(date_split[2])-tup[2])
		d = '/'.join([date_split[0], date_split[1], year])
		date = convert_date(d)

		if than == 'Under':
			filter_under[title] = date
		else:
			output[title] = [date, None]

	u_sort = sorted(filter_under, key=filter_under.get, reverse=True)
	for i in range(len(u_sort)):
		title = u_sort[i]
		date = filter_under[title]
		if i == 0:
			output[title] = [None, date]
		else:
			output[title] = [filter_under[u_sort[i-1]], date]

	return output

def calculate_competitor_grades(tournament, dob, **kwargs):

	gender = kwargs.pop('gender', None)
	age = kwargs.pop('age', 'Date of Birth')

	db = Database(tournament)

	if gender != None:
		g = [db.cursor.execute('''SELECT title FROM template_groups
							WHERE filter="%s"''' % gender).fetchone()[0]]

	if age != 'Date of Birth':
		a = [age]

	else:
		age_filter = calculate_age_filters(tournament)

		a = []

		for filter in age_filter:

			title = filter
			min = age_filter[title][0]
			max = age_filter[title][1]
			if min == None:
				if dob > max:
					a.append(title)
			elif max == None:
				if dob <= min:
					a.append(title)
			else:
				if  dob <= min and dob > max:
					a.append(title)

	a.append('Open')
	c = []

	for y in a:
		for mf in g:
			c.append('%s %s' % (y, mf))

	return g, a, c

def calculate_contest_list(tournament, **kwargs):

	time = kwargs.pop('time', False)
	distance = kwargs.pop('distance', False)
	team = kwargs.pop('team', True)

	db = Database(tournament)

	if team == True:
		command1 = '''SELECT title FROM template_contests 
					WHERE measure="time" AND selection=1'''	

		command2 = '''SELECT title FROM template_contests 
					WHERE measure="distance" AND selection=1'''	
	elif team == False:
		command1 = '''SELECT title FROM template_contests 
					WHERE measure="time" AND selection=1
					AND team!=1'''	

		command2 = '''SELECT title FROM template_contests 
					WHERE measure="distance" AND selection=1
					AND team!=1'''

	t1 = db.cursor.execute(command1).fetchall()		
	d1 = db.cursor.execute(command2).fetchall()

	c_time = []
	for tup in t1:
		c_time.append(tup[0])
	c_distance = [] 
	for tup in d1:
		c_distance.append(tup[0])

	if time == True:
		return c_time
	if distance == True:
		return c_distance
	return c_time + c_distance

def calculate_contest_type(tournament, contest, **kwargs):

	team = kwargs.pop('team', False)
	height = kwargs.pop('height', False) # not figured out how I need this

	db = Database(tournament)

	if team == False:

		return db.cursor.execute('''SELECT measure FROM template_contests
								WHERE title="%s"''' % contest).fetchone()[0]

	if team == True:
		boole =  db.cursor.execute('''SELECT team FROM template_contests
								WHERE title="%s"''' % contest).fetchone()[0]

		return db_boole(boole, db_in=False)

def calculate_event_state(tournament, event, round):

	'''
	This checks to see if the event round has any placings in it, 
	that way it can see if it should be referenced on wristbands or
	in type-lists. None means no placings, True means a placing has 
	been saved
	'''

	db = Database(tournament)

	table = db_table('%s %s' % (event, round))
	print table

	placing = db.cursor.execute('''SELECT pp
									FROM %s''' % table).fetchall()

	check = None

	if placing == None:

		return None

	for tup in placing:

		if tup[0] not in [None, '', 0]:
			check = True

	return check

def calculate_event_list_state(tournament):

	db = Database(tournament)

	grade_list = calculate_grade_list(tournament)
	contest_list = calculate_contest_list(tournament)
	table_list = db.list_tables()

	event_list = []

	def check_table(table):

		placing = db.cursor.execute('''SELECT pp
							FROM %s''' % table).fetchall()
		check = None

		for tup in placing:
			if tup[0] not in [None, '', 0, '0']:
				check = True

		return check

	for c in contest_list:
		for g in grade_list:		
			event = '%s %s' % (g, c)
			table = db_table(event)
			if table not in table_list:
				continue

			main_table = db_table('%s R1' % event)
			if check_table(main_table) == None:

				prog = db.cursor.execute('''SELECT program_time
								FROM template_events
								WHERE title="%s"''' % event).fetchone()[0]

				event_list.append([event, 'R1', prog, event])


			rounds = db.cursor.execute('''SELECT title, round, program_time
								FROM %s
								WHERE round!="R1"
								AND round!="RR"''' % table).fetchall()

			for tup in rounds:

				evt = tup[0]
				rd = tup[1]
				extra_table = db_table('%s %s' % (event, rd))
				if check_table(extra_table) == None:
					event_list.append([event, rd, tup[2], evt])

	return event_list

def calculate_event_team(tournament, event, **kwargs):

	db = Database(tournament)

	contest = db.cursor.execute('''SELECT contest FROM template_events
								WHERE title="%s"''' % event).fetchone()[0]
	return calculate_contest_type(tournament, contest, team=True)

def calculate_grade_events(tournament, grade, **kwargs):
	
	table = kwargs.pop('table', False)
	open = kwargs.pop('open', True)
	team = kwargs.pop('team', True)

	db = Database(tournament)

	if team == True:
		var = calculate_contest_list(tournament)
	elif team == False:
		var = calculate_contest_list(tournament, team=False)

	try:
		type = db.cursor.execute('''SELECT category FROM template_grades
								WHERE title="%s"''' % grade).fetchone()[0]
	except TypeError:
		type = 'Open'

	combined_events = []
	age_events = []
	gender_events = []
	open_events = []

	for contest in var:
	
		op = db.cursor.execute('''SELECT title, contest FROM template_events
							WHERE grade="Open"
							AND contest="%s"''' % contest).fetchall()
		for tup in op:
			open_events.append(tup[0])

	if type == 'gender':

		for contest in var:

			gev = db.cursor.execute('''SELECT title FROM template_events
							WHERE grade="%s" 
							AND contest="%s"''' % (grade, contest)).fetchone()
	
			if gev != None:
				gender_events.append(gev[0])

			open_gender = "Open %s" % grade
			o_events = db.cursor.execute('''SELECT title FROM template_events
								WHERE grade="%s"
						AND contest="%s"''' % (open_gender, contest)).fetchone()
			
			if o_events != None:
				for tup in o_events:
					open_events.append(o_events[0])			

	elif type == 'age':

		for contest in var:
			
			aev = db.cursor.execute('''SELECT title FROM template_events
							WHERE grade="%s"
							AND contest="%s"''' % (grade, contest)).fetchone()
			if aev != None:
				age_events.append(aev[0])

	elif type == 'combined':
		
		genderage = db.cursor.execute('''SELECT gender_grade, age_grade
								FROM template_grades
								WHERE title="%s"''' % grade).fetchall()	

		gender = genderage[0][0]
		age = genderage[0][1]

		for contest in var:

			cev = db.cursor.execute('''SELECT title FROM template_events
							WHERE grade="%s" 
							AND contest="%s"''' % (grade, contest)).fetchone()

			if cev != None:
				combined_events.append(cev[0])


			gev = db.cursor.execute('''SELECT title FROM template_events
							WHERE grade="%s" 
							AND contest="%s"''' % (gender, contest)).fetchone()
	
			if gev != None:
				gender_events.append(gev[0])

			open_gender = "Open %s" % gender
			o_events = db.cursor.execute('''SELECT title FROM template_events
								WHERE grade="%s"
						AND contest="%s"''' % (open_gender, contest)).fetchone()
			
			if o_events != None:
				for tup in o_events:
					open_events.append(o_events[0])

			aev = db.cursor.execute('''SELECT title FROM template_events
							WHERE grade="%s"
							AND contest="%s"''' % (age, contest)).fetchone()

			if aev != None:
				age_events.append(aev[0])

	event_list = combined_events + gender_events + age_events + open_events
	

	if table == False:
		
		return event_list

	elif table == True:
		output = []
		for event in event_list:
			output.append('_'.join(event.split(' ')))

		return output

def calculate_grade_list(tournament, **kwargs):
	
	gender = kwargs.pop('gender', False)
	combined = kwargs.pop('combined', False)
	open = kwargs.pop('open', False)
	age = kwargs.pop('age', False)
	gender_check = kwargs.pop('gender_check', False)

	db = Database(tournament)

	commandg = '''SELECT title FROM template_groups 
					WHERE category="gender"'''
	commanda = '''SELECT title FROM template_groups
					WHERE category="age"'''
	g1 = db.cursor.execute(commandg).fetchall()
	a1 =db.cursor.execute(commanda).fetchall()

	gender_grades = []
	age_grades = []
	combined_grades = []
	open_grades = ['Open']

	for tup in g1:
		gender_grades.append(tup[0])
	for tup in a1:
		a = tup[0]
		if a != 'Open':
			age_grades.append(a)
	for a in age_grades:
		for g in gender_grades:
			grade = '%s %s' % (a, g)
			combined_grades.append(grade)
	for g in gender_grades:		
		open_grades.append('Open %s' % g)

	if combined == True:

		if gender_check == True:
			if len(gender_grades) == 1:
				return age_grades
	
		return combined_grades

	elif age == True:
		return age_grades
	elif gender == True:
		return gender_grades
	elif open == True:
		return open_grades

	return gender_grades + age_grades + combined_grades + open_grades

def calculate_grade_type(tournament, grade, **kwargs):

	db = Database(tournament)

	try:
		type = db.cursor.execute('''SELECT category FROM template_grades
								WHERE title="%s"''' % grade).fetchone()[0]
	except TypeError:
		type = 'Open'

	return type

def calculate_id_list(tournament, **kwargs):

	db = Database(tournament)

	type = kwargs.pop('type', None)
	grade = kwargs.pop('grade', None)
	tup_style = kwargs.pop('tup_style', False)

	grade_type = calculate_grade_type(tournament, grade)
	grade_like = '%{}%'.format(grade)

	# return all from one type i.e. house 
	if type != None:

		if grade != None and grade != 'Open':
			result = db.cursor.execute('''SELECT id FROM individuals
						WHERE type="%s" AND %s_grade LIKE "%s"''' % 
						(type, grade_type, grade_like)).fetchall()
		else:
			result = db.cursor.execute('''SELECT id FROM individuals
										WHERE type="%s"''' % type).fetchall()
	# return all in one grade ()
	elif type == None:
		if grade != None:
			result = db.cursor.execute('''SELECT id FROM individuals
				WHERE type="%s" AND %s_grade LIKE "%s"''' % 
				(type, grade_type, grade_like)).fetchall()

	else:
		result = db.cursor.execute("SELECT * FROM individuals").fetchall()

	if tup_style == False:
		output = []
		for tup in result:
			output.append(tup[0])
		return output

	if tup_style == True:
		return result

def calculate_id_list_v1(tournament, **kwargs):

	db = Database(tournament)

	type = kwargs.pop('type', None)
	grade = kwargs.pop('grade', None)
	tup_style = kwargs.pop('tup_style', False)

	grade_type = calculate_grade_type(tournament, grade)
	grade_like = '%{}%'.format(grade)

	# return all from one type i.e. house 
	if type != None:

		if grade != None and grade != 'Open':
			result = db.cursor.execute('''SELECT id FROM individuals
						WHERE type="%s" AND %s_grade LIKE "%s"''' % 
						(type, grade_type, grade_like)).fetchall()
		else:
			result = db.cursor.execute('''SELECT id FROM individuals
										WHERE type="%s"''' % type).fetchall()
	# return all in one grade ()
	elif type == None:
		if grade != None:
			result = db.cursor.execute('''SELECT id FROM individuals
				WHERE %s_grade LIKE "%s"''' % 
				(grade_type, grade_like)).fetchall()

	else:
		result = db.cursor.execute("SELECT * FROM individuals").fetchall()

	if tup_style == False:
		output = []
		for tup in result:
			output.append(tup[0])
		return output

	if tup_style == True:
		return result

def calculate_type_tables(tournament, type, **kwargs):

	grade = kwargs.pop('grade', False)

	grade_list = calculate_grade_list(tournament, combined=True, 
										gender_check=True)
	if grade_list in [None, []]:
		grade_list = calculate_grade_list(tournament,
											open=True).remove('Open')	

	output = []
	for g in grade_list:

		typelist = type.split(' ')
		table = ('_'.join(typelist+g.split(' ')))

		if grade == False:
			output.append(table)
		elif grade == True:
			output.append([g, table])

	return output

def calculate_ordinal(n, **kwargs):
	'''	
	Converts a number (n) into a ordinal string i.e. int(1) become 1st	
	**kwargs:
		reverse=True reverses this i.e. 2nd becomes int(2)
	(Gareth on codegolf)
	'''
	if n == 13: # 13/01/2017 - Just found out the code below returns 13rd
		return '13th'
	# which means that goes for all numbers ending eleventh, twelfth, thirteenth
	if len(str(n)) > 1:
		if str(n)[-2] == '1':
			if str(n)[-1] in ['1', '2', '3']:			
				return '%sth' % n

	reverse = kwargs.pop('reverse', False)	
	if reverse == True:
		ending = n[-2] + n[-1]
		return int(n.split(ending)[0])					
	ordinal = lambda n: "%d%s" % (n,"tsnrhtdd"[(n/10%10!=1)*(n%10<4)*n%10::4])
	return ordinal(n)

def calculate_str_record(record, type):

	if record in [0, 0.0, '', None]:
		return ''

	record = float(record)
	if record.is_integer():
		record = int(record)

	if type == 'time':

		mins = 0
		secs = record
		while secs > 59:
			mins += 1
			secs -= 60

		if mins == 0:
			return '%ss' % secs
		else:
			return '%sm %ss' % (mins, secs) #'1m 12.5s'

	elif type == 'distance':

		return '%s m' % record

def champollion(demotic):
	'''
	I know this isn't that clever, but it gives the right 'look' of a 
	registration key, it changes based on the year (well did) and frankly I could
	have wasted days trying to come up with more difficult key to crack, 
	but it would naturally be in vain, if someone wants to crack a 
	pass-key they just need time. 
	Why am I telling myself this? 

	'''

	year = '2016'

	table = {0:['X'], 1:['A','J','S'], 2:['B','K','T'], 3:['C','L','U'],
			4:['D','M', 'V'], 5:['E', 'N', 'W'], 6:['F','O', 'X'],
			7:['G','P','Y'], 8:['H','Q','Z'], 9:['I','R']}

	numerology = []
	for i in range(len(demotic)):
		letter = demotic.upper()[i]
		for number in table:
			if number == 0:
				continue
			if letter in table[number]:
				numerology.append(number)

	convert = []
	position = 0
	for number in numerology:
		total = number * int(year[position])
		for t in str(total):
			convert.append(int(t))
		if position != 3:
			position += 1
		else:
			position = 0

	output = []

	output.append(table[convert[0]][0])

	position = 0
	triple = []
	for i in range(1, len(convert)-1):
		if i == 13:
			output.append('Z')
		if position == 2:			
			total = str(sum(triple))
			while len(total) > 1:
				shorten = []
				for i in range(len(total)):
					shorten.append(int(total[i]))
				total = str(sum(shorten))
			output.append(total)
			position = 0
			triple = []
		else:
			triple.append(convert[i])
			position += 1

	output.append(table[convert[-1]][0])

	return ''.join(output)

def convert_date(input, **kwargs):

	error = kwargs.pop('error', False)
	string = kwargs.pop('string', False)

	if string == False: # convert input which is a str() into datetime()

			if input == None:
				return None			
			try:
				return datetime.strptime(input, '%d/%m/%Y').date()
			except ValueError:
				if error == True:
					return False				
				cups = input.split('/')
				try:
					input = '%s/%s/%s' % (cups[1], cups[0], cups[2])
				except IndexError:
					print cups

				return datetime.strptime(input, '%d/%m/%Y').date()

	elif string == True: # convert a datetime() to a str() format 'dd/mm/YY'
				
		return input.strftime('%d/%m/%Y')

def convert_score(input, measure):

	if input in [None, 0, 0.0]:
		return ''

	if measure == 'time':

		value = float(input)
		minutes = 0
		while value >= 60:
			value -= 60
			minutes += 1
		
		min = 0
		sec = 0

		if minutes > 0:
			min = str(minutes)
		if value >= 0:
			if value.is_integer():
				sec = str(int(value))
			else:
				sec = str(value)

		if min != 0:
			return '%sm %ss' % (min, sec)
		else:
			if sec != 0:
				return '%ss' % sec
			else:
				return ''

	elif measure == 'distance':

		return '%s m' % input

def db_boole(input, **kwargs):

	db_in = kwargs.pop('db_in', True)

	if db_in == True: # converting from boole to binary, to go into db
		if input == False:
			return 0
		elif input == True:
			return 1

	elif db_in == False:
		if input in [None, 0, '']:
			return False
		elif input == 1:
			return True

def db_colour(input, **kwargs):

	db_in = kwargs.pop('db_in', True)

	if db_in == True: # take a colour list and convert for db
		output = '%s#%s#%s#%s' % (input[0], input[1], input[2], input[3])

	elif db_in == False: # converts from db to colour list
		preoutput = input.split('#')
		output = []
		for i in preoutput:
			output.append(float(i))

	return output

def db_date(input, **kwargs):

	db_in = kwargs.pop('db_in', True)

	if input == None:
		return None

	if db_in == True: # convert date() to str()
		return input.strftime('%d/%m/%Y')

	elif db_in == False: # convert str() to date()
		return datetime.strptime(input, '%d/%m/%Y').date()

def db_list(input, **kwargs):
	
	db_in = kwargs.pop('db_in', True)

	if db_in == True: 
		return '#'.join(input)
	elif db_in == False: 
		return input.split('#')

def db_height(input, **kwargs):

	# input = [[' ',' ',' ']]

	db_in = kwargs.pop('db_in', True) # assumes its going into db

	if db_in == True:

		output = []

		for t in input:
			output.append('%s%s%s' % (t[0], t[1], t[2]))

		return '#'.join(output)

	elif db_in == False:

		if input == None:
			return [[' ',' ',' '], [' ',' ',' '], [' ',' ',' '],
					[' ',' ',' '], [' ',' ',' '], [' ',' ',' ']]

		output = []

		for t in input.split('#'):
			output.append([t[0], t[1], t[2]])

		return output

def db_score(input, contest, contest_type, **kwargs):

	db_in = kwargs.pop('db_in', True)

	if contest_type == 'time':
		return input
	if contest_type == 'distance':

		if contest in ['High Jump', 'Pole Vault']:
			return db_height(input, db_in=db_in)

		if db_in == True: 
			output = []
			for i in input:
				output.append(str(i))
			return '#'.join(output)
		elif db_in == False: 
			output = []
			for i in input.split('#'):
				if i == '':
					output.append(0.0)
				else:
					output.append(float(i))
			return output	

def db_table(input, **kwargs):

	db_in = kwargs.pop('db_in', True) # adds the '_' part

	if db_in == True:
		return '_'.join(input.split(' '))
	elif db_in == False:
		return ' '.join(input.split('_'))

def db_vertical(input, tournament, event, event_round, **kwargs):

	db = Database(tournament)
	best = kwargs.pop('best', False)

	event_table = db_table(event)

	h_ref = db.cursor.execute('''SELECT height
							FROM %s
							WHERE title="%s"''' % (event_table,
													event_round)).fetchone()[0]
	
	if h_ref in [None, ()]:
		return None

	height_ref = db_score(h_ref, False, 'distance', db_in=False)

	heights = []
	for i in range(6):
		hgt = height_ref[i]
		if 'O' in input[i]:
			heights.append(hgt)

	if heights == []:

		return None

	elif best == True:

		return sorted(heights)[-1]

	else:

		return heights

def file_copy(old_file, new_file):

	if platform.system() == "Darwin": # Mac
		subprocess.check_call(["copy", old_file, new_file])				
	elif platform.system() == "Windows": # PC
		os.system("copy %s %s" % (old_file, new_file)) # does this require resource_path()

def file_open(file):

	if platform.system() == "Darwin": # Mac
		subprocess.check_call(["open", file])				
	elif platform.system() == "Windows": # PC
		os.startfile(file) 	 # does this require resource_path()

def make_tournament_list(**kwargs):

	copy = kwargs.pop('copy', False)
	add_sport = kwargs.pop('add_sport', False) # returns 'title - sport'
	add_stage = kwargs.pop('add_stage', False) # returns ['title', 'stage']

	## Database
	db = Database('main')
	tournaments = db.select(table='tournaments')

	output = []

	for pair in tournaments:

		db1 = Database(pair[0])
		stage = db1.cursor.execute("SELECT stage FROM details").fetchone()[0]

		if copy == True:
			if stage == 'Template':
				continue

			if add_sport == True:
				title = '%s - %s' % (pair[0], pair[1])
				output.append(title)			

			else:
				output.append(pair[0])

		elif add_sport == True:
			title = '%s - %s' % (pair[0], pair[1])
			if add_stage == True:
				output.append([title, stage])
			else:
				output.append(title)

		elif add_stage == True:

			output.append([pair[0], stage])

		else:
			output.append(pair[0])

	return output

def resource_path(relative_path):
    """ Gets absolute path to resource, works for dev and for PyInstaller """

    try:
    	base_path = sys._MEIPASS
    	return os.path.join(base_path, relative_path)
    except:
    	base_path = os.path.abspath(".")
    	return os.path.join(base_path, relative_path)

#-----------------------------------------------------------------------------#

### DATABASE CODE ###

class Database():
	'''
	Use to access connection and cursor for creating/updating/deleting data,
	also has built in methods which allow for quick request of stored data
	'''

	def __init__(self, file, *args, **kwargs):

		self.file = file
		filename = os.path.join(os.path.abspath("."), '%s.db' % self.file)
		
		try:
			self.connection = sqlite3.connect(filename) # allows for TIMESTAMP
	
		except Exception:

			base_path = sys._MEIPASS # exists in temp dif
			tempfilename = os.path.join(base_path, '%s.db' % self.file)

			connection = sqlite3.connect(tempfilename) # access temp file
			copyfile(tempfilename, filename) # copy temp file to path
			connection.close() # close temp file
			self.connection = sqlite3.connect(filename) # open path file

		self.cursor = self.connection.cursor()

	def list_columns(self, table):

		cursor = self.connection.execute("SELECT * FROM %s;" % table)
		return list(map(lambda x: x[0], cursor.description))

	def list_tables(self):

		command = "SELECT name FROM sqlite_master WHERE type='table';"
		tables = self.cursor.execute(command).fetchall()
		output = []
		for table in tables:
			output.append(table[0])
		return output

	def select(self, **kwargs):

		table = kwargs.pop('table', None) # String Property
		cols = kwargs.pop('columns', None) # List Property

		if cols != None:
			columns = ', '.join(cols)
		else:
			columns = '*'

		command = "SELECT %s FROM %s;" % (columns, table)

		return self.cursor.execute(command).fetchall()		

class DataWarning():
	'''
	Class called to warn if an attempt to update a database fails because
	that database is being used already
	'''
	def __init__(self, *args, **kwargs):

		PopBox().showwarning(title='Error: Database Issue',
			message='There is a database access issue.\n\nUsually this can be easily solved by closing the program down and reopening it.\n\nIf the problem persists, contact us directly at info@zo-sports.com.\n\nThis version is very much still in development so there may be the occasional bug, thank you for your understanding')

class DataEmail():

	def __init__(self, *args, **kwargs):

		## Settings
		self.tournament = kwargs.pop('tournament', 'main')
		self.style = kwargs.pop('style', 'tournaments')
		self.feedback = kwargs.pop('feedback', 'None')
		self.db = Database('main')
		details = self.db.cursor.execute('''SELECT title, organisation  
										FROM details''').fetchone()
		self.title = details[0]
		self.organisation = details[1]

		settings = self.db.cursor.execute('''SELECT address, email
										FROM settings''').fetchone()
		self.address = settings[0]
		self.email = settings[1]

		## Engine
		if self.style == 'tournaments':
			self.send_tournament()
		elif self.style == 'feedback':
			self.send_feedback()

	def send_tournament(self):

		fromaddr = "zosportsnz@gmail.com"
		toaddr = "tournaments@zo-sports.com"
 
		msg = MIMEMultipart()
 
		msg['From'] = fromaddr
		msg['To'] = toaddr
		msg['Subject'] = "%s - %s" % (self.title, self.tournament)
 
		body = "Title: %s\nOrg: %s\nAddress: %s\nEmail: %s" % (self.title,
							self.organisation, self.address, self.email)
 
		msg.attach(MIMEText(body, 'plain'))
 
		filename = "%s.db" % self.tournament
		attachment = open("%s.db" % self.tournament, "rb")
 
		part = MIMEBase('application', 'octet-stream')
		part.set_payload((attachment).read())
		encoders.encode_base64(part)
		part.add_header('Content-Disposition', 
						"attachment; filename= %s" % filename)
 
		msg.attach(part)
 
		server = smtplib.SMTP('smtp.gmail.com', 587)
		server.starttls()
		server.login(fromaddr, "vetinari13")
		text = msg.as_string()
		server.sendmail(fromaddr, toaddr, text)
		server.quit()

	def send_feedback(self):

		fromaddr = "zosportsnz@gmail.com"
		toaddr = "feedback@zo-sports.com"
 
		msg = MIMEMultipart()
 
		msg['From'] = fromaddr
		msg['To'] = toaddr
		msg['Subject'] = "%s - %s" % (self.title, self.tournament)
 
		body = "Title: %s\nOrg: %s\nAddress: %s\nEmail: %s\n\n%s" % (self.title,
							self.organisation, self.address, self.email, self.feedback)
 
		msg.attach(MIMEText(body, 'plain'))
 
		server = smtplib.SMTP('smtp.gmail.com', 587)
		server.starttls()
		server.login(fromaddr, "vetinari13")
		text = msg.as_string()
		try:
			server.sendmail(fromaddr, toaddr, text)
			server.quit()
		except:
			pass

class EntryAddRemove():

	def __init__(self, *args, **kwargs):
		
		## Settings
		self.style = kwargs.pop('style', 'Add')
		self.tournament = kwargs.pop('tournament', '')
		self.typename = kwargs.pop('typename', None)
		self.event = kwargs.pop('event', None)
		self.event_round = kwargs.pop('event_round', None)
		self.event_table = kwargs.pop('event_table', None)
		self.widget = kwargs.pop('widget', None)

		self.db = Database(self.tournament)
		details = self.db.cursor.execute('''SELECT *
										FROM details''').fetchone()

		self.type = details[3]
		self.type_list = []
		type_list = self.db.cursor.execute('''SELECT title 
								FROM template_groups
								WHERE category="type"''').fetchall()
		for tup in type_list:
			
			self.type_list.append(tup[0])

		self.age = details[4]
		self.age_list = []
		if self.age != 'Date of Birth':
			age_list = self.db.cursor.execute('''SELECT title
								FROM template_groups
								WHERE category="age"''').fetchall()
			for tup in age_list:
				self.age_list.append(tup[0])

		## Database
		self.menu = {}

		## Display
		self.content = GridLayout(cols=1, spacing=5)
		self.build_title()
		self.build_search()
		self.build_results()

		## Engine
		self.pop = Popup(title=self.tournament, content=self.content, 
							auto_dismiss=False,
							size_hint=(None, None), size=(800, 500))
		self.pop.open()

	### FUNCTIONS ###

	def command_add(self, details):

		last_id = self.db.cursor.execute('''SELECT id FROM individuals
											ORDER BY id DESC''').fetchone()[0]

		id = 'ID%s' % str((int(last_id.split('ID')[1]) + 1))
		fn = details[1]
		sn = details[2]
		dob = db_date(details[3])
		gender = details[4]
		type = details[5]
		year = details[6]

		if year == None:
			gender_grade, age_grade, combined_grade = calculate_competitor_grades(self.tournament,
								details[3], gender=gender)
		else:
			gender_grade, age_grade, combined_grade = calculate_competitor_grades(self.tournament,
								details[3], gender=gender, age=year)

		gg = db_list(gender_grade)
		ag = db_list(age_grade)
		cg = db_list(combined_grade)

		events = ''
		if self.event != None:
			events = db_list([self.event])

		try:
			var = [id, fn, sn, dob, gender, type, year, gg, ag, cg]

			self.db.cursor.execute('''INSERT INTO individuals
							VALUES(?,?,?,?,?,?,?,?,?,?)''', var)
		except:
			var = [id, fn, sn, dob, gender, type, year, gg, ag, cg, '']

			self.db.cursor.execute('''INSERT INTO individuals
							VALUES(?,?,?,?,?,?,?,?,?,?,?)''', var)			

		self.db.connection.commit()

		if self.event == None: # Entry - Events

			grade_list = calculate_grade_list(tournament=self.tournament,
										combined=True, gender_check=True)
			for grade in grade_list:
				t = '_'.join([type] + grade.split(' '))
				id_list = calculate_id_list(self.tournament, type=type,
											grade=grade, tup_style=True)

				for tup in id_list:
					if id in tup:
						self.db.cursor.execute('''INSERT INTO %s (id)
											VALUES ('%s')''' % (t, id))
						self.db.connection.commit()		

		else: # Competition
			self.command_add_entry(id)
		## Engine
		self.pop.dismiss()

	def command_add_entry(self, i):

		id = i
		if i == None: # means its from a selection, see self.menu
			for m in self.menu:
				if self.menu[m].active == True:
					id = m

			if id == None:
				return

		self.ref = id
		
		names = self.db.cursor.execute('''SELECT firstname, surname
									FROM individuals
									WHERE id="%s"''' % id).fetchone()
		name = '%s %s' % (names[0], names[1])

		event_grade = self.db.cursor.execute('''SELECT grade
								FROM template_events
								WHERE title="%s"''' % self.event).fetchone()[0]

		id_grades = self.db.cursor.execute('''SELECT gender_grade, 
											age_grade, combined_grade
									FROM individuals
									WHERE id="%s"''' % id).fetchall()[0]
		id_grade = []
		for grade in id_grades:

			try:
				grade_list = db_list(grade, db_in=False)
			except AttributeError:
				grade_list = grade

			for g in grade_list:
				id_grade.append(g)

		if event_grade not in id_grade:

			mes = "%s is a %s event and %s is not in %s. Are you sure you still want to add them to this event?" % ( self.event,
				event_grade, name, event_grade)

			PopBox().askyesno(title='Grade Error', message=mes,
								function=self.command_entry)
			return

		else:
			self.command_entry(True)

	def command_clear(self, i):

		names = [self.fn, self.sn]
		if self.style != 'Add':
			names.append(self.id)
		for n in names:
			n.text = ''

		self.dob.date = None
		self.dob.show.text = 'Date'

		select = [self.gender_choice, self.type_choice]
		if self.age != 'Date of Birth':
			select.append(self.age_choice)
		for s in select:
			s.text = 'Select'

		self.results.clear_widgets()
		self.ref = None
		self.menu = {}

	def command_entry(self, boole):

		if boole != True:
			return

		# check if already in event round
		table = self.event_table

		check = self.db.cursor.execute('''SELECT * FROM %s
							WHERE id="%s"''' % (table, self.ref)).fetchone()

		if check != None:
			PopBox().showwarning(title='Entry Error',
				message="Competitor already entered in %s" % self.event_round)
			self.pop.dismiss()
			return

		self.db.cursor.execute('''INSERT INTO %s (id)
								VALUES ("%s")''' % (table, self.ref))
	
		self.db.connection.commit()

		## Engine
		self.widget.build_display_structure(self.event_round)
		self.pop.dismiss()

	def command_remove(self, *args):

		id = None
		for i in self.menu:
			if self.menu[i].active == True:
				id = i

		if id == None:
			PopBox().showwarning(title='Error',
				message='No Competitor has been chosen for removal')
			return

		# individual table
		type = self.db.cursor.execute('''SELECT type FROM individuals
										WHERE id="%s"''' % id).fetchone()[0]

		self.db.cursor.execute('''DELETE FROM individuals
								WHERE id="%s"''' % id)
		# house table
		table_list = calculate_type_tables(self.tournament, type)
		for table in table_list:
			self.db.cursor.execute('''DELETE FROM %s
								WHERE id="%s"''' % (table, id))
		# commit
		self.db.connection.commit()
			
	def command_search(self, i):

		self.results.clear_widgets()

		if self.style == 'Add':

			details = [None] # None represents the id number

			check = True
			# check for all details
			for n in [self.fn, self.sn]:
				if n.text == '':
					check = False
				else:
					details.append(n.text.title())
			if self.dob.date == None:
				check = False
			else:
				details.append(self.dob.date)
			select = [self.gender_choice, self.type_choice]
			if self.age != 'Date of Birth':
				select.append(self.age_choice)
			for s in select:
				if s.text == 'Select':
					check = False
				else:
					details.append(s.text)
			if len(select) == 2:
				details.append(None)
			if check == False:
				PopBox().showwarning(title='Error',
					message='All competitor details must be entered')
				return

			self.match(details)

		elif self.style in ['Remove', 'Entry']:

			#settings
			id = 'ID%s' % self.id.text
			fn = self.fn.text
			sn = self.sn.text
			dob = self.dob.date
			gender = self.gender_choice.text
			type = self.type_choice.text

			if fn != '':
				fn = fn.title()
			if sn != '':
				sn = sn.title()

			details = [id, fn, sn, dob, gender, type]
			if self.age != 'Date of Birth':
				age = self.age_choice.text
				details.append(age)
			else:
				details.append(None)

			self.match(details)

	def command_select(self, obj, value, id):

		if value == True:
			for ref in self.menu:
				if ref != id:
					self.menu[ref].active = False

	def match(self, details):

		dob = db_date(details[3])

		# id match
		if details[0] not in [None, '', 'ID']:
			match0 = self.db.cursor.execute('''SELECT * FROM individuals
							WHERE id="%s"''' % details[0]).fetchall()
		else:
	
			match0 = [[]]

		# fn + sn
		if details[1] != '' and details[2] != '':

			match1 = self.db.cursor.execute('''SELECT * FROM individuals
							WHERE firstname="%s" 
							AND surname="%s"''' % (details[1], 
										details[2])).fetchall()
		else:
		
			match1 = []

		# sn + dob + gender
		if details[2] != '' and dob != None and details[4] != 'Select':
			match2 = self.db.cursor.execute('''SELECT * FROM individuals
							WHERE surname="%s"
							AND dob="%s"
							AND gender="%s"''' % (details[2], dob,
													details[4])).fetchall()
		else:
	
			match2 = []

		matches = []

		for mat in [match0, match1, match2]:
			if mat in [[], None]:
				continue
			for m in mat:
				matches.append(m)

		if len(matches) == 0:

				self.build_no_match(details)

		elif matches == [[]]:

				self.build_no_match(details)

		elif self.style == 'Add':

			message = LabelWrap(size_hint_y=None, height=70,
			text="The following competitors already match some or all of those details (unmatched are in red).\nPress 'Add Competitor' if you still wish to add this new competitor")
			self.results.add_widget(message)

			heading = BoxLayout(size_hint_y=None, height=40)
			header = ['ID', 'First Name', 'Surname', 'Date of Birth',
						'Gender', '%s' % self.type]
			if self.age != 'Date of Birth':
				header.append(self.age)

			for h in header:
				heading.add_widget(Label(text=h))
			self.results.add_widget(heading)

			scroll = ScrollView(size_hint_y=None, height=80)
			self.results.add_widget(scroll)

			display = GridLayout(cols=1, size_hint_y=None, height=0)
			scroll.add_widget(display)

			id_check = []
			for m in matches:
				if len(m) == 0:
					continue				
				if m[0] not in id_check:
					box = self.build_match_box(m, details)
					display.add_widget(box)
					display.height += 30
					id_check.append(m[0])

			self.build_add_action(details)

		elif self.style == 'Remove':

			message = LabelWrap(size_hint_y=None, height=70,
			text="The following competitors match some or all of those details (unmatched are in red).\nSelect the one to remove and press 'Remove Competitor' to do so")
			self.results.add_widget(message)

			heading = BoxLayout(size_hint_y=None, height=40)
			header = ['ID', 'First Name', 'Surname', 'Date of Birth',
						'Gender', '%s' % self.type]
			if self.age != 'Date of Birth':
				header.append(self.age)
			header.append('Select')
			for h in header:
				heading.add_widget(Label(text=h))
			self.results.add_widget(heading)

			scroll = ScrollView(size_hint_y=None, height=80)
			self.results.add_widget(scroll)

			display = GridLayout(cols=1, size_hint_y=None, height=0)
			scroll.add_widget(display)

			id_check = []
			for m in matches:
				if len(m) == 0:
					continue
				if m[0] not in id_check:
					if m[0] in match0[0]:
						box = self.build_match_box(m, details, id=True)
					else:
						box = self.build_match_box(m, details)
					display.add_widget(box)
					display.height += 30

			self.build_add_action(details)

		elif self.style == 'Entry':

			message = LabelWrap(size_hint_y=None, height=70,
			text="The following competitors match some or all of those details (unmatched are in red).\nSelect the one to add to the event and press 'Add Entry'")
			self.results.add_widget(message)

			heading = BoxLayout(size_hint_y=None, height=40)
			header = ['ID', 'First Name', 'Surname', 'Date of Birth',
						'Gender', '%s' % self.type]
			if self.age != 'Date of Birth':
				header.append(self.age)
			header.append('Select')
			for h in header:
				heading.add_widget(Label(text=h))
			self.results.add_widget(heading)

			scroll = ScrollView(size_hint_y=None, height=80)
			self.results.add_widget(scroll)

			display = GridLayout(cols=1, size_hint_y=None, height=0)
			scroll.add_widget(display)

			id_check = []
			for m in matches:
				if len(m) == 0:
					continue
				if m[0] not in id_check:
					if m[0] in match0[0]:
						box = self.build_match_box(m, details, id=True)
					else:
						box = self.build_match_box(m, details)
					display.add_widget(box)
					display.height += 30

			self.build_add_action(details)

	### BUILd ###

	def build_title(self):

		if self.style == 'Add':
			title = 'Add Competitor'
		elif self.style == 'Remove':
			title = 'Remove Competitor'
		elif self.style == 'Entry':
			title = 'Add Entry - %s' % self.event_round

		t = Label(text=title, font_size=20, size_hint_y=None, height=40)
		self.content.add_widget(t)

	def build_search(self):

		col = 4
		if self.style == 'Add': # no id
			col = 3
		elif self.age == 'Date of Birth':
			col = 3

		grid = GridLayout(cols=col, size_hint_y=None, height=120)

		# id
		if self.style in ['Entry', 'Remove']:
			box1 = BoxLayout(orientation='vertical')
			box1.add_widget(Label(text='ID Number'))
			self.id = TextInputC(input_filter='int')
			box1.add_widget(self.id)
			grid.add_widget(box1)

		# fn
		box2 = BoxLayout(orientation='vertical')
		box2.add_widget(Label(text='First Name'))
		self.fn = TextInputC()
		box2.add_widget(self.fn)
		grid.add_widget(box2)

		# sn
		box3 = BoxLayout(orientation='vertical')
		box3.add_widget(Label(text='Surname'))
		self.sn = TextInputC()
		box3.add_widget(self.sn)
		grid.add_widget(box3)

		# dob
		box4 = BoxLayout(orientation='vertical')
		box4.add_widget(Label(text='Date of Birth'))
		self.dob = DatePicker()
		box4.add_widget(self.dob)
		grid.add_widget(box4)

		# blank
		if col == 4:
			grid.add_widget(BoxLayout())

		# gender
		box5 = BoxLayout(orientation='vertical')
		box5.add_widget(Label(text='Gender'))

		glist = self.db.cursor.execute('''SELECT title, filter 
									FROM template_groups
									WHERE category="gender"''').fetchall()
		gender_list = []
		for g in glist:
			gender_list.append(g[1])
		gender_list = sorted(gender_list)

		self.gender_choice = OptionMenu(options=gender_list)
		box5.add_widget(self.gender_choice)
		grid.add_widget(box5)

		# type
		box6 = BoxLayout(orientation='vertical')
		box6.add_widget(Label(text=self.type))
		self.type_choice = OptionMenu(options=self.type_list)
		box6.add_widget(self.type_choice)
		grid.add_widget(box6)

		# age
		if self.age != 'Date of Birth':
			box7 = BoxLayout(orientation='vertical')
			box7.add_widget(Label(text=self.age))
			self.age_choice = OptionMenu(options=self.age_list)
			box7.add_widget(self.age_choice)
			grid.add_widget(box7)

		self.content.add_widget(grid)

		## action

		action = BoxLayout(size_hint_y=None, height=40)

		btn1 = Button(text='Close', font_size=20)
		btn1.bind(on_press=lambda i :self.pop.dismiss())
		action.add_widget(btn1)

		btn2 = Button(text='Clear', font_size=20)
		btn2.bind(on_press=self.command_clear)
		action.add_widget(btn2)

		btn3 = Button(text='Search', font_size=20)
		btn3.bind(on_press=self.command_search)
		action.add_widget(btn3)

		self.content.add_widget(action)

	def build_results(self):

		self.results = GridLayout(cols=1)
		self.content.add_widget(self.results)

	def build_no_match(self, details):

		if self.style == 'Add':

			box = LabelWrap(size_hint_y=None, height=70,
		text="No matches found.\nPress 'Add Competitor' to add their details")

			self.results.add_widget(box)
			self.build_add_action(details)

		elif self.style == 'Remove':

			box = LabelWrap(size_hint_y=None, height=70,
					text="No matches found\nCheck search details")

			self.results.add_widget(box)

		elif self.style == 'Entry':

			message = "No matches found. Check search details or go to the Competitor page and entr a new competitor"

			box = LabelWrap(size_hint_y=None, height=70,
						text=message)

			self.results.add_widget(box)
			#

	def build_add_action(self, details):

		box = BoxLayout()

		action = BoxLayout(size_hint_y=None, height=40)
	
		btn1 = Button(text='Close', font_size=20)
		btn1.bind(on_press=lambda i :self.pop.dismiss())
		action.add_widget(btn1)

		btn2 = Button(text='Clear', font_size=20)
		btn2.bind(on_press=self.command_clear)
		action.add_widget(btn2)
	
		if self.style == 'Add':

			btn = Button(text='Add Competitor', font_size=20)
			btn.bind(on_press=lambda i, details=details: self.command_add(details))
			action.add_widget(btn)

		elif self.style == 'Remove':

			btn = Button(text='Remove Competitor', font_size=20)
			btn.bind(on_press=self.command_remove)
			action.add_widget(btn)

		elif self.style == 'Entry':

			btn = Button(text='Add Entry', font_size=20)
			btn.bind(on_press=lambda i:self.command_add_entry(None))
			action.add_widget(btn)

		box.add_widget(action)
		self.results.add_widget(box)

	def build_match_box(self, tup, details, **kwargs):

		first_match = kwargs.pop('id', False)

		box = BoxLayout(size_hint_y=None, height=30)

		if self.style == 'Add':

			count = len(details)
			if self.age == 'Date of Birth':
				count -= 1

			for i in range(count):

				if i == 3:
					if db_date(tup[3], db_in=False) != details[3]:
						box.add_widget(LabelC(text=tup[3], bg=[1, 0, 0, 1]))
					else:
						box.add_widget(Label(text=tup[3]))

				elif tup[i] != details[i]:
					box.add_widget(LabelC(text=tup[i], bg=[1, 0, 0, 1]))
				else:
					box.add_widget(Label(text=tup[i]))		

			return box			

		if self.style in ['Remove', 'Entry']:

			count = len(details)
			if self.age == 'Date of Birth':
				count -= 1

			for i in range(count):

				if i == 3:
					if db_date(tup[3], db_in=False) != details[3]:
						box.add_widget(LabelC(text=tup[3], bg=[1, 0, 0, 1]))
					else:
						box.add_widget(Label(text=tup[3]))

				elif tup[i] != details[i]:
					box.add_widget(LabelC(text=tup[i], bg=[1, 0, 0, 1]))
				else:
					box.add_widget(Label(text=tup[i]))		

			id = tup[0]
			if first_match == True:
				chk = CheckBoxA(active= True)
			else:
				chk = CheckBoxA(active=False)
			chk.bind(active=lambda obj,
				value, id=id:self.command_select(obj, value, id))
			box.add_widget(chk)
			self.menu[id] = chk

			return box				

class EntryAddTeam():

	def __init__(self, *args, **kwargs):
		
		## Settings
		self.tournament = kwargs.pop('tournament', '')
		self.typename = kwargs.pop('typename', None)
		self.event = kwargs.pop('event', None)
		self.event_round = kwargs.pop('event_round', None)
		self.event_table = kwargs.pop('event_table', None)
		self.widget = kwargs.pop('widget', None)

		self.db = Database(self.tournament)
		details = self.db.cursor.execute('''SELECT *
										FROM details''').fetchone()
		self.type = details[3]
		self.type_list = []
		type_list = self.db.cursor.execute('''SELECT title 
								FROM template_groups
								WHERE category="type"''').fetchall()
		for tup in type_list:
			
			self.type_list.append(tup[0])
		self.type_colours = {}
		type_colours = self.db.cursor.execute('''SELECT title, 
								colour_bg, colour_text
								FROM template_groups
								WHERE category="type"''').fetchall()
		for tup in type_colours:
			self.type_colours[tup[0]] = [db_colour(tup[1], db_in=False),
									db_colour(tup[2], db_in=False)]
		events = self.db.cursor.execute('''SELECT grade, contest
								FROM template_events
								WHERE title="%s"''' % self.event).fetchone()							
		self.grade = events[0]
		self.contest = events[1]
		self.contest_type = calculate_contest_type(self.tournament, 
													self.contest)
		self.age = details[4]
		self.age_list = []
		if self.age != 'Date of Birth':
			age_list = self.db.cursor.execute('''SELECT title
								FROM template_groups
								WHERE category="age"''').fetchall()
			for tup in age_list:
				self.age_list.append(tup[0])

		## Database
		self.menu = {}

		## Display
		self.build_content()
		self.build_name()
		self.build_action()

		## Engine
		self.pop = Popup(title=self.event_round, content=self.content, 
							auto_dismiss=False,
							size_hint=(None, None), size=(800, 500))
		self.pop.open()

	### FUNCTIONS ###

	def command_add_team(self, i):

		if self.type_choice.text == 'Select':
			return

		name = self.example.text
		type = self.type_choice.text

		# check unique name 
		check = self.db.cursor.execute('''SELECT * FROM %s
								WHERE team="%s"''' % (self.event_table,
														name)).fetchone()
		if check != None:
			PopBox().showwarning(title='Error',
				message="%s is already the name of a team in this round, choose a unique team name" % name)
			return

		self.db.cursor.execute('''INSERT INTO %s
					VALUES (?, 0, 0, '', '', '', ?)''' % self.event_table, [name, type])

		self.db.connection.commit()
		self.widget.build_display_structure(self.event_round)
		self.pop.dismiss()

	def command_adjust(self, value):

		if self.type_choice.text == 'Select':
			self.adjust.text = ''
			return

		self.example.text = '%s %s' % (self.type_choice.text, self.adjust.text)

	def command_type_choice(self, obj, value):

		self.adjust.text = ''

		bg = self.type_colours[value][0]
		txt = self.type_colours[value][1]

		self.example.text = value
		self.example.bg = bg
		self.example.color = txt

	### BUILD ###

	def build_content(self):
		
		self.content = GridLayout(cols=1, spacing=5)
		self.content.add_widget(Label(text='Add Team', font_size=20, 
			size_hint_y=None, height=40))

	def build_name(self):

		box = GridLayout(cols=3, size_hint_y=None, height=80)

		box.add_widget(Label(text='Choose %s' % self.type))
		box.add_widget(Label(text='Adjust Team Name'))
		box.add_widget(Label(text='Team Name'))

		self.type_choice = OptionMenu(options=self.type_list)
		self.type_choice.bind(text=self.command_type_choice)
		box.add_widget(self.type_choice)

		self.adjust = TextInputC(function=self.command_adjust)
		box.add_widget(self.adjust)

		self.example = LabelC()
		box.add_widget(self.example)

		self.content.add_widget(box)

	def build_action(self):

		box = BoxLayout(size_hint_y=None, height=40)

		btn1 = Button(text='Close', font_size=20)
		btn1.bind(on_press=lambda i:self.pop.dismiss())
		box.add_widget(btn1)

		btn2 = Button(text='Add Team', font_size=20)
		btn2.bind(on_press=self.command_add_team)
		box.add_widget(btn2)

		self.content.add_widget(box)

class TryExcept():
	'''
	Just an example code to get try/except info
	'''
	try:
		pass
	except Exception as inst:
		print type(inst)
		print inst.args
		print inst

#-----------------------------------------------------------------------------#

### APP CLASS ###

class MainApp(App):

	version = StringProperty('Login') #
	title = 'ZO-SPORTS'

	def __init__(self, *args, **kwargs):
		super(MainApp, self).__init__(**kwargs)

		## Settings
		Window.size = (1000, 600)
		self.title = 'ZO-SPORTS'

	def build(self):
		
		global zo
		zo = MainWindow()
		zologin = ProgramLogin(style=self.version)
		return zo #

#-----------------------------------------------------------------------------#

### BASE CLASSES ###

class Label90(Label):

	pass

class LabelB(Label):

	pass

class LabelC(Label):

	bg = ListProperty([0.5, 0.5, 0.5, 1])

class LabelTitle(Label):

	height_max = NumericProperty(70)

class LabelWrap(Label):

	pass

class LabelWrapC(LabelC, LabelWrap):

	pass

class Tab(Button):

	width_min = NumericProperty(100)
	font_size = 16

	def __init__(self, *args, **kwargs):
		super(Tab, self).__init__(**kwargs)

		self.size_hint_x = None
		self.width = self.width_min

	def resize(self, obj, value):
		if self.texture_size[0] > self.width - 10:
			self.size[0] = self.texture_size[0] + 20

class TextInputC(TextInput):
	'''
	This Widget auto halign & valign to center.
	self.run_status means any change in text i.e. add or backspace
	will trigger a self.status = False. This can be linked so the program
	will recognise changes in order to save them

	'''
	status = BooleanProperty(True)
	run_status = BooleanProperty(False)
	function = ObjectProperty(None)

	def __init__(self, *args, **kwargs):
		super(TextInputC, self).__init__(**kwargs)

		
	### FUNCTIONS ###

	def update_padding(self, *args):

		text_width = self._get_text_width(
			self.text, self.tab_width, self._label_cached
			)
		self.padding_x = (self.width - text_width)/2

		if self.run_status == True:
			self.insert_text = self.insert_text_status
			self.do_backspace = self.do_backspace_status

		if self.function != None:
			self.function(self.text)

	def update_padding_size(self, *args):

		text_width = self._get_text_width(
			self.text, self.tab_width, self._label_cached
			)
		self.padding_x = (self.width - text_width)/2

		if self.run_status == True:
			self.insert_text = self.insert_text_status
			self.do_backspace = self.do_backspace_status

	def insert_text(self, substring, from_undo=False):

		s = substring
		if s in ['-', '#', '%']:
			return

		return super(TextInputC, self).insert_text(s, 
													from_undo=from_undo)

	def insert_text_status(self, substring, from_undo=False):
		if self.status != False:
			self.status = False
		s = substring

		# try for TextInput24
		try: 
			if self.chronos == True:
				pass
			if len(self.text) == 0:
				if int(s) <= int(self.chronos[0]):
					return super(TextInput24, self).insert_text(s, 
															from_undo=from_undo)				
			elif len(self.text) == 1:
				if int(self.chronos[1]) == 0 or int(self.text) == 1:
					return super(TextInput24, self).insert_text(s, 
															from_undo=from_undo)
				elif int(s) <= int(self.chronos[1]):
					return super(TextInput24, self).insert_text(s, 
															from_undo=from_undo)
				elif self.text == '0': # 09:30
					return super(TextInput24, self).insert_text(s, 
															from_undo=from_undo)
		except AttributeError:
			return super(TextInputC, self).insert_text(s, 
													from_undo=from_undo)

	def do_backspace_status(self, from_undo=False, mode='bkspc'):
		if self.status != False:
			self.status = False
		try:
			return super(TextInputC, self).do_backspace(from_undo=from_undo,
													mode='bkspc')
		except:
			return

class TextInputE(TextInput):
	'''
	This Widget auto halign & valign to center.
	self.run_status means any change in text i.e. add or backspace
	will trigger a self.status = False. This can be linked so the program
	will recognise changes in order to save them

	'''
	status = BooleanProperty(True)
	run_status = BooleanProperty(False) 

	def __init__(self, *args, **kwargs):
		super(TextInputE, self).__init__(**kwargs)

		Clock.schedule_once(self.update_padding, 0)
		
	### FUNCTIONS ###

	def update_padding(self, *args):

		text_width = self._get_text_width(
			self.text, self.tab_width, self._label_cached
			)
		self.padding_x = (self.width - text_width)/2

		if self.run_status == True:
			self.insert_text = self.insert_text_status
			self.do_backspace = self.do_backspace_status

	def update_padding_size(self, *args):

		text_width = self._get_text_width(
			self.text, self.tab_width, self._label_cached
			)
		self.padding_x = (self.width - text_width)/2

		if self.run_status == True:
			self.insert_text = self.insert_text_status
			self.do_backspace = self.do_backspace_status

	def insert_text_status(self, substring, from_undo=False):
		if self.status != False:
			self.status = False
		s = substring

		# try for TextInput24
		try: 
			if self.chronos == True:
				pass
			if len(self.text) == 0:
				if int(s) <= int(self.chronos[0]):
					return super(TextInput24, self).insert_text(s, 
															from_undo=from_undo)				
			elif len(self.text) == 1:
				if int(self.chronos[1]) == 0:
					return super(TextInput24, self).insert_text(s, 
															from_undo=from_undo)
				elif int(s) <= int(self.chronos[1]):
					return super(TextInput24, self).insert_text(s, 
															from_undo=from_undo)
				elif self.text == '0': # 09:30
					return super(TextInput24, self).insert_text(s, 
															from_undo=from_undo)
		except AttributeError:
			return super(TextInputE, self).insert_text(s, 
													from_undo=from_undo)

	def do_backspace_status(self, from_undo=False, mode='bkspc'):
		if self.status != False:
			self.status = False
		try:
			return super(TextInputE, self).do_backspace(from_undo=from_undo,
													mode='bkspc')
		except:
			return

class TextInputL(TextInput):
	'''
	Left align, height padded, status effects on
	'''
	status = BooleanProperty(True)
	run_status = BooleanProperty(False) 

	def __init__(self, *args, **kwargs):
		super(TextInputL, self).__init__(**kwargs)

		self.insert_text = self.insert_text_status
		self.do_backspace = self.do_backspace_status

	def insert_text_status(self, substring, from_undo=False):
		if self.status != False:
			self.status = False
		s = substring
		return super(TextInputL, self).insert_text(s, 
													from_undo=from_undo)

	def do_backspace_status(self, from_undo=False, mode='bkspc'):
		if self.status != False:
			self.status = False
		return super(TextInputL, self).do_backspace(from_undo=from_undo,
													mode='bkspc')

class TextInput24(TextInputC):

	chronos = StringProperty('24')

class FileIconView(FileChooserIconView):

	function = ObjectProperty(None)

	def on_selection(self, *args, **kwargs):

		self.function(self.selection)

#-----------------------------------------------------------------------------#

### DISPLAY CLASSES ### 

class FeedBackForm(GridLayout):

	cols = 1

	def __init__(self, *args, **kwargs):
		super(FeedBackForm, self).__init__(**kwargs)

		## Settings
		self.function = kwargs.pop('function', None)
		self.tournament = kwargs.pop('tournament', None)

		## Display
		self.build_instructions()
		self.build_input()
		self.build_action()

		## Engine
		self.pop = Popup(title='Feedback Form', content=self, auto_dismiss=False,
							size_hint=(None, None), size=(500, 500))
		self.pop.open()

	### FUNCTIONS ###

	def command_action(self, i):

		self.pop.dismiss()

		if i.text == 'No Feedback':

			self.function(False)

		elif i.text == 'Send':

			DataEmail(style='feedback', tournament=self.tournament,
						feedback=self.input.text)
			self.function(True)

	### BUILD ###

	def build_instructions(self):

		instructions = LabelWrap(text="We hope that you are finding ZO-SPORTS everything that you need to run this tournament. We are still very much in a developmental stage and to that end any and all feedback you can provide us will guide future versions. Please let us know what you think so far by typing a message and clicking 'Send', your feedback will be greatly appreciated",
			size_hint_y=None, height=150)
		self.add_widget(instructions)

	def build_input(self):

		self.input = TextInputC(size_hint_y=None, height=160)
		self.add_widget(self.input)

	def build_action(self):

		box = BoxLayout(size_hint_y=None, height=40)
		self.add_widget(box)

		btn1 = Button(text='No Feedback')
		btn1.bind(on_press=self.command_action)
		box.add_widget(btn1)

		btn2 = Button(text='Send')
		btn2.bind(on_press=self.command_action)
		box.add_widget(btn2)

class FrameAddMinus(GridLayout):

	style = StringProperty('') # type/gender/age/contest
	tournament = StringProperty('')

	def __init__(self, *args, **kwargs):
		super(FrameAddMinus, self).__init__(**kwargs)

		## Settings
		self.unsave = kwargs.pop('unsave', False)
		self.db = Database(self.tournament)
		self.scroll_var = 80

		## Database
		self.menu = []
		self.save_menu = {}
		self.unsaved_changes = []

		## Engine
		if self.style == 'type':
			self.build_type()
		elif self.style == 'gender':
			self.build_gender()
		elif self.style == 'age dob':
			self.build_age_dob()
		elif self.style == 'age':
			self.build_age()
		elif self.style == 'contest':
			self.build_contest()
		elif self.style == 'points':
			self.build_points()
		elif self.style == 'restrictions':
			self.build_restrictions()
		elif self.style == 'special':
			self.build_special()

	### FUNCTIONS ###

	def change_scroll_height(self, obj, value):

		self.ids.scroll.height = self.height - self.scroll_var

	def command_action(self, i):

		if self.style == 'type':

			if i.text == '-':
				if len(self.menu) == 2:
					return
				ref = self.menu[-1]
				self.ids.display.remove_widget(ref[0])
				self.ids.display.height -= 70
				self.menu.remove(ref)
			elif i.text == '+':
				self.build_type_box()

		elif self.style == 'gender':
			if i.text == '-':
				if len(self.menu) == 1:
					return
				ref = self.menu[-1]
				self.ids.display.remove_widget(ref[0])
				self.ids.display.height -= 70
				self.menu.remove(ref)
			elif i.text == '+':
				if len(self.menu) == 2:
					return
				self.build_gender_box()

		elif self.style == 'age dob':
			if i.text == '-':
				if len(self.menu) == 1:
					return
				ref = self.menu[-1]
				self.ids.display.remove_widget(ref[0])
				self.ids.display.height -= 70
				self.menu.remove(ref)
			elif i.text == '+':
				if len(self.menu) > 1:
					current = self.menu[-1][4].date
					self.build_age_dob_box(current=current)
				else:
					self.build_age_dob_box()			

		elif self.style == 'age':
			if i.text == '-':
				if len(self.menu) == 1:
					return
				ref = self.menu[-1]
				self.ids.display.remove_widget(ref[0])
				self.ids.display.height -= 70
				self.menu.remove(ref)
			elif i.text == '+':
				self.build_age_box()	

		elif self.style == 'contest':
			if i.text == '-':
				if len(self.menu) == 0:
					return
				ref = self.menu[-1]
				self.ids.display.remove_widget(ref[0])
				self.ids.display.height -= 70
				self.menu.remove(ref)
			elif i.text == '+':
				self.build_contest_box()

		elif self.style == 'points':
			if i.text == '-':
				if len(self.menu) == 4: # pp and 3 min
					return
				ref = self.menu[-1]
				self.ids.display.remove_widget(ref[0])
				self.ids.display.height -= 70
				self.menu.remove(ref)
				try:
					self.unsave(False)
				except:
					pass
			elif i.text == '+':
				#28/01/2017 removed limit
				#if len(self.menu) == 11: # pp and 10 max
				#	return
				rank = len(self.menu)
				self.build_points_box(rank=rank)

		elif self.style == 'special':
			if i.text == '-':
				if len(self.menu) == 0: 
					return
				ref = self.menu[-1]
				self.ids.display.remove_widget(ref[0])
				self.ids.display.height -= 70
				self.menu.remove(ref)
			elif i.text == '+':
				if len(self.menu) == 11: # pp and 10 max
					return
				self.build_special_box(None, None, None)

	def command_save(self, value, what):

		if what == 'special':
			if self.unsave != False:
				self.unsave(False)

		if value == False:
			if what not in self.unsaved_changes:
				self.unsaved_changes.append(what)

			if self.unsave != False:
				self.unsave(False)

	### BUILD ###

	def build_type(self):

		# header
		self.ids.header.add_widget(Label(text='Title', font_size=20))
		self.ids.header.add_widget(Label(text='Colour', font_size=20))

		# database
		c = '''SELECT title, colour_bg, colour_text 
				FROM template_groups
				WHERE category="type";'''		
		rows = self.db.cursor.execute(c).fetchall()	

		# display
		if rows == []: # First time
			self.build_type_box()
			self.build_type_box()
		else:
			for row in rows:
				title = row[0]
				bg = row[1]
				txt = row[2]
				self.build_type_box(title=title, bg=bg, txt=txt)

	def build_type_box(self, **kwargs):

		title = kwargs.pop('title', '')
		bg = kwargs.pop('bg', '0.5#0.5#0.5#1')
		txt = kwargs.pop('txt', '1#1#1#1')
		bg = db_colour(bg, db_in=False)
		txt = db_colour(txt, db_in=False)

		box = BoxLayout(size_hint_y=None, height=70)
		t = TextInputC(text=title, font_size=20, padding=[20, 22, 0, 22],
						tab_width=0, multiline=False, write_tab=False)
		box.add_widget(t)
		cp = ColourPicker(bg=bg, txt=txt)
		box.add_widget(cp)

		ref = []
		ref.append(box)
		ref.append(t)
		ref.append(cp)

		# engine
		self.menu.append(ref)
		self.ids.display.add_widget(box)
		self.ids.display.height += 70

	def build_gender(self):

		# header
		self.ids.header.add_widget(Label(text='Title', font_size=20))
		self.ids.header.add_widget(Label(text='Gender', font_size=20))

		# database
		c = '''SELECT title, filter
				FROM template_groups
				WHERE category="gender";'''		
		rows = self.db.cursor.execute(c).fetchall()	
		
		# display
		if rows == []: # First time
			self.build_gender_box()
		else:
			for row in rows:
				title = row[0]
				filter = row[1]
				self.build_gender_box(title=title, filter=filter)

	def build_gender_box(self, **kwargs):

		title = kwargs.pop('title', '')
		filter = kwargs.pop('filter', 'Select')

		box = BoxLayout(size_hint_y=None, height=70)
		t = TextInputC(text=title, font_size=20, padding=[20, 22, 0, 22],
						tab_width=0, multiline=False, write_tab=False)
		box.add_widget(t)
		f = OptionMenu(options=['Male', 'Female'], main=filter)
		box.add_widget(f)

		ref = []
		ref.append(box)
		ref.append(t)
		ref.append(f)

		# engine
		self.menu.append(ref)
		self.ids.display.add_widget(box)
		self.ids.display.height += 70

	def build_age_dob(self):

		# header
		self.ids.header.add_widget(Label(text='Title', font_size=20))
		self.ids.header.add_widget(Label(text='Under/Over', font_size=20))
		self.ids.header.add_widget(Label(text='this Age', font_size=20))
		self.ids.header.add_widget(Label(text='as of this Date', font_size=20))

		# database
		c = '''SELECT title, under, year, date
				FROM template_groups
				WHERE category="age";'''		
		rows = self.db.cursor.execute(c).fetchall()	

		# display
		if rows == []:
			self.build_age_dob_box(title='Open')

		for row in rows:
			title = row[0]
			under = row[1]
			age = str(row[2]) # saved as int()
			date = db_date(row[3], db_in=False)
			self.build_age_dob_box(title=title, under=under, 
									age=age, date=date)

	def build_age_dob_box(self, **kwargs):

		title = kwargs.pop('title', '')
		under = kwargs.pop('under', 'Select')
		age = kwargs.pop('age', '')
		date = kwargs.pop('date', None)
		current = kwargs.pop('current', None)

		box = BoxLayout(size_hint_y=None, height=70)

		if title == 'Open':
			ref = ['Open']
			ot = TextInputC(text='Open', font_size=20, multiline=False,
							readonly=True)
			ut = Button(text='N/A')
			at = TextInputC(text='N/A', font_size=20, multiline=False,
							readonly=True)
			dt = Label(text='N/A')
			for widget in [ot, ut, at, dt]:
				box.add_widget(widget)

			self.menu.append(ref)
			self.ids.display.add_widget(box)
			self.ids.display.height += 70
			return

		t = TextInputC(text=title, font_size=20, multiline=False,)
		box.add_widget(t)
		u = OptionMenu(options=['Under', 'Over'], main=under)
		box.add_widget(u)
		a = TextInputC(text=age, font_size=30, multiline=False,
							input_filter='int')
		box.add_widget(a)
		if current == None:
			d = DatePicker(date=date)
		else:
			d = DatePicker(date=current)
		box.add_widget(d)

		ref = []
		ref.append(box)
		ref.append(t)
		ref.append(u)
		ref.append(a)
		ref.append(d)

		# engine
		self.menu.append(ref)
		self.ids.display.add_widget(box)
		self.ids.display.height += 70
		
	def build_age(self):

		# header
		self.ids.header.add_widget(Label(text='Title', font_size=20))

		# database
		c = '''SELECT title
				FROM template_groups
				WHERE category="age";'''		
		rows = self.db.cursor.execute(c).fetchall()	
		
		# display
		for row in rows:
			title = row[0]
			self.build_age_box(title=title)

	def build_age_box(self, **kwargs):

		title = kwargs.pop('title', '')

		box = BoxLayout(size_hint_y=None, height=70)

		if title == 'Open':
			t = TextInputC(text=title, font_size=20, multiline=False,
							readonly=True)
		else:
			t = TextInputC(text=title, font_size=20, multiline=False)
		box.add_widget(t)


		ref = []
		ref.append(box)
		ref.append(t)

		# engine
		self.menu.append(ref)
		self.ids.display.add_widget(box)
		self.ids.display.height += 70

	def build_contest(self):
		
		# header
		self.ids.header.add_widget(Label(text='Contest', font_size=20))
		self.ids.header.add_widget(Label(text='Measured by', font_size=20))
		self.ids.header.add_widget(Label(text='Competition', font_size=20))
		self.ids.header.add_widget(Label(text='Select', font_size=20))

		# database
		c = '''SELECT title, measure, team, selection 
				FROM template_contests
				WHERE standard=0;'''		
		rows = self.db.cursor.execute(c).fetchall()	

		# display
		for row in rows:
			title = row[0]
			measure = row[1].title()
			if db_boole(row[2], db_in=False) == True:
				comp = 'Team'
			else:
				comp = 'Individual'
			select = db_boole(row[3], db_in=False)

			self.build_contest_box(title=title, measure=measure, 
									comp=comp, select=select)

	def build_contest_box(self, **kwargs):

		title = kwargs.pop('title', '')
		measure = kwargs.pop('measure', 'Select')
		comp = kwargs.pop('comp', 'Select')
		select = kwargs.pop('select', True)

		box = BoxLayout(size_hint_y=None, height=70)

		t = TextInputC(text=title, font_size=20, multiline=False)
		box.add_widget(t)

		m = OptionMenu(options=['Time', 'Distance'], main=measure)
		box.add_widget(m)

		c = OptionMenu(options=['Individual', 'Team'], main=comp)
		box.add_widget(c)

		s = CheckBoxA(active=select)
		box.add_widget(s)

		ref = [box, t, m, c, s]

		# engine
		self.menu.append(ref)
		self.ids.display.add_widget(box)
		self.ids.display.height += 70

	def build_points(self):
		
		# header
		self.ids.header.add_widget(Label(text='Rank', font_size=20))
		self.ids.header.add_widget(Label(text='Individual Events',font_size=20))
		self.ids.header.add_widget(Label(text='Team Events', font_size=20))

		# database
		c = '''SELECT rank, individual, team
				FROM template_points'''	
		rows = self.db.cursor.execute(c).fetchall()	

		# display
		for row in rows:
			rank = row[0]
			individual = row[1]
			team = row[2]

			self.build_points_box(rank=rank, individual=individual,
									team=team)

	def build_points_box(self, **kwargs):

		rank = kwargs.pop('rank', None)
		if rank in ['Participation', 'participation']:
			rank = rank.title()
			rk = rank
		else:
			rank = int(rank)
			rk = calculate_ordinal(rank)
		indiv = kwargs.pop('individual', '')
		team = kwargs.pop('team', '')

		box = BoxLayout(size_hint_y=None, height=70)
		ref = [box, rank]

		r = Label(text=rk, font_size=20)
		box.add_widget(r)
		ind = TextInputC(text=str(indiv), font_size=20, run_status=True, 
						multiline=False, input_filter='int')
		tm = TextInputC(text=str(team), font_size=20, run_status=True,
						 multiline=False, input_filter='int')

		for widget in [ind, tm]:
			box.add_widget(widget)
			ref.append(widget)
			widget.bind(status=lambda obj, value, 
					rank=rank:self.command_save(value, rank))

		self.save_menu[rank] = [ref[2], ref[3]]

		# engine
		self.menu.append(ref)
		self.ids.display.add_widget(box)
		self.ids.display.height += 70

	def build_restrictions(self):

		# settings
		self.remove_widget(self.ids.action)
		self.scroll_var = 40

		# header
		self.ids.header.add_widget(Label(text='Entry Minimum/Maximum', 
											font_size=20))
		self.ids.header.add_widget(Label(text='Restriction',font_size=20))

		# database
		c = '''SELECT * FROM template_restrictions'''	
		rows = self.db.cursor.execute(c).fetchall()	

		# display
		for row in rows:
			entry = row[0]
			r = row[1]

			ref = []
			box = BoxLayout(size_hint_y=None, height=70)

			l1 = Label(text=entry.title(), font_size=20)
			box.add_widget(l1)
			ref.append(l1)

			t1 = TextInputC(text=str(r), font_size=20, run_status=True,
							multiline=False, input_filter='int')
			t1.bind(status=lambda obj, value, 
					entry=entry:self.command_save(value, entry))

			self.save_menu[entry] = [t1]
			box.add_widget(t1)
			ref.append(t1)

			# engine
			self.menu.append(ref)
			self.ids.display.add_widget(box)
			self.ids.display.height += 70			

	def build_special(self):
		
		# settings
		details = self.db.cursor.execute("SELECT * FROM details").fetchone()
		self.type = details[3]

		# header
		self.ids.header.add_widget(Label(text='Title', font_size=20))
		self.ids.header.add_widget(Label(text=self.type, font_size=20))
		self.ids.header.add_widget(Label(text='Points', font_size=20))
		self.ids.header.add_widget(Label(text='Select', font_size=20))

		# database

		type_list = self.db.cursor.execute('''SELECT title
							FROM template_groups
							WHERE category="type"''').fetchall()
		self.type_list = []
		for t in type_list:
			self.type_list.append(t[0])

		try:
			c = '''SELECT title, type, points
				FROM special_events'''
			rows = self.db.cursor.execute(c).fetchall()					
		except:
			self.db.cursor.execute('''CREATE TABLE special_events
						(title TEXT PRIMARY KEY, type TEXT, points REAL)''')
			self.db.connection.commit()
			rows = []

		# display
		for row in rows:
			self.build_special_box(row[0], row[1], row[2])

	def build_special_box(self, title, type, points):

		box = BoxLayout(size_hint_y=None, height=70)
		ref = [box]

		if title == None:
			title = ''
		t = TextInputC(text=title, multiline=True, run_status=True)
		box.add_widget(t)
		ref.append(t)

		if type == None:
			o = OptionMenu(options=sorted(self.type_list))
		else:
			o = OptionMenu(options=sorted(self.type_list), main=type)
		box.add_widget(o)
		ref.append(o)

		if points == None:
			points = ''
		else:
			if points.is_integer():
				points = str(int(points))
			else:
				points = str(points)
		p = TextInputC(text=points, input_filter='float', run_status=True)
		box.add_widget(p)
		ref.append(p)

		s = CheckBoxA(active=True)
		box.add_widget(s)
		ref.append(s)
		s.bind(active=lambda obj, value:self.command_save(value, 'special'))

		o.bind(text=lambda obj, value:self.command_save(value, 'special'))
		for widget in [t, p]:
			widget.bind(status=lambda obj, value:self.command_save(value, 
																	'special'))

		# engine
		self.menu.append(ref)
		self.ids.display.add_widget(box)
		self.ids.display.height += 70

	### PROCESS ###

	def process_type(self):

		# conditions
		unique = []
		for ref in self.menu:
			if ref[1].text == '':
				return None
			if len(ref[1].text) < 3:
				return None
			if ref[1].text.title() in unique:
				return None
			for r in ref[1].text:
				if r.isalnum() == False:
					if r != ' ':
						return None
			if ref[1].text[1] in [0,1,2,3,4,5,6,7,8,9]:
				return None
			unique.append(ref[1].text.title())
		
		# delete all category="type" 
		command = "DELETE FROM template_groups WHERE category='type'"
		self.db.cursor.execute(command)
		self.db.connection.commit()

		# insert in alphabetical order & alter titles
		data = {}
		for ref in self.menu:
			data[ref[1].text] = ref
		alpha = sorted(data)
		var = []
		for t in alpha:
			ref = data[t]
			title = t.title()
			bg = db_colour(ref[2].bg)
			txt = db_colour(ref[2].txt) # same widget different data
			var.append((title, 'type', bg, txt))

		command = '''INSERT INTO template_groups
					(title, category, colour_bg, colour_text)
					VALUES (?, ?, ?, ?)'''
		self.db.cursor.executemany(command, var)
		self.db.connection.commit()

		return True

	def process_gender(self):
		
		# conditions
		unique = []
		unique1 = []
		for ref in self.menu:
			if ref[1].text == '':
				return None
			if ref[1].text.title() in unique:
				return None
			unique.append(ref[1].text.title())
			if ref[2].text == 'Select':
				return None
			if ref[2].text in unique1:
				return None
			for r in ref[1].text:
				if r.isalnum() == False:
					if r != ' ':
						return None
			if ref[1].text[1] in [0,1,2,3,4,5,6,7,8,9]:
				return None		
			unique1.append(ref[2].text)

		# delete all category="gender" 
		command = "DELETE FROM template_groups WHERE category='gender'"
		self.db.cursor.execute(command)
		self.db.connection.commit()

		# insert in alphabetical order & alter titles
		data = {}
		for ref in self.menu:
			data[ref[1].text] = ref
		alpha = sorted(data)
		var = []
		for t in alpha:
			ref = data[t]
			title = t.title()
			filter = ref[2].text
			var.append((title, 'gender', filter))

		command = '''INSERT INTO template_groups
					(title, category, filter)
					VALUES (?, ?, ?)'''
		self.db.cursor.executemany(command, var)
		self.db.connection.commit()

		return True

	def process_age_dob(self):
		
		# conditions
		unique = ['Open']
		for ref in self.menu:
			if ref[0] == 'Open':
				continue
			if ref[1].text == '':
				return None
			if ref[2].text == 'Select':
				return None
			if ref[3].text == '':
				return None
			if ref[1].text.title() in unique:
				return None
			for r in ref[1].text:
				if r.isalnum() == False:
					if r != ' ':
						return None
			if ref[1].text[1] in [0,1,2,3,4,5,6,7,8,9]:
				return None				
			unique.append(ref[1].text.title())


		# delete all category="gender" 
		command = "DELETE FROM template_groups WHERE category='age'"
		self.db.cursor.execute(command)
		self.db.connection.commit()

		# insert 
		var = []
		for ref in self.menu:

			if ref[0] == 'Open':
				command = '''INSERT INTO template_groups
					(title, category)
					VALUES ('Open', 'age')'''
				self.db.cursor.execute(command)
				self.db.connection.commit()
				continue
			title = ref[1].text.title()
			under = ref[2].text
			age = ref[3].text
			date = db_date(ref[4].date)
			var.append((title, 'age', under, age, date))

		comm = '''INSERT INTO template_groups
					(title, category, under, year, date)
					VALUES (?, ?, ?, ?, ?)'''
		self.db.cursor.executemany(comm, var)
		self.db.connection.commit()

		return True

	def process_age(self):
		
		# conditions
		unique = []
		for ref in self.menu:
			if ref[1].text == '':
				return None
			if ref[1].text in unique:
				return None
			for r in ref[1].text:
				if r.isalnum() == False:
					if r != ' ':
						return None
			if ref[1].text[1] in [0,1,2,3,4,5,6,7,8,9]:
				return None				
			unique.append(ref[1].text)

		# delete all category="age" 
		command = "DELETE FROM template_groups WHERE category='age'"
		self.db.cursor.execute(command)
		self.db.connection.commit()

		# insert 
		var = []
		for ref in self.menu:
			title = ref[1].text.title()
			var.append((title, 'age'))

		command = '''INSERT INTO template_groups
					(title, category)
					VALUES (?, ?)'''
		self.db.cursor.executemany(command, var)
		self.db.connection.commit()

		return True

	def process_contest(self):
		
		# conditions
		unique = []
		for ref in self.menu:
			if ref[1].text == '':
				return None
			if ref[1].text.title() in unique:
				return None
			if ref[2].text == 'Select':
				return None
			if ref[3].text == 'Select':
				return None
			if ref[2].text == 'Distance':
				if ref[3].text == 'Team':
					return False
			for r in ref[1].text:
				if r.isalnum() == False:
					if r != ' ':
						return None
			if ref[1].text[1] in [0,1,2,3,4,5,6,7,8,9]:
				return None					
			unique.append(ref[1].text.title())
		
		# delete all category="type" 
		command = "DELETE FROM template_contests WHERE standard=0"
		self.db.cursor.execute(command)
		self.db.connection.commit()

		# insert in measure/alphabetical order & alter titles&measure
		data_time = {}
		data_distance = {}
		for ref in self.menu:
			if ref[-1].active == False:
				continue
			if ref[2].text == 'Time':
				data_time[ref[1].text] = ref
			else:
				data_distance[ref[1].text] = ref
		
		alpha_time = sorted(data_time)
		var = []
		for t in alpha_time:
			ref = data_time[t]
			title = t.title()
			if ref[3].text == 'Individual':
				comp = 0
			else:
				comp = 1
			var.append((title, 'time', comp, 0, 1))
		command = '''INSERT INTO template_contests
					(title, measure, team, standard, selection)
					VALUES (?, ?, ?, ?, ?)'''
		self.db.cursor.executemany(command, var)
		self.db.connection.commit()

		alpha_distance = sorted(data_distance)
		var = []
		for t in alpha_distance:
			ref = data_distance[t]
			title = t.title()
			if ref[3].text == 'Individual':
				comp = 0
			else:
				comp = 1
			var.append((title, 'distance', comp, 0, 1))
		command = '''INSERT INTO template_contests
					(title, measure, team, standard, selection)
					VALUES (?, ?, ?, ?, ?)'''
		self.db.cursor.executemany(command, var)
		self.db.connection.commit()

		return True

	def process_points(self):
		
		# conditions really only ''. 0 would have to work for none
		for ref in self.menu:
			if ref[2].text == '':
				return None
			if ref[3].text == '':
				return None

		# delete all category="gender" 
		command = "DELETE FROM template_points"
		self.db.cursor.execute(command)
		self.db.connection.commit()

		# insert 
		var = []
		for ref in self.menu:
			rank = ref[1]
			indiv = int(ref[2].text)
			team = int(ref[3].text)
			var.append((rank, indiv, team))

			if self.unsave != False:
				ref[2].status = True
				ref[3].status = True

		if self.unsave != False:
			self.unsave(True)
			self.unsaved_changes = []

		command = '''INSERT INTO template_points
					(rank, individual, team)
					VALUES (?, ?, ?)'''
		self.db.cursor.executemany(command, var)
		self.db.connection.commit()

		return True

	def process_restrictions(self):
		
		# conditions: a 0 value means no, not none. ie. 0 max is no events
		for ref in self.menu:
			if ref[1].text in ['', '0']:
				value = ''
			else:
				value = int(ref[1].text)
			restriction = ref[0].text

			if self.unsave != False:
				ref[1].status = True

			# update
			self.db.cursor.execute('''UPDATE template_restrictions
				SET value = ? WHERE restriction = ?''', (value, restriction))
			self.db.connection.commit()

		if self.unsave != False:
			self.unsave(True)
			self.unsaved_changes = []

		return True

	def process_special(self):

		title_list = []
		error_list = []
		
		var = []
		for ref in self.menu:
			if ref[-1].active == False:
				continue

			if ref[1].text == '' or ref[1].text in title_list:
				error_list.append('All Special Events need a unique Title')
			else:
				title_list.append(ref[1].text)

			if ref[2].text == 'Select':
				error_list.append('All Special Events need a %s chosen' % self.type)

			if ref[3].text == '':
				error_list.append("All Special Events need Points entered")

			if error_list == []:
				var.append([ref[1].text, ref[2].text, float(ref[3].text)])

		if error_list != []:
			message = '\n\n'.join(set(error_list))
			PopBox().showwarning(title='Error', message=message)
			return

		self.db.cursor.execute('''DROP TABLE IF EXISTS special_events''')
		self.db.cursor.execute('''CREATE TABLE special_events
					(title TEXT PRIMARY KEY, type TEXT, points REAL)''')
		
		self.db.cursor.executemany('''INSERT INTO special_events
									VALUES(?, ?, ?)''', var)
		self.db.connection.commit()

		if self.unsave != False:
			self.unsave(True)

		return True

class FrameSelect(GridLayout):

	style = StringProperty('') # type/gender/age/contest
	tournament = StringProperty('')

	def __init__(self, *args, **kwargs):
		super(FrameSelect, self).__init__(**kwargs)

		## Settings
		self.db = Database(self.tournament)
		## Database
		self.menu = []
		
		## Display

		## Engine
		if self.style == 'contest':
			self.build_contest()
		elif self.style == 'events':
			self.build_events()

	### FUNCTIONS ###

	def change_scroll_height(self, obj, value):

		self.ids.scroll.height = self.height - self.ids.header.height

	### BUILD ###

	def build_contest(self):

		# header
		self.ids.header.add_widget(Label(text='Contest', font_size=20))
		self.ids.header.add_widget(Label(text='Measured by', font_size=20))
		self.ids.header.add_widget(Label(text='Competition', font_size=20))
		self.ids.header.add_widget(Label(text='Select', font_size=20))

		# database
		c = '''SELECT title, measure, team, selection 
				FROM template_contests
				WHERE standard=1;'''		
		rows = self.db.cursor.execute(c).fetchall()	

		# display
		for row in rows:
			title = row[0]
			measure = row[1].title()
			if db_boole(row[2], db_in=False) == True:
				comp = 'Team'
			else:
				comp = 'Individual'
			select = db_boole(row[3], db_in=False)

			box = BoxLayout(size_hint_y=None, height=70)
			for text in [title, measure, comp]:
				box.add_widget(Label(text=text))
			s = CheckBoxA(active=select)
			box.add_widget(s)

			ref = []
			ref.append(title)
			ref.append(s)

			# engine
			self.menu.append(ref)
			self.ids.display.add_widget(box)
			self.ids.display.height += 70

	def build_events(self):

		# heading
		grade_list = calculate_grade_list(self.tournament)

		self.ids.header.height = 200
		self.ids.header.add_widget(Label(text=''))
		for grade in grade_list:
			gl = Label90(text=grade, size_hint_x=None, width=40)
			self.ids.header.add_widget(gl)

		# display
		contests = calculate_contest_list(self.tournament)
		for contest in contests:
			self.build_events_box(contest, grade_list)

	def build_events_box(self, contest, grade_list):

		box = BoxLayout(size_hint_y=None, height=40)
		ref = [contest]

		c = Label(text=contest)
		box.add_widget(c)

		command = '''SELECT grade FROM template_events
				WHERE contest="%s"''' % contest
		g = self.db.cursor.execute(command).fetchall()

		e1 = []
		for tup in g:
			e1.append(tup[0])
		first_grades = calculate_grade_list(self.tournament, combined=True, 
												gender_check=True)

		if first_grades == []:
			first_grades = calculate_grade_list(self.tournament, age=True)
		for grade in grade_list:
			# check if its already been selected
			select = CheckBoxA(active=False, size_hint_x=None, width=40)
			if grade in e1:
				select.active = True
			if e1 == []: # first time
				if grade in first_grades:
					select.active = True
			box.add_widget(select)
			ref.append(select)

		self.menu.append(ref)
		self.ids.display.add_widget(box)
		self.ids.display.height += 40

	### PROCESS ###

	def process_contest(self):

		# update
		var = []
		for ref in self.menu:
			var.append((db_boole(ref[1].active), ref[0]))

		command = '''UPDATE template_contests
					SET selection=?
					WHERE title=?;'''
		self.db.cursor.executemany(command, var)
		self.db.connection.commit()

		return True

	def process_events(self):

		grade_list = calculate_grade_list(self.tournament, process=True)	
		# get event selection
		event_dict = {}
		for ref in self.menu:
			contest = ref[0]
			for i in range(len(grade_list)):
				grade = grade_list[i]
				if ref[i+1].active == True:
					event = '%s %s' % (grade, contest)
					event_dict[event] = [contest, grade]

		# check event selection against what is currently in template_events
		command1 = "SELECT title FROM template_events"
		current_tup = self.db.cursor.execute(command1).fetchall()
		current = []
		for tup in current_tup:
			current.append(tup[0])

		delete_var = []
		add_var = []
		for event in current:
			if event not in event_dict:
				delete_var.append([event])
		for event in event_dict:
			if event not in current:
				add_var.append((event))

		# delete 
		command2 = '''DELETE FROM template_events 
					WHERE title=?'''
		if delete_var != []:
			self.db.cursor.executemany(command2, delete_var)
			self.db.connection.commit()
		# insert
		var = []
		for event in add_var:
			ref = event_dict[event]
			var.append((event, ref[0], ref[1]))
		command3 = '''INSERT INTO template_events 
					(title, contest, grade, 
					record_what, record_who, record_when,
					program_number, program_time)
					VALUES (?, ?, ?, 0, '', '', '', '');'''
		if var != []:
			self.db.cursor.executemany(command3, var)
			self.db.connection.commit()

		return True

class FrameTab(GridLayout):

	style = StringProperty('') 
	tournament = StringProperty('')

	def __init__(self, *args, **kwargs):
		super(FrameTab, self).__init__(**kwargs)

		## Settings
		self.unsave = kwargs.pop('unsave', False)
		self.db = Database(self.tournament)
		
		## Database
		self.menu = {}
		self.panel_menu = {}
		self.scroll_menu = []
		self.panel = None
		self.unsaved_changes = []

		## Display

		## Engine
		if self.style == 'records':
			self.build_event_tab('records')
		elif self.style == 'program':
			self.build_event_tab('program')
		elif self.style == 'archive':
			self.build_archive_tab()

	### FUNCTIONS ###

	def change_scroll_height(self, obj, value):

		for scroll in self.scroll_menu:

			if self.style == 'archive':
				scroll.height = self.height - 80
			else:
				scroll.height = self.height - 120 # 3x40 panel/title/header

	def change_scroll_width(self, obj, value):

		self.ids.scroll.width = self.width

	def change_panel(self, i):

		panel = self.panel_menu[i.text]

		if self.panel != None:
			self.ids.display.remove_widget(self.panel)

		self.ids.display.add_widget(panel)
		self.panel = panel

	def command_save(self, value, event):

		if value == False:
			if event not in self.unsaved_changes:
				self.unsaved_changes.append(event)

			if self.unsave != False:
				self.unsave(False)

	### BUILD ###

	def build_event_tab(self, type):

		# settings
		grade_list = calculate_grade_list(self.tournament)
		contest_list = calculate_contest_list(self.tournament)

		for grade in grade_list:
			event_details = []
			for contest in contest_list:
				command = '''SELECT * FROM template_events
							WHERE grade="%s" AND contest="%s"''' % (grade, contest)
				res = self.db.cursor.execute(command).fetchone()
				if res != None:
					event_details.append(res)
			if len(event_details) != 0:
				if type == 'records':
					self.build_record_panel(grade, event_details)
				elif type == 'program':
					self.build_program_panel(grade, event_details)

		# set first panel
		if len(self.panel_menu) > 0:
			panel = None
			for grade in grade_list:
				if grade in self.panel_menu:
					panel = self.panel_menu[grade]
					break
			self.ids.display.add_widget(panel)
			self.panel = panel

	def build_record_panel(self, grade, event_details):

		# tab
		tab = Tab(text=grade)
		tab.bind(on_press=self.change_panel)
		self.ids.tab_panel.add_widget(tab)
		self.ids.tab_panel.width += tab.width

		# panel
		panel = GridLayout(cols=1)

		# title
		title = Label(text=grade, font_size=20, size_hint_y=None, height=40)
		panel.add_widget(title)

		# header
		header = BoxLayout(size_hint_y=None, height=40)
		panel.add_widget(header)
		for h in ['Event', 'Record', 'Holder']:
			header.add_widget(Label(text=h, font_size=20))
		header.add_widget(Label(text='Year', font_size=20, size_hint_x=0.5))

		# scroll
		scroll = ScrollView(size_hint_y=None, height=100)
		panel.add_widget(scroll)
		self.scroll_menu.append(scroll)
		display = GridLayout(cols=1, size_hint_y=None, height=0)
		scroll.add_widget(display)

		# body
		for tup in event_details:
			event = tup[0]
			contest_type = calculate_contest_type(self.tournament, tup[1])

			box = BoxLayout(size_hint_y=None, height=50)
			event_label = LabelWrap(text=event)
			box.add_widget(event_label)

			if contest_type == 'time':

				record_what = ScoreTime(input=float(tup[3]), run_status=True)
			elif contest_type == 'distance':

				record_what = ScoreDistance(input=float(tup[3]), 
											run_status=True)
			box.add_widget(record_what)

			record_who = TextInputC(text=tup[4], font_size=20, run_status=True,
										status=True, multiline=False)
			box.add_widget(record_who)

			record_when = TextInputC(text=str(tup[5]), font_size=20, size_hint_x=0.5,
						input_filter='int', multiline=False, run_status=True)
			box.add_widget(record_when)

			ref = [record_what, record_who, record_when]
			for record in ref:
				record.bind(status=lambda obj, value, 
					event=event:self.command_save(value, event))

			self.menu[event] = ref
			display.add_widget(box)
			display.height += 50

		# engine
		self.panel_menu[grade] = panel

	def build_program_panel(self, grade, event_details):

		# tab
		tab = Tab(text=grade)
		tab.bind(on_press=self.change_panel)
		self.ids.tab_panel.add_widget(tab)
		self.ids.tab_panel.width += tab.width

		# panel
		panel = GridLayout(cols=1)

		# title
		title = Label(text=grade, font_size=20, size_hint_y=None, height=40)
		panel.add_widget(title)

		# header
		header = BoxLayout(size_hint_y=None, height=40)
		panel.add_widget(header)
		for h in ['Event', 'Event Number', 'Event Time']:
			header.add_widget(Label(text=h, font_size=20))

		# scroll
		scroll = ScrollView(size_hint_y=None, height=100)
		panel.add_widget(scroll)
		self.scroll_menu.append(scroll)
		display = GridLayout(cols=1, size_hint_y=None, height=0)
		scroll.add_widget(display)

		# body
		for tup in event_details:
			event = tup[0]
			contest_type = calculate_contest_type(self.tournament, tup[1])

			box = BoxLayout(size_hint_y=None, height=50)
			event_label = LabelWrap(text=event)
			box.add_widget(event_label)

			program_number = TextInputC(text=str(tup[6]), font_size=20, 
										run_status=True, multiline=False)

			program_time = Time24(input=tup[7], font_size=20, 
									run_status=True)

			ref = [program_number, program_time]
			for prog in ref:
				prog.bind(status=lambda obj, value, 
					event=event:self.command_save(value, event))
				box.add_widget(prog)

			self.menu[event] = ref
			display.add_widget(box)
			display.height += 50

		# engine
		self.panel_menu[grade] = panel

	def build_archive_tab(self):

		# settings
		grade_list = calculate_grade_list(self.tournament)
		contest_list = calculate_contest_list(self.tournament)

		for grade in grade_list:
			event_details = []
			for contest in contest_list:
				command = '''SELECT * FROM archive_records
						WHERE grade="%s" AND contest="%s"''' % (grade, contest)
				res = self.db.cursor.execute(command).fetchone()
				if res != None:
					event_details.append(res)
			if len(event_details) != 0:
				self.build_record_panel(grade, event_details)

		# set first panel
		if len(self.panel_menu) > 0:
			panel = None
			for grade in grade_list:
				if grade in self.panel_menu:
					panel = self.panel_menu[grade]
					break
			self.ids.display.add_widget(panel)
			self.panel = panel

	### PROCESS ###

	def process_records(self):

		var = []
		for event in self.unsaved_changes:		
			ref = self.menu[event]
			what = ref[0].output()
			who = ref[1].text
			when = ref[2].text
			var.append((what, who, when, event))

			if self.unsave != False:
				ref[0].status = True
				ref[1].status = True
				ref[2].status = True

		if self.unsave != False:
			self.unsave(True)
			self.unsaved_changes = []

		command = '''UPDATE template_events
					SET record_what=?, record_who=?, record_when=?
					WHERE title=? '''
		self.db.cursor.executemany(command, var)
		self.db.connection.commit()

		return True

	def process_program(self):

		var = []
		for event in self.unsaved_changes:		
			ref = self.menu[event]
			prog_number = ref[0].text
			prog_time = ref[1].output()
			var.append((prog_number, prog_time, event))

			if self.unsave != False:
				ref[0].status = True
				ref[1].status = True

		if self.unsave != False:
			self.unsave(True)
			self.unsaved_changes = []

		command = '''UPDATE template_events
					SET program_number=?, program_time=?
					WHERE title=? '''
		self.db.cursor.executemany(command, var)
		self.db.connection.commit()

		return True

class FrameLogin(BoxLayout):

	orientation = 'vertical'
	add_help = StringProperty('')

	def __init__(self, *args, **kwargs):
		super(FrameLogin, self).__init__(**kwargs)

		## Settings
		self.db = Database('main')
		self.list = kwargs.pop('t_list', None)
		self.dict = kwargs.pop('t_dict', None)

		## Database
		self.build_database()

		## Display
		self.build_header()
		self.build_heading()
		self.build_tournament()

	### FUNCTIONS ###

	def command_action(self, i):

		tournament = self.pick_tournament.text
		if tournament == 'Select':
			return
		title = self.pick_role.text
		if title == 'Select':
			return
		cipher = self.menu[tournament][title]
		pw = self.password.text

		if cipher in [None, ''] or cipher == pw:

			role = title.split(' - ')[0]
			category = title.split(' - ')[1]

			if role == 'Captain':
				zohome = ProgramCaptain(tournament=tournament,
										category=category)

			elif role == 'Official':
				zohome = ProgramOfficial(tournament=tournament,
										category=category)

			elif role == 'Public':
				zohome = ProgramPublic(tournament=tournament)

	def command_help(self, i):

		PopBox().help_screen(title='Tournament Login', message=self.add_help)

	def command_role(self, obj, value):

		if value == 'Select':
			return

		options = []

		for ref in self.menu[value]:
			options.append(ref)

		options = sorted(options)

		self.pick_role.text = 'Select'
		self.pick_role.change_all_options(options)

	### BUILD ###

	def build_database(self):

		self.menu = {}

		for t in self.dict:
			self.menu[t] = {}
			for ref in self.dict[t]:
				role = '%s - %s' % (ref[0], ref[1])
				pw = ref[2]
				self.menu[t][role] = pw

	def build_header(self):

		title_box = BoxLayout(size_hint_y=None, height=50)
		self.add_widget(title_box)

		title = Label(text='Tournament Login', font_size=30)
		title_box.add_widget(title)

		btn = Button(text='?', font_size=30, size_hint_x=None, width=50)
		btn.bind(on_press=self.command_help)
		title_box.add_widget(btn)

	def build_heading(self):

		heading = BoxLayout(size_hint_y=None, height=40)
		self.add_widget(heading)

		heading.add_widget(Label(text='Tournament', font_size=20))
		heading.add_widget(Label(text='Access', font_size=20))
		heading.add_widget(Label(text='Password', font_size=20, 
									size_hint_x=None, width=120))
		heading.add_widget(Label(text='Confirm', font_size=20,
									size_hint_x=None, width=120))

	def build_tournament(self):

		body = BoxLayout(size_hint_y=None, height=70)
		self.add_widget(body)

		self.pick_tournament = OptionMenu(options=self.list)
		self.pick_tournament.bind(text=self.command_role)
		self.pick_role = OptionMenu(options=[])
		self.password = TextInputC(multiline=False, size_hint_x=None, width=120)
		
		btn = Button(text='Enter', font_size=14, size_hint_x=None, width=120)
		btn.bind(on_press=self.command_action)

		body.add_widget(self.pick_tournament)
		body.add_widget(self.pick_role)
		body.add_widget(self.password)
		body.add_widget(btn)

class MainWindow(BoxLayout):

	def __init__(self, *args, **kwargs):
		super(MainWindow, self).__init__(**kwargs)

		## Settings
		self.sidebar = self.ids.sidebar
		self.body = self.ids.body
		self.title = self.ids.title
		self.ids.sidebody.bind(height=self.command_scroll_height)

		## Database
		self.page = None
		self.sidemenu = {}

	### FUNCTIONS ###

	def command_button(self, obj, value):

		pass

	def command_sidebar(self, i, menu):

		if self.sidemenu[menu][0] == True:
			for btn in self.sidemenu[menu][1]:
				menu.remove_widget(btn)
				menu.height -= 30
				self.sidebar.height -= 30
			self.sidemenu[menu][0] = False
		else:
			for m in self.sidemenu:
				if m != menu:
					if self.sidemenu[m][0] == False:
						continue
					for btn in self.sidemenu[m][1]:
						m.remove_widget(btn)
						m.height -= 30
						self.sidebar.height -= 30
						self.sidemenu[m][0] = False

			for btn in self.sidemenu[menu][1]:
				menu.add_widget(btn)
				menu.height += 30
				self.sidebar.height += 30
				self.sidemenu[menu][0] = True

	def command_scroll_height(self, *args):
		
		self.ids.scroll.height = Window.height - self.title.height

	def resize_font(self, *args):

		if self.title.texture_size[1] > self.title.size[1]:
			self.title.font_size -= 1
			self.title.texture_update()

	def resize_label(self, *args):

		if self.title.texture_size[1] > self.title.size[1]:

			if self.title.height == self.title.height_max:
				self.title.size[1] += 1
				self.title.size[1] -= 1
			elif self.title.texture_size[1] > self.title.height_max:
				self.title.size[1] = self.title.height_max
				self.title.texture_update()
			else:
				self.title.size[1] = self.title.texture_size[1]
				self.title.texture_update()

	### PROCESS ###	

	def add_sidemenu(self, **kwargs):
		
		heading = kwargs.pop('heading', '')
		subheadings = kwargs.pop('subheadings', [])
		function = kwargs.pop('function', None)

		menu = BoxLayout(orientation='vertical', size_hint_y=None, height=0)

		h = Button(text=heading, size_hint_y=None, height=50)
		h.bind(size=self.command_button)
		h.bind(on_press=lambda i, menu=menu:self.command_sidebar(i, menu))
		menu.add_widget(h)
		menu.height = 50
		self.sidebar.height += 50

		box = []
		for sub in subheadings:
			s = Button(text=sub, size_hint=(0.9, None), height=30)
			if function != None:
				s.bind(on_press=lambda i, sub=sub, 
						heading=heading:function(heading, sub))
			box.append(s)
		
		self.sidemenu[menu] = [False, box]
		self.sidebar.add_widget(menu)

	def change_page(self, page):

		if self.page != None:
			self.body.remove_widget(self.page)

		self.body.add_widget(page)
		self.page = page

	def change_title(self, title):
		
		self.title.text = ''
		self.title.font_size = 30
		self.title.text = title

	def clear_sidemenu(self):
		
		for menu in self.sidemenu:
			self.sidebar.remove_widget(menu)
			
		self.sidebar.height = 0 # prevents permanent growth to display
		self.sidemenu = {}

	def open_menu(self, heading):

		for menu in self.sidemenu:     
			if menu.children[-1].text == heading:
				self.command_sidebar(None, menu)

class Ribbon(BoxLayout):

	title = StringProperty('')
	question = StringProperty('')
	answer = ObjectProperty()
	reset = BooleanProperty(None) # None/True/False
	help = StringProperty('')

	def __init__(self, *args, **kwargs):
		super(Ribbon, self).__init__(**kwargs)

		## Settings
		self.function = kwargs.pop('function', None)
		self.reset_position = kwargs.pop('reset_position', None)

		## Database

		## Display
		if self.reset == None:
			self.ids.action.remove_widget(self.ids.reset)
		self.ids.answer.add_widget(self.answer)

		## Engine

	### FUNCTIONS ###

	def command_confirm(self, *args):

		self.function(self)

	def command_footer(self, *args):
		
		if self.ids.footer_label.text == '':
			self.ids.footer_label.text = self.help
		else:
			self.ids.footer_label.text = ''

	def command_reset(self, *args):
		
		if self.reset == True:	
			try:	
				self.answer.text = self.answer.main
			except AttributeError:
				try:
					self.answer.option.text = 'Select'
				except AttributeError:
					self.answer.text = ''

			if self.reset_position != None:
				self.reset_position(self)

	### BUILD ###

### PAGE CLASSES ### 

class Page(GridLayout):

	cols = 1

	title = StringProperty('') #
	add_reset = BooleanProperty(False) # 
	add_help = StringProperty('') # accesses help db?


	def __init__(self, *args, **kwargs):
		super(Page, self).__init__(**kwargs)

		## Settings
		self.reset = None

		## Database
		self.display = None

		## Display
		self.build_header()
		self.body = self.ids.body

		## Engine		

	### FUNCTIONS ###

	def command_help(self, *args):

		PopBox().help_screen(title=self.title, message=self.add_help)

	def command_reset(self, *args):
		''' 
		This function can be overwritten in sub-classes, or connected to       
		in add_display()
		'''
		if self.reset != None:
			self.reset

	def resize_font(self, obj, value):

		w = self.width

		if w == 100: # (100,100) is the pre-display default size
			return

		if self.add_help != '':
			w -= 50
		if self.add_reset == True:
			w -= 50

		if value[0]> w:
			self.ids.title.font_size -= 1
			self.ids.title.texture_update()

	### BUILD ###

	def build_header(self):

		if self.add_reset == False:
			self.ids.header.remove_widget(self.ids.reset)

		if self.add_help == '':
			self.ids.header.remove_widget(self.ids.help)

	### DISPLAY ###

	def add_display(self, display, **kwargs):

		self.reset = kwargs.pop('reset', None) # why is this here?

		if self.display != None:
			self.body.remove_widget(self.display)

		self.body.add_widget(display)
		self.display = display

class PageComplete(Page):

	title = StringProperty('Complete')
	add_help = StringProperty('Help')

	tournament = StringProperty('')

	def __init__(self, *args, **kwargs):
		super(PageComplete, self).__init__(**kwargs)

		## Settings
		self.db = Database(self.tournament)

		## Database
		self.complete = None

		## Display
		self.build_display()

	### FUNCTIONS ###

	def command_check(self, i):

		if self.complete != None:
			self.display.remove_widget(self.complete)

		# conditions
		unfinished = []
		eve = self.db.cursor.execute('''SELECT title
							FROM template_events''').fetchall()
		for e in eve:

			table = db_table('%s RR' % e[0])

			try:
				check = self.db.cursor.execute('''SELECT *
							FROM %s''' % table).fetchall()
			except:
				unfinished.append(e[0])
	
		if unfinished != []:

			if len(unfinished) < 10:
				events = '\n'.join(unfinished)
				message = 'The following events have not been completed:\n\n%s\n\nDo you still want to complete this Tournament?' % events
			else:
				message = '%s Events have not been completed.\n\nDo you still want to complete this Tournament?' % len(unfinished)

			PopBox().askyesno(title='Tournament Results', 
				function=self.command_confirm,
				message=message)
			return

		self.command_confirm(True)

	def command_confirm(self, boole):

		if boole == True:

			self.complete = Button(text='Complete Tournament', font_size=20,
				size_hint_y=None, height=150)
			self.complete.bind(on_release=self.command_complete)
			self.display.add_widget(self.complete)

	def command_complete(self, *args):

		try:

			try: 
				self.db.cursor.execute('''CREATE TABLE archive_records AS
					SELECT * FROM template_events''')
			except:
				pass

			self.db.cursor.execute('''UPDATE details
								SET stage="Complete"''')
		
			# turn off house/official access
			self.db.cursor.execute('''UPDATE settings
									SET active=0
									WHERE role="Captain"''')
			self.db.cursor.execute('''UPDATE settings
									SET active=0
									WHERE role="Official"''')
			self.db.cursor.execute('''UPDATE settings
									SET active=1
									WHERE role="Public"''')
			self.db.connection.commit()

			## Email
			import smtplib
			from email.MIMEMultipart import MIMEMultipart
			from email.MIMEText import MIMEText
			from email.MIMEBase import MIMEBase
			from email import encoders
			 
			fromaddr = "zosportsnz@gmail.com"
			toaddr = "tournaments@zo-sports.com"
			
			db = Database('main')

			sport = self.db.cursor.execute("SELECT sport FROM details").fetchone()[0]
			organisation = db.cursor.execute('''SELECT title, organisation
											FROM details''').fetchone()
			settings = db.cursor.execute('''SELECT address, email
											FROM settings''').fetchall()[0]
			address = settings[0]
			email = settings[1]

			msg = MIMEMultipart()
			 
			msg['From'] = fromaddr
			msg['To'] = toaddr
			msg['Subject'] = "%s - %s - %s" % (sport, organisation, title)
			 
			body = "%s\n%s\n%s" % (address, email, self.tournament)
			 
			msg.attach(MIMEText(body, 'plain'))

			filename = os.path.join(os.path.abspath("."), "%s.db" % self.tournament)
			attachment = open(filename, "rb")
			 
			part = MIMEBase('application', 'octet-stream')
			part.set_payload((attachment).read())
			encoders.encode_base64(part)
			part.add_header('Content-Disposition', "attachment; filename= %s" % filename)
			 
			msg.attach(part)
			 
			server = smtplib.SMTP('smtp.gmail.com', 587)
			server.starttls()
			server.login(fromaddr, "vetinari13")
			text = msg.as_string()
			server.sendmail(fromaddr, toaddr, text)
			server.quit()

			## zo
			self.db.connection.close()			
			zotournament = ProgramTournament(title=self.tournament)

		except Exception as inst:
			print type(inst)
			print inst.args
			print inst
			DataWarning()

	### BUILD ###

	def build_display(self):

		self.display = GridLayout(cols=1, padding=50, spacing=50)

		btn = Button(text='Check Tournament', font_size=20,
						size_hint_y=None, height=150)
		btn.bind(on_release=self.command_check)
		self.display.add_widget(btn)

		self.add_display(self.display)

class PageEntry(Page):

	tournament = StringProperty('')
	typename = StringProperty('')

	def __init__(self, *args, **kwargs):
		super(PageEntry, self).__init__(**kwargs)

		## Settings
		self.db = Database(self.tournament)
		details = self.db.cursor.execute("SELECT * FROM details").fetchone()
		self.type = details[3]

		type_details = self.db.cursor.execute('''SELECT * 
								FROM template_groups
								WHERE title="%s"''' % self.typename).fetchone()
		self.bg = db_colour(type_details[6], db_in=False)
		self.txt = db_colour(type_details[7], db_in=False)

		self.grade_tables = calculate_type_tables(self.tournament, 
										self.typename, grade=True)

		self.restrictions = self.db.cursor.execute('''SELECT * 
								FROM template_restrictions''').fetchall()

		## Database
		self.current = None

		self.grade = None
		self.menu = {}
		self.unsaved = []
		self.event_list = None
		self.contest_type = []
		self.save = None

		## Display
		self.frame = GridLayout(cols=1)
		self.add_display(self.frame)
		
		self.build_panel()
		for grade in self.grade_tables:
			self.build_tab(grade[0])

		## Engine
		self.build_default()

	### FUNCTIONS ###

	def change_grade_header(self, value, box, scroll):

		scroll.height = box.height - value - 60

	def change_grade_height(self, value, header, scroll):

		scroll.height = value - header.height - 60

	def change_header(self, value, header):

		if value[0] > header.height:

			header.height = value[0] + 10

	def change_panel_width(self, obj, value):

		self.tab_panel.width += (value - 100)

	def change_scroll_width(self, obj, value):

		self.panel.width = self.width

	def command_check(self, obj, value, id, i):

		# check restrictions: only matter if entering 
		if value == True:

			ctype = self.contest_type[i]
			for res in self.restrictions:

				if res[1] in [None, '']:

					continue

				if res[0] == 'Tournament Maximum':

					count = 0
					for chk in self.menu[id]:
						if chk.active == True:
							count += 1

					if count > res[1]:
						obj.active = False
						obj.react = False
						self.pop_restriction('max')
						return

				type_count = 0
				for i in range(len(self.contest_type)):
					if self.contest_type[i] == ctype:
						if self.menu[id][i].active == True:
							type_count += 1

				if res[0] == 'Track Event Maximum':
					if ctype == 'time':
						if type_count > res[1]:
							obj.active = False
							obj.react = False
							self.pop_restriction('track')
							return

				elif res[0] == 'Field Event Maximum':
					if ctype =='distance':
						if type_count > res[1]:
							obj.active = False
							obj.react = False
							self.pop_restriction('field')
							return
					
		# check not a restriction response callback
		if obj.react == False:
			obj.react = True
			return

		# mark as unsaved
		self.save.text = 'Unsaved'
		self.save.bg = [1, 0, 0, 1]

		if id not in self.unsaved:
			self.unsaved.append(id)

	def command_tab(self, i):

		# check current
		if self.unsaved != []:
			PopBox().askyesno(title='Unsaved', function=self.command_tab_check,
			message='Unsaved changes were made to %s, continue anyway?' % self.grade)
			self.tab_ref = i.text
			return

		self.frame.remove_widget(self.current)
		self.build_grade(i.text)

	def command_tab_check(self, boole):

		if boole == True: 
			self.unsaved = []
			self.frame.remove_widget(self.current)
			self.build_grade(self.tab_ref)

	def pop_restriction(self, style):

		if style == 'max':
			res = self.restrictions[1][1]
			message = 'Tournament Event entry maximum is %d' % res
		elif style == 'track':
			res = self.restrictions[2][1]
			message = 'Track Event entry maximum is %d' % res
		elif style == 'field':
			res = self.restrictions[3][1]
			message = 'Field Event entry maximum is %d' % res

		title = 'Event Entry Restriction'

		PopBox().showwarning(title=title, message=message)

	### BUILD ###

	def build_panel(self):

		self.panel = ScrollView(size_hint=(None, None), height=50, width=0)
		self.bind(size=self.change_scroll_width)
		self.frame.add_widget(self.panel)

		self.tab_panel = BoxLayout(size_hint=(None, None), height=50, width=0)
		self.panel.add_widget(self.tab_panel)

		self.frame.add_widget(Label(size_hint_y=None, height=10))

	def build_tab(self, grade):

		tab = Tab(text=grade, size_hint_y=None, height=40)
		tab.bind(on_press=self.command_tab)
		tab.bind(width=self.change_panel_width)
		self.tab_panel.add_widget(tab)
		self.tab_panel.width += 100

	def build_grade(self, grade):

		## Settings - Reset
		self.grade = grade
		self.menu = {}
		self.unsaved = []
		self.event_list = None
		self.contest_type = []
		self.save = None

		box = GridLayout(cols=1, padding=[10, 10, 10, 10])

		# header
		header = BoxLayout(size_hint_y=None, height=150)
		box.add_widget(header)
		header.add_widget(LabelWrapC(text=grade, font_size=30,
			bg=self.bg, color=self.txt))

		self.event_list = calculate_grade_events(self.tournament, grade, 
													team=False)

		for event in self.event_list:
			contest = self.db.cursor.execute('''SELECT contest, grade
										FROM template_events
										WHERE title="%s"''' % event).fetchone()
			
			cteam = calculate_contest_type(self.tournament, contest[0],
											team=True)
			if cteam == True:
				continue
			if contest[1] == grade:
				event = contest[0]
			l9 = Label90(text=event, size_hint_x=None, width=40)
			l9.bind(texture_size=lambda obj, value, 
				header=header:self.change_header(value, header))
			header.add_widget(l9)

			# add contest type 
			ctype = calculate_contest_type(self.tournament, contest[0])
			self.contest_type.append(ctype)
			
		# body
		scroll = ScrollView(size_hint_y=None, height=250)
		box.add_widget(scroll)

		body = GridLayout(cols=1, size_hint_y=None, height=0)
		scroll.add_widget(body)

		# id
		self.build_grade_id(grade, body)

		# action
		action = BoxLayout(size_hint_y=None, height=40)
		box.add_widget(action)

		self.save = LabelC(size_hint_x=None, width=100)
		action.add_widget(self.save)

		btn = Button(text='Save', font_size=30)
		btn.bind(on_press=self.process_save)
		action.add_widget(btn)

		## Engine
		box.bind(height=lambda obj, value, header=header,
				scroll=scroll:self.change_grade_height(value, header, scroll))
		header.bind(height=lambda obj, value, box=box,
				scroll=scroll:self.change_grade_header(value, box, scroll))
		self.frame.add_widget(box)
		self.current = box

	def build_grade_id(self, grade, body):

		# id
		typelist = self.typename.split(' ')
		table = ('_').join(typelist + grade.split(' '))

		id_entry = self.db.cursor.execute('''SELECT * 
											FROM %s''' % (table)).fetchall()

		id_var = {}
		for entry in id_entry:
			id = entry[0]
			id_var[id] = entry[1:]
		
		data = {}
		id_names = {}
		for id in id_var:
			title = self.db.cursor.execute('''SELECT firstname, surname 
							FROM individuals
							WHERE id="%s"''' % id).fetchone()

			data[id] = '%s %s' % (title[1], title[0])
			id_names[id] = '%s %s' % (title[0], title[1])

		for id in sorted(data, key=data.get):

			name = id_names[id]

			id_box = self.build_competitor_box(id, name, id_var[id])
			body.add_widget(id_box)
			body.height += 40

	def build_competitor_box(self, id, name, entry):

		## Settings
		box = BoxLayout(size_hint_y=None, height=40)
		box.add_widget(LabelWrap(text=name))

		ref = []
		for i in range(len(entry)):

			if entry[i] in [None, '']:
				active = 0
			else:
				active = int(entry[i])

			entered = db_boole(active, db_in=False)

			chk = CheckBoxA(active=entered, size_hint_x=None, width=40)
			chk.bind(active=lambda obj, value, i=i,
							id=id:self.command_check(obj, value, id, i))
			box.add_widget(chk)
			ref.append(chk)

		self.menu[id] = ref

		return box

	def build_default(self):

		box = GridLayout(cols=1)

		frame = GridLayout(cols=1, padding=[100, 100, 100, 0])
		box.add_widget(frame)

		default = Label(text="Select Grade", font_size=20, 
							size_hint_y=None, height=150)

		frame.add_widget(default)

		self.frame.add_widget(box)
		self.current = box

	### PROCESS ###

	def process_save(self, *args):

		typelist = self.typename.split(' ')
		table = ('_').join(typelist + self.grade.split(' '))

		for id in self.unsaved:
			ref = self.menu[id]
			for i in range(len(self.event_list)):
				event = '_'.join(self.event_list[i].split(' '))
				entry = db_boole(ref[i].active)

				try:
					self.db.cursor.execute('''UPDATE %s
							SET %s="%s"
							WHERE id="%s"''' % (table, event, entry, id))
				except:
					DataWarning()
					return

		self.db.connection.commit()

		self.save.text = ''
		self.save.bg = [0.5, 0.5, 0.5, 1]		
		self.unsaved = []

class PageEntryComplete(Page):

	title = StringProperty('Entry Section Complete')
	tournament = StringProperty('')

	def __init__(self, *args, **kwargs):
		super(PageEntryComplete, self).__init__(**kwargs)

		## Setting
		self.db = Database(self.tournament)
		details = self.db.cursor.execute('''SELECT * 
										FROM details''').fetchone()
		self.sport = details[1]
		self.type = details[3]
		self.age = details[4]
		self.type_list = self.db.cursor.execute('''SELECT title
									FROM template_groups
									WHERE category="type"''').fetchall()
		## Database
		self.individuals = {}
		self.events = {}
		self.team_events = []
		events = self.db.cursor.execute('''SELECT title, contest
										FROM template_events''').fetchall()
		
		no_team_contests = calculate_contest_list(self.tournament, team=False)
		for tup in events:
			if tup[1] in no_team_contests:
				self.events[tup[0]] = []
			else:
				self.team_events.append(tup[0])

		## Display
		self.build_box()

	### FUNCTIONS ###

	def command_check(self, i):

		## Settings
		self.confirm.clear_widgets()

		## Check Conditions
		tournament_min = self.db.cursor.execute('''SELECT value
						FROM template_restrictions
					WHERE restriction="Tournament Minimum"''').fetchone()[0]
		if tournament_min != '':
			
			min = int(tournament_min)

			type_tables = []
			for tup in self.type_list:
				tables = calculate_type_tables(self.tournament, tup[0])
				type_tables += tables

			restrict = []

			for table in type_tables:
				ids = self.db.cursor.execute('''SELECT * 
											FROM %s''' % table).fetchall()
				for tup in ids:
					count = 0
					for t in tup:
						if t == '1':
							count += 1
					if count < min:
						restrict.append(table)
						break

			if restrict != []:

				type_restrict = []
				for r in restrict:
					type_restrict.append(' '.join(r.split('_')))
				type_restrict = ', '.join(type_restrict)

				PopBox().showwarning(title='Tournament Minimums',
					message='Tournament Minimum entry requirement of %s, not met in %s.\n\nCheck those entries or change the minimum requirement' % (min, type_restrict))
				return

		## Process first half of information - to speed up time later
		self.process_check()

	def command_confirm(self, i):

		## Process second half of information
		self.process_confirm()

	def command_empty(self, boole):

		if boole == True:
			self.build_confirm()

	def command_feedback(self, boole):

		zotournament = ProgramTournament(title=self.tournament)

	### BUILD ###

	def build_confirm(self):

		btn = Button(text='Save & Continue', font_size=20)
		btn.bind(on_press=self.command_confirm)
		self.confirm.add_widget(btn)

	def build_box(self):

		box = GridLayout(cols=1, padding=50, spacing=50)

		complete = LabelWrap(text='''This will finish the Entry section.\n\nYou will still be able to add late entries during the Competition section as well as change Event Restrictions, Event Records and Event Number & Times.\n\nPress 'Entry Check' which will do a final check of all entered details, then the 'Save and Continue' button will appear.''', 
				font_size=20, size_hint_y=None, height=200)

		box.add_widget(complete)

		box1 = BoxLayout(size_hint_y=None, height=150)
		box.add_widget(box1)

		self.check = Button(text='Entry Check', font_size=20)
		self.check.bind(on_press=self.command_check)
		box1.add_widget(self.check)

		self.confirm = BoxLayout()
		box1.add_widget(self.confirm)

		self.add_display(box)

	### PROCESS ###

	def process_check(self):

		# 1st only get id from those that have entered anything
		type_tables = []
		for tup in self.type_list:
			tables = calculate_type_tables(self.tournament, '_'.join(tup[0].split(' ')), grade=True)
			type_tables += tables

		for table in type_tables:
			
			# event list
			grade = table[0]
			event_list = calculate_grade_events(self.tournament, grade,
												team=False, table=False)

			# id list
			all_id = self.db.cursor.execute('''SELECT * 
											FROM %s''' % table[1]).fetchall()
			for id_tup in all_id:

				if '1' not in id_tup:
					continue 

				entered_events = []

				for i in range(len(event_list)):
					boole = id_tup[i+1]
					if boole == '1':
						event = event_list[i]
						self.events[event].append(id_tup[0]) 
						entered_events.append(event)
				
				altered_events = []
				for e in entered_events:
					altered_events.append(''.join(e.split(' ')))

				self.individuals[id_tup[0]] = db_list(altered_events)
				# no need to add those ids of un-entered competitors

		empty = []
		for event in self.events:
			if len(self.events[event]) == 0:
				empty.append(event)
		if len(empty) > 0:
			unentered = ', '.join(empty)
			PopBox().askyesno(title='Event/s Unentered',
				message='%s have no competitors entered\n\nContinue anyway?' % unentered,
				function=self.command_empty)
			return
		self.build_confirm()

	def process_confirm(self):

		# update individuals
		var = []
		for id in self.individuals:
			var.append([self.individuals[id], id])

		try:
			self.db.cursor.execute('''ALTER TABLE individuals
								ADD events TEXT''')
		except:
			pass

		self.db.cursor.executemany('''UPDATE individuals
								SET events=?
								WHERE id=?''', var)
		self.db.connection.commit()

		# create event table/s
		for event in self.events:
		
			table = '_'.join(event.split(' '))
			table_r = '%s_R1' % table
			self.db.cursor.execute('''DROP TABLE IF EXISTS %s''' % table)
			self.db.cursor.execute('''CREATE TABLE %s
						(round TEXT, title TEXT, height TEXT, 
						program_number INT, program_time TEXT)''' % table)
			self.db.cursor.execute('''INSERT INTO %s
								VALUES('R1', "%s", '', '', '')''' % (table, event))
			self.db.cursor.execute('''DROP TABLE IF EXISTS %s''' % table_r)			
			self.db.cursor.execute('''CREATE TABLE %s
					(id TEXT, pp INT, 
					time_measure INT, length TEXT, height TEXT,
					placing INT)''' % table_r)
			self.db.connection.commit()

			var = []
			for id in self.events[event]:
				var.append([id])
			self.db.cursor.executemany('''INSERT INTO %s (id)
										VALUES (?)''' % table_r, var)	
			self.db.connection.commit()

		# create team event_tables
		for event in self.team_events:
		
			table = '_'.join(event.split(' '))
			table_r = '%s_R1' % table

			self.db.cursor.execute('''DROP TABLE IF EXISTS %s''' % table)
			self.db.cursor.execute('''CREATE TABLE %s
						(round TEXT, title TEXT, height TEXT, 
						program_number INT, program_time TEXT)''' % table)
			self.db.cursor.execute('''INSERT INTO %s
								VALUES('R1', "%s", '', '', '')''' % (table, event))
			self.db.cursor.execute('''DROP TABLE IF EXISTS %s''' % table_r)
			self.db.cursor.execute('''CREATE TABLE %s
						(team TEXT PRIMARY KEY, pp INT, 
						time_measure INT, length TEXT, height TEXT,
						placing INT, type TEXT)''' % table_r)				
			## change stage

		# change stage
		self.db.cursor.execute('''UPDATE details
								SET stage="Competition"''')

		# add official access
		self.db.cursor.execute('''INSERT INTO settings
							VALUES("Official", 1, 'All', "")''')

		# turn off house access and on public
		self.db.cursor.execute('''UPDATE settings
								SET active=0
								WHERE role="Captain"''')
		self.db.cursor.execute('''UPDATE settings
								SET active=1
								WHERE role="Public"''')

		if self.sport == 'Athletics':
			self.db.cursor.execute('''INSERT INTO settings
							VALUES("Official", 1, "Track", "")''')				
			self.db.cursor.execute('''INSERT INTO settings
							VALUES("Official", 1, "Field", "")''')	

		self.db.connection.commit()
		self.db.connection.close()

		self.process_feedback()

	def process_feedback(self):

		FeedBackForm(function=self.command_feedback, tournament=self.tournament)

class PageEvent(GridLayout):

	cols = 1

	tournament = StringProperty('')
	contest = StringProperty('')
	add_help = StringProperty('')

	def __init__(self, *args, **kwargs):
		super(PageEvent, self).__init__(**kwargs)

		## Settings
		self.db = Database(self.tournament)
		details = self.db.cursor.execute("SELECT * FROM details").fetchone()
		self.type = details[3]
		self.contest_type = calculate_contest_type(self.tournament, 
													self.contest)
		self.team = calculate_contest_type(self.tournament, 
											self.contest, team=True)

		grade_list = []
		table_list = self.db.list_tables()
		event_list = self.db.cursor.execute('''SELECT title, grade
							FROM template_events
							WHERE contest="%s"''' % self.contest).fetchall()
		for tup in event_list:
			if '_'.join(tup[0].split(' ')) in table_list:
				grade_list.append(tup[1])
		self.grade_list = []
		for grade in calculate_grade_list(self.tournament):
			if grade in grade_list:
				self.grade_list.append(grade) # puts them in order


		## Database
		self.current = None
		self.tab_ref = None
		self.temp_header = None

		## Display
		self.build_panel()
		for grade in self.grade_list:
			self.build_tab(grade)
		self.build_display()

		## Engine
		self.build_default()

	### FUNCTIONS ###

	def change_display(self, display):

		self.display.clear_widgets()
		self.display.add_widget(display)
		self.current = display

	def change_grade_height(self, value, header, scroll):

		scroll.height = value - header.height - 60

	def change_header(self, value, header):

		if value[0] > header.height:

			header.height = value[0] + 10

	def change_panel_width(self, obj, value):

		self.tab_panel.width += (value - 100)#

	def change_scroll_width(self, obj, value):

		self.panel.width = self.width - 50

	def command_help(self, *args):

		PopBox().help_screen(title=self.contest, message=self.add_help,
							large=True)

	def command_tab(self, i):

		self.build_grade(i.text)

	def command_tab_check(self, boole):

		if boole == True: 
			self.build_grade(self.tab_ref)

	### BUILD ###

	def build_panel(self):

		self.temp_header = BoxLayout(size_hint_y=None, height=50)
		self.temp_header.add_widget(Label(text=self.contest,
								font_size=30))
		self.add_widget(self.temp_header)

		panel = BoxLayout(size_hint_y=None, height=50)
		self.add_widget(panel)

		self.panel = ScrollView(size_hint=(None, None), height=50, width=0)
		self.bind(size=self.change_scroll_width)
		panel.add_widget(self.panel)

		btn = Button(text='?', font_size=30, size_hint_x=None, width=50)
		btn.bind(on_press=self.command_help)
		panel.add_widget(btn)

		self.tab_panel = BoxLayout(size_hint=(None, None), height=50, width=0)
		self.panel.add_widget(self.tab_panel)

		self.add_widget(Label(size_hint_y=None, height=10))#

	def build_tab(self, grade):

		tab = Tab(text=grade, size_hint_y=None, height=40)
		tab.bind(on_press=self.command_tab)
		tab.bind(width=self.change_panel_width)
		self.tab_panel.add_widget(tab)
		self.tab_panel.width += 100

	def build_display(self):

		self.display = GridLayout(cols=1, padding=10)
		self.add_widget(self.display)

	def build_default(self):

		box = GridLayout(cols=1, padding=100)

		box.add_widget(Label(text='Select Grade', font_size=20, 
								size_hint_y=None, height=150))

		self.change_display(box)

	def build_grade(self, grade):

		## Settings - Reset
		if self.temp_header != None:
			self.remove_widget(self.temp_header)
			self.temp_header = None

		if self.contest_type == 'time':

			if self.team == True:
				display = EventTimeTeam(tournament=self.tournament,
									contest=self.contest, grade=grade)			
			else:
				display = EventTime(tournament=self.tournament, 
									contest=self.contest, grade=grade)	

		elif self.contest_type == 'distance':

			if self.team == True:
				display = EventDistanceTeam(tournament=self.tournament, 
									contest=self.contest, grade=grade)	
			elif self.contest in ['High Jump', 'Pole Vault']:
				display = EventDistanceHeight(tournament=self.tournament, 
									contest=self.contest, grade=grade)	
			else:
				display = EventDistance(tournament=self.tournament, 
									contest=self.contest, grade=grade)	

		self.change_display(display)

class PageLogin(BoxLayout):

	style = StringProperty('Login')
	orientation = 'vertical'
	padding = [50, 50, 50, 50]
	spacing = 50

	def __init__(self, *args, **kwargs):
		super(PageLogin, self).__init__(**kwargs)

		## Settings
		self.size_hint_y = None
		self.height = 260
		if self.style == 'Login':
			self.db = Database('main')
			details = self.db.cursor.execute('''SELECT title, organisation
												 FROM details''').fetchone()
			self.title = details
			self.organisation = details[1]

		## Display
		if self.style == 'Login':
			self.build_admin()
			self.build_tournament()
		elif self.style == 'Main':
	
			self.build_main()
		elif self.style == 'Register':

			self.build_register()

	### FUNCTIONS ###

	def command_admin(self, pw):

		password = pw.text

		regkey = self.db.cursor.execute('''SELECT registration 
										FROM details''').fetchone()

		cipher = self.db.cursor.execute('''SELECT cipher 
										FROM settings''').fetchone()

		if cipher not in [None, '']:
			if cipher[0] == password or regkey[0] == password: 
				zohome = ProgramHome()
			else:
				pw.text = ''
				PopBox().showwarning(title='Password Incorrect',
					message='Password Incorrect')
		else:
			zohome = ProgramHome()

	def command_help(self, i):

		if self.style == 'Login':
			
			if i == 'Admin':

				mes = "If you have forgotten your password, Admin access can be gained using the Registration Key.\nIf you do not have this key, contact info@zo-sports.com"

		elif i == "Main":

			mes = "Your Username will be the name of your School/Club and the Registration Key will be the alpha-numeric password given to you.\nThe Registration Key is case sensitive.\nIf you have any problems, contact info@zo-sports.com"

		elif i == "Register":

			mes = "Your registration time has run out, contact info@zo-sports.com with your Username, to be issued this years Registration Key"

		PopBox().help_screen(title="Login", message=mes)

	def command_register(self, *args):

		if self.username.text == '' or self.reg_key.text == '':
			return

		elif champollion(self.username.text) == self.reg_key.text:

			PopBox().register(function=self.create_main)
			return

		else:
			PopBox().showwarning(title='Incorrect Username or Registration Key',
			message='Username or Registration Key incorrect, try again or contact info@zo-sports.com')
			return

	def command_reregister(self, key):

		pass

	def create_main(self, organisation, address, email):
		
		import string
		title = string.capwords(self.username.text)
		organisation = organisation
		registration = self.reg_key.text

		# create 'main'
		db = Database('main')

		# create 'details', 'settings', 'tournaments'
		db.cursor.execute('''CREATE TABLE details
						(title TEXT, organisation TEXT, registration TEXT)''')
		db.cursor.execute('''INSERT INTO details
							VALUES(?,?,?)''', (title, 
													organisation,
													registration))
		db.connection.commit()

		db.cursor.execute('''CREATE TABLE settings
						(bg_colour TEXT, text_colour TEXT, 
						crest TEXT, motto TEXT, cipher TEXT,
						address TEXT, email TEXT)''')

		bg = '0.5#0.5#0.5#1'
		txt = '1#1#1#1'

		db.cursor.execute('''INSERT INTO settings
							VALUES(?,?,?,?,?,?,?)''', (bg, txt, '', '', '',
								address, email))
		db.connection.commit()

		db.cursor.execute('''CREATE TABLE tournaments
							(title TEXT, sport TEXT)''')
		db.connection.commit()
		db.connection.close()

		zohome = ProgramHome(first=True)

	### BUILD ###

	def build_admin(self):

		box = BoxLayout(orientation='vertical', size_hint_y=None, height=160)
		
		title_box = BoxLayout(size_hint_y=None, height=50)
		box.add_widget(title_box)

		title = Label(text='Admin Login', font_size=30)
		title_box.add_widget(title)

		admin_btn = Button(text='?', font_size=30, size_hint_x=None, width=50)
		admin_btn.bind(on_press=lambda i:self.command_help('Admin'))
		title_box.add_widget(admin_btn)

		heading = BoxLayout(size_hint_y=None, height=40)
		box.add_widget(heading)

		heading.add_widget(Label(text=self.organisation, font_size=20))
		heading.add_widget(Label(text='Password', font_size=20))
		heading.add_widget(Label(text='Confirm', font_size=20))

		body = BoxLayout(size_hint_y=None, height=70)
		box.add_widget(body)

		user = LabelWrap(text=self.title[0], font_size=20)
		body.add_widget(user)

		pw = TextInputC(multiline=False)
		body.add_widget(pw)

		btn = Button(text='Enter')
		btn.bind(on_press=lambda i, pw=pw:self.command_admin(pw))
		body.add_widget(btn)

		self.add_widget(box)

	def build_tournament(self):
		
		tournament_dict = {}
		new_list = []

		try:
			tournament_list = self.db.select(table='tournaments')

			if len(tournament_list) > 0:

				for tournament in tournament_list:

					db = Database(tournament[0])
					stage = db.cursor.execute('''SELECT stage 
										FROM details''').fetchone()[0]
					details = db.cursor.execute('''SELECT role, category, password
											FROM settings
											WHERE active=1
											AND role="Public"''').fetchall()


					if stage not in ['Competition', 'Complete', 'Archive']:
						d1 = db.cursor.execute('''SELECT role, category, password
											FROM settings
											WHERE active=1
											AND role="Captain"''').fetchall()
						details += d1

					else:
						d1 = db.cursor.execute('''SELECT role, category, password
											FROM settings
											WHERE active=1
											AND role="Official"''').fetchall()

						details += d1						

					if details != []:
						tournament_dict[tournament[0]] = details
						new_list.append(tournament[0])

		except sqlite3.OperationalError:
		
			pass

		if tournament_dict != {}:
			mes = "Tournament Login is used to access those tournaments that have Public/Captain/Official access made available\n\nPublic Access only lets you see the tournament main page, which during Competition stage is the Scoreboard.\n\nCaptain access is for the Club/House/School leaders to enter their competitors in events.\n\nOfficial access is for the Marshals\Recorders to run the events"
			frame = FrameLogin(t_list=new_list, t_dict=tournament_dict,
					add_help=mes)
			self.add_widget(frame)
			self.height = 470

	def build_register(self):

		box = BoxLayout(orientation='vertical', size_hint_y=None, height=160)
		
		title_box = BoxLayout(size_hint_y=None, height=50)
		box.add_widget(title_box)

		title = Label(text='Registration Expired', font_size=30)
		title_box.add_widget(title)

		admin_btn = Button(text='?', font_size=30, size_hint_x=None, width=50)
		admin_btn.bind(on_press=lambda i:self.command_help('Register'))
		title_box.add_widget(admin_btn)

		heading = BoxLayout(size_hint_y=None, height=40)
		box.add_widget(heading)

		heading.add_widget(Label(text=self.organisation, font_size=20))
		heading.add_widget(Label(text='New Registration', font_size=20))
		heading.add_widget(Label(text='Confirm', font_size=20))

		body = BoxLayout(size_hint_y=None, height=70)
		box.add_widget(body)

		user = LabelWrap(text=self.title[0], font_size=20)
		body.add_widget(user)

		pw = TextInputC(multiline=False)
		body.add_widget(pw)

		btn = Button(text='Enter')
		btn.bind(on_press=lambda i, pw=pw:self.command_reregister(pw))
		body.add_widget(btn)

		self.add_widget(box)

	def build_main(self):

		## Settings
		box = BoxLayout(orientation='vertical',
						size_hint_y=None, height=160)

		title_box = BoxLayout(size_hint_y=None, height=50)
		box.add_widget(title_box)

		title = Label(text='Registration', font_size=30)
		title_box.add_widget(title)

		btn = Button(text='?', font_size=30, size_hint_x=None, width=50)
		btn.bind(on_press=lambda i, :self.command_help('Main'))
		title_box.add_widget(btn)

		heading = BoxLayout(size_hint_y=None, height=40)
		box.add_widget(heading)

		heading.add_widget(Label(text='Username', font_size=20))
		heading.add_widget(Label(text='Registration Key', font_size=20))
		heading.add_widget(Label(text='Confirm', font_size=20,
									size_hint_x=None, width=100))

		body = BoxLayout(size_hint_y=None, height=70)
		box.add_widget(body)

		self.username = TextInputC(multiline=False)
		self.reg_key = TextInputC(multiline=False)

		btn = Button(text='Enter', font_size=14, size_hint_x=None, width=100)
		btn.bind(on_press=self.command_register)

		body.add_widget(self.username)
		body.add_widget(self.reg_key)
		body.add_widget(btn)

		self.add_widget(box)

	def build_advert(self):

		self.height += 150
		self.add_widget(AdvertMainFooter())

class PageMain(GridLayout):

	tournament = StringProperty('main')
	motto = StringProperty('')
	cols = 1

	def __init__(self, *args, **kwargs):
		super(PageMain, self).__init__(**kwargs)

		## Settings
		self.db = Database(self.tournament)
		details = self.db.cursor.execute('''SELECT title
										FROM details''').fetchone()
		self.title = details[0]

		if self.tournament != 'main':
			self.stage = self.db.cursor.execute('''SELECT stage
										FROM details''').fetchone()[0]
		else:
			self.stage = None

		## Database
		self.display = None

		## Display
		if self.tournament == 'main':
			self.build_main_header()
		else:
			self.build_tournament_header()
		self.build_body()
		self.build_footer()

	### FUNCTIONS ###

	def access_settings(self):

		settings = self.db.cursor.execute('''SELECT bg_colour, text_colour,
											crest, motto 
											FROM settings''').fetchone()

		return settings

	### BUILD ###

	def build_main_header(self):
		
		## Settings
		settings = self.access_settings()
		self.bg_colour = db_colour(settings[0], db_in=False)
		self.text_colour = db_colour(settings[1], db_in=False)
		self.crest_source = settings[2] # prob need more for this
		self.motto = settings[3]

		## Display
		box = BoxLayout(padding=10, size_hint_y=None, height=150)

		if True == True:
			crest_box = BoxLayout(size_hint_x=None, width=150)
			box.add_widget(crest_box)
			crest = Image(source=resource_path('sponsor_wristband.png'), allow_stretch=True)
			crest_box.add_widget(crest)

			title = BoxLayout(orientation='vertical', padding=[5, 0, 5, 0], 
								spacing=10)
			t = LabelWrapC(text=self.title, font_size=50, bg=self.bg_colour,
								color=self.text_colour)
			title.add_widget(t)

		else:
			title = BoxLayout(orientation='vertical', spacing=10)

			t = LabelWrapC(text=self.title, font_size=50, bg=self.bg_colour,
								color=self.text_colour)
			title.add_widget(t)
		
		if self.motto not in ['', None]:
			motto = LabelC(text=self.motto, size_hint_y=None, height=30)
			title.add_widget(motto)

		box.add_widget(title)

		self.add_widget(box)

	def build_tournament_header(self):

		## Settings
		if self.motto == '':
			main_title = Database('main').cursor.execute('''SELECT title
												FROM details''').fetchone()[0]
		else:
			main_title = self.motto

		self.logo_source = Database('main').cursor.execute('''SELECT crest
												FROM settings''').fetchone()[0]

		## Display
		box = BoxLayout(padding=10, size_hint_y=None, height=150)

		if True == True:
			logo_box = BoxLayout(size_hint_x=None, width=150)
			box.add_widget(logo_box)
			logo = Image(source=resource_path('sponsor_wristband.png'), allow_stretch=True)
			logo_box.add_widget(logo)

			title = BoxLayout(orientation='vertical', padding=[5, 0, 5, 0], 
								spacing=10)
			t = LabelWrapC(text=self.title, font_size=50)
			title.add_widget(t)

		else:
			title = BoxLayout(orientation='vertical', spacing=10)

			t = LabelWrapC(text=self.title, font_size=50)
			title.add_widget(t)

		main = LabelC(text=main_title, size_hint_y=None, height=30)
		title.add_widget(main)

		box.add_widget(title)

		self.add_widget(box)

	def build_body(self):

		self.body = GridLayout(cols=1, padding=[10, 10, 10, 0])
		box = GridLayout(cols=1, padding=40)
	
		if self.stage == 'Template':

			self.body.add_widget(box)
			inst = LabelWrap(text="This is the Template stage.\n\nTo start building or to check/amend this tournament template, click the 'Template' button on the sidebar to the left, and then click 'Creation'.",
				font_size=20, size_hint_y=None, height=150)
			box.add_widget(inst)

		elif self.stage == 'Entry - Competitors':

			self.body.add_widget(box)
			inst = LabelWrap(text="This is the Entry - Competitors stage.\n\nIn this stage you will enter all the competitor details, you can do this by clicking 'Entry' in the sidebar and then 'Competitors.\n\nIf you would like to continue adding to the Event Records/Program or to change Restrictions/Points you can do this from the 'Template' menu on the sidebar",
				font_size=20, size_hint_y=None, height=200)
			box.add_widget(inst)
		
		elif self.stage == 'Entry - Events':

			self.type = self.db.cursor.execute('''SELECT type
									FROM details''').fetchone()[0]

			self.body.add_widget(box)
			inst = LabelWrap(text="This is the Entry - Events stage.\n\nNow that all the competitors have been placed into their %ss and Grades, you can select which events they want to compete in\n\nOnce you are done, click 'Complete' in the 'Entry' sidebar menu and follow the instructions." % self.type,
				font_size=20, size_hint_y=None, height=200)
			box.add_widget(inst)

		self.add_widget(self.body)

	def build_footer(self):

		return
		box = AdvertMainFooter()
		self.add_widget(box)

	### DISPLAY ###

	def add_display(self, display):

		if self.display != None:
			self.body.remove_widget(self.display)

		self.body.add_widget(display)
		self.display = display

class PageSaveCheck(Page):

	tournament = StringProperty('')
	style = StringProperty('')

	def __init__(self, *args, **kwargs):
		super(PageSaveCheck, self).__init__(**kwargs)

		## Display
		self.build_body()
		self.build_frame()
		self.sheet.add_widget(self.frame)

	### FUNCTIONS ###

	def command_save(self, *args):

		if self.style == 'points':
			self.frame.process_points()
		elif self.style == 'restrictions':
			self.frame.process_restrictions()
		elif self.style == 'records':
			self.frame.process_records()
		elif self.style == 'program':
			self.frame.process_program()
		elif self.style == 'special':
			self.frame.process_special()

	def command_unsave(self, boole):

		if boole == True:
			self.save_label.text = ''
			self.save_label.bg = [0.5, 0.5, 0.5, 1]

		elif boole == False:
			self.save_label.text = 'Unsaved'
			self.save_label.bg = [1, 0, 0, 1]

	### BUILD ###

	def build_body(self):

		box = GridLayout(cols=1)

		self.sheet = GridLayout(cols=1, padding=10)
		box.add_widget(self.sheet)

		action = BoxLayout(size_hint_y=None, height=50)
		box.add_widget(action)

		self.save_label = LabelC(bg=[0.5, 0.5, 0.5, 1], 
								size_hint_x=None, width=100)
		action.add_widget(self.save_label)

		self.save_btn = Button(text='Save', font_size=30)
		self.save_btn.bind(on_press=self.command_save)
		action.add_widget(self.save_btn)


		if self.style == 'archive':
			box.remove_widget(action)

		self.add_display(box)

	def build_frame(self):

		if self.style in ['points', 'restrictions', 'special']:
			self.frame = FrameAddMinus(tournament=self.tournament, 
									style=self.style,
									unsave=self.command_unsave)
		elif self.style in ['records', 'program', 'archive']:
			self.frame = FrameTab(tournament=self.tournament, 
									style=self.style,
									unsave=self.command_unsave)

class PageSetting(Page):

	tournament = StringProperty('main')

	def __init__(self, *args, **kwargs):
		super(PageSetting, self).__init__(**kwargs)

		## Settings
		self.db = Database(self.tournament)
		self.bind(size=self.change_scroll_height)
		
		## Database


		## Display
		self.build_grid()
		if self.tournament == 'main':
			self.build_main_details()
			self.build_main_settings()

		else:
			self.build_tournament_access()

	### FUNCTIONS ###

	def change_scroll_height(self, obj, value):

		self.scroll.height = self.height - 100

	def command_access(self, i, ref):

		role = ref.text.split(' - ')[0]
		category = ref.text.split(' - ')[1]

		if i.text == 'On':
		
			self.db.cursor.execute('''UPDATE settings
						SET active=%s
						WHERE role="%s" 
						AND category="%s"''' % (0, role, category))
			self.db.connection.commit()
			i.text = 'Off'

		elif i.text == 'Off':

			self.db.cursor.execute('''UPDATE settings
						SET active=%s
						WHERE role="%s" 
						AND category="%s"''' % (1, role, category))
			self.db.connection.commit()
			i.text = 'On'

	def command_crest(self, boole, file):

		if boole == True:
			self.crest_box.clear_widgets()
			img = Image(source=file, allow_stretch=True)
			self.crest_box.add_widget(img)

	def command_main(self, style, ref):

		if style == 'password':

			self.db.cursor.execute('''UPDATE settings
									SET cipher="%s"''' % ref.text)

		elif style == 'crest':

			FileDialog(function=self.command_crest)

		elif style == 'motto':

			self.db.cursor.execute('''UPDATE settings
									SET motto="%s"''' % ref.text)


		self.db.connection.commit()

	def command_password(self, ref1, ref2):

		role = ref1.text
		password = ref2.text

	### BUILD ###

	def build_grid(self):

		box = GridLayout(cols=1, padding=10)
		self.add_display(box)

		header = BoxLayout(size_hint_y=None, height=40)
		box.add_widget(header)

		if self.tournament == 'main':
			header.add_widget(Label(text='Feature', font_size=20))
			header.add_widget(Label(text='Setting', font_size=20))
		else:
			header.add_widget(Label(text='Role', font_size=20))
			header.add_widget(Label(text='Access', font_size=20,
										size_hint_x=None, width=100))
			header.add_widget(Label(text='Password', font_size=20))
			header.add_widget(Label(text='Save', font_size=20,
										size_hint_x=None, width=100))

		self.scroll = ScrollView(size_hint_y=None, height=0)
		box.add_widget(self.scroll)

		self.grid = GridLayout(cols=1, size_hint_y=None, height=0)
		self.scroll.add_widget(self.grid)

	def build_main_details(self):

		## Settings
		self.grid.height += 250

		details = self.db.cursor.execute("SELECT * FROM details").fetchone()

		title = details[0]
		self.organisation = details[1]
		regkey = details[2]
		
		settings = self.db.cursor.execute('''SELECT address, email
											 FROM settings''').fetchone()


		address = settings[0]
		email = settings[1]

		## Display
		box = GridLayout(cols=2, size_hint_y=None, height=250)

		detail_list = ['Organisation', title, 'Organisation Type',
						self.organisation, 'Registration Key', regkey,
						'Address', address, 'Email', email]

		for d in detail_list:
			box.add_widget(Label(text=d))
	
		## Engine
		self.grid.add_widget(box)

	def build_main_settings(self):

		## Settings
		self.grid.height += 300

		settings = self.db.cursor.execute("SELECT * FROM settings").fetchone()

		crest = settings[2]
		motto = settings[3]
		password = settings[4]

		## Display
		box = GridLayout(cols=1, size_hint_y=None, height=250)

		# password
		b1 = BoxLayout(size_hint_y=None, height=50)
		b1.add_widget(Label(text='Admin Password'))

		a1 = BoxLayout()
		b1.add_widget(a1)

		pw = TextInputC(text=password)
		a1.add_widget(pw)

		btn1 = Button(text='Save', size_hint_x=None, width=100)
		btn1.bind(on_press=lambda i, pw=pw:self.command_main('password', pw))
		a1.add_widget(btn1)

		box.add_widget(b1)

		# crest - removed

		# motto
		b3 = BoxLayout(size_hint_y=None, height=50)
		b3.add_widget(Label(text='%s Motto' % self.organisation))

		a3 = BoxLayout()
		b3.add_widget(a3)

		mt = TextInputC(text=motto)
		a3.add_widget(mt)

		btn3 = Button(text='Save', size_hint_x=None, width=100)
		btn3.bind(on_press=lambda i, mt=mt:self.command_main('motto', mt))
		a3.add_widget(btn3)

		box.add_widget(b3)

		## Engine
		self.grid.add_widget(box)

	def build_tournament_access(self):
		
		settings = self.db.cursor.execute('''SELECT role, category, 
													active, password
											FROM settings''').fetchall()

		if settings != []:
			for tup in sorted(settings):

				title = '%s - %s' % (tup[0], tup[1])

				self.grid.add_widget(self.build_access_box(title, 
															tup[2], tup[3]))
				self.grid.height += 50

	def build_access_box(self, title, access, password):

		## Settings
		if access == 1:
			
			access = 'On'
		else:
			
			access = 'Off'

		## Display
		box = BoxLayout(size_hint_y=None, height=50)

		role = Label(text=title)
		box.add_widget(role)

		btn1 = Button(text=access, size_hint_x=None, width=100)
		btn1.bind(on_press=lambda i, role=role:self.command_access(i, role))
		box.add_widget(btn1)

		pw = TextInputC(text=password)
		box.add_widget(pw)

		btn2 = Button(text='Enter', size_hint_x=None, width=100)
		btn2.bind(on_press=lambda i, 
				pw=pw, role=role:self.command_password(role, pw))
		box.add_widget(btn2)

		## Engine
		return box

class PageTemplateCreate(GridLayout):

	tournament = StringProperty('Example')
	example = BooleanProperty(False)

	def __init__(self, *args, **kwargs):
		super(PageTemplateCreate, self).__init__(**kwargs)

		## Settings
		self.db = Database(self.tournament)
		details = self.db.select(table='details')[0]
		self.type = details[3]
		self.age = details[4]

		## Database
		self.frame = None

		## Display
		self.build_type()

	### FUNCTIONS ###

	def command_action(self, i):

		if self.ids.title.text == '%ss' % self.type:

			if i.text == 'Save & Continue':
				action = self.frame.process_type()	

				if action == None:
					PopBox().showwarning(title='Error', 
					message='%s groups each require a unique Title\nTitles must only be alphanumeric, cannot start with a number and must be at least 3 characters long' % self.type)
					return

				self.build_gender()

		elif self.ids.title.text == 'Gender Grades':

			if i.text == 'Back':
				self.build_type()
			else:
				action = self.frame.process_gender()
				if action == None:
					PopBox().showwarning(title='Error', 
				message='Gender groups each require a unique Title & Gender\nTitles must only be alphanumeric and cannot start with a number')
					return
				if self.age == 'Date of Birth':
					self.build_age_dob()
				else:
					self.build_age()			

		elif self.ids.title.text == 'Age Grades - Date of Birth':

			if i.text == 'Back':
				self.build_gender()
			else:
				action = self.frame.process_age_dob()
				if action == None:
					PopBox().showwarning(title='Error', 
						message='All Age Groups fields must be entered\nTitles must only be alphanumeric and cannot start with a number')
					return
				self.build_contests()

		elif self.ids.title.text == 'Age Grades - %s' % self.age:

			if i.text == 'Back':
				self.build_gender()
			else:
				action = self.frame.process_age()
				if action == None:
					PopBox().showwarning(title='Error', 
						message='Age Groups require a unique Title\nTitles must only be alphanumeric and cannot start with a number')
					return
				self.build_contests()

		elif self.ids.title.text == 'Select Standard Contests':

			if i.text == 'Back':
				if self.age == 'Date of Birth':
					self.build_age_dob()
				else:
					self.build_age()

			elif i.text == 'Save & Continue':

				action = self.frame.process_contest()

				self.build_add_contest()

		elif self.ids.title.text == 'Add Contests':

			if i.text == 'Back':
				self.build_contests()

			elif i.text == 'Save & Continue':

				action = self.frame.process_contest()
				if action == None:
					PopBox().showwarning(title='Error', 
					message='All contest fields must be complete and selected\nTitles must only be alphanumeric and cannot start with a number')
					return
				elif action == False:
					PopBox().showwarning(title='Error', 
					message="Distance contests can't be Team events")
					return					
				self.build_events()

		elif self.ids.title.text == 'Select Events':

			if i.text == 'Back':
				self.build_add_contest()

			elif i.text == 'Save & Continue':

				action = self.frame.process_events()
				if action == None:
					PopBox().showwarning(title='Error', 
						message='No Events selected')
					return
				self.build_points()

		elif self.ids.title.text == 'Tournament Points':

			if i.text == 'Back':
				self.build_events()

			elif i.text == 'Save & Continue':

				action = self.frame.process_points()
				if action == None:
					PopBox().showwarning(title='Error', 
						message='All fields require a value')
					return
				self.build_restrictions()

		elif self.ids.title.text == 'Tournament Entry Restrictions':

			if i.text == 'Back':
				self.build_points()

			elif i.text == 'Save & Continue':
				action = self.frame.process_restrictions()
				self.build_records()

		elif self.ids.title.text == 'Event Records':

			if i.text == 'Back':
				self.build_restrictions()

			elif i.text == 'Save & Continue':
				action = self.frame.process_records()				

				self.build_program()

		elif self.ids.title.text == 'Event Program':
			
			if i.text == 'Back':
				self.build_records()

			elif i.text == 'Save & Continue':
				action = self.frame.process_program()				

				self.build_template_complete()

		elif self.ids.title.text == 'Template Complete':

			if i.text == 'Back':
				self.build_program()

	def command_help(self, *args):

		page = self.ids.title.text # determins which help screen to open

		if page == '%ss' % self.type:

			mes = "Here you say which %ss will be competing in this tournament.\nEach %s has to have a unique name, and if you want you can choose the %s colours.\nBe aware if you do not choose a %s colour here, you cannot add one after Template Stage" % (self.type, self.type, self.type, self.type)

		elif page == 'Gender Grades':
		
			mes = "Choose the title for your gender grades.\nIf this is a single-sex tournament, a gender still has to be chosen\n\ni.e. Boys = Male, Girls = Female"

		elif page == 'Age Grades - Date of Birth':

			mes = "Create your age grades with a unique title, and choose if that grade is under or over a certain age on a certain date.\n'Over' is technically 'Over or Equal to'\nHint: If you only have one cut off Date, enter that date and then add a row, this will save you having to reenter it each time"

		elif page == 'Age Grades - Year':

			mes = "Create a title for each Year based age grade"

		elif page == 'Select Standard Contests':

			mes = "These are some of the standard contests in this sport.\nSelect which ones will be run in this tournament\nIf you have a contest that isn't here, you can create it on the next page."

		elif page == "Add Contests":

			mes = "Add a contest by entering a unique Contest title and choosing how it is measured i.e. by Time or by Distance and then by Indivdual/Team Competition\nDistance contests cannot be Team Competitions\nIf you change your mind, either unselect that contest or use the - button"

		elif page == "Select Events":

			mes = "Choose which grades each contest will be run in, this combination creates Events"

		elif page == 'Tournament Points':

			mes = "This page decides how competitors and their %ss are awarded points for participation and placing.\nPlacing or Competition Points are limited to 10th place.\nTeam Participation Points are usually 0\nNote: This Page will always be available for you to change" % self.type

		elif page == 'Tournament Entry Restrictions':

			mes = "Here you can decide if there is a minimum requirement or maximum limit to how many events a competitor can enter.\nNote: This Page will always be available for you to change"

		elif page == "Event Records":

			mes = "If you have records, enter them here.\nThis template can be updated after your tournament and then next year if you copy this template it will have the records already stored.\nNote: This Page will always be available for you to change"

		elif page == "Event Program":

			mes = "If you have a set time/number plan for your tournament, enter those details here.\nFor events that will have Finals or for multi-day tournaments, those program details can be added in the Competition Stage\nNote: This Page will always be available for you to change"

		elif page == "Template Complete":

			mes = "This is the final page in the Template stage.\nOnce you press 'Save and Continue' you will not be able to change %ss or Grades, or add Contests or Events" % self.type

		PopBox().help_screen(title=page, message=mes)

	def command_stage_entry(self, *args):

		# update template_grades

		gender_list = calculate_grade_list(self.tournament, gender=True)
		age_list = calculate_grade_list(self.tournament, age=True)

		var = []
		for g in gender_list:
			var.append((g, 'gender', g, ''))
			for a in age_list:
				cg = "%s %s" % (a, g)
				var.append((cg, 'combined', g, a))
		for a in age_list:
			var.append((a, 'age', '', a))

		command = '''INSERT INTO template_grades
					VALUES(?, ?, ?, ?)'''

		self.db.cursor.executemany(command, var)
		self.db.connection.commit

		# update stage
		self.db.cursor.execute('''UPDATE details
								SET stage="Entry - Competitors"
								''')
		self.db.connection.commit()
		self.db.connection.close()

		zotournament = ProgramTournament(title=self.tournament) 

	### BUILD ###

	def build_type(self):

		## Settings
		self.ids.title.text = '%ss' % self.type
		self.ids.left.text = ''
		self.ids.right.text = 'Save & Continue'

		## Display
		frame = FrameAddMinus(tournament=self.tournament, style='type')

		## Engine
		self.change_frame(frame)

	def build_gender(self):

		## Settings
		self.ids.title.text = 'Gender Grades'
		self.ids.left.text = 'Back'
		self.ids.right.text = 'Save & Continue'

		## Display
		frame = FrameAddMinus(tournament=self.tournament, style='gender')

		## Engine
		self.change_frame(frame)

	def build_age_dob(self):

		## Settings
		self.ids.title.text = 'Age Grades - Date of Birth'
		self.ids.left.text = 'Back'
		self.ids.right.text = 'Save & Continue'

		## Display
		frame = FrameAddMinus(tournament=self.tournament, style='age dob')

		## Engine
		self.change_frame(frame)

	def build_age(self):

		## Settings
		self.ids.title.text = 'Age Grades - %s' % self.age
		self.ids.left.text = 'Back'
		self.ids.right.text = 'Save & Continue'

		## Display
		frame = FrameAddMinus(tournament=self.tournament, style='age')

		## Engine
		self.change_frame(frame)

	def build_contests(self):

		## Settings
		self.ids.title.text = 'Select Standard Contests'
		self.ids.left.text = 'Back'
		self.ids.right.text = 'Save & Continue'

		## Display
		frame = FrameSelect(tournament=self.tournament, style='contest')

		## Engine
		self.change_frame(frame)

	def build_add_contest(self):
		
		## Settings
		self.ids.title.text = 'Add Contests'
		self.ids.left.text = 'Back'
		self.ids.right.text = 'Save & Continue'

		## Display
		frame = FrameAddMinus(tournament=self.tournament, style='contest')

		## Engine
		self.change_frame(frame)

	def build_events(self):

		## Settings
		self.ids.title.text = 'Select Events'
		self.ids.left.text = 'Back'
		self.ids.right.text = 'Save & Continue'

		## Display
		frame = FrameSelect(tournament=self.tournament, style='events')

		## Engine
		self.change_frame(frame)

	def build_points(self):

		## Settings
		self.ids.title.text = 'Tournament Points'
		self.ids.left.text = 'Back'
		self.ids.right.text = 'Save & Continue'

		## Display
		frame = FrameAddMinus(tournament=self.tournament, style='points')

		## Engine
		self.change_frame(frame)

	def build_restrictions(self):

		## Settings
		self.ids.title.text = 'Tournament Entry Restrictions'
		self.ids.left.text = 'Back'
		self.ids.right.text = 'Save & Continue'

		## Display
		frame = FrameAddMinus(tournament=self.tournament, style='restrictions')

		## Engine
		self.change_frame(frame)

	def build_records(self):

		## Settings
		self.ids.title.text = 'Event Records'
		self.ids.left.text = 'Back'
		self.ids.right.text = 'Save & Continue'

		## Display
		frame = FrameTab(tournament=self.tournament, style='records')

		## Engine
		self.change_frame(frame)

	def build_program(self):

		## Settings
		self.ids.title.text = 'Event Program'
		self.ids.left.text = 'Back'
		self.ids.right.text = 'Save & Continue'

		## Display
		frame = FrameTab(tournament=self.tournament, style='program')

		## Engine
		self.change_frame(frame)

	def build_template_complete(self):

		## Settings
		self.ids.title.text = 'Template Complete'
		self.ids.left.text = 'Back'
		self.ids.right.text = ''

		## Display
		frame = GridLayout(cols=1, padding=[0, 20, 0, 0],
							spacing=20)

		info = '''This completes the Template section.\n\nYou will still be able to alter the Event Points, Event Restrictions, Event Records and Event Number & Times in later sections.\n\nPress Save & Continue if you ready to move to the Entry section.'''

		info_label = LabelWrap(text=info, font_size=20, size_hint_y=None,
								height=200)
		frame.add_widget(info_label)

		box = BoxLayout(size_hint_y=None, height=150)
		box.add_widget(BoxLayout(size_hint_x=0.2))
		btn = Button(text='Save & Continue', font_size=20)
		btn.bind(on_press=self.command_stage_entry)
		box.add_widget(btn)
		box.add_widget(BoxLayout(size_hint_x=0.2))	
		frame.add_widget(box)	

		## Engine
		self.change_frame(frame)

	### FRAME ###

	def change_frame(self, frame):

		if self.frame != None:
			self.ids.display.remove_widget(self.frame)

		self.ids.display.add_widget(frame)
		self.frame = frame

class PageTournamentCreate(Page):

	def __init__(self, *args, **kwargs):
		super(PageTournamentCreate, self).__init__(**kwargs)

		## Settings
		self.body.padding = 10
		self.db = Database('main')

		## Database
		self.menu = []
		self.tournament_list = make_tournament_list(copy=True)
		## Display

		## Engine
		self.build_display()
		self.build_template()

	### FUNCTIONS ###

	def command_engine(self, ribbon):

		title = ribbon.title
		answer = ribbon.answer.text

		if answer in ['Select', '']: # no entry made
			return
		if self.menu[-1] != ribbon: # not the last ribbon
			return
		if ribbon.reset == True: # change in answer would change flowchart
			ribbon.answer.readonly = True
		elif ribbon.reset == False:
			if len(answer) > 19:
				PopBox().showwarning(title='Error',
					message='Text needs to be less than 20 characters')
				return
			elif len(answer) < 4:
				PopBox().showwarning(title='Error',
					message='Text needs to be greater than 3 characters')
				return
			else:
				try:
					ribbon.answer.option.readonly = True
				except AttributeError:
					pass

		if title == 'Template':
			if answer in self.tournament_list:
				c = "SELECT sport FROM tournaments WHERE title='%s';" % answer
				sport = self.db.cursor.execute(c).fetchone()[0]
				self.build_sport(sport=sport)
			else:
				self.build_sport()

		elif title == 'Sport':
			if ribbon.answer.main == 'Select': # create new template
				self.build_inter()
			else:
				self.build_title()

		elif title == 'Inter -':
			self.build_age()

		elif title == 'Age Grades':
			self.build_title()

		elif title == 'Title':
			
			if answer in self.tournament_list:
				PopBox().showwarning(title='Error',
					message='New tournament must have a unique name, %s already exists' % answer)
				return

			if self.menu[0].answer.text == 'Create New Template':
				template = self.menu[0].answer.text.title()
				sport = self.menu[1].answer.text
				inter = self.menu[2].answer.text
				age = self.menu[3].answer.text
				title = answer

				message = 'Template: %s\nSport: %s\nCompetition: Inter-%s\nAge Grades: %s\nTitle: %s\n\nConfirm:' % (template,
					sport, inter, age, title)
				
				PopBox().askyesno(function=self.create_template,
					title='Create New Template', message=message)
			else:
				template = self.menu[0].answer.text
				sport = self.menu[1].answer.text
				title = answer.title()

				message = 'Template: %s\nSport: %s\nTitle: %s\n\nConfirm:' % (template,
					sport, title)

				PopBox().askyesno(function=self.copy_template,
					title='Copy %s Template' % template, message=message)

	def command_reset(self, *args):

		for rbn in self.menu:
			self.ribbon_sheet.remove_widget(rbn)
		self.menu = self.menu
		self.ribbon_sheet.height = 0
		self.build_template()

	def command_scroll_height(self, obj, value):

		self.scroll.height = self.body.height

	### BUILD ###

	def build_display(self):

		self.scroll = ScrollView(size_hint_y=None, height=self.body.height)
		self.body.bind(height=self.command_scroll_height)
		self.add_display(self.scroll)
		self.ribbon_sheet = GridLayout(cols=1, spacing=5, 
										size_hint_y=None, height=0)
		self.scroll.add_widget(self.ribbon_sheet)

	def build_ribbon(self, *args, **kwargs):

		title = kwargs.pop('title', '')
		question = kwargs.pop('question', '')
		answer = kwargs.pop('answer', None)
		reset = kwargs.pop('reset', None)
		help = kwargs.pop('help', '')
		reset_position = kwargs.pop('reset_position', None)

		rbn = Ribbon(title=title, question=question, answer=answer,
						reset=reset, help=help, function=self.command_engine,
						reset_position=self.reset_position)

		self.menu.append(rbn)
		self.ribbon_sheet.add_widget(rbn)
		self.ribbon_sheet.height += 100

	def build_template(self):

		self.ribbon_sheet.height = 150

		question = 'Choose Template'
		options = ['Create New Template'] + self.tournament_list
		answer = OptionMenu(options=options)
		help = 'Copy a template from another tournament, or create a new one'

		self.build_ribbon(title='Template', question=question, 
					answer=answer, reset=True, help=help)

	def build_sport(self, **kwargs):

		sport = kwargs.pop('sport', None)

		question = 'Choose Sport'
		if sport == None:
			answer = OptionMenu(options=['Athletics', 'Swimming'])
			help = 'Choose the type of sport that the tournament will run'
		else:
			answer = OptionMenu(options=['Athletics', 'Swimming'],
								main=sport)
			help = 'Choose the same sport, or keep the general template but use another sport'

		self.build_ribbon(title='Sport', question=question, 
					answer=answer, reset=True, help=help)

	def build_inter(self):

		question = 'Choose Competition'
		answer = OptionMenu(options=['Club', 'House', 'School'])
		help = 'Choose the type of competition, i.e. inter-house, inter-club'

		self.build_ribbon(title='Inter -', question=question, answer=answer,
					help=help, reset=True)

	def build_age(self):

		question = 'Choose Age Grade'
		answer = OptionMenu(options=['Date of Birth', 'Year'])
		help = 'Choose how age grades will be decided, i.e. by date of birth, year-group etc'
		
		self.build_ribbon(title='Age Grades', question=question, help=help,
					answer=answer, reset=False)

	def build_title(self):

		question = 'Choose Tournament Title'
		answer = TextInputC()
		help = 'Choose a unique name for your tournament i.e. Athletics Day 2017'

		self.build_ribbon(title='Title', question=question, answer=answer,
					help=help, reset=False)

	### PROCESS ###

	def reset_position(self, ribbon):
		'''
		When the reset alters the flowchart, 
		this removes ribbons placed after the argument ribbon
		'''
		index = self.menu.index(ribbon)
		remove_list = self.menu[index+1:]
		for rbn in remove_list:
			self.ribbon_sheet.remove_widget(rbn)
			self.ribbon_sheet.height -= 120
		self.menu = self.menu[:index+1]
		if ribbon.reset == True:
			ribbon.answer.readonly = False
		elif ribbon.reset == False:
			ribbon.answer.option.readonly = False

	def create_template(self, boole):

		## Settings
		if boole == False:
			return
		
		template = self.menu[0].answer.text.title()
		sport = self.menu[1].answer.text
		inter = self.menu[2].answer.text
		age = self.menu[3].answer.text
		title = self.menu[4].answer.text.title()

		## PROCESS

		# insert reference into main.db - tournaments table
		self.db = Database('main')
		c = '''INSERT INTO tournaments 
				VALUES('%s', '%s');''' % (title, sport)
		self.db.cursor.execute(c)
		self.db.connection.commit()
		self.db.connection.close()

		# reset the self.db variable and also open a new database
		self.db = Database('%s' % title)

		# table - details
		c1 = ['''CREATE TABLE details
				(title TEXT, sport TEXT, stage TEXT, type TEXT, age TEXT);''',
			'''INSERT INTO details 
				VALUES('%s','%s','Template', '%s', '%s');''' % (title, sport,
																inter, age),
			'''CREATE TABLE settings
				(role TEXT, active INT, category TEXT, password TEXT);''']
		for command in c1:
			self.db.cursor.execute(command)
		self.db.connection.commit() 

		# table - settings
		self.db.cursor.execute('''INSERT INTO settings
								VALUES('Public', 0, 'All', '')''')
		self.db.connection.commit()

		# table - template_groups
		var1 = "title TEXT PRIMARY KEY"
		var2 = "category TEXT"
		var3 = "filter TEXT" # gender i.e. Male/Female 
		var4 = "under TEXT" # dob i.e. Under/Over
		var5 = "year INT" # dob 
		var5a = "date TEXT" # dob i.e. db_date
		var6 = "colour_bg TEXT" # have to convert to str then back to list
		var7 = "colour_text TEXT" # as about

		c2 = '''CREATE TABLE template_groups
				(%s,%s,%s,%s,%s,%s,%s,%s);''' % (var1, var2, var3, 
												var4, var5, var5a, var6, var7)
		self.db.cursor.execute(c2)
		self.db.connection.commit()

		c4 = '''INSERT INTO template_groups 
				(title, category) 
				VALUES('Open', 'age');'''
		self.db.cursor.execute(c4)
		self.db.connection.commit()

		# table - template_contests
		c5 = '''CREATE TABLE template_contests
				(title TEXT PRIMARY KEY, measure TEXT, measure_var TEXT,
				team INT, standard INT, selection INT);'''
		self.db.cursor.execute(c5)
		var9 = zosportsdb[sport]
		c6 = '''INSERT INTO template_contests
				VALUES(?, ?, ?, ?, ?, 0);'''
		self.db.cursor.executemany(c6, var9)
		self.db.connection.commit()

		# table - template_grades
		c7 = '''CREATE TABLE template_grades
				(title TEXT PRIMARY KEY, category TEXT, 
				gender_grade TEXT, age_grade TEXT);'''
		self.db.cursor.execute(c7)
		self.db.connection.commit()

		# table - template_events
		# events are either in or out, if admin change, all saved info is lost
		c8 = '''CREATE TABLE template_events
				(title TEXT PRIMARY KEY, contest TEXT, grade TEXT, 
				record_what REAL, record_who TEXT, 
				record_when INT, program_number INT, program_time TEXT);'''
		self.db.cursor.execute(c8)
		self.db.connection.commit()

		# table - template_points
		c9 = ''' CREATE TABLE template_points
				(rank PRIMARY KEY, individual INT, team INT)'''
		self.db.cursor.execute(c9)
		var10 = [('Participation', 1, 0), (1, 10, 20), (2, 7, 14), (3, 5, 10)]
		c10 = '''INSERT INTO template_points
				VALUES(?, ?, ?);'''
		self.db.cursor.executemany(c10, var10)
		self.db.connection.commit()

		# table - template_restrictions

		c11 = '''CREATE TABLE template_restrictions
				(restriction, value);'''
		self.db.cursor.execute(c11)

		if sport == 'Athletics':
			var11 = [('Tournament Minimum', ''), ('Tournament Maximum', 10), 
					('Track Event Maximum', 5), ('Field Event Maximum', 5)]
		else:
			var11 = [('Tournament Minimum', ''), ('Tournament Maximum', 10)]
		c12 = '''INSERT INTO template_restrictions
				VALUES(?, ?);'''
		self.db.cursor.executemany(c12, var11)
		self.db.connection.commit()
		self.db.connection.close()

		## Engine
		zotournament = ProgramTournament(title=title)

	def copy_template(self, boole):
	
		## Settings
		if boole == False:
			return

		template = self.menu[0].answer.text.title()
		sport = self.menu[1].answer.text
		title = self.menu[2].answer.text.title()

		## Database

		# open new db
		new_db = Database(title)

		# main tournament
		self.db = Database('main')
		c = '''INSERT INTO tournaments 
				VALUES('%s', '%s');''' % (title, sport)
		self.db.cursor.execute(c)
		self.db.connection.commit()


		# attach
		new_db.cursor.execute("ATTACH '%s.db' AS db1" % template)

		# details/settings
		new_db.cursor.execute('''CREATE TABLE details
								 AS SELECT * FROM db1.details''')
		new_db.cursor.execute('''UPDATE details 
						SET title="%s", stage="Template"'''% title)
		new_db.cursor.execute('''CREATE TABLE settings
				(role TEXT, active INT, 
				category TEXT, password TEXT);''')
		new_db.cursor.execute('''INSERT INTO settings
								VALUES('Public', 0, 'All', '')''')
		new_db.connection.commit()

		# groups
		new_db.cursor.execute('''CREATE TABLE template_groups
						AS SELECT * FROM db1.template_groups''')
		new_db.connection.commit()

		# points
		new_db.cursor.execute('''CREATE TABLE template_points
						AS SELECT * FROM db1.template_points''')
		new_db.connection.commit()				

		# grades - empty
		new_db.cursor.execute('''CREATE TABLE template_grades
							(title TEXT PRIMARY KEY, category TEXT, 
							gender_grade TEXT, age_grade TEXT);''')
		new_db.connection.commit()

		# check if same sport
		c = "SELECT sport FROM tournaments WHERE title='%s';" % template
		newsport = self.db.cursor.execute(c).fetchone()[0]
		self.db.connection.close()

		if sport == newsport: # i.e. both Athletics
			
			# table - template_contests
			new_db.cursor.execute('''CREATE TABLE template_contests
						AS SELECT * FROM db1.template_contests''')
			new_db.connection.commit()

			# table - template_events
			new_db.cursor.execute('''CREATE TABLE template_events
						AS SELECT * FROM db1.template_events''')
			new_db.connection.commit()

			# table - template_restrictions
			new_db.cursor.execute('''CREATE TABLE template_restrictions
						AS SELECT * FROM db1.template_restrictions''')
			new_db.connection.commit()

		else: # i.e template=Athletics, new=Swimming

			# table - template_contests
			new_db.cursor.execute('''CREATE TABLE template_contests
				(title TEXT PRIMARY KEY, measure TEXT, measure_var TEXT,
				team INT, standard INT, selection INT);''')
			var9 = zosportsdb[sport]
			c6 = '''INSERT INTO template_contests
				VALUES(?, ?, ?, ?, ?, 0);'''
			new_db.cursor.executemany(c6, var9)
			new_db.connection.commit()

			# table - template_events
			new_db.cursor.execute('''CREATE TABLE template_events
				(title TEXT PRIMARY KEY, contest TEXT, grade TEXT, 
				record_what REAL, record_who TEXT, 
				record_when INT, program_number INT, program_time TEXT);''')
			new_db.connection.commit()			

			# table - restrictions
			new_db.cursor.execute('''CREATE TABLE template_restrictions
									(restriction, value);''')

			if sport == 'Athletics':
				var11 = [('Tournament Minimum', ''), ('Tournament Maximum', 10), 
						('Track Event Maximum', 5), ('Field Event Maximum', 5)]
			else:
				var11 = [('Tournament Minimum', ''), ('Tournament Maximum', 10)]
			c12 = '''INSERT INTO template_restrictions
				VALUES(?, ?);'''
			new_db.cursor.executemany(c12, var11)
			new_db.connection.commit()
			new_db.connection.close()

		## Engine
		zotournament = ProgramTournament(title=title) 

class PageUpdateRecords(GridLayout):

	cols = 1

	tournament = StringProperty('')

	def __init__(self, *args, **kwargs):
		super(PageUpdateRecords, self).__init__(**kwargs)

		## Settings
		self.db = Database(self.tournament)
		details = self.db.cursor.execute("SELECT * FROM details").fetchone()
		self.type = details[3]

		grade_list = []
		event_list = self.db.cursor.execute('''SELECT grade
							FROM template_events''').fetchall()
		for tup in event_list:
			grade_list.append(tup[0])

		self.grade_list = []
		for grade in calculate_grade_list(self.tournament):
			if grade in grade_list:
				self.grade_list.append(grade) # puts them in order

		self.contest_order = calculate_contest_list(self.tournament)

		## Database
		self.current = None
		self.tab_ref = None
		self.temp_header = None
		self.unsaved = {}

		## Display
		self.build_panel()
		for grade in self.grade_list:
			self.build_tab(grade)
		self.build_display()

		## Engine
		self.build_default()

	### FUNCTIONS ###

	def change_display(self, display):

		self.display.clear_widgets()
		self.display.add_widget(display)
		self.current = display

	def change_grade_height(self, value, scroll):

		scroll.height = value - 120

	def change_header(self, value, header):

		if value[0] > header.height:

			header.height = value[0] + 10

	def change_panel_width(self, obj, value):

		self.tab_panel.width += (value - 100)#

	def change_scroll_width(self, obj, value):

		self.panel.width = self.width

	def command_results(self, event, arch, contest, record, holder, year):

		# settings

		try:
			res = self.db.cursor.execute('''SELECT id, best
				FROM %s''' % db_table('%s RR' % event)).fetchall()
		except sqlite3.OperationalError:

			try:
				res = self.db.cursor.execute('''SELECT team, best
					FROM %s''' % db_table('%s RR' % event)).fetchall()				
			except:
				PopBox().showwarning(title='No Results',
					message="No results available for this event")
				return				

		except Exception as inst:
			print type(inst)
			print inst.args
			print inst

			PopBox().showwarning(title='No Results',
				message="No results available for this event")
			return

		# functions
		def command(i):

			if i.text == 'Close':

				self.pop.dismiss()

			elif i.text == 'Update':

				if self.selected != None:

					record.input = self.selected[0]
					record.convert()
					record.status = False
					holder.text = self.selected[1]
					year.text = ''
					year.text = str(self.selected[2])

					self.selected = None

				self.pop.dismiss()

		def select(obj, value, ref):

			if value == True:

				for s in select_list:
					if s != obj:
						s.active = False
				self.selected = ref

		# database
		select_list = []
		self.selected = None

		# display
		id_dict = {}
		for tup in res:
			id = tup[0]
			best = tup[1]

			if best in [0, 0.0, None, '']:
				continue

			try:
				n = self.db.cursor.execute('''SELECT firstname, surname
					FROM individuals
					WHERE id="%s"''' % id).fetchone()

				name = '%s %s' % (n[0], n[1])
			except:
				name = id
			id_dict[name] = best

		ctype = calculate_contest_type(self.tournament, contest)
		if ctype == 'time':
			ranked = sorted(id_dict, key=id_dict.get)
		elif ctype == 'distance':
			ranked = sorted(id_dict, key=id_dict.get, reverse=True)

		content = GridLayout(cols=1)

		heading = BoxLayout(size_hint_y=None, height=40)
		heading.add_widget(Label(text='Name'))
		heading.add_widget(Label(text='Best'))
		heading.add_widget(Label(text='Select', size_hint_x=None, width=50))
		content.add_widget(heading)

		scroll = ScrollView(size_hint_y=None, height=130)
		content.add_widget(scroll)

		display = GridLayout(cols=1, size_hint_y=None, height=0)
		scroll.add_widget(display)

		for name in ranked:

			box = BoxLayout(size_hint_y=None, height=30)

			best = id_dict[name]
			if ctype == 'time':

				if best < arch:

					box.add_widget(LabelWrap(text='%s (R)' % name, 
						color=[1, 0, 0, 1]))

				else: 

					box.add_widget(LabelWrap(text=name))
			elif ctype == 'distance':

				if best > arch:

					box.add_widget(LabelWrap(text='%s (R)' % name, 
						color=[1, 0, 0, 1]))

				else: 

					box.add_widget(LabelWrap(text=name))
			box.add_widget(Label(text=convert_score(best, ctype)))

			ref = [best, name, datetime.now().year]

			s = CheckBoxA(active=False, size_hint_x=None, width=50)
			s.bind(active=lambda obj, value, 
				ref=ref:select(obj, value, ref))
			select_list.append(s)

			box.add_widget(s)

			display.add_widget(box)
			display.height += 30

		action = BoxLayout(size_hint_y=None, height=40)
		content.add_widget(action)

		btn1 = Button(text='Close')
		btn1.bind(on_press=command)
		action.add_widget(btn1)

		btn2 = Button(text='Update')
		btn2.bind(on_press=command)
		action.add_widget(btn2)

		# engine
		self.pop = Popup(title='Update %s Record' % event,
			content=content, auto_dismiss=False, size_hint=(None, None),
			size=(600, 300))
		self.pop.open()

	def command_save(self, i):

		if self.unsaved != {}:

			for event in self.unsaved:

				ref = self.unsaved[event]

				what = ref[0].output()
				who = ref[1].text
				when = ref[2].text

				if when == '':
					when = '2017'


				self.db.cursor.execute('''UPDATE template_events
						SET record_what="%s",
						record_who="%s",
						record_when="%s"
						WHERE title="%s"''' % (what, who, when, event))

				for r in ref:
					r.status = True

			self.db.connection.commit()

			# reset saved
			self.saved.text = ''
			self.saved.bg = [0.5, 0.5, 0.5, 1]
			self.unsaved = {}

	def command_tab(self, i):

		self.build_grade(i.text)

	def command_tab_check(self, boole):

		if boole == True: 
			self.build_grade(self.tab_ref)

	def command_unsave(self, value, event, ref):

		if value == True:
			return

		if self.saved.text == '':

			self.saved.text = 'Unsaved'
			self.saved.bg = [1, 0, 0, 1]

		self.unsaved[event] = ref

	### BUILD ###

	def build_panel(self):

		self.panel = ScrollView(size_hint=(None, None), height=50, width=0)
		self.bind(size=self.change_scroll_width)
		self.add_widget(self.panel)

		self.tab_panel = BoxLayout(size_hint=(None, None), height=50, width=0)
		self.panel.add_widget(self.tab_panel)

		self.add_widget(Label(size_hint_y=None, height=10))#

	def build_tab(self, grade):

		tab = Tab(text=grade, size_hint_y=None, height=40)
		tab.bind(on_press=self.command_tab)
		tab.bind(width=self.change_panel_width)
		self.tab_panel.add_widget(tab)
		self.tab_panel.width += 100

	def build_display(self):

		self.display = GridLayout(cols=1, padding=10)
		self.add_widget(self.display)

	def build_default(self):

		box = GridLayout(cols=1, padding=100)

		box.add_widget(Label(text='Select Grade', font_size=20, 
								size_hint_y=None, height=150))

		self.change_display(box)

	def build_grade(self, grade):

		## Settings - Reset

		frame = self.build_grade_frame(grade)

		scroll = ScrollView(size_hint_y=None, height=300)
		frame.add_widget(scroll)
		frame.bind(height=lambda obj, value, 
			scroll=scroll:self.change_grade_height(value, scroll))

		display = self.build_grade_display(grade)
		scroll.add_widget(display)

		action = self.build_grade_action()
		frame.add_widget(action)

		self.change_display(frame)

	def build_grade_frame(self, grade):

		frame = GridLayout(cols=1)

		header = Label(text=grade, font_size=20, size_hint_y=None, height=40)
		frame.add_widget(header)

		heading = BoxLayout(size_hint_y=None, height=40)
		frame.add_widget(heading)

		for h in ['Event', 'Record', 'Holder']:
			heading.add_widget(Label(text=h, font_size=20))

		heading.add_widget(Label(text='Year', font_size=20, size_hint_x=0.5))
		heading.add_widget(Label(size_hint_x=None, width=70))

		return frame

	def build_grade_display(self, grade):

		## Settings

		template = self.db.cursor.execute('''SELECT *
							FROM template_events
							WHERE grade="%s"''' % grade).fetchall()

		archive = self.db.cursor.execute('''SELECT *
							FROM archive_records
							WHERE grade="%s"''' % grade).fetchall()
		## Display
		display = GridLayout(cols=1, size_hint_y=None, height=0)

		for contest in self.contest_order:

			t = None
			a = None

			for tup in template:

				if tup[1] == contest:
					t = tup

			if t == None:
				continue

			for tup in archive:
				if tup[1] == contest:
					a = tup

			box = self.grade_box(t[0], contest, [str(t[3]),str(t[4]),str(t[5])], 
								[str(a[3]),str(a[4]),str(a[5])])

			display.add_widget(box)
			display.height += 70

		return display

	def grade_box(self, event, contest, t, a):

		box = BoxLayout(size_hint_y=None, height=70)

		box.add_widget(LabelWrap(text=event))

		rec = BoxLayout(orientation='vertical')

		ctype = calculate_contest_type(self.tournament, contest)

		if ctype == 'time':

			record = ScoreTime(input=float(t[0]), run_status=True,
						size_hint_y=None, height=40)
		elif ctype == 'distance':

			record = ScoreDistance(input=float(t[0]), run_status=True,
						size_hint_y=None, height=40)

		rec.add_widget(record)

		archive = float(a[0])

		arch = convert_score(float(a[0]), ctype)
		rec.add_widget(LabelWrap(text=arch))
		box.add_widget(rec)

		hold = BoxLayout(orientation='vertical')
		holder = TextInputC(text=t[1], size_hint_y=None, height=40, 
							run_status=True, multiline=False)
		hold.add_widget(holder)
		hold.add_widget(LabelWrap(text=a[1]))
		box.add_widget(hold)

		yr = BoxLayout(orientation='vertical', size_hint_x=0.5)
		year = TextInputC(text=t[2], input_filter='int', size_hint_y=None,
						height=40, run_status=True, multiline=False)
		yr.add_widget(year)
		yr.add_widget(LabelWrap(text=a[2]))
		box.add_widget(yr)

		results = Button(text='Results', size_hint_x=None, width=70)
		results.bind(on_press=lambda i, event=event, contest=contest,
			record=record, holder=holder, archive=archive,
			year=year:self.command_results(event, archive, contest,
				record, holder, year))
		box.add_widget(results)

		# saving
		ref = [record, holder, year]
		for r in ref:

			r.bind(status=lambda obj, value, ref=ref,
				event=event:self.command_unsave(value, event, ref))

		return box

	def build_grade_action(self):

		action = BoxLayout(size_hint_y=None, height=40)

		self.saved = LabelC(size_hint_x=None, width=100, color=[0, 0, 0, 1])
		action.add_widget(self.saved)

		save = Button(text='Save', font_size=20)
		save.bind(on_press=self.command_save)
		action.add_widget(save)

		return action

### SPONSORSHIP CLASSES ### (Not currently in use)

class AdvertMainFooter(BoxLayout):

	orientation = 'vertical'

	thanks = StringProperty('ZO-SPORTS for 2017 is sponsored by')

	def __init__(self, *args, **kwargs):
		super(AdvertMainFooter, self).__init__(**kwargs)

		return 

		## Settings
		self.size_hint_y = None
		self.height = 150
		self.padding = [10, 0, 10, 10]
		self.spacing = 10

		## Database
		self.ad = None

		## Display
		self.build_thanks()
		self.build_sponsors()

		## Engine
		Clock.schedule_interval(self.update, 10)

	### FUNCTIONS ###

	def update(self, *args):

		self.ad_box.remove_widget(self.ad)

		if self.ad == self.ad1:
			self.ad_box.add_widget(self.ad2)
			self.ad = self.ad2
		elif self.ad == self.ad2:
			self.ad_box.add_widget(self.ad1)
			self.ad = self.ad1

	### BUILD ###
	
	def build_thanks(self):

		self.add_widget(Label(text=self.thanks, size_hint_y=None, height=20))

	def build_sponsors(self):

		box = BoxLayout()

		left_box = BoxLayout(size_hint_x=None, width=200)
		left_sponsor = Image(source=resource_path('\sponsor1_logo.jpg'))
		left_box.add_widget(left_sponsor)
		box.add_widget(left_box)

		box.add_widget(self.build_advert())
		
		right_box = BoxLayout(size_hint_x=None, width=200)
		right_sponsor = Image(source=resource_path('sponsor2_logo.jpg'))
		right_box.add_widget(right_sponsor)
		box.add_widget(right_box)

		self.add_widget(box)

	def build_advert(self):

		self.ad_box = BoxLayout()

		self.ad1 = Image(source=resource_path('sponsor2_ad1.jpg'))
		self.ad2 = Image(source=resource_path('.\sponsor1_ad1.jpg'))

		self.ad = self.ad1
		self.ad_box.add_widget(self.ad1)

		return self.ad_box

### WIDGET CLASSES ### 

class BoxC(BoxLayout):

	bg = ListProperty([1, 1, 1, 1])

class CalendarWidget(BoxLayout):

    def __init__(self, **kwargs):
        super(CalendarWidget, self).__init__(**kwargs)
        self.date = kwargs.pop('date', date.today())
        self.orientation = "vertical"
        
        self.weekdays = ['Mon', 'Tues', 'Wed', 'Thurs', 
        					'Fri', 'Sat', 'Sun']
        self.months = ['January', 'February', 'March', 'April', 'May', 'June',
        				'July', 'August', 'September', 'October',
        				'November', 'December']
        
        self.header = BoxLayout(orientation = 'horizontal', 
                                size_hint = (1, 0.2))
        self.week = GridLayout(cols = 7, size_hint=(1, 0.2))
        self.body = GridLayout(cols = 7)
        self.add_widget(self.header)
        self.add_widget(self.week)
        self.add_widget(self.body)

        self.populate_body()
        self.populate_week()
        self.populate_header()

    def populate_header(self):
        self.header.clear_widgets()
        
        self.previous_year = Button(text='<<', size_hint_x=None, width=40)
        self.previous_year.bind(on_release=self.move_previous_year)
        self.previous_month = Button(text="<", size_hint_x=None, width=40)
        self.previous_month.bind(on_release=self.move_previous_month)
        self.next_month = Button(text = ">", size_hint_x=None, width=40)
        self.next_month.bind(on_release=self.move_next_month)
        self.next_year = Button(text='>>', size_hint_x=None, width=40)
        self.next_year.bind(on_release=self.move_next_year)
        
        self.current_month = Label(text=self.calculate_date(), 
                                   size_hint = (2, 1))
        self.header.add_widget(self.previous_year)
        self.header.add_widget(self.previous_month)
        self.header.add_widget(self.current_month)
        self.header.add_widget(self.next_month)
        self.header.add_widget(self.next_year)
        
    def calculate_date(self):
	
		year = self.date.year
		month = self.months[self.date.month-1]
		day = self.ordinal_conversion(self.date.day)
		wkday = self.weekdays[self.date.weekday()]
		
		return "%s, %s %s %s" % (wkday, day, month, year)
		   
    def populate_week(self):		
		for day in self.weekdays:
			self.week.add_widget(Label(text=day)) 

    def populate_body(self):
        self.body.clear_widgets()
        date_cursor = date(self.date.year, self.date.month, 1)
        	
        for blank in range(0, date_cursor.weekday()):
        	self.body.add_widget(LabelB())
        	
        while date_cursor.month == self.date.month:
            self.date_label = Button(text = str(date_cursor.day))
            self.date_label.bind(on_release=self.set_date)
            
            self.body.add_widget(self.date_label)
            date_cursor += timedelta(days = 1)
            
        if date_cursor.weekday() != 0:
        	for blank in range(date_cursor.weekday(), 7):
        		self.body.add_widget(LabelB())           
  
    def set_date(self, i):
    	day = int(i.text)
        self.date = date(self.date.year, self.date.month, day)
        self.populate_body()
        self.populate_header()
        
    def move_next_month(self, i):
		if self.date.month == 12:
			self.date = date(self.date.year + 1, 1, 1)
		else:
			self.date = date(self.date.year, self.date.month + 1, 1)
		self.populate_header()
		self.populate_body()
        
    def move_previous_month(self, i):

        if self.date.month == 1:
            self.date = date(self.date.year - 1, 12, 31)
        else:
        	date_cursor = date(self.date.year, self.date.month-1, 28)
        	final = None
        	while date_cursor.month == self.date.month-1: 
        		final = date_cursor.day
        		date_cursor += timedelta(days=1)
        	self.date = date(self.date.year, self.date.month -1, final)
        self.populate_header()
        self.populate_body()
        self.calculate_date()

    def move_next_year(self, i):
    	try:
       		self.date = date(self.date.year + 1, self.date.month, self.date.day)
       	except ValueError:
       		self.date = date(self.date.year + 1, self.date.month, 1)
        self.populate_header()
        self.populate_body()
        
    def move_previous_year(self, i):
    	try:
        	self.date = date(self.date.year - 1, self.date.month, self.date.day)
        except ValueError:
        	self.date = date(self.date.year - 1, self.date.month, 1)
        self.populate_header()
        self.populate_body()
        
    def ordinal_conversion(self, n):
    	
    	return calculate_ordinal(n)
  
class CheckBoxA(BoxLayout):	

	react = BooleanProperty('True')
	active = BooleanProperty('False')

	def __init__(self, *args, **kwargs):
		super(CheckBoxA, self).__init__(**kwargs)
	
		self.bind(active=self.change_box_active)
		
	def change_active(self, obj, value):
		self.active = value
	
	def change_box_active(self, obj, value):
		self.ids.box.active = value

class ColourPicker(BoxLayout):

	bg = ListProperty([0.5, 0.5, 0.5, 1])
	txt = ListProperty([1, 1, 1, 1])

	def __init__(self, *args, **kwargs):
		super(ColourPicker, self).__init__(**kwargs)

		self.pick = Button(text='^', font_size=20, size_hint_x=0.25)
		self.show = LabelC(text='Colour', bg=self.bg, color=self.txt)
		
		self.pick.bind(on_press=self.pop_open)
		
		self.add_widget(self.pick)
		self.add_widget(self.show)
		
	def pop_open(self, i):
	
		self.box = BoxLayout(orientation='vertical')
		
		self.pop = Popup(title="Choose Background and Text Colours",
							content=self.box, size_hint=(None, None),
							size=(500, 500), auto_dismiss=False)
		#					
		self.colour = ColorPicker(a=1)
		self.colour.color = self.show.bg
		
		#
		self.box1 = BoxLayout(size_hint_y=None, height='30dp')
		
		self.example = LabelC(text='Colour')
		self.example.color = self.show.color
		
		self.bw = Button(text='Text', size_hint_x=0.25)
		
		self.box1.add_widget(self.bw)
		self.box1.add_widget(self.example)
		
		#
		self.box2 = BoxLayout(size_hint_y=None, height=30)
		self.clr = Button(text="Clear")
		self.dis = Button(text="Confirm")
		
		self.box2.add_widget(self.clr)
		self.box2.add_widget(self.dis)
		
		## Bindings
		self.colour.bind(color=self.bg_colour)
		self.bw.bind(on_press=self.txt_colour)
		self.clr.bind(on_press=self.clear)
		self.dis.bind(on_press=self.confirm)
		
		#
		self.box.add_widget(self.colour)
		self.box.add_widget(self.box1)
		self.box.add_widget(self.box2)
		
		self.pop.open()
		
	def bg_colour(self, i, value):
	
		self.example.bg = value			

	def txt_colour(self, i):
		
		if self.example.color == [1, 1, 1, 1]:
			self.example.color = [0, 0, 0, 1]
		elif self.example.color == [0, 0, 0, 1]:
			self.example.color = [1, 1, 1, 1]

	def clear(self, i):
		
		self.example.bg = [0.5, 0.5, 0.5, 1]
		self.example.color = [1, 1, 1, 1]
		
	def confirm(self, i):
	
		self.show.bg = self.example.bg
		self.bg = self.example.bg
		self.show.color = self.example.color
		self.txt = self.example.color
		self.pop.dismiss() 

class DatePicker(BoxLayout):

	def __init__(self, *args, **kwargs):
		super(DatePicker, self).__init__(**kwargs)
		
		self.date = kwargs.pop('date', None)
		self.months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'June',
        				'July', 'Aug', 'Sept', 'Oct',
        				'Nov', 'Dec']		
		self.pick = Button(text='^', font_size=20, size_hint_x=0.25)
		self.pick.bind(on_press=self.pop_open)
		if self.date == None:
			self.show = Label(text='Date')
		else:
			self.show = Label(text=self.calculate_date())	
		
		self.add_widget(self.pick)
		self.add_widget(self.show)
		
	def pop_open(self, i):
	
		self.box = BoxLayout(orientation='vertical')
		
		self.pop = Popup(title="Choose Date",
							content=self.box, size_hint=(None, None),
							size=(500, 500), auto_dismiss=False)
		if self.date == None:
			self.calendar = CalendarWidget()
		else:	
			self.calendar = CalendarWidget(date=self.date)
				
		self.action = BoxLayout(size_hint_y=None, height='30dp')
		
		self.reset = Button(text="Clear")
		self.reset.bind(on_press=self.clear)
		
		self.dismiss = Button(text="Confirm")
		self.dismiss.bind(on_press=self.confirm)
		
		self.action.add_widget(self.reset)
		self.action.add_widget(self.dismiss)
		
		self.box.add_widget(self.calendar)
		self.box.add_widget(self.action)
		
		self.pop.open()
		
	def confirm(self, i):
	
		self.date = self.calendar.date
		self.show.text = self.calculate_date()
		self.pop.dismiss()

	def clear(self, i):
	
		self.date = None
		self.show.text = 'Date'
		self.pop.dismiss()
	
	def calculate_date(self):
	
		year = self.date.year
		month = self.months[self.date.month-1]
		day = self.ordinal_conversion(self.date.day)
		
		return "%s %s %s" % (day, month, year)
		
	def ordinal_conversion(self, n):
		ordinal = lambda n: "%d%s" % (n,"tsnrhtdd"[(n/10%10!=1)*(n%10<4)*n%10::4])
		return ordinal(n)

class FileDialog(Popup):

	title = 'Choose File'
	auto_dismiss = False

	function = ObjectProperty(None)
	filename = StringProperty('main_crest')
	copy = BooleanProperty(True)
	specify_type = ListProperty(['JPG','PNG'])

	def __init__(self, *args, **kwargs):
		super(FileDialog, self).__init__(**kwargs)

		## Settings
		self.size_hint = (None, None)
		self.size = (700, 500)


		## Display
		self.build_content()

		## Engine
		self.open()

	### FUNCTIONS ###

	def command_action(self, i):

		if i.text == 'Close':
			self.dismiss()
			return

		if self.path.text == '':
			return

		path = self.path.text
		file = path.split('\\')[-1]
		filetype = file.split('.')[1]

		if filetype.upper() not in self.specify_type:
			type = ', '.join(self.specify_type)
			self.title_bar.text = 'File must be %s type' % type
			return

		new_file = '%s.%s' % (self.filename, filetype)
		old_file = self.dialog.selection[0]

		if self.copy == True:
			file_copy(old_file, new_file)

			if self.filename == 'main_crest':
				db = Database('main')
				db.cursor.execute('''UPDATE settings
									SET crest="%s"''' % new_file)
				db.connection.commit()
				self.function(True, new_file)

			self.dismiss()

	def on_selection(self, *args, **kwargs):

		self.path.text = ''

		if self.dialog.selection != []:
			self.path.text = self.dialog.selection[0]

	### BUILD ###

	def build_content(self):

		box = BoxLayout(orientation='vertical')

		self.title_bar = Label(font_size=20, size_hint_y=None, height=30)
		box.add_widget(self.title_bar)

		self.dialog = FileIconView(function=self.on_selection)
		box.add_widget(self.dialog)

		self.path = TextInput(size_hint_y=None, height=30, readonly=True)
		box.add_widget(self.path)

		action = BoxLayout(size_hint_y=None, height=40)
		box.add_widget(action)

		nobtn = Button(text='Close', font_size=20)
		nobtn.bind(on_press=self.command_action)
		action.add_widget(nobtn)

		btn = Button(text='Confirm', font_size=20)
		btn.bind(on_press=self.command_action)
		action.add_widget(btn)

		self.content = box

class OptionMenu(Button):

	options = ListProperty()
	main = StringProperty('Select')

	def __init__(self, *args, **kwargs):
		super(OptionMenu, self).__init__(**kwargs)
	
		self.readonly = False
		self.dropdown = DropDown(padding=5, spacing=5)
		self.text = self.main
		
		for option in self.options:
			btn = Button(text=option, size_hint_y=None, height='30dp')
			btn.bind(on_release=lambda btn: self.dropdown.select(btn.text))
			self.dropdown.add_widget(btn)
				
		self.bind(on_release=self.open_close)
		
		self.dropdown.bind(on_select=lambda instance, x: setattr(self, 'text', x))	

	def open_close(self, ref):
	
		if self.readonly == False:
			self.dropdown.open(ref)	

	def add_option(self, option, **kwargs):			
		remove = kwargs.pop('remove', None)		
		if remove == None:
			if option in self.options:
				return
			self.options.append(option)	
		
		btn = Button(text=option, size_hint_y=None, height='30dp')
		btn.bind(on_release=lambda btn: self.dropdown.select(btn.text))
		self.dropdown.add_widget(btn)

	def remove_options(self, delete_list):
		# delete options
		for option in delete_list:
			if option in self.options:
				self.options.remove(option)
		# clear all buttons
		self.dropdown.clear_widgets()
		# reload new list
		for option in self.options:
			self.add_option(option, remove=True) 

	def change_all_options(self, change_list):
		self.dropdown.clear_widgets()
		self.options = change_list
		for option in self.options:
			self.add_option(option, remove=True)

class OptionMenuText(BoxLayout):

	options = ListProperty([])
	text = StringProperty('')
	
	def __init__(self, *args, **kwargs):
		super(OptionMenuText, self).__init__(**kwargs)
				
		if '(enter text)' not in self.options:
			self.options.append('(enter text)')		
		self.option = OptionMenu(options=self.options)
		self.option.bind(text=self.command_text)
		self.add_widget(self.option)
		
		self.state = None
		self.bind(text=self.command_text)

	def command_text(self, obj, value):
		
		if value != '(enter text)':
			self.text = value
			if self.state != None:
				self.remove_widget(self.state)
				self.state = None			
		else:
			enter_text = TextInput()
			enter_text.bind(text=self.command_enter_text)
			self.add_widget(enter_text)
			self.state = enter_text
	
	def command_enter_text(self, obj, value):
		self.text = value

class PopBox():

	def __init__(self):
		pass
	
	def dismiss(self, *args):

		self.pop.dismiss()

	def askyesno(self, *args, **kwargs):

		self.title = kwargs.pop('title', 'Ask Yes No')
		self.message = kwargs.pop('message', 'Insert Message Here')
		self.function = kwargs.pop('function', None)
		
		self.content = BoxLayout(orientation='vertical')
		
		self.main = LabelWrap(text=self.message, multiline=True)
		self.content.add_widget(self.main)
		
		def output(i):
		
			if self.function == None:
				self.pop.dismiss()
				return
		
			self.pop.dismiss()
		
			if i.text == 'Yes':
				self.function(True)
			
			elif i.text == 'No':
				self.function(False)
								
		self.box = BoxLayout(size_hint_y=None, height='50dp')
		self.yes = Button(text='Yes')
		self.yes.bind(on_press=output)
		self.no = Button(text='No')
		self.no.bind(on_press=output)		
		self.box.add_widget(self.no)
		self.box.add_widget(self.yes)
		self.content.add_widget(self.box)	

		self.pop = Popup(title=self.title, content=self.content, auto_dismiss=False,
							size_hint=(None, None), size=(500, 500))
		self.pop.open()

	def showwait(self, *args, **kwargs):

		function = kwargs.pop('function', None)
		content = LabelWrap(text="Processing... Please Wait", font_size=20)

		self.pop = Popup(title='...processing', content= content,
					size_hint=(None, None), size=(400, 400))
		self.pop.open()

		if function != None:
			function(True)

	def showwarning(self, *args, **kwargs):
	
		self.title = kwargs.pop('title', 'Warning')
		self.message = kwargs.pop('message', 'Warning Message')
		
		self.content = LabelWrap(text=self.message)

		self.pop = Popup(title=self.title, content=self.content, size_hint=(None, None),
							size=(400, 400))
	
		self.content.bind(on_press=self.pop.dismiss)
		self.pop.open()					

	def showok(self, *args, **kwargs):
	
		self.title = kwargs.pop('title', 'Welcome')
		self.message = kwargs.pop('message', 'Click to continue...')
		self.function = kwargs.pop('function', None)
		
		def command(*args):

			self.pop.dismiss()
			self.function
			print self.function

		self.content = Button(text=self.message)
		self.content.bind(on_press=command)

		self.pop = Popup(title=self.title, content=self.content, size_hint=(None, None),
							size=(400, 400), auto_dismiss=False)
	
		self.content.bind(on_press=self.pop.dismiss)
		self.pop.open()		

	def help_screen(self, *args, **kwargs):

		large = kwargs.pop('large', False)

		self.title = kwargs.pop('title', '')
		self.message = kwargs.pop('message', '')
		
		self.content = LabelWrap(text=self.message, multiline=True, padding_x=3)
		
		if large == False:
			self.pop = Popup(title='Help Screen: %s' % self.title, 
							content=self.content, size_hint=(None, None), 
							size=(400, 400))
		elif large == True:
			self.pop = Popup(title='Help Screen: %s' % self.title, 
							content=self.content, size_hint=(None, None), 
							size=(700, 500))
		
		self.pop.open()

	def register(self, *args, **kwargs):

		## Settings
		self.function = kwargs.pop('function', None)

		## Functions
		def command(i):

			if i.text == 'Close':
				self.pop.dismiss()

			elif i.text == 'Confirm':

				organisation = None
				address = None
				email = None

				if self.options.text != 'Select':
					organisation = self.options.text
					if self.address.text != '':
						address = self.address.text
						if self.email.text != '':
							email = self.email.text
							self.function(organisation, address, email)
							self.pop.dismiss()
				
				return

		## Display
		box = BoxLayout(orientation='vertical')

		box.add_widget(Label(text='Choose Organisation Type', 
			font_size=20, size_hint_y=None, height=40))

		self.options = OptionMenu(options=['Club','School'], font_size=20)
		box.add_widget(self.options)

		box.add_widget(Label(text='Address', 
			font_size=20, size_hint_y=None, height=40))

		self.address = TextInputC()
		box.add_widget(self.address)

		box.add_widget(Label(text='Email Address', 
			font_size=20, size_hint_y=None, height=40))

		self.email = TextInputC()
		box.add_widget(self.email)

		action = BoxLayout(size_hint_y=None, height=50)

		nobtn = Button(text='Close', font_size=20)
		nobtn.bind(on_press=command)
		action.add_widget(nobtn)

		btn = Button(text='Confirm', font_size=20)
		btn.bind(on_press=command)
		action.add_widget(btn)

		box.add_widget(action)

		## Engine
		self.pop = Popup(title='Registration', content=box, 
						auto_dismiss=False,
						size_hint=(None, None), size=(400, 350))
		self.pop.open()

class ScoreDistance(BoxLayout):
	'''
	This just displays the score of a distance event,
	actual recording of them is in ScoreLength and ScoreHeight
	'''

	status = BooleanProperty(True)
	run_status = BooleanProperty(False)

	input = NumericProperty(0)
	font_size = NumericProperty(20)

	def __init__(self, *args, **kwargs):
		super(ScoreDistance, self).__init__(**kwargs)

		## Engine
		if self.input not in [0, '', None]:
			self.ids.metre.text = str(self.input)

		if self.run_status == True:
			self.ids.metre.run_status = True
			self.ids.metre.bind(status=self.change_status)

		self.bind(status=self.change_kv_status)

	### FUNCTIONS ###

	def change_status(self, obj, value):
		if value == False:
			self.status = False

	def change_kv_status(self, obj, value):		
		if value == True:
			self.ids.metre.status = True

	def convert(self):

		self.ids.metre.multiline = False
		self.ids.metre.text = ''
		self.ids.metre.text = str(self.input)

	### OUTPUT ###

	def output(self):
	
		value = self.ids.metre.text
		
		if value == '':
			return 0
		else:
			value = float(value)
			if value.is_integer():
				return int(value)
			return value 

class ScoreHeight(BoxLayout):

	height_ref = ObjectProperty()
	status = BooleanProperty(True)
	run_status = BooleanProperty(True)

	input = ListProperty([[' ',' ',' '],[' ',' ',' '],[' ',' ',' '],
						[' ',' ',' '],[' ',' ',' '],[' ',' ',' ']])

	def __init__(self, *args, **kwargs):
		super(ScoreHeight, self).__init__(**kwargs)

		## Settings
		self.height_ref.bind(best=self.command_best)

		## Database
		self.id_list = [self.height_ref.ids.first, self.height_ref.ids.second, 
						self.height_ref.ids.third, self.height_ref.ids.fourth, 
						self.height_ref.ids.fifth, self.height_ref.ids.sixth]		
		self.menu = []

		## Display
		for triple in self.input:
			self.build_triple(triple)
		self.build_best()

	### FUNCTIONS ###

	def command_check(self, ref, i):

		if self.status == True:
			self.status = False

		value = i.text
		menu_pos = self.menu.index(ref)
		pos = ref.index(i) # index

		if menu_pos != 0:
			if self.menu[menu_pos-1][-1].text == 'X':
				return
			elif self.menu[menu_pos-1][-1].text == ' ':
				blank = True
				for r in range(0, menu_pos):
					box = self.menu[r]
					for b in box:
						if b.text != ' ':
							blank = False
				if blank == False:
					return
				else:
					if pos == 1:
						if self.menu[menu_pos][0].text == ' ':
							return
					elif pos == 2:
						if self.menu[menu_pos][1].text == ' ':
							return
					else: # its the first in the box
						for r in range(0, menu_pos):
							box = self.menu[r]
							for b in box:
								b.text = 'P'
						i.text = 'X'
						return
	
		# no result values
		if value == '-':
			return
		if value == 'P':
			for b in self.menu[menu_pos]:
				if b.text == 'P':
					b.text = ' '
			if menu_pos != 5:
				for box in range(menu_pos+1, 6):
					for b in self.menu[box]:
						b.text = ' '
			return

		# check on previous boxes
		if pos == 1:
			if ref[0].text == ' ':
				return
		if pos == 2:
			if ref[1].text == ' ':
				return
		# otherwise
		if value == ' ':
			# count back on
			miss = 0
			for r in range(0, menu_pos):
				box = self.menu[r]
				for b in box:
					if b.text == 'X':
						miss += 1
					elif b.text == 'O':
						miss = 0
			for b in ref:
				if b.text == 'X':
					miss += 1
				elif b.text == 'O':
					miss = 0
				if miss >= 3:
					return
			i.text = 'X'

		elif value == 'X':
			i.text = 'O'
			if pos != 2:
				for b in range(pos+1, 3):
					ref[b].text = '-'
		elif value == 'O':
			for b in range(pos, 3):
				ref[b].text = 'P'

		# check new best
		self.command_best()

	def command_best(self, *args):

		heights = []
		for i in range(6):
			box = self.menu[i]
			for b in box:
				if b.text == 'O':
					if self.id_list[i].text not in ['', '0', '0.0']:
						heights.append(float(self.id_list[i].text))

		sorted_heights = sorted(heights)
		if heights != []:
			if sorted_heights[-1] not in [0.0, 0]:
				self.best.text = str(sorted_heights[-1])

	### BUILD ###

	def build_triple(self, triple):

		box = BoxLayout()
		ref = []

		for t in triple:
			btn = Button(text=t)
			btn.bind(on_press=lambda i, 
				ref=ref:self.command_check(ref, i))
			box.add_widget(btn)
			ref.append(btn)

		self.menu.append(ref)
		self.add_widget(box)

	def build_best(self):

		self.best = Label(text='', size_hint_x=None, width=50)
		self.add_widget(self.best)

		heights = []
		for i in range(6):
			if 'O' in self.input[i]:
				if self.id_list[i].text not in ['', '0', '0.0']:
					heights.append(float(self.id_list[i].text))

		sorted_heights = sorted(heights)
		if heights != []:
			if sorted_heights[-1] not in [0.0, 0]:
				self.best.text = str(sorted_heights[-1])

	### OUTPUT ###

	def output(self, **kwargs):
	
		best = kwargs.pop('best', False)

		output = []
		for ref in self.menu:
			output.append([ref[0].text, ref[1].text, ref[2].text])
		
		if best == False:
			return output

		elif best == True:
			self.command_best()
			if self.best.text == '':
				return 0
			else:
				return float(self.best.text)

class ScoreLength(BoxLayout):

	status = BooleanProperty(True)
	run_status = BooleanProperty(False)

	input = ListProperty([0, 0, 0, 0, 0, 0])
	font_size = NumericProperty(14)

	def __init__(self, *args, **kwargs):
		super(ScoreLength, self).__init__(**kwargs)

		## Settings
		self.id_list = [self.ids.first, self.ids.second, self.ids.third,
						self.ids.fourth, self.ids.fifth, self.ids.sixth]

		## Display
		for i in range(0, 6):
			if self.input[i] in [0.0, 0]:
				score = ''
			else:
				score = float(self.input[i])
			self.id_list[i].text = str(score)
			
		self.command_best(True, True)

		## Engine
		if self.run_status == True:

			for id in self.id_list:
				id.run_status = True
				id.bind(status=self.change_status)
				id.bind(text=self.command_best)

		self.bind(status=self.change_kv_status)
		
	### FUNCTIONS ###

	def change_status(self, obj, value):
		if value == False:
			self.status = False

	def change_kv_status(self, obj, value):
		if value == True:
			for id in self.id_list:
				id.status = True

	def command_best(self, obj, value):

		score_list = []
		for id in self.id_list:
			if id.text != '':
				score_list.append(float(id.text))
			else:
				score_list.append(0.0)

		best = float(sorted(score_list)[-1])
		if best == 0.0:
			best = ''

		self.ids.best.text = str(best)		

	### OUTPUT ###

	def output(self, **kwargs):
	
		best = kwargs.pop('best', False)

		output = []

		for id in self.id_list:

			value = id.text
			if id.text == '':
				value = 0
			else:
				value = float(value)

			output.append(value)

		if best == True:

			return sorted(output)[-1]

		else:

			return output

class ScoreVertical(BoxLayout):

	best = BooleanProperty(True)
	status = BooleanProperty(True)
	run_status = BooleanProperty(True)

	input = ListProperty([0, 0, 0, 0, 0, 0])
	font_size = NumericProperty(14)

	def __init__(self, *args, **kwargs):
		super(ScoreVertical, self).__init__(**kwargs)

		## Settings
		self.id_list = [self.ids.first, self.ids.second, self.ids.third,
						self.ids.fourth, self.ids.fifth, self.ids.sixth]

		if self.input == []:
			self.input = [0, 0, 0, 0, 0, 0]
		## Display
		for i in range(0, 6):
			if self.input[i] in [0.0, 0]:
				score = ''
			else:
				score = self.input[i]

			self.id_list[i].text = str(score)

		## Engine
		if self.run_status == True:

			for id in self.id_list:
				id.run_status = True
				id.bind(status=self.change_status)

		self.bind(status=self.change_kv_status)

	### FUNCTIONS ###

	def change_status(self, obj, value):
		if value == False:
			self.status = False
			self.best = False

	def change_kv_status(self, obj, value):
		if value == True:
			for id in self.id_list:
				id.status = True
			self.best = True

	### OUTPUT ###

	def output(self, **kwargs):

		output = []

		for id in self.id_list:

			value = id.text
			if id.text == '':
				value = 0
			else:
				value = float(value)

			output.append(value)

		return output

class ScoreTime(BoxLayout):

	status = BooleanProperty(True)
	run_status = BooleanProperty(False)

	input = NumericProperty(0)
	font_size = NumericProperty(20)

	def __init__(self, *args, **kwargs):
		super(ScoreTime, self).__init__(**kwargs)

		## Engine
		if self.input not in [0, '', None]:
			self.convert()

		if self.run_status == True:
			self.ids.min.run_status = True
			self.ids.min.bind(status=self.change_status)
			self.ids.sec.run_status = True
			self.ids.sec.bind(status=self.change_status)

		self.bind(status=self.change_kv_status)

	### FUNCTIONS ###

	def change_status(self, obj, value):
		if value == False:
			if self.status == True:
				self.status = False
				self.ids.min.status = False
				self.ids.sec.status = False

	def change_kv_status(self, obj, value):		
		if value == True:
			self.ids.min.status = True
			self.ids.min.status = True

	def convert(self):

		value = float(self.input)
		
		minutes = 0
		while value >= 60:
			value -= 60
			minutes += 1
		
		if minutes > 0:
			self.ids.min.text = str(minutes)
		if value >= 0:
			if value.is_integer():
				self.ids.sec.text = str(int(value))
			else:
				self.ids.sec.text = str(value)

	### OUTPUT ###

	def output(self, **kwargs):

		if self.ids.min.text == '':
			minutes = 0
		else:
			minutes = int(self.ids.min.text)
		
		if self.ids.sec.text == '':
			seconds = 0
		else:
			seconds = float(self.ids.sec.text)			
		
		return (minutes*60) + seconds

class Time24(BoxLayout):

	status = BooleanProperty(True)
	run_status = BooleanProperty(False)

	font_size = NumericProperty(30)

	def __init__(self, *args, **kwargs):
		super(Time24, self).__init__(**kwargs)

		## Settings
		self.input = kwargs.pop('input', None)

		## Engine
		self.convert()

		if self.run_status == True:
			self.ids.hour.run_status = True
			self.ids.hour.bind(status=self.change_status)
			self.ids.min.run_status = True
			self.ids.min.bind(status=self.change_status)

		self.bind(status=self.change_kv_status)

	### FUNCTIONS ###

	def change_status(self, obj, value):
		if value == False:
			if self.status == True:
				self.status = False
				self.ids.hour.status = False
				self.ids.min.status = False

	def change_kv_status(self, obj, value):		
		if value == True:
			self.ids.hour.status = True
			self.ids.min.status = True

	def convert(self):

		if self.input in ['', None]:
			return

		var = self.input.split(':')
		self.ids.hour.text = var[0]
		self.ids.min.text = var[1]

	### OUTPUT ###

	def output(self):

		return '%s:%s' % (self.ids.hour.text, self.ids.min.text)

#-----------------------------------------------------------------------------#

### EVENT CLASSES - ADD ### 

class EventAddRound():

	def __init__(self, *args, **kwargs):

		## Settings

		self.tournament = kwargs.pop('tournament', None)
		self.event = kwargs.pop('event', None)
		self.current_title = kwargs.pop('current_title', None)
		self.current_table = kwargs.pop('current_table', None)
		self.widget = kwargs.pop('widget', None)

		self.db = Database(self.tournament)


		groups = self.db.cursor.execute('''SELECT title, colour_bg, colour_text
										FROM template_groups
										WHERE category="type"''').fetchall()
		self.type_dict = {}
		for tup in groups:
			self.type_dict[tup[0]] = [db_colour(tup[1], db_in=False), 
									db_colour(tup[2], db_in=False)]

		event = self.db.cursor.execute('''SELECT grade, contest
								FROM template_events
								WHERE title="%s"''' % self.event).fetchone()
		self.grade = event[0]
		self.contest = event[1]
		self.contest_type = calculate_contest_type(self.tournament, 
													self.contest)


		## Database
		self.menu = {}


		## Display
		self.build_content()
		self.build_competitors()

		## Engine
		self.pop = Popup(title=self.tournament, content=self.content, 
					auto_dismiss=False, size_hint=(None, None), size=(800, 600))
		self.pop.open()

	### FUNCTIONS ###

	def command_add_round(self, i):

		if self.example.text == '':
			PopBox().showwarning(title="Error",
			message="Please create a unique title for this round")
			return

		# check if round title already exists
		table = db_table(self.event)

		check = self.db.cursor.execute('''SELECT * FROM %s
									WHERE title="%s"''' % (table, 
												self.example.text)).fetchone()
		if check != None:
			PopBox().showwarning(title='Error',
			message="This event already has a round called %s, choose a unique title" % self.example.text)
			return

		id_list = []
		for id in self.menu:
			if self.menu[id].active == True:
				id_list.append(id)

		# check minimum
		if len(id_list) < 2:
			PopBox().showwarning(title='Error',
			message="Not enough competitors have been selected\nNeed a minimum of 2, have %s" % str(len(id_list)))
			return

		# update event table
		rounds = self.db.cursor.execute('''SELECT round
										FROM %s''' % table).fetchall()
		rd = []
		for r in rounds:
			if r[0] == 'RR':
				continue
			else:
				rd.append(int(r[0][1]))
		num = 'R%s' % str(sorted(rd)[-1] + 1)

		## 
		if self.program_number.text != '':
			program_number = int(self.program_number.text)
		else:
			program_number = ''

		if self.program_time.output() in [None, 0, 0.0, '']:
			program_time = ''
		else:
			program_time = self.program_time.output()
		##
		try:
			self.db.cursor.execute('''ALTER TABLE %s
						ADD program_number INT''' % table)
			self.db.cursor.execute('''ALTER TABLE %s
						ADD program_time TEXT''' % table)
		except:
			pass

		var = [num, self.example.text, '0#0#0#0#0#0', 
				program_number, program_time]

		self.db.cursor.execute('''INSERT INTO %s
							VALUES(?,?,?,?,?)''' % table, var)

		# create round table
		nt = '%s %s' % (self.event, num)
		new_table = db_table(nt)

		self.db.cursor.execute('''DROP TABLE IF EXISTS %s''' % new_table)
		self.db.cursor.execute('''CREATE TABLE %s
						(id TEXT, pp INT, 
						time_measure INT, length TEXT, height TEXT,
						placing INT)''' % new_table)
		var1 = []
		for id in id_list:
			var1.append([id, 0, 0, '', '', 0])

		self.db.cursor.executemany('''INSERT INTO %s
						VALUES(?, ?, ?, ?, ?, ?)''' % new_table, var1)

		self.db.connection.commit()
		self.widget.command_change_round(True, self.example.text)
		self.pop.dismiss()

	def command_create(self, value):

		self.opt.text = 'Select'
		if self.create.text == '':
			self.example.text = ''
		else:
			self.example.text = '%s %s' % (self.event, value)

	def command_example(self, obj, value):

		if value != 'Select':
			self.create.text = ''
			self.example.text = '%s %s' % (self.event, value)

	def command_select(self, obj, value):

		if value == True:
			self.select.text = str(int(self.select.text) + 1)
		elif value == False:
			self.select.text = str(int(self.select.text) - 1)			

	### BUILD ###

	def build_content(self):

		self.content = GridLayout(cols=1, spacing=5)

		self.content.add_widget(Label(text='Add Round', font_size=20,
									size_hint_y=None, height=40))

		# select

		select_box = BoxLayout(orientation='vertical', 
						size_hint_y=None, height=180)
		self.content.add_widget(select_box)
		select_box.add_widget(LabelWrap(size_hint_y=None, height=30,
			text='Select which competitors will go through to the next round'))
		header = BoxLayout(size_hint_y=None, height=40)
		select_box.add_widget(header)

		header.add_widget(Label(text='House', size_hint_x=None, width=60))
		header.add_widget(Label(text='Competitor'))
		header.add_widget(Label(text='Best Score', size_hint_x=None, width=100))
		header.add_widget(Label(text='Placing', size_hint_x=None, width=100))
		self.select = Label(text='0', size_hint_x=None, width=100)
		header.add_widget(self.select)

		scroll = ScrollView(size_hint_y=None, height=110)
		select_box.add_widget(scroll)

		self.display = GridLayout(cols=1, size_hint_y=None, height=0)
		scroll.add_widget(self.display)

		# name and create

		name_box = GridLayout(cols=1, size_hint_y=None, height=110)
		self.content.add_widget(name_box)

		name_box.add_widget(LabelWrap(size_hint_y=None, height=30,
			text='Choose from the options or create your own title for this new round'))
		heading = BoxLayout(size_hint_y=None, height=40)
		name_box.add_widget(heading)

		heading.add_widget(Label(text='Choose Title'))
		heading.add_widget(Label(text='Create Title'))
		heading.add_widget(Label(text='Title'))

		choose = BoxLayout(size_hint_y=None, height=40)
		name_box.add_widget(choose)

		options = ['Quarter-Final', 'Semi-Final', 'Final']

		self.opt = OptionMenu(options=options)
		self.opt.bind(text=self.command_example)
		choose.add_widget(self.opt)

		self.create = TextInputC(function=self.command_create)
		choose.add_widget(self.create)

		self.example = LabelWrap()
		choose.add_widget(self.example)

		self.result = BoxLayout(size_hint_y=None, height=30)
		self.content.add_widget(self.result)

		# program box
		prog_box = GridLayout(cols=1, size_hint_y=None, height = 110)
		self.content.add_widget(prog_box)

		prog_box.add_widget(LabelWrap(size_hint_y=None, height=30,
			text='Add Event Number and Event Time for this new Round (not required)'))
		headp = BoxLayout(size_hint_y=None, height=40)
		prog_box.add_widget(headp)

		headp.add_widget(Label(text='Event Number'))
		headp.add_widget(Label(text='Event Time'))

		prog = BoxLayout(size_hint_y=None, height=40)
		prog_box.add_widget(prog)
		self.program_number = TextInputC(input_filter='int')
		prog.add_widget(self.program_number)
		self.program_time = Time24(font_size=20, run_status=True)
		prog.add_widget(self.program_time)

		# action
		action = BoxLayout(size_hint_y=None, height=40)
		self.content.add_widget(action)

		btn1 = Button(text='Close', font_size=20)
		btn1.bind(on_press=lambda i:self.pop.dismiss())
		action.add_widget(btn1)

		btn2 = Button(text='Add Round', font_size=20)
		btn2.bind(on_press=self.command_add_round)
		action.add_widget(btn2)

	def build_competitors(self):

		data = self.db.cursor.execute('''SELECT * 
									FROM %s''' % self.current_table).fetchall()	

		id_dict = {}

		rank = {}
		best = {}
		present = []
		rest = []

		# split
		for tup in data:
			id = tup[0]
			pp = tup[1]
			if self.contest_type == 'time':
	
				score = tup[2]
			else:
				if self.contest in ['High Jump', 'Pole Vault']:
					if tup[4] != None:
						scr = db_score(tup[4], 
										'High Jump', 'distance', db_in=False)
						try:
							score = db_vertical(scr, self.tournament, self.event, 
								self.current_title, best=True)
						except:
							PopBox().showwarning(title='Heights',
							message="Check you have entered all the height values")
							return

					else:
						score = None
				else:
					if tup[3] != None:
						scr = db_score(tup[3], None, 'distance', db_in=False)
						score = sorted(scr)[-1]
					else:
						score = None
			placing = tup[5]

			details = self.db.cursor.execute('''SELECT firstname, surname,
														type
									FROM individuals
									WHERE id="%s"''' % id).fetchone()
			fn = details[0]
			sn = details[1]
			type = details[2]

			id_dict[id] = [pp, score, placing, fn, sn, type]

			if placing not in [None, 0, '']:
				rank[id] = placing
			elif score not in [None, 0, '']:
				best[id] = score
			elif pp == 1:
				present.append(id) 
			else:
				rest.append(id)

		id_list = sorted(rank, key=rank.get)

		# by best
		if self.contest_type == 'time':
			id_list += sorted(best, key=best.get)
		else:
			id_list += sorted(best, key=best.get, reverse=True)
		# by present (alphabetical)
		data1 = {}
		for id in present:
			data1[id] = '%s %s' % (id_dict[id][4], id_dict[id][3])
		id_list += sorted(data1, key=data1.get)
		data2 = {}
		for id in rest:
			data2[id] = '%s %s' % (id_dict[id][4], id_dict[id][3])
		id_list += sorted(data2, key=data2.get)		

		for id in id_list:

			self.build_competitor_box(id, id_dict[id])

	def build_competitor_box(self, id, details):

		box = BoxLayout(size_hint_y=None, height=30)

		if details[1] in [None, 0]:
			best = ''
		else:
			best = str(details[1])

		if details[2] in [None, 0]:
			rank = ''
		else:
			rank = str(details[2])

		type_title = details[-1].upper()[:3]
		bg = self.type_dict[details[-1]][0]
		txt = self.type_dict[details[-1]][1]

		box.add_widget(LabelC(text=type_title, bg=bg, color=txt, 
								size_hint_x=None, width=60))
		box.add_widget(Label(text="%s %s" % (details[3], details[4])))
		box.add_widget(Label(text=best, size_hint_x=None, width=100))
		box.add_widget(Label(text=rank, size_hint_x=None, width=100))

		chk = CheckBoxA(active=False, size_hint_x=None, width=100)
		chk.bind(active=self.command_select)
		box.add_widget(chk)

		self.menu[id] = chk
		self.display.add_widget(box)
		self.display.height += 30

class EventAddRoundTeam(EventAddRound):

	### FUNCTIONS ###

	def command_add_round(self, i):

		if self.example.text == '':
			PopBox().showwarning(title="Error",
			message="Please create a unique title for this round")
			return

		# check if round title already exists
		table = db_table(self.event)

		check = self.db.cursor.execute('''SELECT * FROM %s
									WHERE title="%s"''' % (table, 
												self.example.text)).fetchone()
		if check != None:
			PopBox().showwarning(title='Error',
			message="This event already has a round called %s, choose a unique title" % self.example.text)
			return

		id_list = []
		for id in self.menu:
			if self.menu[id].active == True:
				id_list.append(id)

		# check minimum
		if len(id_list) < 2:
			PopBox().showwarning(title='Error',
			message="Not enough teams have been selected\nNeed a minimum of 2, have %s" % str(len(id_list)))
			return

		# update event table
		rounds = self.db.cursor.execute('''SELECT round
										FROM %s''' % table).fetchall()
		rd = []
		for r in rounds:
			if r[0] == 'RR':
				print 'complete already'
				continue
			else:
				rd.append(int(r[0][1]))
		num = 'R%s' % str(sorted(rd)[-1] + 1)

		## 
		if self.program_number.text != '':
			program_number = int(self.program_number.text)
		else:
			program_number = ''

		if self.program_time.output() in [None, 0, 0.0, '']:
			program_time = ''
		else:
			program_time = self.program_time.output()
		##
		try:
			self.db.cursor.execute('''ALTER TABLE %s
						ADD program_number INT''' % table)
			self.db.cursor.execute('''ALTER TABLE %s
						ADD program_time TEXT''' % table)
		except:
			pass

		var = [num, self.example.text, '0#0#0#0#0#0', program_number,
				program_time]

		self.db.cursor.execute('''INSERT INTO %s
							VALUES(?,?,?,?,?)''' % table, var)

		# create round table
		nt = '%s %s' % (self.event, num)
		new_table = db_table(nt)
		print new_table

		self.db.cursor.execute('''DROP TABLE IF EXISTS %s''' % new_table)
		self.db.cursor.execute('''CREATE TABLE %s
						(team TEXT, pp INT, 
						time_measure INT, length TEXT, height TEXT,
						placing INT, type TEXT)''' % new_table)
		var1 = []
		for id in id_list:
			type = self.menu_type[id]
			var1.append([id, 0, 0, '', '', '', type])

		self.db.cursor.executemany('''INSERT INTO %s
						VALUES(?, ?, ?, ?, ?, ?, ?)''' % new_table, var1)

		self.db.connection.commit()
		self.widget.command_change_round(True, self.example.text)
		self.pop.dismiss()

	### BUILD ###

	def build_content(self):

		self.content = GridLayout(cols=1, spacing=5)

		self.content.add_widget(Label(text='Add Round', font_size=20,
									size_hint_y=None, height=40))

		# select

		select_box = BoxLayout(orientation='vertical', 
						size_hint_y=None, height=180)
		self.content.add_widget(select_box)
		select_box.add_widget(LabelWrap(size_hint_y=None, height=30,
			text='Select which teams will go through to the next round'))
		header = BoxLayout(size_hint_y=None, height=40)
		select_box.add_widget(header)

		header.add_widget(Label(text='House', size_hint_x=None, width=60))
		header.add_widget(Label(text='Team'))
		header.add_widget(Label(text='Best Score', size_hint_x=None, width=100))
		header.add_widget(Label(text='Placing', size_hint_x=None, width=100))
		self.select = Label(text='0', size_hint_x=None, width=100)
		header.add_widget(self.select)

		scroll = ScrollView(size_hint_y=None, height=110)
		select_box.add_widget(scroll)

		self.display = GridLayout(cols=1, size_hint_y=None, height=0)
		scroll.add_widget(self.display)

		# name and create

		name_box = GridLayout(cols=1)
		self.content.add_widget(name_box)

		name_box.add_widget(LabelWrap(size_hint_y=None, height=30,
			text='Choose from the options or create your own title for this new round'))
		heading = BoxLayout(size_hint_y=None, height=40)
		name_box.add_widget(heading)

		heading.add_widget(Label(text='Choose Title'))
		heading.add_widget(Label(text='Create Title'))
		heading.add_widget(Label(text='Title'))

		choose = BoxLayout(size_hint_y=None, height=40)
		name_box.add_widget(choose)

		options = ['Quarter-Final', 'Semi-Final', 'Final']

		self.opt = OptionMenu(options=options)
		self.opt.bind(text=self.command_example)
		choose.add_widget(self.opt)

		self.create = TextInputC(function=self.command_create)
		choose.add_widget(self.create)

		self.example = LabelWrap()
		choose.add_widget(self.example)

		self.result = BoxLayout(size_hint_y=None, height=30)
		self.content.add_widget(self.result)
		
		# program box
		prog_box = GridLayout(cols=1, size_hint_y=None, height = 110)
		self.content.add_widget(prog_box)

		prog_box.add_widget(LabelWrap(size_hint_y=None, height=30,
			text='Add Event Number and Event Time for this new Round (not required)'))
		headp = BoxLayout(size_hint_y=None, height=40)
		prog_box.add_widget(headp)

		headp.add_widget(Label(text='Event Number'))
		headp.add_widget(Label(text='Event Time'))

		prog = BoxLayout(size_hint_y=None, height=40)
		prog_box.add_widget(prog)
		self.program_number = TextInputC(input_filter='int')
		prog.add_widget(self.program_number)
		self.program_time = Time24(font_size=20, run_status=True)
		prog.add_widget(self.program_time)

		# action
		action = BoxLayout(size_hint_y=None, height=40)
		self.content.add_widget(action)

		btn1 = Button(text='Close', font_size=20)
		btn1.bind(on_press=lambda i:self.pop.dismiss())
		action.add_widget(btn1)

		btn2 = Button(text='Add Round', font_size=20)
		btn2.bind(on_press=self.command_add_round)
		action.add_widget(btn2)

	def build_competitors(self):

		self.menu_type = {}

		data = self.db.cursor.execute('''SELECT * 
									FROM %s''' % self.current_table).fetchall()	

		id_dict = {}

		rank = {}
		best = {}
		present = []
		rest = []

		# split
		for tup in data:
			id = tup[0]
			pp = tup[1]
			if self.contest_type == 'time':
	
				score = tup[2]
			else:
				if self.contest in ['High Jump', 'Pole Vault']:
					if tup[4] != None:
						scr = db_score(tup[4], 
										'High Jump', 'distance', db_in=False)
						score = db_vertical(scr, self.tournament, self.event, 
							self.current_title, best=True)
					else:
						score = None
				else:
					if tup[3] != None:
						scr = db_score(tup[3], None, 'distance', db_in=False)
						score = sorted(scr)[-1]
					else:
						score = None
			placing = tup[5]
			type = tup[6]

			id_dict[id] = [pp, score, placing, type]

			if placing not in [None, 0, '']:
				rank[id] = placing
			elif score not in [None, 0, '']:
				best[id] = score
			elif pp == 1:
				present.append(id) 
			else:
				rest.append(id)

		id_list = sorted(rank, key=rank.get)

		# by best
		if self.contest_type == 'time':
			id_list += sorted(best, key=best.get)
		else:
			id_list += sorted(best, key=best.get, reverse=True)
		# by present (alphabetical)
		data1 = {}
		for id in present:
			data1[id] = '%s %s' % (id_dict[id][3])
		id_list += sorted(data1, key=data1.get)
		data2 = {}
		for id in rest:
			data2[id] = '%s %s' % (id_dict[id][3])
		id_list += sorted(data2, key=data2.get)		

		for id in id_list:

			self.build_competitor_box(id, id_dict[id])

	def build_competitor_box(self, id, details):

		box = BoxLayout(size_hint_y=None, height=30)

		if details[1] in [None, 0]:
			best = ''
		else:
			best = str(details[1])

		if details[2] in [None, 0]:
			rank = ''
		else:
			rank = str(details[2])

		type_title = details[-1].upper()[:3]
		bg = self.type_dict[details[-1]][0]
		txt = self.type_dict[details[-1]][1]

		box.add_widget(LabelC(text=type_title, bg=bg, color=txt, 
								size_hint_x=None, width=60))
		box.add_widget(Label(text=id))
		box.add_widget(Label(text=best, size_hint_x=None, width=100))
		box.add_widget(Label(text=rank, size_hint_x=None, width=100))

		chk = CheckBoxA(active=False, size_hint_x=None, width=100)
		chk.bind(active=self.command_select)
		box.add_widget(chk)

		self.menu[id] = chk
		self.menu_type[id] = details[-1] # type
		self.display.add_widget(box)
		self.display.height += 30

### EVENT CLASSES - BASE ###

class EventBase(GridLayout):

	cols = 1

	tournament = StringProperty('')
	contest = StringProperty('')
	grade = StringProperty('')
	official = BooleanProperty(False)

	def __init__(self, *args, **kwargs):
		super(EventBase, self).__init__(**kwargs)

		## Settings
		self.spacing = 5
		self.db = Database(self.tournament)
		details = self.db.cursor.execute("SELECT * FROM details").fetchone()
		self.type = details[3]

		self.event = '%s %s' % (self.grade, self.contest)
		self.event_table = db_table(self.event)
		self.event_details = self.db.cursor.execute('''SELECT * 
									FROM template_events
								WHERE title="%s"''' % self.event).fetchall()[0]
		self.contest_type = calculate_contest_type(self.tournament,
													self.contest)
		self.team = calculate_contest_type(self.tournament, 
										self.contest, team=True)

		self.typec = {}
		colours = self.db.cursor.execute('''SELECT title, colour_bg, colour_text
									FROM template_groups
									WHERE category="type"''').fetchall()
		for c in colours:
			self.typec[c[0]] = [db_colour(c[1], db_in=False), 
								db_colour(c[2], db_in=False)]

		## Database
		self.current_display = None
		self.current_round = None
		self.current_table = None
		self.current_body = None
		self.menu = {}
		self.unsaved = []
		self.scroll_menu = []

		## Display
		self.build_base_title()
		self.build_base_display()

		## Engine
		self.build_display_structure(self.current_round)

	### FUNCTIONS ###

	def change_display(self):

		self.base_display.clear_widgets()
		self.base_display.add_widget(self.current_display)

	def command_add_entry(self, i):

		if self.official == False:

			EntryAddRemove(style='Entry', tournament=self.tournament,
				event=self.event, event_round=self.current_round,
				event_table=self.current_table, widget=self)

	def command_add_round(self, i):

		self.command_save(None)
		
		EventAddRound(tournament=self.tournament, event=self.event,
						current_title=self.current_round,
						current_table=self.current_table,
						widget=self)

	def command_calculate(self, i):

		if i.text == 'Calculate':
			self.command_calculate_placing()
			i.text = 'Clear'
		elif i.text == 'Clear':
			for id in self.menu:
				if self.menu[id][2].text != '':
					self.menu[id][2].text = ''
					if id not in self.unsaved:
						self.unsaved.append(id)

			if self.save.text != 'Unsaved':
				self.save.text = 'Unsaved'
				self.save.bg = [1, 0, 0, 1]				
			i.text = 'Calculate'

	def command_calculate_placing(self):

		data = {}
		last = []
		for id in self.menu:
			ref = self.menu[id]
			ref[2].text = ''
			if ref[0].active == True:
				score = ref[1].output(best=True)
				if score in [0.0, 0]:
					last.append(id)
				else:
					data[id] = score
				if id not in self.unsaved:
					self.unsaved.append(id)
					if self.save.text != 'Unsaved':
						self.save.text = 'Unsaved'
						self.save.bg = [1, 0, 0, 1]

		if self.contest_type == 'time':
			rank_list = sorted(data, key=data.get)
		else:
			rank_list = sorted(data, key=data.get, reverse=True)

		rank = 1
		tie = 1
		for i in range(len(rank_list)):

			id = rank_list[i]
			score = data[id]
			plc = self.menu[id][2]

			plc.text = str(rank)

			if id == rank_list[-1]: # last id
				rank += tie
				break

			if score == data[rank_list[i+1]]:
				tie += 1
			else:
				rank += tie
				tie = 1

		for id in last:
			self.menu[id][2].text = str(rank)

	def command_change_round(self, obj, value):

		if obj == True: # check re-run
			self.build_display_structure(value)
			self.base_title.add_option(value)
			self.base_title.text = value
			return

		# check for unsaved
		#if self.unsaved != []:
		#	PopBox().askyesno(title='Unsaved', 
		#						function=self.command_change_round,
		#		message='Unsaved changes have been made.\n\nDo you still wish to change rounds?')
		#	return
		
		self.build_display_structure(value)

	def command_complete(self, i):

		if i == False:
			return

		# check conditions
		check = []
		if i != True:
			check_round = self.db.cursor.execute('''SELECT round
									FROM %s''' % self.event_table).fetchall()
			count = 0
			for tup in check_round:
				if tup[0] == 'RR':
					check.append('This event has already been completed')
					continue
				count += 1
			if count > 1:
				check.append('There is more than one round in this event, if you complete here you will complete using the placings from %s' % self.current_round)

			if check != []:
				check.append('Do you wish to continue?')
				message = '\n\n'.join(check)
				PopBox().askyesno(title='Complete %s' % self.event,
					function=self.command_complete,
					message=message)
				return

		# save 
		self.process_save()

		# CP
		points = self.db.cursor.execute('''SELECT *
								FROM template_points''').fetchall()

		point_ref = 1
		if self.team == True:
			point_ref = 2

		if points[0][point_ref] not in [None, '']:
			ppp = points[0][point_ref]
		else:
			ppp = 0

		current = self.db.cursor.execute('''SELECT id, pp, placing, 
								time_measure, length, height
								FROM %s''' % self.current_table).fetchall()
		id_dict = {}
		rank_dict = {}

		first_rank = False

		for tup in current:

			id = tup[0]
			pp = tup[1]

			if pp  == 0:
				continue # no participation

			placing = tup[2]
			
			if placing == 1: # checks there is a winner
				first_rank = True

			dict = {'best':0, 'rank':placing, 'PP':0, 'CP':0}

			if pp == 1:
				dict['PP'] += ppp

			if placing < len(points) and placing != None:
				rank_dict[id] = placing

			if self.contest_type == 'time':
				dict['best'] = tup[3]
			elif self.contest_type == 'distance':
				if self.contest in ['High Jump', 'Pole Vault']:

					if tup[5] != None:

						input = db_height(tup[5], db_in=False)
						try:
							best = db_vertical(input, self.tournament, self.event, 
										self.current_round, best=True)
						except:
							PopBox().showwarning(title='Heights',
							message="Check you have entered all the height values")
							return

						dict['best'] = best

				else:
					if tup[4] != None:
						dict['best'] = sorted(db_score(tup[4], None, 
									'distance', db_in=False))[-1]

			id_dict[id] = dict

		rank_order = sorted(rank_dict, key=rank_dict.get)
		
		rank = 1
		score = 0.0
		temp_list = []
		for i in range(len(rank_order)):

			id = rank_order[i]
			placing = rank_dict[id]

			if id != rank_order[-1]:
				next_placing = rank_dict[rank_order[i+1]]

				if placing == next_placing:
					temp_list.append(id)
					try:
						score += points[rank][point_ref]
						rank += 1
						continue
					except:
						continue

			if temp_list != []:
				temp_list.append(id)
				try:
					score += points[rank][point_ref]
				except IndexError:
					pass
				rank += 1

				share = len(temp_list)
				cp = round(float(score/share), 2)

				if cp.is_integer():
					cp = int(cp)

				for id in temp_list:
					id_dict[id]['CP'] = cp

				score = 0.0
				temp_list = []

			elif temp_list == []:

				try:
					cp = points[rank][point_ref]
					id_dict[id]['CP'] = cp
					rank += 1
				except IndexError:
					PopBox().showwarning(title='Placing Error',
						message='Check placings, there may be a tie and you have entered the wrong next placing\n\ni.e. if you have a tie for 2nd but you have also put down a 3rd place')
					return
		# PP
		table_ref = self.db.cursor.execute('''SELECT round, title FROM %s
							WHERE round!="RR"
							AND title!="%s"''' % (self.event_table,
										self.current_round)).fetchall()
		for tup in table_ref:

			table = db_table('%s %s' % (self.event, tup[0]))
			round_name = tup[1]

			id_list = self.db.cursor.execute('''SELECT id, time_measure,
									length, height FROM %s
							WHERE pp=1''' % table).fetchall()

			for tupp in id_list:

				id = tupp[0]
				if id not in id_dict:
					id_dict[id] = {'PP':ppp, 'CP':0, 'best':0, 'rank':None}
				else:
					id_dict[id]['PP'] += ppp


				if self.contest_type == 'time':
					best = tupp[1]

					if best not in [None, 0]:
						if id_dict[id]['best'] == 0 or best < id_dict[id]['best']:
							id_dict[id]['best'] = best
				elif self.contest_type == 'distance':
					if self.contest in ['High Jump', 'Pole Vault']:

						if tupp[3] != None:

							input = db_height(tupp[3], db_in=False)
							best = db_vertical(input, self.tournament, self.event, 
											round_name, best=True)	

							if best not in [None, 0] and best > id_dict[id]['best']:
								id_dict[id]['best'] = best

					else:
						if tupp[2] != None:

							best = sorted(db_score(tupp[4], None, 
											'distance', db_in=False))[-1]
							if best not in [0.0, 0] and best > id_dict[id]['best']:
								id_dict[id]['best'] = best

		#
		if first_rank != True:
			PopBox().showwarning(title='Placing Error',
				message='No 1st place has been awarded')
			return




		# db update
		r_check = self.db.cursor.execute('''SELECT title FROM %s
						WHERE round="RR"''' % db_table(self.event)).fetchone()

		if r_check == None:
			self.db.cursor.execute('''INSERT INTO %s
			VALUES("RR", "%s Results", "", "", "")''' % (db_table(self.event),
														self.event))
			self.db.connection.commit()

		table = db_table('%s RR' % self.event)
		self.db.cursor.execute('''DROP TABLE IF EXISTS %s''' % table)
		self.db.cursor.execute('''CREATE TABLE %s
					(id TEXT, best REAl, rank INT, pp INT, cp REAL)''' % table)

		var = []
		for id in id_dict:

			var.append([id, id_dict[id]['best'], id_dict[id]['rank'], 
							id_dict[id]['PP'], id_dict[id]['CP']])
		self.db.cursor.executemany('''INSERT INTO %s
							VALUES(?, ?, ?, ?, ?)''' % table, var)
		self.db.connection.commit()

		# results table
		self.db.cursor.execute('''CREATE TABLE IF NOT EXISTS results 
									(event TEXT PRIMARY KEY)''')
		
		result_check = self.db.cursor.execute('''SELECT *
						FROM results
						WHERE event="%s"''' % self.event).fetchone()

		if result_check in [None, []]:
			self.db.cursor.execute('''INSERT INTO results 
								VALUES ("%s")''' % self.event)
			self.db.connection.commit()

		# engine
		value = '%s Results' % self.event
		self.build_display_structure(value)
		self.base_title.add_option(value)
		self.base_title.text = value

	def command_save(self, i):

		if len(self.unsaved) > 0:
			self.process_save() # checks record on saving

		self.save.bg = [0.5, 0.5, 0.5, 1]
		self.save.text = ''

	def command_scroll_height(self, obj, value, **kwargs):

		results = kwargs.pop('results', False)

		if results == True:
			for scroll in self.scroll_menu:
				scroll.height = value - 50
			return


		for scroll in self.scroll_menu:
			scroll.height = value - 100

	def command_unsaved(self, obj, value, id):

		if id == 'vert':
			pass
		elif obj != self.menu[id][0]: # sets the PP
			self.menu[id][0].active = True

		if id not in self.unsaved:
			self.unsaved.append(id)

		if self.save.text != 'Unsaved':
			self.save.text = 'Unsaved'
			self.save.bg = [1, 0, 0, 1]

	def record_check(self, score, label, name):

		if score in [None, 0, 0.0, '']:
			return

		record = self.record_what
		if record in [None, 0, 0.0, '']:
			return

		if self.contest_type == 'time':

			if score < record:
				label.text = '%s (R)' % name
				label.color = [1, 0, 0, 1]
				return

		elif self.contest_type == 'distance':

			if self.contest in ['High Jump', 'Pole Vault']:

				height_ref = self.height_ref.output()
				heights = []
				for i in range(6):
					if 'O' in score[i]:
						heights.append(height_ref[i])
				if heights != []:
					best = sorted(heights)[-1]
				else:
					best = 0

				if best not in [None, 0, 0.0, '']:
			
					if best > record:
						label.text = '%s (R)' % name
						label.color = [1, 0, 0, 1]
						return

			else:

				if sorted(score)[-1] > record:

					label.text = '%s (R)' % name
					label.color = [1, 0, 0, 1]
					return

	def record_results(self, score):
		
		record = self.record_what

		for r in [record, score]:
			if r in [None, '', 0, 0.0]:
				return False

		if self.contest_type == 'time':

			if score < record:
				return True

		elif self.contest_type == 'distance':

			if score > record:
				return True

	### BUILD - BASE BUILDS ###

	def build_base_title(self):

		box = BoxLayout(size_hint_y=None, height=100)

		# program
		event_box = BoxLayout(orientation='vertical', 
								size_hint_x=None, width=100)


		event_box.add_widget(Label(text='Event Number'))

		self.title_number = Label()
		event_box.add_widget(self.title_number)

		event_box.add_widget(Label(text='Event Time'))
		
		self.title_time = Label()
		event_box.add_widget(self.title_time)

		box.add_widget(event_box)

		# title
		rounds = self.db.cursor.execute('''SELECT * FROM %s
						ORDER BY round ASC''' % db_table(self.event)).fetchall()
		options = []
		for r in rounds:
			options.append(r[1])
			if r == rounds[-1]:
				title = r[1]

		self.current_round = title # sets default

		self.base_title = OptionMenu(font_size=30, options=options, main=title)
		self.base_title.bind(text=self.command_change_round)
		box.add_widget(self.base_title)

		# records
		self.record_what = self.event_details[3]
		record_what = calculate_str_record(self.record_what, self.contest_type)
		record_who = str(self.event_details[4])
		record_when = str(self.event_details[5])
		record_box = BoxLayout(orientation='vertical', 
								size_hint_x=None, width=100)
		for r in ['Record', record_what, record_who, record_when]:
			record_box.add_widget(Label(text=r))
		box.add_widget(record_box)

		# engine
		self.add_widget(box)

	def build_base_display(self):

		self.base_display = GridLayout(cols=1)
		self.add_widget(self.base_display)

	def build_display_structure(self, value):

		## Settings - Reset
		self.current_round = value

		rd = self.db.cursor.execute('''SELECT round, height FROM %s
					WHERE title="%s"''' % (db_table(self.event),
											self.current_round)).fetchone()
		self.current_table = db_table('%s %s' % (self.event, rd[0]))
		if rd[1] not in [None, '', 0]:
			self.height_list = db_score(rd[1], False, 'distance', db_in=False)
		else:
			self.height_list = [0, 0, 0, 0, 0, 0]
		self.menu = {}
		self.unsaved = []
		self.scroll_menu = []		

		if rd[0] == 'RR':
			self.title_number.text = ''
			self.title_time.text = ''
			self.build_results()
			return

		# program
		if self.current_round == self.event:
			event_number = str(self.event_details[6])
			event_time = str(self.event_details[7])
		else:
			try:
				prog = self.db.cursor.execute('''SELECT program_number, program_time
								FROM %s
								WHERE title="%s"''' % (self.event_table,
												self.current_round)).fetchone()
				event_number = str(prog[0])
				event_time = str(prog[1])
			except:
				event_number = ''
				event_time = ''
		self.title_number.text = event_number
		self.title_time.text = event_time

		# record

		structure = GridLayout(cols=1)
		structure.bind(height=self.command_scroll_height)

		header = BoxLayout(size_hint_y=None, height=50)
		structure.add_widget(header)
		
		scroll = ScrollView(size_hint_y=None, height=200)
		self.scroll_menu.append(scroll)
		structure.add_widget(scroll)
		self.current_body = GridLayout(cols=1, size_hint_y=None, height=0)
		scroll.add_widget(self.current_body)

		footer = BoxLayout(size_hint_y=None, height=50)
		structure.add_widget(footer)

		# header - house, name, pp, score, placing

		type_box = BoxLayout(orientation='vertical', size_hint_x=None, width=60)
		type_box.add_widget(Label(text=self.type))
		type_btn = Button(text='Add Entry', font_size=10, size_hint_y=None,
							height=20)
		type_btn.bind(on_press=self.command_add_entry)
		type_box.add_widget(type_btn)

		header.add_widget(type_box)
		if self.team == False:
			header.add_widget(Label(text='Competitor'))
		elif self.team == True:
			header.add_widget(Label(text='Team'))
		header.add_widget(Label(text='PP', size_hint_x=None, width=40))
		header.add_widget(self.build_score())

		place_box = BoxLayout(orientation='vertical', size_hint_x=None, width=60)
		place_box.add_widget(Label(text='Placing'))
		place_btn = Button(text='Calculate', font_size=10, size_hint_y=None,
							height=20)
		place_btn.bind(on_press=self.command_calculate)
		place_box.add_widget(place_btn)
		header.add_widget(place_box)

		# body
		self.build_body()

		# footer - unsaved, save, 
		self.save = LabelC(size_hint_x=None, width=100)
		footer.add_widget(self.save)
		save_btn = Button(text='Save')
		save_btn.bind(on_press=self.command_save)
		footer.add_widget(save_btn)
		add_btn = Button(text='Add Round', size_hint_x=None, width=100)
		add_btn.bind(on_press=self.command_add_round)
		footer.add_widget(add_btn)	
		final = Button(text='Complete', size_hint_x=None, width=100)
		final.bind(on_press=self.command_complete)
		footer.add_widget(final)

		self.current_display = structure
		self.change_display()

	def build_score(self):

		if self.contest_type == 'time':
			return Label(text='Time')
		elif self.contest_type == 'distance':
			if self.contest in ['High Jump', 'Pole Vault']:
				box = BoxLayout(orientation='vertical', size_hint_x=None, width=450)
				box.add_widget(Label(text='Heights (m)', size_hint_y=None, height=20))
				
				vert = ScoreVertical(input=self.height_list, size_hint_x=None, width=450)
				vert.bind(status=lambda obj, value: self.command_unsaved(obj, value, 'vert'))
				self.height_ref = vert
				box.add_widget(vert)

				return box

			else:
				box = BoxLayout(orientation='vertical', size_hint_x=None,
									width=420)
				box.add_widget(Label(text='Attempts (m)'))

				att = BoxLayout()
				box.add_widget(att)

				for i in range(1, 7):
					attempt = calculate_ordinal(i)
					att.add_widget(Label(text=attempt, size_hint_x=None,
									width=60))
				att.add_widget(Label(text='Best', size_hint_x=None, width=60))

				return box

	def build_body(self):

		# this really directs the creation of body_boxes and their position

		round_details = self.db.cursor.execute('''SELECT *
							FROM %s''' % self.current_table).fetchall()
		id_dict = {} # id, name, type, pp, score, placing
		data = {}
		for r in round_details:

			id_details = self.db.cursor.execute('''SELECT *
							FROM individuals
							WHERE id="%s"''' % r[0]).fetchone()
			id = r[0]
			fn = id_details[1]
			sn = id_details[2]
			type = id_details[5]
			pp = r[1]
			if self.contest_type == 'time':
				score = r[2]
			elif self.contest_type == 'distance':

				if self.contest in ['Pole Vault', 'High Jump']:
					score = r[4]
				else:
					score = r[3]

				if score not in [0, '', None]:
					score = db_score(score, self.contest, self.contest_type,
										db_in=False)

			placing = r[5]

			id_dict[id] = ['%s %s' % (fn, sn), type, pp, score, placing]
			data[id] = '%s %s' % (sn, fn)

		id_list = sorted(data, key=data.get)

		for id in id_list:
			self.build_body_box(id, id_dict[id])

	def build_body_box(self, id, details):

		name = details[0]
		type = details[1]
		col = self.typec[type]
		typename = type.upper()[:3]
		
		pp = db_boole(details[2], db_in=False)
		score = details[3]
		placing = details[4]
		if placing in [0, None]:
			placing = ''
		else:
			placing = str(placing)

		box = BoxLayout(size_hint_y=None, height=40)
		ref = []

		# Type/Name
		box.add_widget(LabelC(text=typename, bg=col[0], color=col[1], 
				size_hint_x=None, width=60))
		name_label = Label(text=name)
		box.add_widget(name_label)

		self.record_check(score, name_label, name)

		# PP
		chk = CheckBoxA(active=pp, size_hint_x=None, width=40)
		chk.bind(active=lambda obj, value,
					id=id:self.command_unsaved(obj, value, id))
		box.add_widget(chk)
		ref.append(chk)

		# Score
		scr = self.build_score_box(score)
		scr.bind(status=lambda obj, value, 
					id=id:self.command_unsaved(obj, value, id))
		box.add_widget(scr)
		ref.append(scr)

		# Placing
		plc = TextInputE(text=placing, run_status=True, font_size=20, multiline=False,
						input_filter='int', size_hint_x=None, width=60)
		plc.bind(status=lambda obj, value, 
					id=id:self.command_unsaved(obj, value, id))
		box.add_widget(plc)
		ref.append(plc)
		ref += [name_label, name]

		## Engine
		self.current_body.add_widget(box)
		self.current_body.height += 40
		self.menu[id] = ref

	def build_score_box(self, score):

		return Label()

	def build_results(self):

		table = db_table('%s RR' % self.event)
		res = self.db.cursor.execute('''SELECT * FROM %s''' % table).fetchall()
		results = {}
		var = []
		for tup in res:
			results[tup[0]] = [tup[1], tup[2], tup[3], tup[4]]

		indiv = []

		for id in results:

			id_tup = self.db.cursor.execute('''SELECT id, firstname,
													surname, type
													FROM individuals
											WHERE id="%s"''' % id).fetchone()
			indiv.append(id_tup)

		for tup in indiv:
			results[tup[0]] += [tup[1], tup[2], tup[3]]

		# structure
		structure = BoxLayout(spacing=10)

		# summary
		summary = self.build_results_summary(results)

		# points
		type_scores = {}
		for type in self.typec:
	
			type_scores[type] = 0
		for id in results:
			pp = 0
			cp = 0

			if results[id][2] != None:
				pp = results[id][2]
			if results[id][3] != None:
				cp = results[id][3]
			type = results[id][6]
			type_scores[type] += (pp + cp)
		scores = self.build_results_points(type_scores)
 
 		## Display
		structure.add_widget(summary)
		structure.add_widget(scores)
		
		## Engine
		self.current_display = structure
		self.change_display()

	def build_results_summary(self, results):

		box = GridLayout(cols=1)

		# header
		header = BoxLayout(size_hint_y=None, height=50)
		box.add_widget(header)

		header.add_widget(Label(text='Rank', size_hint_x=None, width=50))
		header.add_widget(Label(text='Competitor'))
		header.add_widget(Label(text='Best', size_hint_x=None, width=100))
		header.add_widget(LabelWrap(text='Participation Points', 
								size_hint_x=None, width=100))
		header.add_widget(LabelWrap(text='Competition Points',
								size_hint_x=None, width=100))
		header.add_widget(LabelWrap(text='Total Points',
								size_hint_x=None, width=100))

		# scroll
		scroll = ScrollView(size_hint_y=None, height=300)
		box.add_widget(scroll)
		self.scroll_menu.append(scroll)

		display = GridLayout(cols=1, size_hint_y=None, height=0)
		scroll.add_widget(display)

		# 
		menu = {}
		rank_data = {}
		best_data = {}
		alpha_data = {}

		for id in results:
			ref = results[id]

			best = ref[0]
			rk = ref[1]
			rank = ''
			pp = 0
			if ref[2] != None:
				pp = ref[2]
			cp = 0
			if ref[3] != None:
				cp = ref[3]
			total = pp + cp
			fn = ref[4]
			sn = ref[5]
			name = '%s %s' % (fn, sn)
			type = ref[6]
			bg = self.typec[type][0]
			txt = self.typec[type][1]

			if rk not in [None, 0, '']:
				rank_data[id] = rk
				rank = calculate_ordinal(rk)

			elif best not in [None, 0, 0.0]:
				best_data[id] = best

			else:
				alpha_data[id] = '%s %s' % (sn, fn)

			comp = BoxLayout(size_hint_y=None, height=40)

			comp.add_widget(LabelC(text=rank, font_size=20, bg=bg, color=txt, 
						size_hint_x=None, width=50))
			
			rec = self.record_results(best)

			if rec == True:
				comp.add_widget(Label(text='%s (R)' % name, color=[1, 0, 0, 1]))
			else:
				comp.add_widget(Label(text=name))
			best = convert_score(best, self.contest_type)
			for h in [best, pp, cp, total]:

				if h in [None, 0, 0.0]:
					h = ''
				else:
					try:
						if h.is_integer():
							h = str(int(h))
						else:
							h = str(h)
					except:
						h = str(h)
				comp.add_widget(Label(text=h, size_hint_x=None, width=100))

			menu[id] = comp

		rank = sorted(rank_data, key=rank_data.get)
		for id in rank:

			display.add_widget(menu[id])
			display.height += 40

		if self.contest_type == 'time':
			best = sorted(best_data, key=best_data.get)
		elif self.contest_type == 'distance':
			best = sorted(best_data, key=best_data.get, reverse=True)
		for id in best:

			display.add_widget(menu[id])
			display.height += 40

		for id in sorted(alpha_data, key=alpha_data.get):

			display.add_widget(menu[id])
			display.height += 40

		return box

	def build_results_points(self, type_scores):

		type_order = sorted(type_scores, key=type_scores.get, reverse=True)

		box = GridLayout(cols=1, size_hint_x=None, width=100)

		box.add_widget(LabelWrap(text='%s Points' % self.type, font_size=20,
					size_hint_y=None, height=50))

		scroll = ScrollView(size_hint_y=None, height=300)
		box.add_widget(scroll)
		self.scroll_menu.append(scroll)
		box.bind(height=lambda obj, 
			value:self.command_scroll_height(obj, value, results=True))

		display = GridLayout(cols=1, size_hint_y=None, height=10, 
								padding=[0, 10, 0, 0], spacing=10)
		scroll.add_widget(display)

		for type in type_order:
			score = type_scores[type]

			try:
				if score.is_integer():
					score = int(score)
			except:
				pass

			score = str(score)
			bg = self.typec[type][0]
			txt = self.typec[type][1]

			point = BoxLayout(orientation='vertical',
								size_hint_y=None, height=70)

			point.add_widget(LabelC(text=type, bg=bg, color=txt, 
									size_hint_y=None, height=20))
			point.add_widget(LabelC(text=score, font_size=20, bg=bg, color=txt))

			display.add_widget(point)
			display.height += 80

		return box

	### PROCESS ###

	def process_save(self):

		var = []
		for id in self.unsaved:

			ref = self.menu[id] # ref = [id, pp, score, placing, label, name]

			for r in range(3):
				ref[r].status = True

			pp = db_boole(ref[0].active)

			score = db_score(ref[1].output(), self.contest, self.contest_type)

			## Record Check
			self.record_check(ref[1].output(), ref[3], ref[4])

			if pp == 0:
				ref[2].text = ''

			placing = ref[2].text
			if placing != '':
				placing = int(placing)

			var.append([pp, score, placing, id])

		if self.contest_type == 'time':
			measure = 'time_measure'
		else:
			if self.contest in ['High Jump', 'Pole Vault']:
				measure = 'height'
			else:
				measure = 'length'

		self.db.cursor.executemany('''UPDATE %s
							SET pp=?, %s=?, placing=?
							WHERE id=?''' % (self.current_table, measure), var)
		self.db.connection.commit()
		self.unsaved = []

class EventTime(EventBase):

	def __init__(self, *args, **kwargs):
		super(EventTime, self).__init__(**kwargs)

	### FUNCTIONS ###

	### BUILD - OVERRIDE ###

	def build_score_box(self, score):

		if score in [None, '']:
			score = 0

		return ScoreTime(input=score, run_status=True)

	### PROCESS - OVERRIDE ###

class EventTimeTeam(EventTime):

	### FUNCTIONS ###

	def command_add_entry(self, i):

		if self.official == False:

			EntryAddTeam(style='Entry', tournament=self.tournament,
				event=self.event, event_round=self.current_round,
				event_table=self.current_table, widget=self)

	def command_add_round(self, i):

		self.command_save(None)
		
		EventAddRoundTeam(tournament=self.tournament, event=self.event,
						current_title=self.current_round,
						current_table=self.current_table,
						widget=self)

	def command_complete(self, i):

		if i == False:
			return

		# check conditions
		check = []
		if i != True:
			check_round = self.db.cursor.execute('''SELECT round
									FROM %s''' % self.event_table).fetchall()
			count = 0
			for tup in check_round:
				if tup[0] == 'RR':
					check.append('This event has already been completed')
					continue
				count += 1
			if count > 1:
				check.append('There is more than one round in this event, if you complete here you will complete using the placings from %s' % self.current_round)

			if check != []:
				check.append('Do you wish to continue?')
				message = '\n\n'.join(check)
				PopBox().askyesno(title='Complete %s' % self.event,
					function=self.command_complete,
					message=message)
				return

		# save 
		self.process_save()

		# CP
		points = self.db.cursor.execute('''SELECT *
								FROM template_points''').fetchall()

		point_ref = 1
		if self.team == True:
			point_ref = 2

		if points[0][point_ref] not in [None, '']:
			ppp = points[0][point_ref]
		else:
			ppp = 0

		current = self.db.cursor.execute('''SELECT team, pp, placing, 
								time_measure, length, height, type
								FROM %s''' % self.current_table).fetchall()
		id_dict = {}
		rank_dict = {}

		for tup in current:

			id = tup[0]
			pp = tup[1]
			type = tup[6]

			if pp  == 0:
				continue # no participation

			placing = tup[2]

			dict = {'best':0, 'rank':placing, 'PP':0, 'CP':0, 'type':type}

			if pp == 1:
				dict['PP'] += ppp

			if placing < len(points) and placing != None:
				rank_dict[id] = placing

			if self.contest_type == 'time':
				dict['best'] = tup[3]
			elif self.contest_type == 'distance':
				if self.contest in ['High Jump', 'Pole Vault']:

					if tup[5] != None:

						input = db_height(tup[5], db_in=False)
						best = db_vertical(input, self.tournament, self.event, 
										self.current_round, best=True)	
						dict['best'] = best

				else:
					if tup[4] != None:
						dict['best'] = sorted(db_score(tup[4], None, 
									'distance', db_in=False))[-1]

			id_dict[id] = dict

		rank_order = sorted(rank_dict, key=rank_dict.get)
		
		rank = 1
		score = 0.0
		temp_list = []
		for i in range(len(rank_order)):

			id = rank_order[i]
			placing = rank_dict[id]

			if id != rank_order[-1]:
				next_placing = rank_dict[rank_order[i+1]]

				if placing == next_placing:
					temp_list.append(id)

					try:
						score += points[rank][point_ref]
						rank += 1
						continue
					except:
						continue



			if temp_list != []:
				temp_list.append(id)
				try:
					score += points[rank][point_ref]
				except IndexError:
					pass
				rank += 1

				share = len(temp_list)
				cp = round(float(score/share), 2)

				if cp.is_integer():
					cp = int(cp)

				for id in temp_list:
					id_dict[id]['CP'] = cp

				score = 0.0
				temp_list = []

			elif temp_list == []:

				try:
					cp = points[rank][point_ref]
					id_dict[id]['CP'] = cp
					rank += 1
				except IndexError:
					PopBox().showwarning(title='Placing Error',
						message='Check placings, there may be a tie and you have entered the wrong next placing\n\ni.e. if you have a tie for 2nd but you have also put down a 3rd place')
					return


		# PP
		table_ref = self.db.cursor.execute('''SELECT round FROM %s
							WHERE round!="RR"
							AND title!="%s"''' % (self.event_table,
										self.current_round)).fetchall()
		for tup in table_ref:

			table = db_table('%s %s' % (self.event, tup[0]))

			id_list = self.db.cursor.execute('''SELECT team, time_measure,
									length, height, type FROM %s
							WHERE pp=1''' % table).fetchall()

			for tupp in id_list:
				id = tupp[0]
				type = tupp[4]
				if id not in id_dict:
					id_dict[id] = {'PP':ppp, 'CP':0, 'best':0, 'rank':None,
									'type':type}
				else:
					id_dict[id]['PP'] += ppp


				if self.contest_type == 'time':
					best = tupp[1]
					if best not in [None, 0] and best < id_dict[id]['best']:
						id_dict[id]['best'] = best
				elif self.contest_type == 'distance':
					if self.contest in ['High Jump', 'Pole Vault']:

						if tupp[3] != None:

							input = db_height(tupp[3], db_in=False)
							best = db_vertical(input, self.tournament, self.event, 
											self.current_round, best=True)	
							if best not in [None, 0] and best > id_dict[id]['best']:
								id_dict[id]['best'] = best

					else:
						if tupp[4] != None:

							best = sorted(db_score(tupp[4], None, 
											'distance', db_in=False))[-1]
							if best not in [0.0, 0] and best > id_dict[id]['best']:
								id_dict[id]['best'] = best

		# db update
		r_check = self.db.cursor.execute('''SELECT title FROM %s
						WHERE round="RR"''' % db_table(self.event)).fetchone()

		if r_check == None:
			self.db.cursor.execute('''INSERT INTO %s
			VALUES("RR", "%s Results", "", "", "")''' % (db_table(self.event),
														self.event))
			self.db.connection.commit()

		table = db_table('%s RR' % self.event)
		self.db.cursor.execute('''DROP TABLE IF EXISTS %s''' % table)
		self.db.cursor.execute('''CREATE TABLE %s
		(team TEXT, best REAl, rank INT, pp INT, cp REAL, type TEXT)''' % table)

		var = []
		for id in id_dict:

			var.append([id, id_dict[id]['best'], id_dict[id]['rank'], 
							id_dict[id]['PP'], id_dict[id]['CP'],
							id_dict[id]['type']])
		self.db.cursor.executemany('''INSERT INTO %s
							VALUES(?, ?, ?, ?, ?, ?)''' % table, var)
		self.db.connection.commit()

		# results table
		self.db.cursor.execute('''CREATE TABLE IF NOT EXISTS results 
									(event TEXT PRIMARY KEY)''')
		
		try:
			self.db.cursor.execute('''INSERT OR IGNORE INTO results 
								VALUES ("%s")''' % self.event)
			self.db.connection.commit()
		except:
			pass # Database Issue (maybe clock reschedule?) 

		# engine
		value = '%s Results' % self.event
		self.build_display_structure(value)
		self.base_title.add_option(value)
		self.base_title.text = value

	### BUILD ###

	def build_body(self):

		round_details = self.db.cursor.execute('''SELECT *
							FROM %s''' % self.current_table).fetchall()
		id_dict = {} # team, pp, time, length, height, placing, type
		data = {}
		for r in round_details:

			team = r[0]
			type = r[6]

			pp = r[1]
			if self.contest_type == 'time':
				score = r[2]
			elif self.contest_type == 'distance':
				if self.contest in ['Pole Vault', 'High Jump']:
					score = r[4]
				else:
					score = r[3]

				if score not in [0, '', None]:
					score = db_score(score, self.contest, self.contest_type,
										db_in=False)

			placing = r[5]

			id_dict[team] = [type, pp, score, placing]
			data[team] = type

		team_list = sorted(data, key=data.get)

		for team in team_list:
			self.build_body_box(team, id_dict[team])

	def build_body_box(self, name, details):

		type = details[0]
		col = self.typec[type]
		typename = type.upper()[:3]
		
		pp = db_boole(details[1], db_in=False)
		score = details[2]
		placing = details[3]
		if placing in [0, None]:
			placing = ''
		else:
			placing = str(placing)

		box = BoxLayout(size_hint_y=None, height=40)
		ref = []

		# Type/Name
		box.add_widget(LabelC(text=typename, bg=col[0], color=col[1], 
				size_hint_x=None, width=60))
		box.add_widget(LabelC(text=name, bg=col[0], color=col[1]))

		# PP
		chk = CheckBoxA(active=pp, size_hint_x=None, width=40)
		chk.bind(active=lambda obj, value,
					name=name:self.command_unsaved(obj, value, name))
		box.add_widget(chk)
		ref.append(chk)

		# Score
		scr = self.build_score_box(score)
		scr.bind(status=lambda obj, value, 
					name=name:self.command_unsaved(obj, value, name))
		box.add_widget(scr)
		ref.append(scr)

		# Placing
		plc = TextInputE(text=placing, run_status=True, font_size=20, multiline=False,
						input_filter='int', size_hint_x=None, width=60)
		plc.bind(status=lambda obj, value, 
					name=name:self.command_unsaved(obj, value, name))
		box.add_widget(plc)
		ref.append(plc)

		## Engine
		self.current_body.add_widget(box)
		self.current_body.height += 40
		self.menu[name] = ref

	def build_results(self):

		table = db_table('%s RR' % self.event)
		res = self.db.cursor.execute('''SELECT * FROM %s''' % table).fetchall()
		results = {}
		var = []

		for tup in res:
			results[tup[0]] = [tup[1], tup[2], tup[3], tup[4], tup[5]]

		# structure
		structure = BoxLayout(spacing=10)

		# summary
		summary = self.build_results_summary(results)

		# points
		type_scores = {}
		for type in self.typec:
	
			type_scores[type] = 0
		for id in results:
			pp = 0
			cp = 0

			if results[id][2] != None:
				pp = results[id][2]
			if results[id][3] != None:
				cp = results[id][3]
			type = results[id][4]
			type_scores[type] += (pp + cp)
		scores = self.build_results_points(type_scores)
 
 		## Display
		structure.add_widget(summary)
		structure.add_widget(scores)
		
		## Engine
		self.current_display = structure
		self.change_display()

	def build_results_summary(self, results):

		box = GridLayout(cols=1)

		# header
		header = BoxLayout(size_hint_y=None, height=50)
		box.add_widget(header)

		header.add_widget(Label(text='Rank', size_hint_x=None, width=50))
		header.add_widget(Label(text='Team'))
		header.add_widget(Label(text='Best', size_hint_x=None, width=100))
		header.add_widget(LabelWrap(text='Participation Points', 
								size_hint_x=None, width=100))
		header.add_widget(LabelWrap(text='Competition Points',
								size_hint_x=None, width=100))
		header.add_widget(LabelWrap(text='Total Points',
								size_hint_x=None, width=100))

		# scroll
		scroll = ScrollView(size_hint_y=None, height=300)
		box.add_widget(scroll)
		self.scroll_menu.append(scroll)

		display = GridLayout(cols=1, size_hint_y=None, height=0)
		scroll.add_widget(display)

		# 
		menu = {}
		rank_data = {}
		best_data = {}
		alpha_data = {}

		for id in results:
			ref = results[id]

			best = ref[0]
			rk = ref[1]
			rank = ''
			pp = 0
			if ref[2] != None:
				pp = ref[2]
			cp = 0
			if ref[3] != None:
				cp = ref[3]
			total = pp + cp
			name = id
			type = ref[4]
			bg = self.typec[type][0]
			txt = self.typec[type][1]

			if rk not in [None, 0, '']:
				rank_data[id] = rk
				rank = calculate_ordinal(rk)

			elif best not in [None, 0, 0.0]:
				best_data[id] = best

			else:
				alpha_data[id] = type

			comp = BoxLayout(size_hint_y=None, height=40)

			comp.add_widget(LabelC(text=rank, font_size=20, bg=bg, color=txt, 
						size_hint_x=None, width=50))
			
			rec = self.record_results(best)

			if rec == True:
				comp.add_widget(Label(text='%s (R)' % name, color=[1, 0, 0, 1]))
			else:
				comp.add_widget(Label(text=name))
			best = convert_score(best, self.contest_type)
			for h in [best, pp, cp, total]:

				if h in [None, 0, 0.0]:
					h = ''
				else:
					try:
						if h.is_integer():
							h = str(int(h))
						else:
							h = str(h)
					except:
						h = str(h)
				comp.add_widget(Label(text=h, size_hint_x=None, width=100))

			menu[id] = comp

		rank = sorted(rank_data, key=rank_data.get)
		for id in rank:

			display.add_widget(menu[id])
			display.height += 40

		if self.contest_type == 'time':
			best = sorted(best_data, key=best_data.get)
		elif self.contest_type == 'distance':
			best = sorted(best_data, key=best_data.get, reverse=True)
		for id in best:

			display.add_widget(menu[id])
			display.height += 40

		for id in sorted(alpha_data, key=alpha_data.get):

			display.add_widget(menu[id])
			display.height += 40

		return box

	### PROCESS ###

	def process_save(self):

		var = []
		for id in self.unsaved:

			ref = self.menu[id]

			for r in ref:
				r.status = True

			pp = db_boole(ref[0].active)

			score = db_score(ref[1].output(), self.contest, self.contest_type)
			placing = ref[2].text
			if placing != '':
				placing = int(placing)

			var.append([pp, score, placing, id])

		if self.contest_type == 'time':
			measure = 'time_measure'
		else:
			if self.contest in ['High Jump', 'Pole Vault']:
				measure = 'height'
			else:
				measure = 'length'

		self.db.cursor.executemany('''UPDATE %s
							SET pp=?, %s=?, placing=?
						WHERE team=?''' % (self.current_table, measure), var)
		self.db.connection.commit()
		self.unsaved = []

class EventDistance(EventBase):

	def __init__(self, *args, **kwargs):
		super(EventDistance, self).__init__(**kwargs)

	### FUNCTIONS ###

	### BUILD - OVERRIDE ###

	def build_score_box(self, score):

		if score in [None, 0, '']:
			score = [0, 0, 0, 0, 0, 0]

		return ScoreLength(input=score, run_status=True, size_hint_x=None, width=420)

	### PROCESS - OVERRIDE ###

class EventDistanceHeight(EventBase):

	def __init__(self, *args, **kwargs):
		super(EventDistanceHeight, self).__init__(**kwargs)

	### FUNCTIONS ###

	def command_calculate_placing(self):

		data = {}
		for id in self.menu:
			ref = self.menu[id]
			ref[2].text = ''
			if ref[0].active == True:
				score = ref[1].output()
				data[id] = score
				if id not in self.unsaved:
					self.unsaved.append(id)
					if self.save.text != 'Unsaved':
						self.save.text = 'Unsaved'
						self.save.bg = [1, 0, 0, 1]

		done_list = []
		error = False
		rank = 1
		tie = 1
		for i in reversed(range(6)):

			id_list = []
			for id in data:
				ref = data[id][i]
				if 'O' in ref:
					if id not in done_list:
						id_list.append(id)
						done_list.append(id)

			if len(id_list) == 1:
				self.menu[id_list[0]][2].text = str(rank)
				rank += 1 # 1 at this height
				tie = 1 # ( just to make sure)
				continue

			elif id_list == []:

				continue # None at this height

			elif id_list != []: # Multiple at same height

				for r in range(0, 3):
					same = []
					for id in id_list:
						if data[id][i][r] == 'O':
							same.append(id)

					if len(same) == 1: # 1 at this attempt
						self.menu[same[0]][2].text = str(rank)
						rank += 1
						continue

					elif same == []: # None at this attempt

						continue

					elif same != []: # Multiple at this attempt

						dato = {}
						for id in same:
							miss = 0
							for triple in data[id]:
								for mark in triple:
									if mark == 'X':
										miss += 1
							dato[id] = miss

						sort = sorted(dato, key=dato.get)

						for s in range(len(sort)):
							
							id = sort[s]
							self.menu[id][2].text = str(rank)
							
							if sort[s] != sort[-1]:

								id_next = sort[s+1]
								if dato[id] == dato[id_next]: # tie
									tie += 1
									if rank == 1:
										error = True
								else:
									rank += tie
									tie = 1
							
							else:
								rank += tie
								tie = 1
								# work even if tie > 1

		for id in data:
			if id not in done_list:
				self.menu[id][2].text = str(rank)

		if error == True:
			PopBox().showwarning(title='First Place Tie',
				message='Tie for first place, run a jump off and manually change the placing')

	### BUILD - OVERRIDE ###

	def build_score_box(self, score):

		if score in [None, 0, '']:
			score = [[' ',' ', ' '],[' ',' ', ' '],[' ',' ', ' '],
					[' ',' ', ' '],[' ',' ', ' '],[' ',' ', ' ']]

		return ScoreHeight(input=score, run_status=True, 
							size_hint_x=None, width=450, 
							height_ref=self.height_ref)

	### PROCESS - OVERRIDE ###

	def process_save(self):

		var = []
		for id in self.unsaved:

			if id == 'vert':
				height = db_score(self.height_ref.output(), False, 'distance')
				table = db_table(self.event)

				self.db.cursor.execute('''UPDATE %s
					SET height="%s"
					WHERE title="%s"''' % (table, height, self.current_round))
				self.db.connection.commit()
				continue

			ref = self.menu[id]
			pp = db_boole(ref[0].active) # converts True-1, False-0
			score = db_height(ref[1].output())
			rank = ref[2].text

			self.record_check(ref[1].output(), ref[3], ref[4])

			var.append([pp, score, rank, id])

		self.db.cursor.executemany('''UPDATE %s
							SET pp=?, height=?, placing=?
							WHERE id=?''' % self.current_table, var)
		self.db.connection.commit()


		## Engine
		self.unsaved = []

class ScoreBoard(ScrollView):

	tournament = StringProperty('')

	def __init__(self, *args, **kwargs):
		super(ScoreBoard, self).__init__(**kwargs)

		## Setting
		self.size_hint_y = None
		self.height = Window.height - 150
		Window.bind(height=self.change_scroll_height)
		
		self.db = Database(self.tournament)

		details = self.db.cursor.execute('''SELECT * 
											FROM details''').fetchall()[0]
		self.stage = details[2]
		self.type = details[3]
		self.age = details[4]


		self.type_dict = {}
		colours = self.db.cursor.execute('''SELECT title, colour_bg, colour_text
									FROM template_groups
									WHERE category="type"''').fetchall()
		for c in colours:
			self.type_dict[c[0]] = [db_colour(c[1], db_in=False), 
								db_colour(c[2], db_in=False)]

		## Database
		self.score_menu = {}
		self.results_menu = {}
		self.special = False
	
		for type in self.type_dict:
			self.results_menu[type] = {'pp':0, 'cp':0, 'sp':0}
		self.events_menu = {}
		self.access_scores()

		## Display
		self.build_display()

		if self.stage == 'Complete': # next stage is archive
			self.build_score_heading('Results')		
			self.build_results_score()
			self.build_champions_heading()
		else:
			self.build_score_heading('Scoreboard')				
			self.build_scoreboard()
		self.build_results_heading()
		self.build_results()
		if self.special == True:
			self.build_special_heading()
			self.build_special()

	### FUNCTIONS ###

	def change_scroll_height(self, obj, value):

		self.height = value - 150

	def command_help(self, i):

		if i == 'Scoreboard':

			if self.stage == 'Complete':

				mes = "Results show which %s won and where each %s placed.\n\nThe scores will intially show the total; click a score box to show the Competition Points, Participation Points and Special Event Points." % (self.type, self.type)

			else:

				mes = "The Scoreboard will display the Tournament %s points. These can be refreshed by reclicking 'Main'.\n\nThe scores will intially show the total; click a score box to show the Competition Points, Participation Points and Special Event Points." % self.type

		elif i == 'Results':

			mes = "Results will appear as each event completes (though you will need to refresh this page for updates).\n\nOnly 1st, 2nd and 3rd places are shown here, in the case of a tie, up to 3 names per place will be displayed."

		elif i == 'Special':

			mes = "These represent points awarded to %ss for circumstances outside of the normal event points." % self.type

		elif i == 'Champions':

			mes = "This section shows the Grade Champion results.\n\nThese are the competitors who scored the most competition points available to their grade"

		PopBox().help_screen(title='%s' % self.tournament,
			message=mes)

	def command_score(self, obj, value):

		if value.x < obj.pos[0] or value.x > (obj.pos[0]+obj.size[0]):
			return
		if value.y < obj.pos[1] or value.y > (obj.pos[1]+obj.size[1]):
			return

		for type in self.score_menu:

			ref = self.score_menu[type]

			total = ref[0]
			pp = ref[1]
			cp = ref[2]
			scoreboard = ref[3]
			points = ref[4]
			sp = ref[5]

			if points.text == '':

				scoreboard.text = str(cp)
				points.text = 'Competition Points'

			elif points.text == 'Competition Points':

				scoreboard.text = str(pp)
				points.text = 'Participation Points'

			elif points.text == 'Participation Points':

				scoreboard.text = str(sp)
				points.text = 'Special Event Points'

			elif points.text == 'Special Event Points':

				scoreboard.text = str(total)
				points.text = ''

	### ACCESS ###

	def access_scores(self):

		self.db.cursor.execute('''CREATE TABLE IF NOT EXISTS results 
								(event TEXT PRIMARY KEY)''')
		self.db.connection.commit()

		scores = self.db.cursor.execute('''SELECT * FROM results''').fetchall()

		contest_list = calculate_contest_list(self.tournament)
		grade_list = calculate_grade_list(self.tournament)
		event_list = []
		for contest in contest_list:
			for grade in grade_list:
				event_list.append('%s %s' % (grade, contest))

		self.complete_list = []
		for event in event_list:
			for tup in scores:
				if event == tup[0]:
					self.complete_list.append(event)

		for event in self.complete_list:

			team = calculate_event_team(self.tournament, event)

			table = db_table('%s RR' % event)
			results = self.db.cursor.execute('''SELECT *
									FROM %s''' % table).fetchall()

			score_dict = {}

			first = []
			second = []
			third = []

			if team == False:

				for res in results:

					id = res[0]
					indiv = self.db.cursor.execute('''SELECT type, firstname, surname
									FROM individuals
									WHERE id="%s"''' % id).fetchone()

					type = indiv[0]
					fn = indiv[1]
					sn = indiv[2]
					rank = res[2]
					pp = res[3]
					cp = res[4]

					if type not in score_dict:
						score_dict[type] = {'pp':0, 'cp':0}

					if pp not in [None, '', 0, 0.0]:
						score_dict[type]['pp'] += pp
					if cp not in [None, '', 0, 0.0]:
						score_dict[type]['cp'] += cp

					if rank == 1:
						first.append([fn, sn, type])
					elif rank == 2:
						second.append([fn, sn, type])
					elif rank == 3:
						third.append([fn, sn, type])

			elif team == True:

				for res in results:

					team = res[0]
					rank = res[2]
					pp = res[3]
					cp = res[4]
					type = res[5]

					if type not in score_dict:
						score_dict[type] = {'pp':0, 'cp':0}

					if pp not in [None, '', 0, 0.0]:
						score_dict[type]['pp'] += pp
					if cp not in [None, '', 0, 0.0]:
						score_dict[type]['cp'] += cp

					if rank == 1:
						first.append([team, type])
					elif rank == 2:
						second.append([team, type])
					elif rank == 3:
						third.append([team, type])

			# add scores to self.results_menu
			for type in score_dict:

				pp = score_dict[type]['pp']
				cp = score_dict[type]['cp']

				self.results_menu[type]['pp'] += pp
				self.results_menu[type]['cp'] += cp

				try:
					special = self.db.cursor.execute('''SELECT points
								FROM special_events
								WHERE type="%s"''' % type).fetchall()

					count = 0
					for tup in special:
						count += float(tup[0])

					self.results_menu[type]['sp'] = count
					self.special = True

				except:
					self.results_menu[type]['sp'] = 0


			# add event details to self.event_menu
			self.events_menu[event] = {1:first, 2:second, 3:third}


		if self.complete_list == []:

			for type in self.results_menu:
				try:
					special = self.db.cursor.execute('''SELECT points
								FROM special_events
								WHERE type="%s"''' % type).fetchall()

					count = 0
					for tup in special:
						count += float(tup[0])

					self.results_menu[type]['sp'] = count
					self.special = True

				except:
					self.results_menu[type]['sp'] = 0			

	### BUILD ###

	def build_display(self):

		self.display = GridLayout(cols=1, spacing=10,
									size_hint_y=None, height=0)
		self.add_widget(self.display)

	def build_score_heading(self, title):

		box = BoxLayout(size_hint_y=None, height=50)

		box.add_widget(Label(text=title, font_size=30))

		btn = Button(text='?', font_size=30, size_hint_x=None, width=50)
		btn.bind(on_press=lambda i:self.command_help('Scoreboard'))
		box.add_widget(btn)

		self.display.add_widget(box)
		self.display.height += 50

	def build_scoreboard(self):

		type_num = len(self.type_dict)

		board = GridLayout(cols=4, spacing=10, size_hint_y=None, height=0)

		count = 1
		for type in self.type_dict:

			bg = self.type_dict[type][0]
			txt = self.type_dict[type][1]

			pp = int(self.results_menu[type]['pp'])
			cp = int(self.results_menu[type]['cp'])
			sp = int(self.results_menu[type]['sp'])
			total = int(pp + cp + sp)

			box = BoxC(bg=bg, orientation='vertical')

			t = LabelC(text=type, font_size=20, bg=bg, color=txt, 
							size_hint_y=None, height=40)
			s = LabelC(text=str(int(total)), font_size=50, bg=bg, color=txt)
			p = LabelC(text='', bg=bg, color=txt, size_hint_y=None, height=20)

			for label in [t, s, p]:
				box.add_widget(label)

			self.score_menu[type] = [total, pp, cp, s, p, sp]

			box.bind(on_touch_down=self.command_score)

			# engine
			board.add_widget(box)
			if count == 1:
				board.height += 150
				count += 1
			elif count == 4:
			
				count = 1
			else:
				
				count += 1

		# engine
		self.display.add_widget(board)
		self.display.height += board.height + 50

	def build_results_score(self):

		board = BoxLayout(orientation='vertical', padding=[0, 10, 0, 0],
							spacing=10, 
							size_hint_y=None, height=0)

		type_dict = {}
		order_dict = {}
		for type in self.type_dict:

			bg = self.type_dict[type][0]
			txt = self.type_dict[type][1]

			pp = int(self.results_menu[type]['pp'])
			cp = int(self.results_menu[type]['cp'])
			sp = int(self.results_menu[type]['sp'])
			total = (self.results_menu[type]['pp']) + (self.results_menu[type]['cp']) + (self.results_menu[type]['sp'])

			type_box = BoxLayout()
			box = BoxC(bg=bg, orientation='vertical')

			t = LabelC(text=type, font_size=30, bg=bg, color=txt)
			type_box.add_widget(t)
			type_box.add_widget(box)
			s = LabelC(text=str(total), font_size=50, bg=bg, color=txt)
			p = LabelC(text='', bg=bg, color=txt, size_hint_y=None, height=20)

			for label in [s, p]:
				box.add_widget(label)

			self.score_menu[type] = [total, pp, cp, s, p, sp]
			type_dict[type] = type_box
			order_dict[type] = total
			box.bind(on_touch_down=self.command_score)

		ranked = sorted(order_dict, key=order_dict.get, reverse=True)
		rank = 1
		tie = 1
		results = {}
		for i in range(len(ranked)):

			type = ranked[i]
			score = order_dict[type]
		
			if rank not in results:
				results[rank] = [type]
			else:
				results[rank].append(type)


			if type != ranked[-1]:
				next_score = order_dict[ranked[i+1]]
			else:
				next_score = 0

			if score == next_score:
				tie += 1
				continue
			elif tie != 1:
				rank += tie
				tie = 1
			else:
				rank += tie
				tie = 1

		for r in results:

			res = results[r]
			rank = calculate_ordinal(r)
			hgt = len(res) * 150

			box = BoxLayout(size_hint_y=None, height=hgt)
			box.add_widget(Label(text=str(rank), size_hint_x=None, width=100))

			side_box = BoxLayout(orientation='vertical')
			box.add_widget(side_box)
			for type in res:
				side_box.add_widget(type_dict[type])

			board.add_widget(box)
			board.height += hgt + 10

		# engine
		self.display.add_widget(board)
		self.display.height += board.height + 40

	def build_champions_heading(self):

		box = BoxLayout(orientation='vertical', size_hint_y=None, height=90)

		title = BoxLayout(size_hint_y=None, height=50)
		
		title.add_widget(Label(text='Grade Champions', font_size=30))
		btn = Button(text='?', font_size=30, size_hint_x=None, width=50)
		btn.bind(on_press=lambda i:self.command_help('Champions'))
		title.add_widget(btn)

		box.add_widget(title)

		heading = BoxLayout(size_hint_y=None, height=40)

		for h in ['Grade', '1st', '2nd', '3rd']:
			heading.add_widget(Label(text=h, font_size=20))

		box.add_widget(heading)

		body, height = self.build_champions()
		box.add_widget(body)
		box.height += height

		self.display.add_widget(box)
		self.display.height += box.height

	def build_champions(self):

		## Settings

		box = GridLayout(cols=1, size_hint_y=None, height=0)

		grade_list = calculate_grade_list(self.tournament, combined=True,
											gender_check=True)
		if grade_list in [None, []]:
			grade_list = calculate_grade_list(self.tournament,
												open=True).remove('Open')	

		for grade in grade_list:

			gender, age = self.db.cursor.execute('''SELECT gender_grade, 
										age_grade
										FROM template_grades
										WHERE title="%s"''' % grade).fetchone()
		
			var = [grade, 'Open']
			if gender != None:
			
				var.append(gender)
			if age != None:
			
				var.append(age)

			event_list = []
			for v in var:
				events = self.db.cursor.execute('''SELECT title
										FROM template_events
										WHERE grade="%s"''' % v).fetchall()
				for e in events:
					event_list.append(e[0])

			id_list = calculate_id_list_v1(self.tournament, grade=grade)

			res = self.db.cursor.execute('''SELECT * FROM results''').fetchall()
			event_results = []
			for r in res:

				event_results.append(r[0])

			score_dict = {}
			for event in event_list:
				if event in event_results:

					results = self.db.cursor.execute('''SELECT * 
						FROM %s''' % db_table('%s RR' % event)).fetchall()

					for tup in results:

						id = tup[0]
						if id in id_list:

							if tup[3] not in [None, '']:
								pp = tup[3]
							else:
								pp = 0
							if tup[4] not in [None, '']:
								cp = tup[4]
							else:
								cp = 0

							total = pp + cp
							if total != 0:
								if id not in score_dict:
									score_dict[id] = total
								else:
									score_dict[id] += total

			order = sorted(score_dict, key=score_dict.get, reverse=True)
			ranking = {1:[], 2:[], 3:[]}
			rank = 1
			tie = 1
			for i in range(len(order)):

				id = order[i]
				score = score_dict[id]

				if rank > 3:
					break
				if score == 0 or score == 0.0:
					break

				ranking[rank].append(id)
				try:
					next_score = score_dict[order[i+1]]
				except:
					continue

				if score == next_score:
					tie += 1
					continue

				elif tie > 1:
					rank += tie
					tie = 1

				else:
					rank += tie

			grade_box = BoxLayout(size_hint_y=None, height=90)
			box.add_widget(grade_box)
			box.height += 90

			grade_box.add_widget(Label(text=grade))

			if ranking[1] == [] and ranking[2] == [] and ranking[3] == []:
					
					grade_box.add_widget(Label(text='No Grade Champions Found',
												size_hint_x=3))

			else:

				for rank in ranking:

					ll = len(ranking[rank])

					if ll == 0:
						continue

					if rank == 1:
						if ll == 1:
							w = 1
						elif ll == 2:
							w = 2
						elif ll == 3:
							w = 3
					elif rank == 2:
						if ll == 1:
							w = 1
						elif ll == 2 or ll == 3:
							w = 2
					elif rank == 3:
	
						w = 1

					rank_box = BoxLayout(orientation='vertical', size_hint_x=w)

					if len(ranking[rank]) > 3:

						rank_box.add_widget(Label(text='More than 3 competitors tied for %s' % calculate_ordinal(rank)))
					
					else:

						id_rank = {}
						id_alpha = {}

						for id in ranking[rank]:

							details = self.db.cursor.execute('''SELECT type,
												firstname, surname
												FROM individuals
												WHERE id="%s"''' % id).fetchone()
							type = details[0]
							fn = details[1]
							sn = details[2]

							bg = self.type_dict[type][0]
							txt = self.type_dict[type][1]

							label = LabelC(text='%s %s' % (fn, sn), 
											bg=bg, color=txt)

							id_rank[id] = label
							id_alpha[id] = '%s %s' % (sn, fn)

						alpha = sorted(id_alpha, key=id_alpha.get)

						for id in alpha:
							rank_box.add_widget(id_rank[id])

					grade_box.add_widget(rank_box)



		return box, box.height

	def build_results_heading(self):

		self.results = GridLayout(cols=1, size_hint_y=None, height=90)
		self.display.add_widget(self.results)

		box = BoxLayout(size_hint_y=None, height=50)

		box.add_widget(Label(text='Tournament Event Results', font_size=30))

		btn = Button(text='?', font_size=30, size_hint_x=None, width=50)
		btn.bind(on_press=lambda i:self.command_help('Results'))
		box.add_widget(btn)

		self.results.add_widget(box)
		self.display.height += 50

		heading = BoxLayout(size_hint_y=None, height=40)

		for t in ['Event', '1st', '2nd', '3rd']:
			heading.add_widget(Label(text=t, font_size=20))

		self.results.add_widget(heading)
		self.display.height += 40

	def build_results(self):

		for event in self.complete_list:

			row = BoxLayout(size_hint_y=None, height=90)
			row.add_widget(LabelWrap(text=event))

			first = self.events_menu[event][1]
			second = self.events_menu[event][2]
			third = self.events_menu[event][3]

			team = calculate_event_team(self.tournament, event)

			if team == False:

				first = sorted(first, key=itemgetter(1, 0))
				second = sorted(second, key=itemgetter(1, 0))
				third = sorted(third, key=itemgetter(1, 0))

				# first
				if len(first) == 1:
					for ref in first:
						name = '%s %s' % (ref[0], ref[1])
						bg = self.type_dict[ref[2]][0]
						txt = self.type_dict[ref[2]][1]

						if len(second) + len(third) == 0:
							row.add_widget(LabelWrapC(text=name, bg=bg, 
								color=txt, size_hint_x=3))
						else:
							row.add_widget(LabelWrapC(text=name, bg=bg, color=txt))


				elif len(first) == 2:
					box = BoxLayout(orientation='vertical', size_hint_x=2)
					for ref in first:
						name = '%s %s' % (ref[0], ref[1])
						bg = self.type_dict[ref[2]][0]
						txt = self.type_dict[ref[2]][1]

						if len(second) + len(third) == 0:
							box.size_hint_x = 3

						box.add_widget(LabelWrapC(text=name, bg=bg, color=txt))
					row.add_widget(box)	



				elif len(first) == 3:
					box = BoxLayout(orientation='vertical', size_hint_x=3)
					for ref in first:
						name = '%s %s' % (ref[0], ref[1])
						bg = self.type_dict[ref[2]][0]
						txt = self.type_dict[ref[2]][1]
						box.add_widget(LabelWrapC(text=name, bg=bg, color=txt))
					row.add_widget(box)	

				elif len(first) > 3:
					box = LabelWrap(text="More than 3 competitors tied for 1st",
						size_hint_x=3)
					row.add_widget(box)	

				# second
				if len(second) == 1:
					for ref in second:
						name = '%s %s' % (ref[0], ref[1])
						bg = self.type_dict[ref[2]][0]
						txt = self.type_dict[ref[2]][1]
						row.add_widget(LabelC(text=name, bg=bg, color=txt))			

				elif len(second) == 2 or len(second) == 3:
					box = BoxLayout(orientation='vertical', size_hint_x=2)
					for ref in second:
						name = '%s %s' % (ref[0], ref[1])
						bg = self.type_dict[ref[2]][0]
						txt = self.type_dict[ref[2]][1]
						box.add_widget(LabelC(text=name, bg=bg, color=txt))
					row.add_widget(box)	

				elif len(second) > 3:
					box = LabelWrap(text="More than 3 competitors tied for 2nd",
						size_hint_x=2)
					row.add_widget(box)	


				# third
				if len(third) == 0 and len(second) == 1:
					row.add_widget(LabelC(text='No Third Place'))

				if len(third) == 1:
					for ref in third:
						name = '%s %s' % (ref[0], ref[1])
						bg = self.type_dict[ref[2]][0]
						txt = self.type_dict[ref[2]][1]
						row.add_widget(LabelC(text=name, bg=bg, color=txt))			

				elif len(third) == 2 or len(third) == 3:
					box = BoxLayout(orientation='vertical', size_hint_x=1)
					for ref in third:
						name = '%s %s' % (ref[0], ref[1])
						bg = self.type_dict[ref[2]][0]
						txt = self.type_dict[ref[2]][1]
						box.add_widget(LabelC(text=name, bg=bg, color=txt))
					row.add_widget(box)				

				elif len(second) > 3:
					box = LabelWrap(text="More than 3 competitors tied for 3rd")
					row.add_widget(box)	

			elif team == True:

				first = sorted(first, key=itemgetter(0))
				second = sorted(second, key=itemgetter(0))
				third = sorted(third, key=itemgetter(0))
		
				# first
				if len(first) == 1:
					for ref in first:
						name = ref[0]
						bg = self.type_dict[ref[1]][0]
						txt = self.type_dict[ref[1]][1]

						if len(second) + len(third) == 0:
							row.add_widget(LabelC(text=name, bg=bg, 
								color=txt, size_hint_x=3))
						else:
							row.add_widget(LabelC(text=name, bg=bg, color=txt))

				elif len(first) == 2:
					box = BoxLayout(orientation='vertical', size_hint_x=2)
					for ref in first:
						name = ref[0]
						bg = self.type_dict[ref[1]][0]
						txt = self.type_dict[ref[1]][1]
						
						if len(second) + len(third) == 0:
							box.size_hint_x = 3

						box.add_widget(LabelC(text=name, bg=bg, color=txt))
					row.add_widget(box)	

				elif len(first) == 3:
					box = BoxLayout(orientation='vertical', size_hint_x=3)
					for ref in first:
						name = ref[0]
						bg = self.type_dict[ref[1]][0]
						txt = self.type_dict[ref[1]][1]
						box.add_widget(LabelC(text=name, bg=bg, color=txt))
					row.add_widget(box)	

				elif len(first) > 3:
					box = LabelWrap(text="More than 3 teams tied for 1st")
					row.add_widget(box)	

				# second
				if len(second) == 1:
					for ref in second:
						name = ref[0]
						bg = self.type_dict[ref[1]][0]
						txt = self.type_dict[ref[1]][1]
						row.add_widget(LabelC(text=name, bg=bg, color=txt))			

				elif len(second) == 2 or len(second) == 3:
					box = BoxLayout(orientation='vertical', size_hint_x=2)
					for ref in second:
						name = ref[0]
						bg = self.type_dict[ref[1]][0]
						txt = self.type_dict[ref[1]][1]
						box.add_widget(LabelC(text=name, bg=bg, color=txt))
					row.add_widget(box)	

				elif len(second) > 3:
					box = LabelWrap(text="More than 3 teams tied for 2nd")
					row.add_widget(box)	

				# third
				if len(third) == 0 and len(second) == 1:
					row.add_widget(LabelC(text='No Third Place'))

				if len(third) == 1:
					for ref in third:
						name = ref[0]
						bg = self.type_dict[ref[1]][0]
						txt = self.type_dict[ref[1]][1]
						row.add_widget(LabelC(text=name, bg=bg, color=txt))			

				elif len(third) == 2 or len(third) == 3:
					box = BoxLayout(orientation='vertical', size_hint_x=1)
					for ref in third:
						name = ref[0]
						bg = self.type_dict[ref[1]][0]
						txt = self.type_dict[ref[1]][1]
						box.add_widget(LabelC(text=name, bg=bg, color=txt))
					row.add_widget(box)				

				elif len(second) > 3:
					box = LabelWrap(text="More than 3 teams tied for 3rd")
					row.add_widget(box)	

			self.results.add_widget(row)
			self.results.height += 90
			self.display.height += 90

	def build_special_heading(self):

		self.special_events = GridLayout(cols=1, size_hint_y=None, height=90)
		self.display.add_widget(self.special_events)

		box = BoxLayout(size_hint_y=None, height=50)

		box.add_widget(Label(text='Tournament Special Events', font_size=30))

		btn = Button(text='?', font_size=30, size_hint_x=None, width=50)
		btn.bind(on_press=lambda i:self.command_help('Special'))
		box.add_widget(btn)

		self.special_events.add_widget(box)
		self.display.height += 50

		heading = BoxLayout(size_hint_y=None, height=40)

		for t in ['Event', self.type, 'Points']:
			heading.add_widget(Label(text=t, font_size=20))

		self.special_events.add_widget(heading)
		self.display.height += 40

	def build_special(self):

		special = self.db.cursor.execute('''SELECT *
					FROM special_events''').fetchall()

		for tup in special:

			event = tup[0]
			type = tup[1]
			bg = self.type_dict[type][0]
			txt = self.type_dict[type][1]
			points = tup[2]
			try:
				if points.is_integer():
					points = int(points)
			except:
				pass

			box = BoxLayout(size_hint_y=None, height=90)

			box.add_widget(LabelWrap(text=event))
			box.add_widget(LabelC(text=type, bg=bg, color=txt))
			box.add_widget(Label(text=str(points)))

			self.special_events.add_widget(box)
			self.special_events.height += 90
			self.display.height += 90

### EXCEL PRINTS ###

class ExcelEntrySheets():

	def __init__(self, *args, **kwargs):

		## KW Settings
		self.tournament = kwargs.pop('tournament', None)
		self.typename = kwargs.pop('typename', None)
		self.hgt = kwargs.pop('height', 32) # landscape

		## Settings
		self.db = Database(self.tournament)

		details = self.db.cursor.execute('''SELECT * FROM details''').fetchone()
		self.type = details[3]

		types = self.db.cursor.execute('''SELECT title 
										FROM template_groups
										WHERE category="type"''').fetchall()
		self.type_tables = {}
		for tup in types:
			type = tup[0]
			self.type_tables[type] = {}
			tables = calculate_type_tables(self.tournament, type, grade=True)
			for pair in tables:
				self.type_tables[type][pair[0]] = pair[1]

		## Database


		## Display
		try:
			self.wb = Workbook()
			self.wb.save('%s Print.xlsx' % self.tournament)
		except (OSError, IOError):
			PopBox().showwarning(title='Excel Error',
				message='%s Print.xlsx is already open, please close and try again' % self.tournament)
			return

		grade_list = calculate_grade_list(self.tournament)
		if self.typename == None:
			for type in sorted(self.type_tables):
				for grade in grade_list:
					if grade in self.type_tables[type]:
						table = self.type_tables[type][grade]
						self.build_worksheet(type, grade, table)
		else:
			for grade in grade_list:
				if grade in self.type_tables[type]:
					table = self.type_tables[self.typename][grade]
					self.build_worksheet(self.typename, grade, table)			

		## Engine
		self.wb.save('%s Print.xlsx' % self.tournament)
		file_open('%s Print.xlsx' % self.tournament)

	### BUILD ###	

	def build_worksheet(self, type, grade, table):

		## Database
		events = self.db.list_columns(table)
		events.remove('id')

		data = self.db.cursor.execute('''SELECT * FROM %s''' % table).fetchall()
		id_dict = {}
		alpha_dict = {}
		for tup in data:
			id = tup[0]
	
			indiv = self.db.cursor.execute('''SELECT firstname, surname, type
									FROM individuals
									WHERE id="%s"''' % id).fetchone()
			choices = []
			for i in range(1, len(tup)):
				choices.append(tup[i])
			fn = indiv[0]
			sn = indiv[1]
			type = indiv[2]

			id_dict[id] = [fn, sn, type] + choices
			alpha_dict[id] = '%s %s' % (sn, fn)
		alpha = sorted(alpha_dict, key=alpha_dict.get)
		## Sheet Title
		if 'Sheet' in self.wb.sheetnames:
			ws = self.wb.active
			ws.title = '%s - %s' % (type, grade)
		else:
	
			ws = self.wb.create_sheet('%s - %s' % (type, grade))

		## Sheet Page Settings
		ws.page_margins.left = 0
		ws.page_margins.right = 0
		ws.page_margins.top = 0
		ws.page_margins.bottom = 0
		#ws.print_options.horizontalCentered = True

		for c in range(2, (len(events)+2)):
			ws.column_dimensions[get_column_letter(c)].width = 4
		ws.column_dimensions['A'].width = 25

		## Sheet CSS
		css_border = Border(left=Side(style='thin'), right=Side(style='thin'),
								top=Side(style='thin'),bottom=Side(style='thin')) 
		css_fill = PatternFill(start_color='B0E0E6',
              						end_color='B0E0E6', fill_type='solid')	
		## Sheet 
		pages = 1			
		if len(alpha) > 29:
			count = len(alpha) - 29
			while count >= 0:
				pages += 1
				count -= 29 
		max_col = len(events) + 1 # 

		## Sheet
		for p in range(0, pages):		
			
			start_row = (p * 31)  

			ws.row_dimensions[start_row+1].height = 42
			ws.row_dimensions[start_row+2].height = 138
			for r in range(start_row+3, start_row+32):
				ws.row_dimensions[r].height = 21

			# Heading
			cellform = ws['A%s' % (start_row+1)]
			cellform.border = css_border
			ws.merge_cells("A%s:%s%s" % (start_row+1, 
									get_column_letter(max_col), start_row+1))		
			cellform = ws['A%s' % (start_row+1)]
			cellform.value = '%s %s' % (type, grade)
			cellform.fill = css_fill		
			cellform.alignment = Alignment(horizontal='center', 
											vertical='center',
											shrinkToFit=True)
			cellform.font = Font(size=30)

			for c in range(2, (len(events)+2)):
				col = get_column_letter(c)
				ws['%s%s' % (col, start_row+1)].border = css_border
	
			# Events
			cellform = ws['A%s' % (start_row+2)]
			cellform.fill = css_fill
			cellform.border = css_border			
			for i in range(len(events)):
				event = db_table(events[i], db_in=False, grade=grade)

				if grade in event:
					contest = self.db.cursor.execute('''SELECT contest
									FROM template_events
									WHERE title="%s"''' % event).fetchone()[0]
					event = contest

				cellform = ws['%s%s' % (get_column_letter(i+2), start_row+2)]
				cellform.value = event
				cellform.font = Font(size=14)
				cellform.fill = css_fill
				cellform.alignment = Alignment(vertical='center',
												textRotation=90,
												shrinkToFit=True)
				cellform.border = css_border

			# Competitors
			min = p * 29
			max = (p + 1) * 29
			id_list = alpha[min:max]	

			for i in range(len(id_list)):
				id = id_list[i]
				row = start_row + i + 3
				details= id_dict[id]
				entry = details[3:]
				name = '%s %s' % (details[0], details[1])
				
				cellname = ws['A%s' % row]
				cellname.value = name
				cellname.fill = css_fill
				cellname.border = css_border
				cellname.alignment = Alignment(shrinkToFit=True)
				cellname.font = Font(size=12)

				for e in range(len(entry)):
					col = get_column_letter(e+2)					
					cellform = ws['%s%s' % (col, row)]
					if entry[e] == '1':
						cellform.value = 'X'
						cellform.alignment = Alignment(vertical='center',
														horizontal='center')
						cellform.font = Font(size=14)				
					cellform.border = css_border

			extra = 29 - len(id_list)
			if extra > 0:
				for i in range(extra):
					row = start_row + 3 + len(id_list) + i
					cellform = ws['A%s' % row]
					cellform.fill = css_fill
					cellform.border = css_border
					for e in range(len(events)):
						col = get_column_letter(e+2)
						ws['%s%s' % (col, row)].border = css_border						

class ExcelEventsBase():

	def __init__(self, *args, **kwargs):

		## KW Settings
		self.tournament = kwargs.pop('tournament', None)
		self.style = kwargs.pop('style', None)
		self.contest = kwargs.pop('contest', None)
		self.row_max = kwargs.pop('row_max', 54) # 54 portrait, 40 landscape

		## Settings
		self.db = Database(self.tournament)

		details = self.db.cursor.execute('''SELECT * FROM details''').fetchone()
		self.type = details[3]

		types = self.db.cursor.execute('''SELECT title FROM template_groups
										WHERE category="type"''').fetchall()
		type_list = []
		for tup in types:
			type_list.append(tup[0])
		self.type_list = sorted(type_list)

		## Database


		## Display
		try:
			self.wb = Workbook()
			self.wb.save('%s Print.xlsx' % self.tournament)
		except (OSError, IOError):
			PopBox().showwarning(title='Excel Error',
				message='%s Print.xlsx is already open, please close and try again' % self.tournament)
			return

		if self.style == 'time':
			contest_list = calculate_contest_list(self.tournament, time=True)
		elif self.style == 'distance':
			contest_list = calculate_contest_list(self.tournament, 
													distance=True)
		elif self.style == 'All':
			contest_list = calculate_contest_list(self.tournament)
		else:
			contest_list = [self.contest]

		for contest in contest_list:
			events = []

			for grade in calculate_grade_list(self.tournament):

				event_details = self.db.cursor.execute('''SELECT * 
								FROM template_events
								WHERE contest="%s"
								AND grade="%s"''' % (contest, grade)).fetchone()

				if event_details != None:
					events.append(event_details)

			for tup in events:
				self.access_event(tup)

		## Engine
		try:
			self.wb.save('%s Print.xlsx' % self.tournament)
			file_open('%s Print.xlsx' % self.tournament)
		except IOError:
			PopBox().showwarning(title='Excel Error',
				message='%s Print.xlsx is already open, please close and try again' % self.tournament)
			return			

	### ACCESS ###
	
	def access_event(self, db):

		## Settings
		event = db[0]
		contest = db[1]

		contest_type = calculate_contest_type(self.tournament, contest)
		team = calculate_contest_type(self.tournament, contest, team=True)

		event_table = db_table(event)

		event_rounds = self.db.cursor.execute('''SELECT * FROM %s
						WHERE round!="RR"''' % event_table).fetchall()

		for tup in event_rounds:

			if contest_type == 'time':

				if team == False:
				
					self.build_worksheet_time(db, tup, event)
				
				elif team == True:
			
					self.build_worksheet_time_team(db, tup, event)

			if contest_type == 'distance':

				if contest in ['High Jump', 'Pole Vault']:

					self.build_worksheet_height(db, tup, event)
				
				elif team == False:

					self.build_worksheet_length(db, tup, event)

	### BUILD ###	

	def build_worksheet_time(self, event_db, round_db, evt):

		## Settings
		event = event_db[0]
		contest = event_db[1]
	
		grade = event_db[2]
		record_what = convert_score(event_db[3], measure='time')
		record_who = event_db[4]
		record_when = event_db[5]

		if round_db[1] == evt:
			program_number = event_db[6]
			program_time = event_db[7]
		else:
			program_number = round_db[3]
			program_time = round_db[4]

		rd = round_db[0]
		round_name = round_db[1]
		round_height = round_db[2] # for vertical contests

		round_table = db_table('%s %s' % (event, rd))

		indiv = self.db.cursor.execute('''SELECT * 
	
							FROM %s''' % round_table).fetchall()
		id_dict = {}
		alpha_dict = {}
		for tup in indiv:
			id = tup[0]
			pp = tup[1]
			score =tup[2] # time
			placing = tup[5]

			id_details = self.db.cursor.execute('''SELECT firstname, 
								surname, type
								FROM individuals
								WHERE id="%s"''' % id).fetchone()
			fn = id_details[0]
			sn = id_details[1]
			type = id_details[2]

			id_dict[id] = [pp, score, placing, fn, sn, type]
			alpha_dict[id] = '%s %s' % (sn, fn)
		alpha = sorted(alpha_dict, key=alpha_dict.get)

		## Excel Settings
		if len(round_name) > 30:
			round_title = grade
		else:
			round_title = round_name
		if 'Sheet' in self.wb.sheetnames:
			ws = self.wb.active
			ws.title = round_title
		else:
			ws = self.wb.create_sheet(round_title)

		## Contest Type Specific 
		ws.page_margins.left = 0
		ws.page_margins.right = 0
		ws.page_margins.top = 0
		ws.page_margins.bottom = 0
		ws.print_options.horizontalCentered = True

		ws.column_dimensions['A'].width = 6
		ws.column_dimensions['B'].width = 28
		ws.column_dimensions['C'].width = 3.5
		ws.column_dimensions['D'].width = 17.5
		ws.column_dimensions['E'].width = 17.5
		ws.column_dimensions['F'].width = 13
	
		pages = 1			
		if len(alpha) > 49:
			count = len(alpha) - 49
			while count >= 0:
				pages += 1
				count -= 49 
		
		## Page
		for p in range(0, pages): # runs page			
			start_row = 1 + (p * self.row_max)           
			for c in ["A", "B", "C", "D", "E", "F"]:			
				for r in range(start_row, start_row+5):
					cellform = ws["%s%s" % (c, r)]	
					cellform.alignment = Alignment(horizontal="center", 
											vertical="center",
											shrinkToFit=True)
					cellform.fill = PatternFill(start_color='B0E0E6',
              								     end_color='B0E0E6', fill_type='solid')        
					if c == "F":
						if r in range(start_row+1, start_row+4):
							cellform.border = Border(left=Side(style='thin'),
													right=Side(style='thin'))
						else:
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin')) 				
	
					elif c == "A":
						if r == start_row:
							cellform.border = Border(left=Side(style='thin'), 
            						    right=Side(style='thin'), 
               							 top=Side(style='thin'))
						elif r == start_row+1:
							cellform.border = Border(left=Side(style="thin"),
										right=Side(style="thin"))										
						elif r == start_row+2:
							cellform.border = Border(left=Side(style='thin'),
										right=Side(style='thin'),
										top=Side(style='thin'))				
						elif r == start_row+3:
							cellform.border = Border(left=Side(style='thin'),
										right=Side(style='thin'),
										bottom=Side(style='thin'))
						else:
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin')) 																
			
					elif c in ["B", "C", "D", "E"]:
						if r == start_row:
							cellform.border = Border(top=Side(style='thin'))
						elif r == start_row+3:
							cellform.border = Border(bottom=Side(style='thin'))
						elif r == start_row+4:							
							if c == "D":
								cellform.border = Border(left=Side(style="thin"),
														top=Side(style='thin'),
														bottom=Side(style='thin'))					
							elif c == "E":
								cellform.border = Border(right=Side(style="thin"),
														top=Side(style='thin'),
														bottom=Side(style='thin'))								
							else:
								cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin')) 								
				
				for r in range(start_row+4, start_row+self.row_max):
					cellform = ws["%s%s" % (c, r)]		
					if r == start_row+4:
						cellform.alignment = Alignment(horizontal="center",
													vertical="center")		
					else:									
						if c in ["C", "D"]:
							cellform.alignment = Alignment(horizontal="center",
										vertical="center",
										wrapText=True)
							cellform.border = Border(left=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin'))						
						elif c == "E":
							cellform.alignment = Alignment(horizontal="center",
										vertical="center",
										wrapText=True)
							cellform.border = Border(top=Side(style='thin'),
												bottom=Side(style='thin'))	

						elif c == "F":
							cellform.alignment = Alignment(horizontal="center",
										vertical="center",
										wrapText=True)
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin'))								
						else:
							cellform.alignment = Alignment(horizontal="left",
										vertical="center",
										wrapText=True)
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin'))									

			for c in ["A", "B"]:
				for r in range(start_row+5, start_row+self.row_max):
					cellform = ws["%s%s" % (c, r)]
					cellform.fill = PatternFill(start_color='B0E0E6', end_color='B0E0E6',
                   										fill_type='solid')					
				ws["A%r" % start_row].value = "Event"
				ws["A%r" % (start_row+1)].value = program_number				
				ws["A%r" % (start_row+2)].value = "Time"			
				ws["A%r" % (start_row+3)].value = program_time

				ws.merge_cells("B%s:E%s" % (start_row, start_row+3))
				ws["B%s" % start_row].font = Font(size=30)
				ws["B%s" % start_row].value = round_name
				
				ws["F%s" % start_row].value = "Record"				
				ws["F%s" % (start_row+1)].value = record_who # record
				ws["F%s" % (start_row+2)].value = record_what # name
				ws["F%s" % (start_row+3)].value = record_when # year

				ws["A%s" % (start_row+4)].value = self.type
				ws["B%s" % (start_row+4)].value = "Competitor"
				ws["C%s" % (start_row+4)].value = "PP"
				ws.merge_cells("D%s:E%s" % (start_row+4, start_row+4))
				ws["D%s" % (start_row+4)].value = "Time"
				
				ws["F%s" % (start_row+4)].value = "Placing"

				## 																
				min = p * 50
				max = ((p + 1) * 50)	
				page_id_list = alpha[min:max]	
				
				row_num = start_row + 5					
				for id in page_id_list:	
					
					if id == None:
						row_num += 1
						continue
					#[pp, score, placing, fn, sn, type]
					type = id_dict[id][5][:3].upper()
					name = '%s %s' % (id_dict[id][3], id_dict[id][4])
					pp = db_boole(id_dict[id][0], db_in=False)
					if pp in [None, False]:
			
						pp = ''
					else:
			
						pp = 'X' # 
					
					placing = id_dict[id][2]
					if placing in [None, 0]:
						placing = ''
					
					score = convert_score(id_dict[id][1], measure='time')

					ws["A%s" % row_num].value = type	
					ws["B%s" % row_num].value = name
					ws["C%s" % row_num].value =	pp		
					ws["D%s" % row_num].value = score
					ws["F%s" % row_num].value =	placing			
										
					row_num += 1

	def build_worksheet_time_team(self, event_db, round_db, evt):

		## Settings
		event = event_db[0]
		contest = event_db[1]
	
		grade = event_db[2]
		record_what = convert_score(event_db[3], measure='time')
		record_who = event_db[4]
		record_when = event_db[5]

		if round_db[1] == evt:
			program_number = event_db[6]
			program_time = event_db[7]
		else:
			program_number = round_db[3]
			program_time = round_db[4]

		rd = round_db[0]
		round_name = round_db[1]
		round_height = round_db[2] # for vertical contests

		round_table = db_table('%s %s' % (event, rd))

		indiv = self.db.cursor.execute('''SELECT * 
	
							FROM %s''' % round_table).fetchall()
		id_dict = {}
		alpha_dict = {}
		for tup in indiv:
			id = tup[0]
			pp = tup[1]
			score =tup[2] # time
			placing = tup[5]
			type = tup[6]

			id_dict[id] = [pp, score, placing, None, None, type]
			alpha_dict[id] = id
		alpha = sorted(alpha_dict, key=alpha_dict.get)

		## Excel Settings
		if len(round_name) > 30:
			round_title = grade
		else:
			round_title = round_name
		if 'Sheet' in self.wb.sheetnames:
			ws = self.wb.active
			ws.title = round_title
		else:
			ws = self.wb.create_sheet(round_title)

		## Contest Type Specific 
		ws.page_margins.left = 0
		ws.page_margins.right = 0
		ws.page_margins.top = 0
		ws.page_margins.bottom = 0
		ws.print_options.horizontalCentered = True

		ws.column_dimensions['A'].width = 6
		ws.column_dimensions['B'].width = 28
		ws.column_dimensions['C'].width = 3.5
		ws.column_dimensions['D'].width = 17.5
		ws.column_dimensions['E'].width = 17.5
		ws.column_dimensions['F'].width = 13
	
		pages = 1			
		if len(alpha) > 50:
			count = len(alpha) - 50
			while count >= 0:
				pages += 1
				count -= 50 
		
		## Page
		for p in range(0, pages): # runs page			
			start_row = 1 + (p * self.row_max)           
			for c in ["A", "B", "C", "D", "E", "F"]:			
				for r in range(start_row, start_row+5):
					cellform = ws["%s%s" % (c, r)]	
					cellform.alignment = Alignment(horizontal="center", 
											vertical="center",
											shrinkToFit=True)
					cellform.fill = PatternFill(start_color='B0E0E6',
              								     end_color='B0E0E6', fill_type='solid')        
					if c == "F":
						if r in range(start_row+1, start_row+4):
							cellform.border = Border(left=Side(style='thin'),
													right=Side(style='thin'))
						else:
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin')) 				
	
					elif c == "A":
						if r == start_row:
							cellform.border = Border(left=Side(style='thin'), 
            						    right=Side(style='thin'), 
               							 top=Side(style='thin'))
						elif r == start_row+1:
							cellform.border = Border(left=Side(style="thin"),
										right=Side(style="thin"))										
						elif r == start_row+2:
							cellform.border = Border(left=Side(style='thin'),
										right=Side(style='thin'),
										top=Side(style='thin'))				
						elif r == start_row+3:
							cellform.border = Border(left=Side(style='thin'),
										right=Side(style='thin'),
										bottom=Side(style='thin'))
						else:
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin')) 																
			
					elif c in ["B", "C", "D", "E"]:
						if r == start_row:
							cellform.border = Border(top=Side(style='thin'))
						elif r == start_row+3:
							cellform.border = Border(bottom=Side(style='thin'))
						elif r == start_row+4:							
							if c == "D":
								cellform.border = Border(left=Side(style="thin"),
														top=Side(style='thin'),
														bottom=Side(style='thin'))					
							elif c == "E":
								cellform.border = Border(right=Side(style="thin"),
														top=Side(style='thin'),
														bottom=Side(style='thin'))								
							else:
								cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin')) 								
				
				for r in range(start_row+4, start_row+self.row_max):
					cellform = ws["%s%s" % (c, r)]		
					if r == start_row+4:
						cellform.alignment = Alignment(horizontal="center",
													vertical="center")		
					else:									
						if c in ["C", "D"]:
							cellform.alignment = Alignment(horizontal="center",
										vertical="center",
										wrapText=True)
							cellform.border = Border(left=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin'))						
						elif c == "E":
							cellform.alignment = Alignment(horizontal="center",
										vertical="center",
										wrapText=True)
							cellform.border = Border(top=Side(style='thin'),
												bottom=Side(style='thin'))	

						elif c == "F":
							cellform.alignment = Alignment(horizontal="center",
										vertical="center",
										wrapText=True)
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin'))								
						else:
							cellform.alignment = Alignment(horizontal="left",
										vertical="center",
										wrapText=True)
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin'))									

			for c in ["A", "B"]:
				for r in range(start_row+5, start_row+self.row_max):
					cellform = ws["%s%s" % (c, r)]
					cellform.fill = PatternFill(start_color='B0E0E6', end_color='B0E0E6',
                   										fill_type='solid')					
				ws["A%r" % start_row].value = "Event"
				ws["A%r" % (start_row+1)].value = program_number				
				ws["A%r" % (start_row+2)].value = "Time"			
				ws["A%r" % (start_row+3)].value = program_time

				ws.merge_cells("B%s:E%s" % (start_row, start_row+3))
				ws["B%s" % start_row].font = Font(size=30)
				ws["B%s" % start_row].value = round_name
				
				ws["F%s" % start_row].value = "Record"				
				ws["F%s" % (start_row+1)].value = record_who # record
				ws["F%s" % (start_row+2)].value = record_what # name
				ws["F%s" % (start_row+3)].value = record_when # year

				ws["A%s" % (start_row+4)].value = self.type
				ws["B%s" % (start_row+4)].value = "Team"
				ws["C%s" % (start_row+4)].value = "PP"
				ws.merge_cells("D%s:E%s" % (start_row+4, start_row+4))
				ws["D%s" % (start_row+4)].value = "Time"
				
				ws["F%s" % (start_row+4)].value = "Placing"

				## 																
				min = p * 50
				max = ((p + 1) * 50)	
				page_id_list = alpha[min:max]	
				
				row_num = start_row + 5					
				for id in page_id_list:	
					
					if id == None:
						row_num += 1
						continue
					#[pp, score, placing, fn, sn, type]
					type = id_dict[id][5][:3].upper()
					pp = db_boole(id_dict[id][0], db_in=False)
					if pp in [None, False]:
			
						pp = ''
					else:
			
						pp = 'X' # 
					
					placing = id_dict[id][2]
					if placing in [None, 0]:
						placing = ''
					
					score = convert_score(id_dict[id][1], measure='time')

					ws["A%s" % row_num].value = type	
					ws["B%s" % row_num].value = id
					ws["C%s" % row_num].value =	pp		
					ws["D%s" % row_num].value = score
					ws["F%s" % row_num].value =	placing			
										
					row_num += 1

	def build_worksheet_length(self, event_db, round_db, evt):

		## Settings
		event = event_db[0]
		contest = event_db[1]
	
		grade = event_db[2]
		record_what = convert_score(event_db[3], measure='distance')
		record_who = event_db[4]
		record_when = event_db[5]
		if round_db[1] == evt:
			program_number = event_db[6]
			program_time = event_db[7]
		else:
			program_number = round_db[3]
			program_time = round_db[4]


		rd = round_db[0]
		round_name = round_db[1]
		round_height = round_db[2] # for vertical contests

		round_table = db_table('%s %s' % (event, rd))

		indiv = self.db.cursor.execute('''SELECT * 
	
							FROM %s''' % round_table).fetchall()
		id_dict = {}
		alpha_dict = {}
		for tup in indiv:
			id = tup[0]
			pp = tup[1]
			score =tup[3] # length
			placing = tup[5]

			id_details = self.db.cursor.execute('''SELECT firstname, 
								surname, type
								FROM individuals
								WHERE id="%s"''' % id).fetchone()
			fn = id_details[0]
			sn = id_details[1]
			type = id_details[2]

			id_dict[id] = [pp, score, placing, fn, sn, type]
			alpha_dict[id] = '%s %s' % (sn, fn)
		alpha = sorted(alpha_dict, key=alpha_dict.get)

		## Excel Settings
		if len(round_name) > 30:
			round_title = grade
		else:
			round_title = round_name
		if 'Sheet' in self.wb.sheetnames:
			ws = self.wb.active
			ws.title = round_title
		else:
			ws = self.wb.create_sheet(round_title)

		## Contest Type Specific 
		ws.page_margins.left = 0
		ws.page_margins.right = 0
		ws.page_margins.top = 0
		ws.page_margins.bottom = 0
		ws.print_options.horizontalCentered = True

		ws.column_dimensions['A'].width = 6
		ws.column_dimensions['B'].width = 22
		ws.column_dimensions['C'].width = 3.5
		for c in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
			ws.column_dimensions[c].width = 7.5
		ws.column_dimensions['K'].width = 11
		
		## Page
		pages = 1			
		if len(alpha) > 49:
			count = len(alpha) - 49
			while count >= 0:
				pages += 1
				count -= 49 # counts sid to see how many pages			
	
		for p in range(0, pages): # runs page			
			start_row = 1 + (p * self.row_max)         
			for c in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:			
				for r in range(start_row, start_row+5):				
					cellform = ws["%s%s" % (c, r)]							
					cellform.alignment = Alignment(horizontal="center", 
											vertical="center",
											shrinkToFit=True)
					cellform.fill = PatternFill(start_color='B0E0E6',
              								     end_color='B0E0E6', fill_type='solid')        
			
					if c == "A":
						if r == start_row:
							cellform.border = Border(left=Side(style='thin'), 
            						    right=Side(style='thin'), 
               							 top=Side(style='thin'))
						elif r == start_row+1:
							cellform.border = Border(left=Side(style="thin"),
										right=Side(style="thin"))										
						elif r == start_row+2:
							cellform.border = Border(left=Side(style='thin'),
										right=Side(style='thin'),
										top=Side(style='thin'))				
						elif r == start_row+3:
							cellform.border = Border(left=Side(style='thin'),
										right=Side(style='thin'),
										bottom=Side(style='thin'))
						else:
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin')) 		
																										
					elif c in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
						if r == start_row:
							cellform.border = Border(top=Side(style='thin'))
						elif r == start_row+3:
							cellform.border = Border(bottom=Side(style='thin'))
						elif r == start_row+4:							
							cellform.border = Border(bottom=Side(style='thin'),
												left=Side(style='thin'),
												right=Side(style='thin')) 														
												
					elif c == "K":
						if r in range(start_row+1, start_row+4):
							cellform.border = Border(left=Side(style='thin'),
													right=Side(style='thin'))
						else:
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin')) 																								
																		
				for r in range(start_row+4, start_row+self.row_max):
					cellform = ws["%s%s" % (c, r)]		
					if r == start_row+4:
						cellform.alignment = Alignment(horizontal="center",
													vertical="center")								
														
					else:	
					
						if c in ['A', 'B']:					
							cellform.alignment = Alignment(horizontal="left",
										vertical="center",
										shrinkToFit=True)
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin'))						

					
						else:
							cellform.alignment = Alignment(horizontal="center",
										vertical="center",
										shrinkToFit=True)
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin'))								
				
			css = Border(left=Side(style='thin'),
						right=Side(style='thin'),
						top=Side(style='thin'),
						bottom=Side(style='thin'))	

			for c in ["A", "B"]:
				for r in range(start_row+5, start_row+self.row_max):
					cellform = ws["%s%s" % (c, r)]
					cellform.fill = PatternFill(start_color='B0E0E6',
                   										end_color='B0E0E6',
                   										fill_type='solid')	

			ws["A%r" % start_row].value = "Event"
			ws["A%r" % (start_row+1)].value = program_number				
			ws["A%r" % (start_row+2)].value = "Time"			
			ws["A%r" % (start_row+3)].value = program_time

			ws.merge_cells("B%s:J%s" % (start_row, start_row+3))
			ws["B%s" % start_row].font = Font(size=30)
			ws["B%s" % start_row].value = round_name

			ws["K%s" % start_row].value = "Record"							
			ws["K%s" % (start_row+1)].value = record_who # record
			ws["K%s" % (start_row+2)].value = record_what # name
			ws["K%s" % (start_row+3)].value = record_when # year

			ws["A%s" % (start_row+4)].value = "House"
			ws["B%s" % (start_row+4)].value = "Competitor"
			ws["B%s" % (start_row+4)].border = css
			ws["C%s" % (start_row+4)].value = "PP"
			ws["C%s" % (start_row+4)].border = css		
			num = 1
			for c in ["D", "E", "F", "G", "H", "I"]:
				attempt = "%s" % calculate_ordinal(num)
				ws["%s%s" % (c, start_row+4)].value = attempt
				ws["%s%s" % (c, start_row+4)].border = css		
				num += 1				
			ws["J%s" % (start_row+4)].value = "Best"
			ws["J%s" % (start_row+4)].border = css			
			ws["K%s" % (start_row+4)].value = "Placing"

			##
			min = p * 50
			max = ((p + 1) * 50)	

			page_id_list = alpha[min:max]	
				
			row_num = start_row + 5		
			
			for id in page_id_list:	# [pp, score, placing, fn, sn, type]
			
				if id == None:
					row_num += 1
					continue

				type = id_dict[id][5][:3].upper()
				name = '%s %s' % (id_dict[id][3], id_dict[id][4])
				pp = db_boole(id_dict[id][0], db_in=False)
				if pp == True:
					pp = 'X'
				else:
					pp = ''

				if id_dict[id][1]  in [None, '', 0, 0.0]:
					score = [0, 0, 0, 0, 0, 0]
				else:
					score = db_score(id_dict[id][1], None, 'distance', db_in=False)

				ws["A%s" % row_num].value = type
				ws["B%s" % row_num].value = name
				ws["C%s" % row_num].value = pp
				
				letter_list = ['D', 'E', 'F', 'G', 'H', 'I']
				best_list = []
				for i in range(len(letter_list)):
					letter = letter_list[i]
					try:
						attempt = score[i]
					except TypeError:
						attempt = None
					if attempt in [0, None, '']:
						attempt = ''
					else:
						best_list.append(attempt)
						attempt = '%s m' % attempt			
					ws["%s%s" % (letter, row_num)].value = attempt				
				
				if len(best_list) > 0:
					best = '%s m' % sorted(best_list)[-1]
				else:
					best = ''					
				ws['J%s' % row_num].value = best
				
				placing = id_dict[id][2]
				if placing in [0, None, '']:
					placing = ''
				ws['K%s' % row_num].value = placing
				
				row_num += 1	

	def build_worksheet_height(self, event_db, round_db, evt):

		## Settings
		event = event_db[0]
		contest = event_db[1]
	
		grade = event_db[2]
		record_what = convert_score(event_db[3], measure='distance')
		record_who = event_db[4]
		record_when = event_db[5]
		if round_db[1] == evt:
			program_number = event_db[6]
			program_time = event_db[7]
		else:
			program_number = round_db[3]
			program_time = round_db[4]

		rd = round_db[0]
		round_name = round_db[1]
		height_ref = round_db[2] # for vertical contests

		if height_ref in [None, '', 0, 0.0]:
			height_ref = [0, 0, 0, 0, 0, 0]
		else:
			height_ref = height_ref.split('#')

		round_table = db_table('%s %s' % (event, rd))

		indiv = self.db.cursor.execute('''SELECT * 
	
							FROM %s''' % round_table).fetchall()
		id_dict = {}
		alpha_dict = {}
		for tup in indiv:
			id = tup[0]
			pp = tup[1]
			score =tup[4] # height
			placing = tup[5]

			id_details = self.db.cursor.execute('''SELECT firstname, 
								surname, type
								FROM individuals
								WHERE id="%s"''' % id).fetchone()
			fn = id_details[0]
			sn = id_details[1]
			type = id_details[2]

			id_dict[id] = [pp, score, placing, fn, sn, type]
			alpha_dict[id] = '%s %s' % (sn, fn)
		alpha = sorted(alpha_dict, key=alpha_dict.get)

		## Excel Settings
		if len(round_name) > 30:
			round_title = grade
		else:
			round_title = round_name
		if 'Sheet' in self.wb.sheetnames:
			ws = self.wb.active
			ws.title = round_title
		else:
			ws = self.wb.create_sheet(round_title)

		## Contest Type Specific 
		ws.page_margins.left = 0
		ws.page_margins.right = 0
		ws.page_margins.top = 0
		ws.page_margins.bottom = 0
		ws.print_options.horizontalCentered = True

		ws.column_dimensions['A'].width = 6
		ws.column_dimensions['B'].width = 25
		ws.column_dimensions['C'].width = 3.5
		for c in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L',
					'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
			ws.column_dimensions[c].width = 2.5
		ws.column_dimensions['V'].width = 7
		ws.column_dimensions['W'].width = 11

		## Page			
		pages = 1			
		if len(alpha) > 49:
			count = len(alpha) - 49
			while count >= 0:
				pages += 1
				count -= 49 # counts id to see how many pages			
	
		for p in range(0, pages): # runs page			
			start_row = 1 + (p * self.row_max)           
			for c in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", 'L',
						'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']:			
				for r in range(start_row, start_row+5):					
					cellform = ws["%s%s" % (c, r)]							
					cellform.alignment = Alignment(horizontal="center", 
											vertical="center",
											shrinkToFit=True)				
					
					if c in ["D", "E", "F", "G", "H", "I", "J", "K", 'L',
						'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U'] and r == start_row+4:
						pass
					else:
						cellform.fill = PatternFill(start_color='B0E0E6',
              								     end_color='B0E0E6', fill_type='solid')        			
					if c == "A":
						if r == start_row:
							cellform.border = Border(left=Side(style='thin'), 
            						    right=Side(style='thin'), 
               							 top=Side(style='thin'))
						elif r == start_row+1:
							cellform.border = Border(left=Side(style="thin"),
										right=Side(style="thin"))										
						elif r == start_row+2:
							cellform.border = Border(left=Side(style='thin'),
										right=Side(style='thin'),
										top=Side(style='thin'))				
						elif r == start_row+3:
							cellform.border = Border(left=Side(style='thin'),
										right=Side(style='thin'),
										bottom=Side(style='thin'))
						else:
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin')) 																
																									
					elif c == "W":
						if r in range(start_row+1, start_row+4):
							cellform.border = Border(left=Side(style='thin'),
													right=Side(style='thin'))
						else:
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin')) 																								

					else:
						if r == start_row:
							cellform.border = Border(top=Side(style='thin'))
						elif r == start_row+3:
							cellform.border = Border(bottom=Side(style='thin'))
						elif r == start_row+4:
							cellform.border = Border(left=Side(style='thin'),
															right=Side(style='thin'))
						
				for r in range(start_row+4, start_row+self.row_max):
					cellform = ws["%s%s" % (c, r)]		
					if r == start_row+4:
						cellform.alignment = Alignment(horizontal="center",
													vertical="center")								
														
					else:						
						if c in ["A", "B"]:
							cellform.alignment = Alignment(horizontal="left",
										vertical="center",
										shrinkToFit=True)
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin'))								
								
						else:					
							cellform.alignment = Alignment(horizontal="center",
										vertical="center",
										wrapText=True)
							cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin'))	
			
			css = Border(left=Side(style='thin'),
						right=Side(style='thin'),
						top=Side(style='thin'),
						bottom=Side(style='thin'))	


			for c in ["G", "H", "I", "M", "N", "O", "S", "T", "U"]:
				for r in range(start_row+5, start_row+self.row_max):
					cellform = ws["%s%s" % (c, r)]
					cellform.fill = PatternFill(start_color='BAB8BA',
                   										end_color='BAB8BA',
                   										fill_type='solid')														
			for c in ["A", "B"]:
				for r in range(start_row+5, start_row+self.row_max):
					cellform = ws["%s%s" % (c, r)]
					cellform.fill = PatternFill(start_color='B0E0E6',
                   										end_color='B0E0E6',
                   										fill_type='solid')	
                   												
			ws["A%r" % start_row].value = "Event"
			ws["A%r" % (start_row+1)].value = program_number				
			ws["A%r" % (start_row+2)].value = "Time"			
			ws["A%r" % (start_row+3)].value = program_time

			ws.merge_cells("B%s:V%s" % (start_row, start_row+3))
			ws["B%s" % start_row].font = Font(size=30)
			ws["B%s" % start_row].value = round_name

			ws["W%s" % start_row].value = "Record"							
			ws["W%s" % (start_row+1)].value = record_who
			ws["W%s" % (start_row+2)].value = record_what
			ws["W%s" % (start_row+3)].value = record_when

			ws["A%s" % (start_row+4)].value = "House"
			ws["B%s" % (start_row+4)].value = "Competitor"
			ws["B%s" % (start_row+4)].border = css
			ws["C%s" % (start_row+4)].value = "PP"
			ws["C%s" % (start_row+4)].border = css

			pairs = [["D", "F"], ["G", "I"], ["J", "L"],
					["M", "O"], ["P", "R"], ["S", "U"]]
							
			for i in range(len(pairs)):
				pair = pairs[i]
				if height_ref != None:
					if height_ref[i] in [None, 0, '']:
						height = ''
					else:
						height = '%s m' % height_ref[i]
				else:
					height = 'm'
				ws.merge_cells("%s%s:%s%s" % (pair[0], start_row+4, pair[1], start_row+4))
				ws["%s%s" % (pair[0], start_row+4)].value = height
				ws["%s%s" % (pair[0], start_row+4)].border = css				
				ws["%s%s" % (pair[0], start_row+4)].alignment = Alignment(horizontal="center")						
	
			ws["V%s" % (start_row+4)].value = "Best"
			ws["V%s" % (start_row+4)].border = css
			ws["W%s" % (start_row+4)].value = "Place"

			min = p * 50
			max = ((p + 1) * 50)	
			page_id_list = alpha[min:max]	
				
			row_num = start_row + 5					
			for id in page_id_list:	
	
				if id == None:
					row_num += 1
					continue

				type = id_dict[id][5][:3].upper()
				name = '%s %s' % (id_dict[id][3], id_dict[id][4])
				pp = db_boole(id_dict[id][0], db_in=False)
				if pp == True:
					pp = 'X'
				else:
					pp = ''

				if id_dict[id][1]  in [None, '', 0, 0.0]:
					score = None
				else:
					score = db_height(id_dict[id][1], db_in=False)

				ws["A%s" % row_num].value = type
				ws["B%s" % row_num].value = name
				ws["C%s" % row_num].value = pp
				
				triple_list = [["D", "E", "F"], ["G", "H", "I"], ["J", "K", 'L'],
								['M', 'N', 'O'], ['P', 'Q', 'R'], ['S', 'T', 'U']]
				best_list = []
				for i in range(len(triple_list)):
					triple = triple_list[i]
					try:
						box = score[i]
					except TypeError:
						continue
				
					achieved = False
				
					for x in range(3):						
						letter = triple[x]
						if len(box) != 3:
							value = ''
						else:
							value = box[x]
						if value == 'O':
							achieved = True
						ws["%s%s" % (letter, row_num)].value = value		
								
					if achieved == True:
						if height_ref != None:						
							best_list.append(height_ref[i])

				if len(best_list) > 0:
					best = '%s m' % sorted(best_list)[-1]
				else:
					best = ''					
				ws['V%s' % row_num].value = best
				
				placing = id_dict[id][2]
				if placing in [0, None, '']:
					placing = ''
				ws['W%s' % row_num].value = placing
				
				row_num += 1	

class ExcelProgram():

	'''
	28/12/16 if I can make this into a word document, then convert that to 
	a PDF, they can use the PDF print as booklet function 
	(or maybe I can auto-print it)

	'''
	pass

class ExcelResults():

	def __init__(self, *args, **kwargs):

		## Settings
		self.tournament = kwargs.pop('tournament', None)

		self.org = Database('main').cursor.execute('''SELECT title
						FROM details''').fetchone()[0]

		self.db = Database(self.tournament)
		
		details = self.db.cursor.execute('''SELECT * 
											FROM details''').fetchall()[0]
		self.stage = details[2]
		self.type = details[3]
		self.age = details[4]


		self.type_dict = {}
		colours = self.db.cursor.execute('''SELECT title, colour_bg, colour_text
									FROM template_groups
									WHERE category="type"''').fetchall()
		for c in colours:
			self.type_dict[c[0]] = [db_colour(c[1], db_in=False), 
								db_colour(c[2], db_in=False)]

		## Database
		self.score_menu = {}
		self.results_menu = {}
		self.special = False
	
		for type in self.type_dict:
			self.results_menu[type] = {'pp':0, 'cp':0, 'sp':0}
		self.events_menu = {}

		self.access_scores() # self.results_menu
		self.access_database() # self.results_dict, self.records_dict
		self.access_champions() # self.champ_dict

		## Display
		try:
			self.wb = Workbook()
			self.wb.save('%s Print.xlsx' % self.tournament)
		except (OSError, IOError):
			PopBox().showwarning(title='Excel Error',
				message='%s Print.xlsx is already open, please close and try again' % self.tournament)
			return
		self.build_tournament_wb()
		self.build_event_wb()

		## Engine
		self.wb.save('%s Print.xlsx' % self.tournament)
		file_open('%s Print.xlsx' % self.tournament)

	### FUNCTIONS ###


	### ACCESS ###

	def access_scores(self):

		self.db.cursor.execute('''CREATE TABLE IF NOT EXISTS results 
								(event TEXT PRIMARY KEY)''')
		self.db.connection.commit()

		scores = self.db.cursor.execute('''SELECT * FROM results''').fetchall()

		contest_list = calculate_contest_list(self.tournament)
		grade_list = calculate_grade_list(self.tournament)
		event_list = []
		for contest in contest_list:
			for grade in grade_list:
				event_list.append('%s %s' % (grade, contest))

		self.complete_list = []
		for event in event_list:
			for tup in scores:
				if event == tup[0]:
					self.complete_list.append(event)
		self.events_menu = {} # redundant i think
		for event in self.complete_list:

			team = calculate_event_team(self.tournament, event)

			table = db_table('%s RR' % event)
			results = self.db.cursor.execute('''SELECT *
									FROM %s''' % table).fetchall()

			score_dict = {}

			first = []
			second = []
			third = []

			if team == False:

				for res in results:

					id = res[0]
					indiv = self.db.cursor.execute('''SELECT type, firstname, surname
									FROM individuals
									WHERE id="%s"''' % id).fetchone()

					type = indiv[0]
					fn = indiv[1]
					sn = indiv[2]
					rank = res[2]
					pp = res[3]
					cp = res[4]

					if type not in score_dict:
						score_dict[type] = {'pp':0, 'cp':0}

					if pp not in [None, '', 0, 0.0]:
						score_dict[type]['pp'] += pp
					if cp not in [None, '', 0, 0.0]:
						score_dict[type]['cp'] += cp

					if rank == 1:
						first.append([fn, sn, type])
					elif rank == 2:
						second.append([fn, sn, type])
					elif rank == 3:
						third.append([fn, sn, type])

			elif team == True:

				for res in results:

					team = res[0]
					rank = res[2]
					pp = res[3]
					cp = res[4]
					type = res[5]

					if type not in score_dict:
						score_dict[type] = {'pp':0, 'cp':0}

					if pp not in [None, '', 0, 0.0]:
						score_dict[type]['pp'] += pp
					if cp not in [None, '', 0, 0.0]:
						score_dict[type]['cp'] += cp

					if rank == 1:
						first.append([team, type])
					elif rank == 2:
						second.append([team, type])
					elif rank == 3:
						third.append([team, type])

			# add scores to self.results_menu
			for type in score_dict:

				pp = score_dict[type]['pp']
				cp = score_dict[type]['cp']

				self.results_menu[type]['pp'] += pp
				self.results_menu[type]['cp'] += cp

				try:
					special = self.db.cursor.execute('''SELECT points
								FROM special_events
								WHERE type="%s"''' % type).fetchall()

					count = 0
					for tup in special:
						count += float(tup[0])

					self.results_menu[type]['sp'] = count
					self.special = True

				except:
					self.results_menu[type]['sp'] = 0


			# add event details to self.event_menu
			self.events_menu[event] = {1:first, 2:second, 3:third}

		if self.complete_list == []:

			for type in self.results_menu:
				try:
					special = self.db.cursor.execute('''SELECT points
								FROM special_events
								WHERE type="%s"''' % type).fetchall()

					count = 0
					for tup in special:
						count += float(tup[0])

					self.results_menu[type]['sp'] = count
					self.special = True

				except:
					self.results_menu[type]['sp'] = 0			

	def access_database(self):

		contest_list = calculate_contest_list(self.tournament)
		grade_list = calculate_grade_list(self.tournament)

		self.results_dict = {'Index':[]}
		self.records_dict = {}

		for contest in contest_list:
			for grade in grade_list:
				event = '%s %s' % (grade, contest)
			
				# results
				try:
					res = self.db.cursor.execute('''SELECT id, rank
					FROM %s
					WHERE rank!=""''' % db_table('%s RR' % event)).fetchall()

					edict = {1:[], 2:[], 3:[]}

					for tup in res:
						if tup[1] in [1, 2, 3]:
							edict[tup[1]].append(tup[0])

					self.results_dict['Index'].append(event)
					self.results_dict[event] = edict
				except:
	
					continue

				# records
				record = self.db.cursor.execute('''SELECT record_what,
					record_who, record_when
					FROM template_events
					WHERE title="%s"''' % event).fetchone()

				if record[2] in ['2017', 2017]:
					self.records_dict[event] = [record[0], record[1]]


	def access_champions(self):

		## Settings

		grade_list = calculate_grade_list(self.tournament, combined=True,
											gender_check=True)
		if grade_list in [None, []]:
			grade_list = calculate_grade_list(self.tournament,
												open=True).remove('Open')	

		self.champ_dict = {}

		for grade in grade_list:

			gender, age = self.db.cursor.execute('''SELECT gender_grade, 
										age_grade
										FROM template_grades
										WHERE title="%s"''' % grade).fetchone()
		
			var = [grade, 'Open']
			if gender != None:
			
				var.append(gender)
			if age != None:
			
				var.append(age)

			event_list = []
			for v in var:
				events = self.db.cursor.execute('''SELECT title
										FROM template_events
										WHERE grade="%s"''' % v).fetchall()
				for e in events:
					event_list.append(e[0])

			id_list = calculate_id_list_v1(self.tournament, grade=grade)

			res = self.db.cursor.execute('''SELECT * FROM results''').fetchall()
			event_results = []
			for r in res:

				event_results.append(r[0])

			score_dict = {}
			for event in event_list:
				if event in event_results:

					results = self.db.cursor.execute('''SELECT * 
						FROM %s''' % db_table('%s RR' % event)).fetchall()

					for tup in results:

						id = tup[0]
						if id in id_list:

							if tup[3] not in [None, '']:
								pp = tup[3]
							else:
								pp = 0
							if tup[4] not in [None, '']:
								cp = tup[4]
							else:
								cp = 0

							total = pp + cp
							if total != 0:
								if id not in score_dict:
									score_dict[id] = total
								else:
									score_dict[id] += total

			order = sorted(score_dict, key=score_dict.get, reverse=True)
			ranking = {1:[], 2:[], 3:[]}
			rank = 1
			tie = 1
			for i in range(len(order)):

				id = order[i]
				score = score_dict[id]

				if rank > 3:
					break
				if score == 0 or score == 0.0:
					break

				ranking[rank].append(id)
				next_score = score_dict[order[i+1]]

				if score == next_score:
					tie += 1
					continue

				elif tie > 1:
					rank += tie
					tie = 1

				else:
					rank += tie

			if ranking[1] == [] and ranking[2] == [] and ranking[3] == []:
					
					self.champ_dict[grade] = None

			else:

					self.champ_dict[grade] = ranking

	### BUILD ###

	def build_tournament_wb(self):

		## Sheet Title 
		if 'Sheet' in self.wb.sheetnames:
			ws = self.wb.active
			ws.title = 'Tournament Results'
		else:
	
			ws = self.wb.create_sheet('Tournament Results')

		## Sheet Page Settings
		ws.page_margins.left = 0
		ws.page_margins.right = 0
		ws.page_margins.top = 0
		ws.page_margins.bottom = 0
		ws.print_options.horizontalCentered = True

		ws.column_dimensions['A'].width = 20
		for r in ['B', 'C']:
			ws.column_dimensions[r].width = 45

		## Sheet CSS
		self.css_border = Border(left=Side(style='thin'), right=Side(style='thin'),
							top=Side(style='thin'),bottom=Side(style='thin')) 
		self.css_fill = PatternFill(start_color='B0E0E6',
              						end_color='B0E0E6', fill_type='solid')	
		self.css_align = Alignment(horizontal='center',
												vertical='center',
												shrinkToFit=True)
		## Sheet 
		self.build_tournament_heading(ws)
		start_row = self.build_tournament_results(ws)
		self.build_tournament_champions(ws, start_row)

	def build_tournament_heading(self, ws):

		# Heading
		ws['B1'].value = self.org
		ws['B1'].font = Font(size=30)
		ws.merge_cells("B1:C6")
		ws['B7'].value = self.tournament
		ws['B7'].font = Font(size=20)		
		ws.merge_cells("B7:C11")
		for c in ['A', 'B', 'C']:
			for r in range(1, 12):
				cellform = ws['%s%s' % (c, r)]
				cellform.fill = self.css_fill
				cellform.alignment = Alignment(horizontal='center',
												vertical='center',
												shrinkToFit=True)
				if c != 'A':

					if r == 7 or r == 6:
						cellform.border = Border(left=Side(style='thin'), 
									right=Side(style='thin'))
					else:
						cellform.border = self.css_border
		# Logo
		img = Image_pyxl(resource_path('sponsor_wristband.png'))
		ws.add_image(img, 'A1')

	def build_tournament_results(self, ws):

		start_row = 13

		ws['A%s' % start_row].value = 'Tournament %s Results' % self.type
		ws['A%s' % start_row].font = Font(size=20)
		ws['A%s' % start_row].alignment = Alignment(horizontal='center',
												vertical='center',
												shrinkToFit=True)
		ws.merge_cells("A%s:C%s" % (start_row, start_row+2))

		for c in [['A', 'Placing'], ['B', self.type], ['C', 'Points']]:
			ws['%s%s' % (c[0], start_row+3)].value = c[1]
			ws['%s%s' % (c[0], start_row+3)].font = Font(size=16)	
			ws.merge_cells("%s%s:%s%s" % (c[0],start_row+3,
											c[0], start_row+4))

			for r in range(0, 5):
				cellform = ws['%s%s' % (c[0], start_row+r)]
				cellform.fill = self.css_fill
				cellform.border = self.css_border
				cellform.alignment = Alignment(horizontal='center',
												vertical='center',
												shrinkToFit=True)
		#$ self.type: results
		sr = start_row + 5

		order_dict = {}
		for type in self.results_menu:
			score = 0
			for points in self.results_menu[type]:
				score += self.results_menu[type][points]
			order_dict[type] = score
		
		ranked = sorted(order_dict, key=order_dict.get, reverse=True)
		rank = 1
		tie = 1
		results = {}
		for i in range(len(ranked)):

			type = ranked[i]
			score = order_dict[type]
		
			if rank not in results:
				results[rank] = [type]
			else:
				results[rank].append(type)

			if type != ranked[-1]:
				next_score = order_dict[ranked[i+1]]
			else:
				next_score = 0

			if score == next_score:
				tie += 1
				continue
			elif tie != 1:
				rank += tie
				tie = 1
			else:
				rank += tie
				tie = 1

		for r in results:

			res = sorted(results[r]) # puts in alphabetical order
			rank = calculate_ordinal(r)

			for type in res:

				for c in [['A', rank], ['B', type], ['C', order_dict[type]]]:

					cellform = ws['%s%s' % (c[0],sr)]
					cellform.value = c[1]
					cellform.font = Font(size=12)
					cellform.alignment = self.css_align
					ws.merge_cells("%s%s:%s%s" % (c[0], sr, c[0], sr+1))

				sr += 2

		return sr 

	def build_tournament_champions(self, ws, sr):

		start_row = sr

		ws['A%s' % start_row].value = 'Tournament Grade Champions'
		ws['A%s' % start_row].font = Font(size=20)
		ws['A%s' % start_row].alignment = Alignment(horizontal='center',
												vertical='center',
												shrinkToFit=True)
		ws.merge_cells("A%s:C%s" % (start_row, start_row+2))

		for c in [['A', 'Grade'], ['B', 'Name']]:
			ws['%s%s' % (c[0], start_row+3)].value = c[1]
			ws['%s%s' % (c[0], start_row+3)].font = Font(size=16)	
			
			if c[0] != 'A':
				ws.merge_cells("B%s:C%s" % (start_row+3,
											start_row+4))
			else:
				ws.merge_cells("A%s:A%s" % (start_row+3,
											start_row+4))

			for r in range(0, 5):
				cellform = ws['%s%s' % (c[0], start_row+r)]
				cellform.fill = self.css_fill
				cellform.border = self.css_border
				cellform.alignment = Alignment(horizontal='center',
												vertical='center',
												shrinkToFit=True)
				if c[0] == 'B':
					cellform = ws['C%s' % (start_row+r)]
					cellform.fill = self.css_fill
					cellform.border = self.css_border
					cellform.alignment = Alignment(horizontal='center',
												vertical='center',
												shrinkToFit=True)					

		# champion
		sr = start_row + 5

		grade_list = calculate_grade_list(self.tournament, combined=True,
											gender_check=True)

		for grade in grade_list:

			cellform = ws['A%s' % sr]
			cellform.value = grade
			cellform.font = Font(size=12)
			cellform.alignment = self.css_align
			ws.merge_cells("A%s:A%s" % (sr, sr+1))

			if self.champ_dict[grade] == None:

				cellform = ws['B%s' % sr]
				cellform.value = 'No Grade Champion'
				cellform.font = Font(size=12)
				cellform.alignment = self.css_align
				ws.merge_cells("B%s:C%s" % (sr, sr+1))				

			else:

				champ = []

				for id in self.champ_dict[grade][1]:

					fn, sn, type = self.db.cursor.execute('''SELECT firstname, 
										surname, type
										FROM individuals
										WHERE id="%s"''' % id).fetchone()

					champ.append('%s %s (%s)' % (fn, sn, type))

				champion = '&'.join(champ)	

				cellform = ws['B%s' % sr]
				cellform.value = champion
				cellform.font = Font(size=12)
				cellform.alignment = self.css_align
				ws.merge_cells("B%s:C%s" % (sr, sr+1))		

			sr += 2

	def build_event_wb(self):

		## Sheet Title 
		if 'Sheet' in self.wb.sheetnames:
			ws = self.wb.active
			ws.title = 'Event Results'
		else:
	
			ws = self.wb.create_sheet('Event Results')

		## Sheet Page Settings
		ws.page_margins.left = 0
		ws.page_margins.right = 0
		ws.page_margins.top = 0
		ws.page_margins.bottom = 0
		ws.print_options.horizontalCentered = True

		ws.column_dimensions['A'].width = 19
		for r in ['B', 'C', 'D', 'E', 'F']:
			ws.column_dimensions[r].width = 18

		## Sheet CSS
		self.css_border = Border(left=Side(style='thin'), right=Side(style='thin'),
							top=Side(style='thin'),bottom=Side(style='thin')) 
		self.css_fill = PatternFill(start_color='B0E0E6',
              						end_color='B0E0E6', fill_type='solid')	
		self.css_align = Alignment(horizontal='center',
												vertical='center',
												shrinkToFit=True)
		## Sheet 
		self.build_event_heading(ws)
		self.build_event_results(ws)

	def build_event_heading(self, ws):

		# Heading
		ws['B1'].value = self.tournament
		ws['B1'].font = Font(size=30)
		ws.merge_cells("B1:F6")
		ws['B7'].value = 'Event Results'
		ws['B7'].font = Font(size=20)		
		ws.merge_cells("B7:F11")
		for c in ['A', 'B', 'C', 'D', 'E', 'F']:
			for r in range(1, 12):
				cellform = ws['%s%s' % (c, r)]
				cellform.fill = self.css_fill
				cellform.alignment = Alignment(horizontal='center',
												vertical='center',
												shrinkToFit=True)
				if c != 'A':

					if r == 7 or r == 6:
						cellform.border = Border(left=Side(style='thin'), 
									right=Side(style='thin'))
					else:
						cellform.border = self.css_border
		# Logo
		img = Image_pyxl(resource_path('sponsor_wristband.png'))
		ws.add_image(img, 'A1')

		col_list = ['A', 'B', 'C', 'D', 'E', 'F']
		col_headings = ['Event', '1st', '2nd', '3rd', 'Old Record', 'New Record']

		for i in range(0, 6):

			col = col_list[i]
			cellform = ws['%s13' % col_list[i]]
			cellform.value = col_headings[i]
			cellform.font = Font(size=16)
			cellform.alignment = self.css_align

			ws.merge_cells('%s%s:%s%s' % (col,'13',col,'15'))

			for r in range(13, 16):

				cellform = ws['%s%s' % (col, r)]
				cellform.border = self.css_border
				cellform.fill = self.css_fill

	def build_event_results(self, ws):

		sr = 16
		col_list = ['A', 'B', 'C', 'D', 'E', 'F']

		css_align = Alignment(horizontal='center',
								vertical='center',
								wrapText=True, shrinkToFit=True)

		for event in self.complete_list:

			if event in self.events_menu:
				
				res = self.events_menu[event]
				arch = self.db.cursor.execute('''SELECT record_what,
					record_who, record_when, contest
					FROM archive_records
					WHERE title="%s"''' % event).fetchone()
				rec = self.db.cursor.execute('''SELECT record_what,
					record_who, record_when, contest
					FROM template_events
					WHERE title="%s"''' % event).fetchone()

				# top border
				for c in col_list:
					ws['%s%s' % (c, sr)].border = Border(top=Side(style='thin'))

				# event 
				ecell = ws['A%s' % sr]
				ecell.value = event
				ecell.alignment = css_align
				ws.merge_cells('A%s:A%s' % (sr, sr+2))

				team = calculate_contest_type(self.tournament, 
												rec[3], team=True)

				# ranks
				for rank in range(1, 4):

					if res[rank] not in [None, []]:

						for i in range(len(res[rank])):
							
							if team == True:
								name = res[rank][i][0]
							else:
								name = '%s %s' % (res[rank][i][0], 
											res[rank][i][1])

							ws['%s%s' % (get_column_letter(rank+1),
											(sr+i))].value = name
							ws['%s%s' % (get_column_letter(rank+1),
											(sr+i))].alignment = css_align
				# archive
				for i in range(3):
					cell = ws['E%s' % (sr+i)]
					if i == 0:
						ctype = calculate_contest_type(self.tournament, arch[3])
						cell.value = convert_score(arch[0], ctype)
					else:
						cell.value = str(arch[i])
					cell.alignment = css_align
				if rec[2] in ['2017', 2017]:
					for i in range(3):
						cell = ws['F%s' % (sr+i)]
						if i == 0:
							ctype = calculate_contest_type(self.tournament, rec[3])
							cell.value = convert_score(rec[0], ctype)
						else:
							cell.value = str(rec[i])
						cell.alignment = css_align
				else:
					cell = ws['F%s' % sr]
					cell.value = 'Record not broken'
					cell.alignment = css_align

					ws.merge_cells('F%s:F%s' % (sr, sr+2))

				# bottom border
				for c in col_list:
					ws['%s%s' % (c, sr+2)].border = Border(bottom=Side(style='thin')) 

				sr += 3

class ExcelTypeList():
	
	def __init__(self, *args, **kwargs):

		## KW Settings
		self.tournament = kwargs.pop('tournament', None)
		self.typename = kwargs.pop('typename', None)
		self.sponsor = kwargs.pop('sponsor', False)

		## Settings
		self.db = Database(self.tournament)

		details = self.db.cursor.execute('''SELECT * FROM details''').fetchone()
		self.type = details[3]

		types = self.db.cursor.execute('''SELECT title 
										FROM template_groups
										WHERE category="type"''').fetchall()
		self.type_tables = {}
		for tup in types:
			type = tup[0]
			self.type_tables[type] = {}
			tables = calculate_type_tables(self.tournament, type, grade=True)
			for pair in tables:
				self.type_tables[type][pair[0]] = pair[1]

		events = self.db.cursor.execute('''SELECT title, program_time 
										FROM template_events''').fetchall()
		self.event_program = {}
		for tup in events:
			self.event_program[tup[0]] = tup[1]

		## Display
		try:
			self.wb = Workbook()
			self.wb.save('%s Print.xlsx' % self.tournament)
		except (OSError, IOError):
			PopBox().showwarning(title='Excel Error',
				message='%s Print.xlsx is already open, please close and try again' % self.tournament)
			return

		grade_list = calculate_grade_list(self.tournament)

		if self.typename != None:
			for grade in grade_list:
				if grade in self.type_tables[self.typename]:
					table = self.type_tables[self.typename][grade]
					self.build_worksheet(self.typename, grade, table)		
		else:
			for type in sorted(self.type_tables):
				for grade in grade_list:
					if grade in self.type_tables[type]:
						table = self.type_tables[type][grade]
						self.build_worksheet(type, grade, table)

		## Engine
		self.wb.save('%s Print.xlsx' % self.tournament)
		try:
			
			file_open('%s Print.xlsx' % self.tournament)
		except IOError:
			return

	### BUILD ###	

	def build_worksheet(self, type, grade, table):

		## Settings

		events = self.db.list_columns(table)
		event_list = []
		extra_event_list = [] # for extra rounds
		for event in events:
			if event == 'id':
				continue

			extra_rounds = self.db.cursor.execute('''SELECT round, title
								FROM %s
								WHERE round!="RR"''' % event).fetchall()
			
			for extra in extra_rounds:

				extra_event = extra[1]
				extra_table = db_table("%s %s" % (event, extra[0]))
				extra_event_list.append([extra_event, extra_table, event])

		indiv = self.db.cursor.execute("SELECT * FROM %s" % table).fetchall()

		id_dict = {}
		alpha_dict = {}
		for tup in indiv:
			id = tup[0]
			choices = tup[1:]
			name = self.db.cursor.execute('''SELECT firstname, surname
										FROM individuals
										WHERE id="%s"''' % id).fetchone()
			fn = name[0]
			sn = name[1]
			id_dict[id] = [fn, sn, choices] # [Adam, Adamson, [list here]]
			alpha_dict[id] = '%s %s' % (sn, fn)

		alpha = sorted(alpha_dict, key=alpha_dict.get)

		event_dict = {}
		for i in range(len(event_list)):

			event = event_list[i]
			entry = []
			for id in alpha:
				if id_dict[id][2][i] in [1, '1']:
					entry.append(id)
			event_dict[event] = entry

		grade_id_list = calculate_id_list(self.tournament, grade=grade,
											type=type)
		for evt in extra_event_list:

			event = evt[0]
			entry = []

			ex_indiv = self.db.cursor.execute('''SELECT id
								FROM %s''' % evt[1]).fetchall()

			for tup in ex_indiv:
				if tup[0] in grade_id_list:
					entry.append(tup[0])
					if tup[0] not in id_dict:
						id = tup[0]
						choices = []
						name = self.db.cursor.execute('''SELECT firstname, surname
													FROM individuals
													WHERE id="%s"''' % id).fetchone()
						fn = name[0]
						sn = name[1]
						id_dict[id] = [fn, sn, choices] # [Adam, Adamson, [list here]]
						alpha_dict[id] = '%s %s' % (sn, fn)

			event_dict[event] = entry

		# reset alpha after extra rounds
		alpha = sorted(alpha_dict, key=alpha_dict.get)
		for event in extra_event_list:
			event_list.append(event[0])

		## Excel Settings
		if 'Sheet' in self.wb.sheetnames:
			ws = self.wb.active
			ws.title = '%s - %s' % (type, grade)
		else:

			ws = self.wb.create_sheet('%s - %s' % (type, grade))

		## Contest Type Specific 
		ws.page_margins.left = 0
		ws.page_margins.right = 0
		ws.page_margins.top = 0
		ws.page_margins.bottom = 0
		ws.print_options.horizontalCentered = True

		col_list = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]

		for c in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
			ws.column_dimensions[c].width = 10

		pages = 1		
		row = 4
		count = 50
		for event in event_list:

			check = 3
			id_list = event_dict[event]
			if len(id_list) <= 6:
	
				check += 2
			else:
				r = len(id_list) / 3
				if float(r).is_integer():
					check += int(r)
				else:
					check += int(r) + 1

			if count - check < 0:
				pages += 1
				count = 50
				row = 4
		
			else:
				count -= check
		## Page
		event_complete = []
		for p in range(0, pages): # runs page			
			start_row = 1 + (p * 54)     

			ws.merge_cells('A%s:I%s' % (start_row, start_row+3))
			ws['A%s' % start_row].value = '%s - %s' % (type, grade)

			for c in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:			
				for r in range(start_row, start_row+3):
					cellform = ws["%s%s" % (c, r)]	
					cellform.alignment = Alignment(horizontal="center", 
											vertical="center",
											shrinkToFit=True)
					cellform.font = Font(size=30)
					cellform.fill = PatternFill(start_color='B0E0E6',
              						     end_color='B0E0E6', fill_type='solid')        
					cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin'),
												top=Side(style='thin'),
												bottom=Side(style='thin')) 								
		
			css_align = Alignment(horizontal="center", 
									vertical="center",
									shrinkToFit=True)
			css_font = Font(size=30)
			css_fill = PatternFill(start_color='B0E0E6',
              						     end_color='B0E0E6', fill_type='solid')        
			css_border = Border(left=Side(style='thin'),
										right=Side(style='thin'),
										top=Side(style='thin'),
										bottom=Side(style='thin')) 
		
			for c in col_list:
				ws['%s%s' % (c, start_row+4)].border = Border(top=Side(style='thin'))
			for r in range(start_row, start_row+4):
				ws['J%s' % r].border = Border(left=Side(style='thin'))

			row = start_row + 3
			count = 50

			if p == 0: # for Grade Id List

				check = 2
				height = 0
				id_list = alpha
				if len(id_list) <= 6:
		
					check += 2
					height = 2
				else:
					r = len(id_list) / 3
					if float(r).is_integer():
						check += int(r)
						height = int(r)
					else:
						check += int(r) + 1
						height = int(r) + 1
				if count - check - 1 > 0:

					## Build ID Box

					sr = row + 2
					time1 = ws["I%s" % sr]
					time1.alignment = css_align
					time1.fill = css_fill

					time2 = ws["I%s" % (sr+1)]
					time2.alignment = css_align
					time2.fill = css_fill	
				
					ws.merge_cells('A%s:H%s' % (sr, sr+1))
					title = ws['A%s' % sr]
					title.value = "ID List"
					title.alignment = css_align
					title.font = Font(size=20)

					title.fill = css_fill

					for i in range(height): # A (left side border)

						try:
							id = id_list[i]

							name = '%s %s %s' % (id_dict[id][0], 
												id_dict[id][1], id)
							ws['A%s' % (sr+2+i)].value = name
							ws['A%s' % (sr+2+i)].border = Border(left=Side(style='thin'))
						except:
							ws['A%s' % (sr+2+i)].border = Border(left=Side(style='thin'))
							continue
				
					for i in range(height): # D 

						try:
							id = id_list[i + height]

							name = '%s %s %s' % (id_dict[id][0], 
												id_dict[id][1], id)
							ws['D%s' % (sr+2+i)].value = name
						
						except:
							break
				
					for i in range(height): # G 

						try:
							id = id_list[i + height + height]

							name = '%s %s %s' % (id_dict[id][0], 
												id_dict[id][1], id)
							ws['G%s' % (sr+2+i)].value = name

						except:
							break

					# border fixes

					for r in range(sr+2, sr+1+height):
					
						ws['J%s' % r].border = Border(left=Side(style='thin'))			

					for c in col_list:
						if c == 'A':
							ws['%s%s' % (c, sr+2)].border = Border(top=Side(style='thin'),
															left=Side(style='thin'))
						else:
							ws['%s%s' % (c, sr+2)].border = Border(top=Side(style='thin'))
						
						if c == "I":
							ws['%s%s' % (c, sr+1+height)].border = Border(bottom=Side(style='thin'),
																	right=Side(style='thin'))
						else:
							ws['%s%s' % (c, sr+2+height)].border = Border(top=Side(style='thin'))

					time1.border = Border(right=Side(style='thin'),
											top=Side(style='thin'))	
					time2.border = Border(right=Side(style='thin'),
											bottom=Side(style='thin'))	
					title.border = css_border

					# fix settings						
					row += check + 1
					count -= (check + 1)				

			for event in event_list:
				
				if event in event_complete:
					continue
				
				check = 2
				height = 0
				id_list = event_dict[event]
				if len(id_list) <= 6:
		
					check += 2
					height = 2
				else:
					r = len(id_list) / 3
					if float(r).is_integer():
						check += int(r)
						height = int(r)
					else:
						check += int(r) + 1
						height = int(r) + 1
				if count - check - 1 < 0:
				
					break

				## Build Event Box

				for evt in extra_event_list:

					if event == evt[0]:

						program_time = self.db.cursor.execute('''SELECT program_time
							FROM %s
							WHERE title="%s"''' % (evt[2], event)).fetchone()

				if program_time == None:
			
					try:
						program_time = self.db.cursor.execute('''SELECT program_time
									FROM template_events
									WHERE title="%s"''' % event).fetchone()[0]
					except:
						program_time = self.db.cursor.execute('''SELECT program_time
									FROM template_events
									WHERE title="%s"''' % event).fetchone()[0]					
				else:
					try:
						if program_time[0] == '':
							program_time = self.db.cursor.execute('''SELECT program_time
									FROM template_events
									WHERE title="%s"''' % event).fetchone()[0]		
					except:			
						program_time = program_time[0]

				sr = row + 2
				time1 = ws["I%s" % sr]
				time1.value = "Event Time"
				time1.alignment = css_align
				time1.fill = css_fill
				time1.border = Border(left=Side(style='thin'),
										right=Side(style='thin'),
										top=Side(style='thin'))	
				time2 = ws["I%s" % (sr+1)]
				try:
					time2.value = program_time
				except:
					time2.value = program_time[0]
				time2.alignment = css_align
				time2.fill = css_fill	
				time2.border = Border(left=Side(style='thin'),
										right=Side(style='thin'),
										bottom=Side(style='thin'))		
				ws.merge_cells('A%s:H%s' % (sr, sr+1))
				title = ws['A%s' % sr]
				title.value = event
				title.alignment = css_align
				title.font = Font(size=20)
				title.border = css_border
				title.fill = css_fill

				for i in range(height): # A (left side border)

					try:
						id = id_list[i]

						name = '%s %s' % (id_dict[id][0], id_dict[id][1])
						ws['A%s' % (sr+2+i)].value = name
						ws['A%s' % (sr+2+i)].border = Border(left=Side(style='thin'))
					except:
						ws['A%s' % (sr+2+i)].border = Border(left=Side(style='thin'))
						continue
			
				for i in range(height): # D 

					try:
						id = id_list[i + height]

						name = '%s %s' % (id_dict[id][0], id_dict[id][1])
						ws['D%s' % (sr+2+i)].value = name
					
					except:
						break
			
				for i in range(height): # G 

					try:
						id = id_list[i + height + height]

						name = '%s %s' % (id_dict[id][0], id_dict[id][1])
						ws['G%s' % (sr+2+i)].value = name

					except:
						break

				# border fixes

				for r in range(sr+2, sr+1+height):
				
					ws['J%s' % r].border = Border(left=Side(style='thin'))			

				for c in col_list:
					if c == 'A':
						ws['%s%s' % (c, sr+2)].border = Border(top=Side(style='thin'),
														left=Side(style='thin'))
					else:
						ws['%s%s' % (c, sr+2)].border = Border(top=Side(style='thin'))
					
					if c == "I":
						ws['%s%s' % (c, sr+1+height)].border = Border(bottom=Side(style='thin'),
																right=Side(style='thin'))
					else:
						ws['%s%s' % (c, sr+2+height)].border = Border(top=Side(style='thin'))

				# removes the event option
				row += check + 1
				count -= (check + 1)
				event_complete.append(event)

class ExcelWristbands():
	
	def __init__(self, *args, **kwargs):

		## KW Settings
		self.tournament = kwargs.pop('tournament', None)
		self.typename = kwargs.pop('typename', None)
		self.sponsor = kwargs.pop('sponsor', False)
		self.row_max = kwargs.pop('row_max', 40)

		## Settings
		self.db = Database(self.tournament)

		details = self.db.cursor.execute('''SELECT * FROM details''').fetchone()
		self.type = details[3]

		types = self.db.cursor.execute('''SELECT title 
										FROM template_groups
										WHERE category="type"''').fetchall()
		self.type_tables = {}
		for tup in types:
			type = tup[0]
			self.type_tables[type] = {}
			tables = calculate_type_tables(self.tournament, type, grade=True)
			for pair in tables:
				self.type_tables[type][pair[0]] = pair[1]

		events = self.db.cursor.execute('''SELECT title, program_time 
										FROM template_events''').fetchall()
		self.event_program = {}
		for tup in events:
			self.event_program[tup[0]] = tup[1]
		
		## Display
		try:
			self.wb = Workbook()
			self.wb.save('%s Print.xlsx' % self.tournament)
		except (OSError, IOError):
			PopBox().showwarning(title='Excel Error',
				message='%s Print.xlsx is already open, please close and try again' % self.tournament)
			return

		grade_list = calculate_grade_list(self.tournament)
		for grade in grade_list:
			if grade in self.type_tables[type]:
				self.build_worksheet(grade)		

		## Engine
		try:
			self.wb.save('%s Print.xlsx' % self.tournament)
			file_open('%s Print.xlsx' % self.tournament)
		except IOError:
			return

	### FUNCTIONS ###

	### BUILD ###	

	def build_worksheet(self, grade):

		## Database
		id_list = calculate_id_list_v1(self.tournament, grade=grade, 
									type=self.typename)
		event_list = calculate_event_list_state(self.tournament)

		id_dict = {}
		alpha_dict = {}
		for id in id_list:
	
			indiv = self.db.cursor.execute('''SELECT firstname, surname
							FROM individuals
							WHERE id="%s"''' % id).fetchone()

			choices = []

			for evt in event_list:

				event_name = evt[0]
				round = evt[1]
				time = evt[2]
				event = evt[3]
				table = db_table('%s %s' % (event_name, round))

				try:
					check = self.db.cursor.execute('''SELECT id
								FROM %s
								WHERE id="%s"''' % (table, id)).fetchone()
				except:
					# probably a team
					continue

				if check == None:
					continue
				else:
					choices.append([event, time])

			fn = indiv[0]
			sn = indiv[1]

			id_dict[id] = [fn, sn] + choices
			alpha_dict[id] = '%s %s' % (sn, fn)
		alpha = sorted(alpha_dict, key=alpha_dict.get)

		## Sheet Title
		if 'Sheet' in self.wb.sheetnames:
			ws = self.wb.active
			ws.title = '%s - %s' % (self.typename, grade)
		else:
	
			ws = self.wb.create_sheet('%s - %s' % (self.typename, grade))

		## Sheet Page Settings
		ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE				
		ws.page_margins.left = 0
		ws.page_margins.right = 0
		ws.page_margins.top = 0
		ws.page_margins.bottom = 0
		ws.print_options.horizontalCentered = True		

		ws.column_dimensions['A'].width = 1
		letter_list = [['B', 'C', 'D'],['E', 'F', 'G'],['H', 'I', 'J'],
						['K','L','M'], ['N', 'O', 'P'], ['Q', 'R', 'S']]							
		for list in letter_list:
			ws.column_dimensions[list[0]].width = 6.5
			ws.column_dimensions[list[1]].width = 13.5
			ws.column_dimensions[list[2]].width = 1

		##
		count = len(alpha)
		position = 0
		page = 0
		while count > 0:
			sr = (page * self.row_max) # starting row
			for c in range(6):	
				col1 = letter_list[c][0]
				col2 = letter_list[c][1]				
				col3 = letter_list[c][2]				
				
				try:
					id = alpha[position]
				except IndexError:
					break

				# Individual Section
				
				## Caitlin/Connor Code
				bugsy = False
				main = Database('main')
				school = main.cursor.execute('''SELECT title
											FROM details''').fetchone()[0]
				if school == 'Nga Tawa':
					if id_dict[id][0] == 'Caitlin':
						if id_dict[id][1] in ["O'Sullivan", "O'sullivan", 
												"Osullivan", "OSullivan"]:
							bugsy = '(The Muggle)'
				elif school in ["St Patrick's College Silverstream",
								"St Patricks College Silverstream",
								"St Patrick's College, Silverstream"]:
					if id_dict[id][0] == 'Connor':
						if id_dict[id][1] in ["O'Sullivan", "O'sullivan", 
												"Osullivan", "OSullivan"]:
							bugsy = '(Joffrey)'					
			
				for col in [col1, col2]:
					ws['%s%s' % (col,sr+4)].border = Border(bottom=Side(style='thin'))
				for col in [col1, col2]:
					ws['%s%s' % (col,sr+10)].border = Border(top=Side(style='thin'))				
				
				ws.merge_cells('%s%s:%s%s' % (col1,sr+5,col2,sr+6))
				cellform = ws['%s%s' % (col1, sr+5)]
				cellform.value = self.typename
				cellform.font = Font(size=25)
				cellform.alignment = Alignment(horizontal='center',
												vertical='center', shrinkToFit=True)
				for i in range(7, 10):
					ws.merge_cells('%s%s:%s%s' % (col1,sr+i,col2,sr+i))
					cellform = ws['%s%s' % (col1,sr+i)]					
					if i == 7:
						value = id		
					elif i == 8:
						if bugsy != False:
							value = "%s %s" % (id_dict[id][0], bugsy)
						else:
							value = id_dict[id][0]
					elif i == 9:
						value = id_dict[id][1]
					cellform.value = value
					cellform.alignment = Alignment(horizontal='center',
													shrinkToFit=True)
				# Event Section
				choices = id_dict[id][2:]
				es = 10 # event starting
				event_count = 0
				for choice in choices:

					evt = choice[0]
					event_time = choice[1]	
	
					ws.merge_cells('%s%s:%s%s' % (col1,sr+es,col1,sr+es+1))
					cellform = ws['%s%s' % (col1,sr+es)]
					cellform.alignment = Alignment(vertical='center', shrinkToFit=True)
					cellform.value = event_time
					cellform.font = Font(size=16)

					#ws.merge_cells('%s%s:%s%s' % (col2,sr+es,col2,sr+es+1))
					cellform = ws['%s%s' % (col2,sr+es)]
					cellform.value = evt
					cellform.alignment = Alignment(vertical='center', 
													wrapText=True,
													shrinkToFit=True)
					if len(evt) > 26:
						if len(evt) > 29:
							cellform.font = Font(size=8)
						else:
							cellform.font = Font(size=9)
					ws.merge_cells('%s%s:%s%s' % (col2,sr+es,col2,sr+es+1))
			

					if event_count == 15:
						break
					else:
						event_count += 1
						es += 2	

				# option to add a sponsor advert
				if self.sponsor == True:

					if event_count <= 10:

						try:
							img = Image_pyxl(resource_path('sponsor_wristband.png'))
							if es == 10:
								ws.add_image(img, '%s%s' % (col1, sr+12))
							elif event_count == 10:
								ws.add_image(img, '%s%s' % (col1, sr+es))
							else:
								ws.add_image(img, '%s%s' % (col1, sr+es+2))
						except:
							pass



				# outside line
				for r in range(sr+1, ((page+1)*self.row_max)):
					ws['T%s' % r].border = Border(left=Side(style='thin'))

				position += 1
				count -= 1
			page +=1

		## Side Lines
		for p in range(page):
			start_row = 1 + (p * self.row_max)
			for letter in ['A', 'D', 'G', 'J', 'M', 'P', 'S']:
				for sr in range(start_row, (start_row+self.row_max)):
					cellform = ws['%s%s' % (letter,sr)]
					cellform.border = Border(left=Side(style='thin'),
												right=Side(style='thin')) 

### EXCEL PROCESS ### 

class ExcelEntryConfirm():

	def __init__(self, *args, **kwargs):

		## Settings
		self.tournament = kwargs.pop('tournament', None)
		self.type = kwargs.pop('type', None)
		self.age = kwargs.pop('age', None)
		self.db = Database(self.tournament)

		## Engine
		try:
			self.extract_entry_sheet()
		except IOError:
			PopBox().showwarning(title='Excel Error',
				message="%s Entry Sheet not created, click 'Create Entry Sheet' and fill it out first" % self.tournament)
			return

	### FUNCTIONS ###

	def command_datetime(self, boole):

		if boole == True:	
			self.extract_filters()

	### EXTRACT ###

	def extract_entry_sheet(self):

		wb = load_workbook(resource_path("%s Entry Sheet.xlsx" % self.tournament))
		ws = wb.active		
		row_count = ws.max_row

		if self.age == 'Date of Birth':
			col_count = 5
		else:
			col_count = 6

		## Check Page
		errors = []

		# check for spaces/errors
		if row_count <=6:
			errors.append("No details have been entered")

		for r in range(6, row_count+1):
			for c in range(1, col_count+1):
				if ws.cell(row=r, column=c).value == None:
					errors.append("Check details, some rows are missing values")


		err = '\n\n'.join(set(errors))
		if len(errors) > 0:
			PopBox().showwarning(title='Error',
				message=err)
			return			

		## Check Filters
		error_list = []

		self.indiv_dict = {}
			
		self.type_filters = [] 	
		self.gender_filters = []
		self.age_filters = []

		# get filters from columns and create 
		base_number = 1300	
		for r in range(6, (row_count+1)):
			temp_list = []			
			for c in range(1, (col_count+1)):
				column_name = ws.cell(row=4, column=c).value
				student_detail = ws.cell(row=r, column=c).value
				# extract group data
				if column_name in ['First Name','Firstname', 'Surname']:
					temp_list.append(student_detail.title())
				elif column_name == "%s" % self.type:
					try:
						self.type_filters.append(student_detail.title())
					except AttributeError:
						self.age_filters.append(student_detail)
					temp_list.append(student_detail)
				elif column_name == "Gender":
					self.gender_filters.append(student_detail.title())
					temp_list.append(student_detail)
				elif column_name == "Date of Birth":
					if type(student_detail) is datetime.date:
						temp_list.append(student_detail)	
					elif type(student_detail) is datetime:
						d = student_detail.date()
						temp_list.append(d)						
					else:
						d = convert_date(student_detail, error=True)
						if d == False:
							error_list.append(True)
							d = convert_date(student_detail)
						temp_list.append(d)					
				elif column_name == self.age:
					try:
						self.age_filters.append(str(student_detail).title())
					except AttributeError:
						self.age_filters.append(str(student_detail))
					temp_list.append(student_detail)					

			id = "ID%d" % base_number
			self.indiv_dict[id] = temp_list
			base_number += 1

		if len(error_list) > 0:
			PopBox().askyesno(title='Error', function=self.command_datetime,
		message="Not all Date of Birth values seem to be in 'dd/mm/YYYY' format, do you wish to continue anyway?")
			return

		self.extract_filters()

	def extract_filters(self):

		# oracle dicts (because they predict)
		self.gender_oracle = {'Male':['M','Male','Boys','Mens'],
							'Female':['F','Female','Girls','Womens','Ladies']}	
		self.type_oracle = {}
		self.age_oracle = {}

		# exception dicts 
		self.gender_exceptions = []
		self.type_exceptions = []
		self.age_exceptions = []

		# extract the titles (and filters for gender) of the groups
		for filter in ['gender', 'type', 'age']:
				
			if filter == 'age':
				if self.age == 'Date of Birth':
					continue
				
			command = '''SELECT title, filter FROM template_groups
						WHERE category="%s"''' % filter

			db = self.db.cursor.execute(command).fetchall()

			if filter == 'gender': # only one using filter
				for tup in db:
					if tup[0] not in self.gender_oracle[tup[1]]:
						self.gender_oracle[tup[1]].append(tup[0])
			
			elif filter == 'age':
				for tup in db:
					self.age_oracle[tup[0]] = [tup[0]]

			else:
				for tup in db:
					self.type_oracle[tup[0]] = [tup[0]]
			
		# compare with the filters from the entry sheet		
		for filters in [self.gender_filters, self.type_filters, self.age_filters]:

			if filters == self.age_filters:
				if self.age == 'Date of Birth':
					continue

			if filters == self.gender_filters:
				genders = self.gender_oracle['Male']+self.gender_oracle['Female']	
				for filter in filters:
					if filter not in genders:
						if filter not in self.gender_exceptions:
							self.gender_exceptions.append(filter)

			elif filters == self.type_filters:
				types = []
				for title in self.type_oracle:
					types += self.type_oracle[title]
				for filter in filters:
					if filter not in types:
						if filter not in self.type_exceptions:
							self.type_exceptions.append(filter)

			elif filters == self.age_filters:
				ages = []
				for title in self.age_oracle:
					ages += self.age_oracle[title]
				for filter in filters:
					if filter not in ages:
						if filter not in self.age_exceptions:
							self.age_exceptions.append(filter)

		for exception in [self.gender_exceptions, self.type_exceptions,
							self.age_exceptions]:
			if len(exception) > 0:
				self.build_exception_pop()
				return
		
		self.build_individual_table()

	### BUILD ###

	def build_individual_table(self, *args):

		gdb = self.db.cursor.execute('''SELECT filter, title
				FROM template_groups WHERE category="gender"''').fetchall()
		gender_titles = {}
		for tup in gdb:
			gender_titles[tup[0]] = tup[1]

		if self.age == 'Date of Birth':
			age_titles = calculate_age_filters(self.tournament)

		var_list = []
		for ind in self.indiv_dict:

			indiv = self.indiv_dict[ind]

			firstname = indiv[0].title()
			surname = indiv[1].title()
			dob = indiv[2]
			gender = None

			for g in self.gender_oracle:
				if indiv[3] in self.gender_oracle[g]:
					gender = g
			type = None
			for t in self.type_oracle:
				if indiv[4] in self.type_oracle[t]:
					type = t
			year = ''

			for a in self.age_oracle:
				if indiv[5] in self.age_oracle[a] or str(indiv[5]) in self.age_oracle[a]:
					year = a

			try:
				gender_grade = [gender_titles[gender]]
			except:
				continue

			age_grade = ['Open']
			if self.age != 'Date of Birth':
				age_grade.append(year)
			else:
				for t in age_titles:
					title = t
					min = age_titles[t][0]
					max = age_titles[t][1]
					if min == None:
						if dob > max:
							age_grade.append(title)
					elif max == None:
						if dob <= min:
							age_grade.append(title)
					else:
						if  dob <= min and dob > max:
							age_grade.append(title)
			combined_grade = []
			for ag in age_grade:
				for gg in gender_grade:
					cg = "%s %s" % (ag, gg)
					combined_grade.append(cg)

			var = [ind, firstname, surname, convert_date(dob, string=True), 
					gender, type, year,
					db_list(gender_grade), db_list(age_grade), 
					db_list(combined_grade)]

			var_list.append(var)

		self.db.cursor.execute('''DROP TABLE IF EXISTS individuals''')
		self.db.connection.commit()

		self.db.cursor.execute('''CREATE TABLE individuals
					(id PRIMARY KEY, firstname TEXT, surname TEXT,
					dob TEXT, gender TEXT, type TEXT, year TEXT,
					gender_grade TEXT, age_grade TEXT, 
					combined_grade TEXT)''') # events TEXT
		self.db.connection.commit()

		command = '''INSERT INTO individuals
				VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?);'''
		self.db.cursor.executemany(command, var_list)
		self.db.connection.commit()

		self.process_competitors_to_events()

	def build_exception_pop(self):
		
		# functions
		def command_process(*args):

			if title.text == 'Unexpected Details':

				body.remove_widget(args[0])
				self.pop_confirm.text = 'Save & Continue'

				ref = exception_db[0]
				body.add_widget(ref[0])

				if ref[-1] == 'type':
					title.text = self.type
				else:
					title.text = ref[-1].title()
				return

			try:
				ref = exception_db[0]
			except IndexError:
				return

			if ref[2].text == 'Select':
				ref[3].text = 'Please select an option'
				return

			if ref[-1] == 'gender':
				self.gender_oracle[ref[2].text].append(ref[1].text)
				self.gender_exceptions.remove(ref[1].text)
			elif ref[-1] == 'type':
				self.type_oracle[ref[2].text].append(ref[1].text)
				self.type_exceptions.remove(ref[1].text)
			elif ref[-1] == 'age':
				self.age_oracle[ref[2].text].append(ref[1].text)
				self.age_exceptions.remove(ref[1].text)

			body.remove_widget(ref[0])
			exception_db.remove(ref)

			if len(exception_db) > 0:

				ref = exception_db[0]
				body.add_widget(ref[0])

				if ref[-1] == 'type':
					title.text = self.type
				else:
					title.text = ref[-1].title()				

			else:
				self.pop.dismiss()
				self.build_individual_table(0)

		# database
		exception_db = []

		gender_options = ['Male', 'Female']
		type_options = []
		for title in self.type_oracle:
			type_options.append(title)
		age_options = []
		for title in self.age_oracle:
			if title == 'Open':
				continue	
			age_options.append(title)

		for exp in self.gender_exceptions:
			ref = self.build_exception_box(exp, gender_options)
			ref.append('gender')
			exception_db.append(ref)
		for exp in self.type_exceptions:
			ref = self.build_exception_box(exp, type_options)
			ref.append('type')
			exception_db.append(ref)
		for exp in self.age_exceptions:
			ref = self.build_exception_box(exp, age_options)
			ref.append('age')
			exception_db.append(ref)

		# display
		content = GridLayout(cols=1)

		title = Label(text='Unexpected Details', font_size=30,
						size_hint_y=None, height=50)
		content.add_widget(title)

		body = GridLayout(cols=1)
		if len(self.gender_exceptions) == 0:
			exp = "Confirm which Title or Gender, the following filters belong to"
		else:
			exp = "Confirm which Title, the following filters belong to"
		c1 = LabelWrap(text=exp, size_hint_y=None, height=40)
		body.add_widget(c1)
		content.add_widget(body)

		action = BoxLayout(size_hint_y=None, height=50)
		dismiss = Button(text='Close', font_size=20)
		self.pop_confirm = Button(text='Continue', font_size=20)
		self.pop_confirm.bind(on_press=lambda i, c1=c1:command_process(c1))
		action.add_widget(dismiss)
		action.add_widget(self.pop_confirm)
		content.add_widget(action)

		# engine
		self.pop = Popup(title='', content=content, auto_dismiss=False, 
					size_hint=(None, None), size=(700, 500))
		self.pop.open()

		dismiss.bind(on_press=self.pop.dismiss)

	def build_exception_box(self, exception, options):

		box = GridLayout(cols=1)

		help = Label(text='', size_hint_y=None, height=30)
		box.add_widget(help)

		header = BoxLayout(size_hint_y=None, height=40)
		box.add_widget(header)
		header.add_widget(Label(text='This Filter', font_size=20))
		header.add_widget(Label(text='Belongs to', font_size=20))

		pad = GridLayout(cols=1)
		box.add_widget(pad)

		choice = BoxLayout(size_hint_y=None, height=100)
		pad.add_widget(choice)

		f1 = Label(text=exception, font_size=30)
		choice.add_widget(f1)

		o1 = OptionMenu(options=options)
		choice.add_widget(o1)

		return [box, f1, o1, help]

	### PROCESS ###

	def process_competitors_to_events(self, *args):

		# creates all the House-Grade tables

		grade_list = calculate_grade_list(self.tournament, combined=True,
											gender_check=True)
		if grade_list == None or grade_list == []:
			grade_list = calculate_grade_list(self.tournament,
												open=True).remove('Open')
			if grade_list == None or grade_list == []:
				grade_list = ['Open']

		type_list = self.db.cursor.execute('''SELECT title FROM template_groups
					WHERE category="type"''').fetchall()

		for tup in type_list:

			type = tup[0]
			var_tables = []	

			# extracts table and column names
			for grade in grade_list:

				typelist = type.split(' ')
				t = '_'.join(typelist + grade.split(' '))

				events = calculate_grade_events(self.tournament, grade, 
												table=True, team=False)
				var_tables.append([t, events])

			# creates table and colums
			for var in var_tables:
				v = var[0]

				try:
					self.db.cursor.execute('''CREATE TABLE %s(id)''' % v)
					for e in var[1]:
						self.db.cursor.execute('''ALTER TABLE %s
										ADD %s TEXT''' % (v, e))

					self.db.connection.commit()
				except:
					print "process_competitors_to_events"

			# inserts student ids
			for grade in grade_list:
				typelist = type.split(' ')
				t = '_'.join(typelist + grade.split(' '))
				id_list = calculate_id_list(self.tournament, type=type,
											grade=grade, tup_style=True)

				self.db.cursor.executemany('''INSERT INTO %s (id)
										VALUES (?)''' % t, id_list)
				self.db.connection.commit()

		# update and reset
		self.db.cursor.execute('''UPDATE details 
								SET stage="Entry - Events"''')
		self.db.connection.commit()

		# update access settings
		titles = self.db.cursor.execute('''SELECT title 
										FROM template_groups
										WHERE category="type"''').fetchall()
		var = []

		for tup in titles:
			var.append(['Captain', tup[0], 1, ''])
		self.db.cursor.executemany('''INSERT INTO settings
									(role, category, active, password)
									VALUES(?,?,?,?)''', var)
		self.db.connection.commit()

		# entry sheet delete
		file = resource_path("%s Entry Sheet.xlsx" % self.tournament)
		if os.path.isfile(file) == True:
			try:
				os.remove(file) 
			except WindowsError:
				pass

		zotournament = ProgramTournament(title=self.tournament)
		zotournament.page_entry_competitors()
		zo.open_menu('Entry')

#-----------------------------------------------------------------------------#

### PROGRAM CLASSES ### 

class ProgramLogin():

	def __init__(self, *args, **kwargs):
		
		## Settings
		self.style = kwargs.pop('style', 'Login') #

		zo.clear_sidemenu()
		zo.change_title('Login')

		## Display
		self.build_page()

	### BUILD ###

	def build_page(self):
		page = PageLogin(style=self.style)
		zo.change_page(page) #

class ProgramHome():

	def __init__(self, *args, **kwargs):

		## Settings
		self.first = kwargs.pop('first', False)
		self.db = Database('main')
		self.details = self.db.select(table='details')
		self.title = self.details[0][0]
		self.organisation = self.details[0][1]
		self.reg_key = self.details[0][2]

		## Database

		## Display
		zo.clear_sidemenu()
		zo.change_title(self.title)
		self.build_sidebar()
		self.page_main()

		## Engine

	### FUNCTIONS ###

	def command_open(self, tournament):

		if tournament == 'Select':
			return

		self.db.connection.close()
		zotournament = ProgramTournament(title=tournament.split(' - ')[0])

	def email_registration(self, *args):

		## Email Resgistration

		fromaddr = "zosportsnz@gmail.com"
		toaddr = "info@zo-sports.com"
 
		msg = MIMEMultipart()
 
		msg['From'] = fromaddr
		msg['To'] = toaddr
		msg['Subject'] = "Registration"
 
		body = "%s\n%s\n%s" % (self.title, self.organisation, 
								self.reg_key)
 
		msg.attach(MIMEText(body, 'plain'))
 
		server = smtplib.SMTP('smtp.gmail.com', 587)
		server.starttls()
		server.login(fromaddr, "vetinari13")
		text = msg.as_string()
		try:
			server.sendmail(fromaddr, toaddr, text)
			server.quit()
		except:
			pass	

	### SIDEBAR ###

	def command_sidebar(self, heading, subheading):

		if heading == 'Home':

			if subheading == 'Main':
				self.page_main()
			elif subheading == 'Settings':
				self.page_settings()
			elif subheading == 'Logout':
				self.db.connection.close()
				zologin = ProgramLogin()

		elif heading == 'Tournaments':

			if subheading == 'Create':
				self.page_tournaments_create()
			elif subheading == 'Delete':
				self.page_tournaments_delete()
			elif subheading == 'Open':
				self.page_tournaments_open()

	def build_sidebar(self):

		zo.add_sidemenu(heading='Home', function=self.command_sidebar,
						subheadings=['Main', 'Settings', 'Logout'])
		zo.add_sidemenu(heading='Tournaments', function=self.command_sidebar,
						subheadings=['Open', 'Create', 'Delete'])

	### BUILD ###

	def build_main(self):

		main_box = GridLayout(cols=1, padding=[0, 0, 0, 10])

		main_box.add_widget(Label(text='Current Tournaments', font_size=20, 
								size_hint_y=None, height=40))

		def change_scroll_height(scroll, value):

			scroll.height = value[1] - 50


		scroll = ScrollView(size_hint_y=None, height=0)
		main_box.bind(size=lambda i, value,
			scroll=scroll:change_scroll_height(scroll, value))
		main_box.add_widget(scroll)

		box = GridLayout(cols=1, spacing=10, padding=[0, 10, 0, 0],
							size_hint_y=None, height=0)
		scroll.add_widget(box)

		tournament_list = make_tournament_list(add_sport=True, add_stage=True)

		comp = []
		entry_e = []
		entry_c = []
		template = []

		for t in tournament_list:
			if t[1] == 'Template':
				template.append(t[0])
			elif t[1] == 'Entry - Competitors':
				entry_c.append(t[0])
			elif t[1] == 'Entry - Events':
				entry_e.append(t[0])
			elif t[1] == 'Competition':
				comp.append(t[0])

		tournaments = comp + entry_e + entry_c + template

		if len(tournaments) == 0:

			pass

		else:
			for t in tournaments:

				btn = Button(text=t, font_size=30, size_hint_y=None, height=50)
				btn.bind(on_press=lambda i,
						t=t:self.command_open(t))
				box.add_widget(btn)
				box.height += 60

		return main_box

	### PAGES - MAIN ###

	def page_main(self):

		page = PageMain()

		box = self.build_main()
		page.add_display(box)

		zo.change_page(page)

	def page_settings(self):

		settings = "The settings are predominantly decided when you first register, however there are a few other options.\n\nAdding an Admin Password can be helpful, especially when you start to give various access to the tournaments.\nHint: If you forget this Password the Registration Key will also always work"

		page = PageSetting(title='%s Settings' % self.organisation,
							add_help=settings)
		zo.change_page(page)

	### PAGES - TOURNAMENTS ###

	def page_tournaments_create(self):

		help_screen = "This page will create a new Tournament.\nYou can choose to create a new template for this tournament, or to copy another tournaments template."

		## Display
		page = PageTournamentCreate(title='Create New Tournament', 
										add_reset=True, add_help=help_screen)
		## Engine
		zo.change_page(page)		

	def page_tournaments_delete(self):

		## Settings
		help_screen = "This page allows you to delete a tournament.\nThis cannot be undone"
		page = Page(title='Delete Tournament', add_help=help_screen)	
		self.delete_tour = None

		## Functions

		def command(tournament):
			if tournament == 'Select':
				return
			self.delete_tour = tournament
			PopBox().askyesno(title='Delete', function=command_delete,
				message='''This will delete all saved information!\n\nAre you sure?''')

		def command_delete(state):

			if state == False:
				return
			tournament = self.delete_tour.split(' - ')[0]

			try:
				os.remove('%s.db' % tournament)
				command = "DELETE FROM tournaments WHERE title='%s'" % tournament

				file_list = ["%s Entry Sheet.xlsx" % tournament]
				for file in file_list:
					if os.path.isfile(file) == True:
						os.remove(file)

				self.db.cursor.execute(command)
				self.db.connection.commit()
				self.page_tournaments_delete()
			except:
				DataWarning()
				return

		## Database
		option_list = make_tournament_list(add_sport=True)

		## Display
		box = GridLayout(cols=1, padding=50)

		action = BoxLayout(size_hint_y=None, height=150)
		box.add_widget(action)

		option = OptionMenu(options=option_list, font_size=20)
		action.add_widget(option)
		confirm = Button(text='Delete', font_size=20)
		confirm.bind(on_press=lambda i, option=option:command(option.text))
		action.add_widget(confirm)

		page.add_display(box)

		## Engine
		zo.change_page(page)

	def page_tournaments_open(self):

		## Settings
		help_screen = "Most tournaments in progress can be accessed from the Main Page but all tournaments can be opened here.\nThis includes complete or archived tournaments"
		page = Page(title='Open Tournament', add_help=help_screen)	

		## Functions

		## Database
		option_list = make_tournament_list(add_sport=True)

		## Display
		box = GridLayout(cols=1, padding=50)

		action = BoxLayout(size_hint_y=None, height=150)
		box.add_widget(action)

		option = OptionMenu(options=option_list, font_size=20)
		action.add_widget(option)
		confirm = Button(text='Open', font_size=20)
		confirm.bind(on_press=lambda i, 
					option=option:self.command_open(option.text))
		action.add_widget(confirm)

		page.add_display(box)

		## Engine
		zo.change_page(page)

class ProgramTournament():

	def __init__(self, *args, **kwargs):
		
		## Settings
		self.tournament = kwargs.pop('title', None)
		main_db = Database('main')
		main_details = main_db.cursor.execute('''SELECT organisation 
											FROM details''').fetchone()
		self.organisation = main_details[0]

		self.db = Database('%s' % self.tournament)
		details = self.db.cursor.execute('SELECT * FROM details').fetchone()

		self.sport = details[1]
		self.stage = details[2]
		self.type = details[3]
		self.age = details[4]

		## Database
		self.excel_confirm = {}

		## Display
		zo.change_title(self.tournament)
		zo.clear_sidemenu()

		## Engine
		self.build_sidebar()
		self.page_main()

	### FUNCTIONS ###

	def command_competitor(self, i):

		if i.text == 'Add Competitor':
			EntryAddRemove(tournament=self.tournament, style='Add')
		if i.text == 'Remove Competitor':
			EntryAddRemove(tournament=self.tournament, style='Remove')
			
	### EXCEL FUNCTIONS ###

	def excel_entry_open(self, *args):

		# check if already exists
		if os.path.isfile(resource_path('%s Entry Sheet.xlsx' % self.tournament)) == True:			
			file_open(resource_path('%s Entry Sheet.xlsx' % self.tournament))
			return

		# if not, create and then open it
	
		wb = Workbook()
		ws = wb.active
		ws.title = "%s Entry Sheet" % self.tournament
		ws.column_dimensions['A'].width = 20
		ws.column_dimensions['B'].width = 20
		ws.column_dimensions['C'].width = 20
		ws.column_dimensions['D'].width = 20
		ws.column_dimensions['E'].width = 20


		headings = ["First Name", "Surname", "Date of Birth", "Gender"]			
		headings.append(self.type)
		if self.age != 'Date of Birth':
			headings.append(self.age)
			ws.column_dimensions['F'].width = 20
			if self.age == 'Year':
				ws["F5"].value = 'i.e. Y9'

		gender_check = self.db.cursor.execute('''SELECT *
								FROM template_groups
								WHERE filter="Male"''').fetchall()

		if gender_check in [None, []]:
			doe = 'Josephine'
		else:
			doe = 'Joseph'

		ws["A5"].value = 'i.e. %s' % doe
		ws["B5"].value = 'i.e. Bloggs'
		ws["C5"].value = 'dd/mm/YYYY'
		ws["D5"].value = 'i.e. M/F'
		ws["E5"].value = 'i.e. %s Title' % self.type

		# header
		header = ws.merge_cells(start_row=1, start_column=1, 
									end_row=3, end_column=len(headings))
		cellform = ws["A1"]
		cellform.value = 'Competitor Entry Details'	
		cellform.font = Font(size=30)
		cellform.alignment = Alignment(horizontal="center", vertical="center")
		cellform.fill = PatternFill(start_color='B0E0E6',
              						end_color='B0E0E6', fill_type='solid')        

		# columns
		for r in range(4, 6):
			for i in range(len(headings)):
				heading = headings[i]
				cellform = ws.cell(row=r, column=(i+1))
				if r == 4:
					cellform.value = heading
				cellform.font = Font(size=12)
				cellform.alignment = Alignment(horizontal="center", 
											vertical="center")
				cellform.fill = PatternFill(start_color='B0E0E6',
              						end_color='B0E0E6', fill_type='solid')
				cellform.border = Border(left=Side(style='thin'),
									right=Side(style='thin'),
									top=Side(style='thin'),
									bottom=Side(style='thin')) 	             						 		
		# save
		wb.save(resource_path("%s Entry Sheet.xlsx" % self.tournament))			
		# open
		file_open(resource_path("%s Entry Sheet.xlsx" % self.tournament))

	def excel_entry_confirm(self, *args):

		ExcelEntryConfirm(tournament=self.tournament, age=self.age, 
							type=self.type)

	### SIDEBAR ###

	def command_sidebar(self, heading, subheading):

		if heading == 'Tournament':
			if subheading == 'Main':
				self.page_main()
			elif subheading == 'Settings':
				self.page_settings()
			elif subheading == 'Home':
				self.db.connection.close()
				zohome = ProgramHome()

		elif heading == 'Print':

			if subheading == 'Entry Sheets':
				self.page_print_entry()	
			elif subheading == 'Program':
				self.page_print_program()
			elif subheading == 'Wristbands':
				self.page_print_wristbands()
			elif subheading == 'Events':
				self.page_print_events()
			elif subheading == self.type:
				self.page_print_type()
			elif subheading == 'Results':
				self.page_print_results()

		elif heading == 'Template':
			if subheading == 'Creation':
				self.page_template_create()
			elif subheading in ['Event Points', 'Event Restrictions',
				'Event Records', 'Event Program']:
				self.page_tournament_savecheck(subheading)
			elif subheading == 'Records':
				self.page_complete_records()

		elif heading == 'Entry':

			if self.stage == 'Entry - Competitors':
				self.page_entry_competitors()
			else:
				if subheading == 'Competitors':
					self.page_entry_competitors()
				elif subheading == 'Complete':
					self.page_entry_complete()
				else:
					self.page_entry_type(subheading)

		elif heading == 'Events':

			if subheading == 'Special Events':
				self.page_competition_special()
			else:
				self.page_competition_events(subheading)

		elif heading == 'Competition':

			if subheading == 'Complete':
				self.page_competition_complete()

	def build_sidebar(self):

		if self.stage == 'Template':
			zo.add_sidemenu(heading='Tournament', function=self.command_sidebar,
						subheadings=['Main', 'Settings', 'Home'])

			zo.add_sidemenu(heading='Template', function=self.command_sidebar,
							subheadings=['Creation'])

		elif self.stage == 'Entry - Competitors':
			zo.add_sidemenu(heading='Tournament', function=self.command_sidebar,
				subheadings=['Main', 'Settings','Home'])
			zo.add_sidemenu(heading='Template', function=self.command_sidebar,
				subheadings=['Event Points', 'Event Restrictions',
							'Event Records', 'Event Program'])
			zo.add_sidemenu(heading='Entry', function=self.command_sidebar,
							subheadings=['Competitors'])

		elif self.stage == 'Entry - Events':

			zo.add_sidemenu(heading='Tournament', function=self.command_sidebar,
				subheadings=['Main', 'Settings', 'Home'])
			zo.add_sidemenu(heading='Print', function=self.command_sidebar,
				subheadings=['Entry Sheets'])
			zo.add_sidemenu(heading='Template', function=self.command_sidebar,
				subheadings=['Event Points', 'Event Restrictions',
							'Event Records', 'Event Program'])
			db = self.db.cursor.execute('''SELECT title 
					FROM template_groups WHERE category="type"''').fetchall()
			
			options = ['Competitors']
			for tup in db:
				options.append(tup[0].title())
			options.append('Complete')

			zo.add_sidemenu(heading='Entry', function=self.command_sidebar,
							subheadings=options)

		elif self.stage == 'Competition':

			zo.add_sidemenu(heading='Tournament', function=self.command_sidebar,
				subheadings=['Main', 'Settings', 'Home'])
			zo.add_sidemenu(heading='Print', function=self.command_sidebar,
				subheadings=['Program', 'Wristbands', 'Events', self.type])			
			zo.add_sidemenu(heading='Template', function=self.command_sidebar,
				subheadings=['Event Points', 'Event Restrictions',
							'Event Records', 'Event Program'])			
			zo.add_sidemenu(heading='Entry', function=self.command_sidebar,
							subheadings=['Competitors'])

			db = self.db.cursor.execute('''SELECT title
					FROM template_groups WHERE category="type"''').fetchall()
			type_options = []
			for tup in db:
				type_options.append(tup[0].title())

			dbe = self.db.cursor.execute('''SELECT title, contest
										FROM template_events''').fetchall()
			table_list = self.db.list_tables()
			event_list = []
			for tup in dbe:
				table = '_'.join(tup[0].split(' '))
				if table in table_list:
					event_list.append(tup[1])
			contest_list = []
			for contest in calculate_contest_list(self.tournament):
				if contest in event_list:
					contest_list.append(contest)
			contest_list.append('Special Events')
			zo.add_sidemenu(heading='Events', function=self.command_sidebar,
							subheadings=contest_list)

			zo.add_sidemenu(heading='Competition', function=self.command_sidebar,
							subheadings=['Complete'])

		elif self.stage == 'Complete':

			zo.add_sidemenu(heading='Tournament', function=self.command_sidebar,
				subheadings=['Main', 'Settings', 'Home'])
			zo.add_sidemenu(heading='Print', function=self.command_sidebar,
				subheadings=['Results', 'Program', 'Wristbands', 
							'Events', self.type])			
			zo.add_sidemenu(heading='Template', function=self.command_sidebar,
				subheadings=['Event Points', 'Event Restrictions',
							'Records', 
							'Event Program'])			

			db = self.db.cursor.execute('''SELECT title
					FROM template_groups WHERE category="type"''').fetchall()
			type_options = []
			for tup in db:
				type_options.append(tup[0].title())

			dbe = self.db.cursor.execute('''SELECT title, contest
										FROM template_events''').fetchall()
			table_list = self.db.list_tables()
			event_list = []
			for tup in dbe:
				table = '_'.join(tup[0].split(' '))
				if table in table_list:
					event_list.append(tup[1])
			contest_list = []
			for contest in calculate_contest_list(self.tournament):
				if contest in event_list:
					contest_list.append(contest)
			contest_list.append('Special Events')
			zo.add_sidemenu(heading='Events', function=self.command_sidebar,
							subheadings=contest_list)

	### PAGES - MAIN ###

	def page_main(self):
	
		page = PageMain(tournament=self.tournament)
	
		if self.stage in ['Competition', 'Complete']:

			scoreboard = ScoreBoard(tournament=self.tournament)
			page.add_display(scoreboard)

		zo.change_page(page)

	def page_settings(self):

		setting = "This determines the access settings for this tournament.\nThere are 3 types of access:\n\nPublic: Only allows access to the tournament main page\n\nCaptain: Allows access to a specific %s in order to enter those competitors in events\n\nOfficial: Allows various levels of access to events, used to record scores etc" % self.type

		page = PageSetting(tournament=self.tournament, 
							title='Tournament Settings', add_help=setting)
		zo.change_page(page)

	### PAGES - TEMPLATE ###

	def page_template_create(self):

		page = PageTemplateCreate(tournament=self.tournament)
		zo.change_page(page)

	def page_tournament_savecheck(self, sub):

		if sub == 'Event Points':

			mes = "This page decides how competitors and their %ss are awarded points for participation and placing.\nPlacing or Competition Points are limited to 10th place.\nTeam Participation Points are usually 0\nNote: This Page will always be available for you to change" % self.type
			page = PageSaveCheck(title=sub, add_help=mes,
					tournament=self.tournament, style='points')

		elif sub == 'Event Restrictions':
			mes = "Here you can decide if there is a minimum requirement or maximum limit to how many events a competitor can enter.\nNote: This Page will always be available for you to change"

			page = PageSaveCheck(title=sub, add_help=mes,
				tournament=self.tournament, style='restrictions')

		elif sub == 'Event Records':
			mes = "If you have records, enter them here.\nThis template can be updated after your tournament and then next year if you copy this template it will have the records already stored.\nNote: This Page will always be available for you to change"
			page = PageSaveCheck(title=sub, add_help=mes,
				tournament=self.tournament, style='records')

		elif sub == 'Event Program':
			mes = "If you have a set time/number plan for your tournament, enter those details here.\nFor events that will have Finals or for multi-day tournaments, those program details can be added in the Competition Stage in the Print section under Program\nNote: This Page will always be available for you to change"
			page = PageSaveCheck(title=sub, add_help=mes,
				tournament=self.tournament, style='program')

		zo.change_page(page)

	### PAGES - ENTRY ###

	def page_entry_competitors(self):

		## Settings
		if self.stage == 'Entry - Competitors':
			mes = "This is where you add the majority of the competitors details into the program.\n\nClick 'Open Entry Sheet' to open an excel sheet which has columns for you to enter competitor details. Be careful to enter names correctly and don't leave any extra whitespace, also, date of birth needs to be entered in dd/mm/YYYY style.\n\nOnce you have filled this sheet out, click 'Confirm Entry Details', this will check the file, let you know any issues and if required ask for verification about filters. It will then process all the information and therefore this may take a little while."
		elif self.stage == 'Entry - Events':
			mes = "Click 'Add Competitor' and fill out all the details asked for.\n\nThe program will then search for any matchs in the system, and then ask you for confirmation to add the competitor.\n\nOnce added the competitor should automatically be added to the correct %s Entry Sheet" % self.type
		elif self.stage in ['Competition', 'Complete']:
			mes = "Click 'Add Competitor' and fill out all the details asked for.\n\nThe program will then search for any matchs in the system, and then ask you for confirmation to add the competitor.\n\nOnce added the competitors details should be available in any Prints or to enter in Events"

		page = Page(title='Competitor - Tournament Entry', add_help=mes)

		## Display
		box = GridLayout(cols=1, padding=50, spacing=50)

		action = BoxLayout(size_hint_y=None, height=150)
		box.add_widget(action)

		if self.stage == 'Entry - Competitors':
			open_excel = Button(text='Open Entry Sheet', font_size=20)
			open_excel.bind(on_press=self.excel_entry_open)
			action.add_widget(open_excel)

			confirm_excel = Button(text='Confirm Entry Details', font_size=20)
			confirm_excel.bind(on_press=lambda i:self.excel_entry_confirm(False))
			action.add_widget(confirm_excel)

		elif self.stage == 'Entry - Events':
			add_comp = Button(text='Add Competitor', font_size=20)
			add_comp.bind(on_press=self.command_competitor)
			action.add_widget(add_comp)

		elif self.stage == 'Competition':
	
			add_comp = Button(text='Add Competitor', font_size=20)
			add_comp.bind(on_press=self.command_competitor)
			action.add_widget(add_comp)

		page.add_display(box)
		
		zo.change_page(page)

	def page_entry_type(self, type):

		mes = "This is where %s %s competitors are entered into events.\n\nClick the Grade in the panel under the heading (if there are too many, this panel will scroll left to right), this will open the Grade Entry Page.\n\nSelect which events the competitor will enter. If the red 'Unsaved' label appears in the bottom left corner, the current Grade Entry Sheet changes are unsaved, press the 'Save' button to save them" % (type, self.type)

		page = PageEntry(title='%s %s - Event Entry' % (type, 
				self.type), add_help=mes, 
				tournament=self.tournament, typename=type)
		zo.change_page(page)

	def page_entry_complete(self):

		help_screen = '''This will finish the Entry section.\n\nYou will still be able to add late entries during the Competition section as well as change Event Restrictions, Event Records and Event Number & Times.\n\nPress 'Entry Check' which will do a final check of all entered details, then the 'Save and Continue' button will appear.'''

		page = PageEntryComplete(tournament=self.tournament, 
									add_help=help_screen)
		zo.change_page(page)

	### PAGES - EVENTS ###

	def page_competition_events(self, contest):

		contest_type = calculate_contest_type(self.tournament, contest)

		if contest_type == 'time':

			help_screen = "This page is where the %s event is recorded.\n\nThe competitors are in alphabetical order, their %s initials and colour to their left.\n\nPP stands for Participation Points. These will automatically select if you enter any scores or you can select them yourself.\n\nTime is measured in minutes (m) and seconds (s). If you enter 72 seconds, when you next load this page it will automatically convert that to 1m 12s.\n\nPlacing can be manually assigned or you can click the small 'Calculate' button, this will calculate the placing based on the scores currently entered.\n\nIf you need to add a new competitor, this can be done by clicking the 'Add Entry' button under the %s column on the left\n\nNext to the 'Save' button are the 'Add Round' button which will allow you to make a final/semi-final etc and the 'Complete' button. The 'Complete button will finalise the event and create the Results page. Be aware that the Results will finalise based on the placings of the current page.\n\nDifferent rounds and the results page can be found by clicking the main title and choosing from the dropdown menu." % (contest, self.type, self.type)

		elif contest_type == 'distance':

			if contest in ['High Jump', 'Pole Vault']:

				help_screen = "This page is where the %s event is recorded.\n\nThe competitors are in alphabetical order, their %s initials and colour to their left.\n\nPP stands for Participation Points. These will automatically select if you enter any scores or you can select them yourself.\n\nHeights are measured in metres (m) and are entered at the top of each score column. Pass(P), Miss(X), Made(O) are entered by clicking the blank buttons\n\nPlacing can be manually assigned or you can click the small 'Calculate' button, this will calculate the placing based on the scores currently entered.\n\nIf you need to add a new competitor, this can be done by clicking the 'Add Entry' button under the %s column on the left\n\nNext to the 'Save' button are the 'Add Round' button which will allow you to make a final/semi-final etc and the 'Complete' button. The 'Complete button will finalise the event and create the Results page. Be aware that the Results will finalise based on the placings of the current page.\n\nDifferent rounds and the results page can be found by clicking the main title and choosing from the dropdown menu." % (contest, self.type, self.type)

			else:

				help_screen = "This page is where the %s event is recorded.\n\nThe competitors are in alphabetical order, their %s initials and colour to their left.\n\nPP stands for Participation Points. These will automatically select if you enter any scores or you can select them yourself.\n\nDistances are measured in metres(m) and there are up to 6 attempts which can be recorded. The best score will automatically calculate.\n\nPlacing can be manually assigned or you can click the small 'Calculate' button, this will calculate the placing based on the scores currently entered.\n\nIf you need to add a new competitor, this can be done by clicking the 'Add Entry' button under the %s column on the left\n\nNext to the 'Save' button are the 'Add Round' button which will allow you to make a final/semi-final etc and the 'Complete' button. The 'Complete button will finalise the event and create the Results page. Be aware that the Results will finalise based on the placings of the current page.\n\nDifferent rounds and the results page can be found by clicking the main title and choosing from the dropdown menu." % (contest, self.type, self.type)
		
		page = PageEvent(tournament=self.tournament, contest=contest,
							add_help=help_screen)
		zo.change_page(page)

	def page_competition_special(self):

		help_screen = "This page allows the creation of special events which don't fall under the normal tournament events.\n\nAs examples...\n:A student may not be able to compete and instead helps out and for that, his House is awarded some points.\n\n:Part of the tournament may involve an all-in-relay and points are awarded based on how many compete.\n\n:Special events such as Tug-of-War or House Chant etc can also be awarded points on this page.\n\nIf you create a special event that you decide to cancel, you can just unselect it and press 'Save', this will remove it when the page is refreshed."

		page = PageSaveCheck(title='Special Events', add_help=help_screen,
								tournament=self.tournament, style='special')
		zo.change_page(page)

	### PAGE - COMPETITION ###

	def page_competition_complete(self):

		help_screen = "This page will complete this tournament.\n\nThe tournament main page will have an altered Scoreboard to show which %s won and where the rest placed.\n\nCompleting this tournament will remove it from the Current Tournaments on the %s main page but it can be opened under the Open Tournaments page.\n\nPublic Access will remain on unless you turn it off in Settings\n\nTournament completion will also create an Archive of the current Template Records so you can update the new Template with new records (whilst keeping a reference to the old ones)." % (self.type, self.organisation)

		page = PageComplete(tournament=self.tournament, add_help=help_screen)
		zo.change_page(page)
		self.db.connection.close() # thought on how to fix the complete issue

	### PAGES - PRINT ###

	def page_print_entry(self):

		## Settings
		help_screen = "This page creates Excel Sheet versions of the %s Entry Sheets\n\nYou can either open All, or each %s individually.\n\nBe aware with large tournaments the process may take a few seconds, so don't double click the button" % (self.type, self.type)
		page = Page(title='Print - %s Entry Sheets' % self.type,
					add_help=help_screen)

		## Functions
		def command(i):

			if i.text == 'All':
				ExcelEntrySheets(tournament=self.tournament)
			else:
				ExcelEntrySheets(tournament=self.tournament,
									typename=i.text)

		def command_scroll(value, scroll):

			scroll.height = value - 50

		## Display
		box = BoxLayout()

		left = GridLayout(cols=1, padding=[50, 50, 50, 0])
		box.add_widget(left)

		btn1 = Button(text='All', font_size=20, size_hint_y=None, height=150)
		btn1.bind(on_press=command)
		left.add_widget(btn1)

		right = ScrollView(size_hint_y=None, height=600)
		box.add_widget(right)
		page.bind(height=lambda obj, value, 
							right=right:command_scroll(value, right))

		display = GridLayout(cols=1, padding=[50, 50, 50, 0], spacing=10,
								size_hint_y=None, height=0)
		right.add_widget(display)

		type_list = sorted(self.db.cursor.execute('''SELECT title 
								FROM template_groups
								WHERE category="type"''').fetchall())
		for tup in type_list:
			btn = Button(text=tup[0], font_size=20, 
							size_hint_y=None, height=100)
			btn.bind(on_press=command)
			display.add_widget(btn)
			display.height += 110

		## Engine
		page.add_display(box)
		zo.change_page(page)

	def page_print_program(self):

		## Settings
		help_screen = "This page is in construction.\n\nThe intention is to have this automatically produce a Program for the tournament"
		page = Page(title='Print - Program',
					add_help=help_screen)

		## Functions
		def command(i):

			pass

		## Display
		box = GridLayout(cols=1, padding=[50, 50, 50, 0], spacing=50)

		btn1 = Button(text='Program', font_size=20, 
						size_hint_y=None, height=150)
		btn1.bind(on_press=command)
		box.add_widget(btn1)

		page.add_display(box)
		zo.change_page(page)

	def page_print_wristbands(self):

		## Settings
		help_screen = "This page will opens Excel files with printable Wristbands for each %s.\n\nThe Wristbands have the competitors ID Number, Name and the Event and Time for any events that they are entered in.\n\nThe Excel sheet is in landscape orientation and we recommend printing them on paper of the same colour as the %s" % (self.type, self.type)
		page = Page(title='Print - Competitor Wristbands',
					add_help=help_screen)

		## Functions
		def command(i):

			ExcelWristbands(tournament=self.tournament, typename=i.text,
								sponsor=True)

		def command_scroll(value, scroll):

			scroll.height = value - 50

		## Display
		box = ScrollView(size_hint_y=None, height=600)
		page.bind(height=lambda obj, value, 
							box=box:command_scroll(value, box))

		display = GridLayout(cols=1, padding=[50, 50, 50, 0], spacing=10,
								size_hint_y=None, height=50)
		box.add_widget(display)

		type_list = sorted(self.db.cursor.execute('''SELECT title 
								FROM template_groups
								WHERE category="type"''').fetchall())
		for tup in type_list:
			btn = Button(text=tup[0], font_size=20, 
							size_hint_y=None, height=100)
			btn.bind(on_press=command)
			display.add_widget(btn)
			display.height += 110

		## Engine
		page.add_display(box)
		zo.change_page(page)

	def page_print_events(self):

		## Settings
		help_screen = "This page opens printable Excel files of each Event Marshall Sheets.\n\nYou can choose to open All events (or all Track/Field events), or to open all events in one particular contest. If there are many contests, the right hand side list will become scrollable up and down.\n\nBe aware that with large tournaments the process to make these files might take little time."
		page = Page(title='Print - Event Marshall Sheets',
					add_help=help_screen)

		## Functions
		def command(i):

			if i.text == 'All':
				ExcelEventsBase(tournament=self.tournament, 
								style='All')
			elif i.text == 'Track':
				ExcelEventsBase(tournament=self.tournament,
								style='time')
			elif i.text == 'Field':
				ExcelEventsBase(tournament=self.tournament,
								style='distance')
			else:
				ExcelEventsBase(tournament=self.tournament,
								contest=i.text)

		def command_scroll(value, scroll1, scroll2):

			scroll1.height = value - 50
			scroll2.height = value - 50

		## Display
		box = BoxLayout()

		scroll = ScrollView(size_hint_y=None, height=600)
		box.add_widget(scroll)

		left = GridLayout(cols=1, padding=[50, 50, 50, 0], spacing=10,
							size_hint_y=None, height=0)
		scroll.add_widget(left)

		btn1 = Button(text='All', font_size=20, size_hint_y=None, height=150)
		btn1.bind(on_press=command)
		left.add_widget(btn1)
		left.height += 150

		if self.sport == 'Athletics':
			for b in ['Track', 'Field']:
				btn = Button(text=b, font_size=20, size_hint_y=None, height=150)
				btn.bind(on_press=command)
				left.add_widget(btn)
				left.height += 150

		right = ScrollView(size_hint_y=None, height=600)
		box.add_widget(right)
		page.bind(height=lambda obj, value, 
			scroll=scroll, right=right:command_scroll(value, scroll, right))

		display = GridLayout(cols=1, padding=[50, 50, 50, 0], spacing=10,
								size_hint_y=None, height=50)
		right.add_widget(display)

		for contest in calculate_contest_list(self.tournament):
			btn = Button(text=contest, font_size=20, 
							size_hint_y=None, height=100)
			btn.bind(on_press=command)
			display.add_widget(btn)
			display.height += 110

		## Engine
		page.add_display(box)
		zo.change_page(page)

	def page_print_type(self):

		## Settings
		help_screen = "This page opens printable Excel files that list all the %s competitors in each event.\n\nIt gives a complete list of the name and ID number of each competitor in each grade.\n\nIf additional event rounds are run i.e. Semi-Finals or Finals, then this page will update with those details" % self.type
		page = Page(title='Print - %s Entry Lists' % self.type,
					add_help=help_screen)

		## Functions
		def command(i):

			if i.text == 'All':
				ExcelTypeList(tournament=self.tournament)
			else:
				ExcelTypeList(tournament=self.tournament, typename=i.text)

		def command_scroll(value, scroll):

			scroll.height = value - 50

		## Display
		box = BoxLayout()

		left = GridLayout(cols=1, padding=[50, 50, 50, 0])
		box.add_widget(left)

		btn1 = Button(text='All', font_size=20, size_hint_y=None, height=150)
		btn1.bind(on_release=command)
		left.add_widget(btn1)

		right = ScrollView(size_hint_y=None, height=600)
		box.add_widget(right)
		page.bind(height=lambda obj, value, 
							right=right:command_scroll(value, right))

		display = GridLayout(cols=1, padding=[50, 50, 50, 0], spacing=10,
								size_hint_y=None, height=0)
		right.add_widget(display)

		type_list = sorted(self.db.cursor.execute('''SELECT title 
								FROM template_groups
								WHERE category="type"''').fetchall())
		for tup in type_list:
			btn = Button(text=tup[0], font_size=20, 
							size_hint_y=None, height=100)
			btn.bind(on_release=command)
			display.add_widget(btn)
			display.height += 110

		## Engine
		page.add_display(box)
		zo.change_page(page)

	def page_print_results(self):
	
		## Settings
		help_screen = "This page will produce a tournament results sheet. This can be used for general information and/or as a media release\n\nOverall Results of each %s will be listed.\n\nResults will also be listed by event, with 1st, 2nd and 3rd placings (where appropriate).\n\nRecords that have been broken will also be included" % self.type
		page = Page(title='Print - Results',
					add_help=help_screen)

		## Functions
		def command(i):
			
			ExcelResults(tournament=self.tournament)

		## Display
		box = GridLayout(cols=1, padding=[50, 50, 50, 0], spacing=50)

		btn1 = Button(text='Results', font_size=20, 
						size_hint_y=None, height=150)
		btn1.bind(on_press=command)
		box.add_widget(btn1)

		page.add_display(box)
		zo.change_page(page)

	### PAGE - COMPLETE ###

	def page_complete_records(self):

		mes = "This page contains both the old Archive and new Template records.\n\nArchive records refer to the event records that were set as of the creation of this tournament and can't be changed (this allows you to retain a record history).\n\nThe new Template records allow you to update these records now, so that they are already available for future tournaments.\n\nYou can do this manually by changing the values or you can click the 'Results' button and choose from a list of those competitors who broke the archive record during this tournament.\n\nBe aware the automated option only allows you to choose one record"
		page = Page(title='Records', add_help=mes)

		display = PageUpdateRecords(tournament=self.tournament)
		page.add_display(display)

		zo.change_page(page)

### PROGRAM CLASES - ADDITIONAL ###

class ProgramCaptain():

	def __init__(self, *args, **kwargs):
	
		## Settings
		self.tournament = kwargs.pop('tournament', None)
		self.role = 'Captain'
		self.category = kwargs.pop('category', None)
		self.title = 'Captain - %s' % self.category
		self.db = Database(self.tournament)

		details = self.db.cursor.execute("SELECT * FROM details").fetchone()
		self.sport = details[1]
		self.stage = details[2]
		self.type = details[3]

		## Display
		zo.clear_sidemenu()
		zo.change_title(self.tournament)
		self.build_sidebar()
		self.page_main()

		## Engine

	### FUNCTIONS ###

	def command_competitor(self, i):

		if i.text == 'Add Competitor':
			EntryAddRemove(tournament=self.tournament, style='Add',
							typename=self.category)
		if i.text == 'Remove Competitor':
			EntryAddRemove(tournament=self.tournament, style='Remove',
							typename=self.category)

	### SIDEBAR ###

	def command_sidebar(self, heading, subheading):

		if subheading == 'Main':
			self.page_main()

		elif subheading == 'Logout':
			zologin = ProgramLogin()

		if heading == 'Entry':

			if subheading == 'Competitors':
				self.page_entry_competitors()

			elif subheading == self.category:
				self.page_entry_type()

		elif heading == 'Competition':

			pass

	def build_sidebar(self):

		zo.add_sidemenu(heading='Tournament', function=self.command_sidebar,
						subheadings=['Main', 'Logout'])

		if self.stage == 'Entry - Events':

			zo.add_sidemenu(heading='Entry', function=self.command_sidebar,
						subheadings=['Competitors', self.category])

		#elif self.stage == 'Competition':

			#zo.add_sidemenu(heading=self.type, function=self.command_sidebar,
			#			subheadings=[self.category])

	### PAGES ###

	def page_main(self):

		page = PageMain(tournament=self.tournament, 
						motto='Captain - %s %s' % (self.category, self.type))
	
		if self.stage == 'Competition':

			scoreboard = ScoreBoard(tournament=self.tournament)
			page.add_display(scoreboard)

		zo.change_page(page)

	def page_entry_competitors(self):

		mes = "Click 'Add Competitor' and fill out all the details asked for.\n\nThe program will then search for any matchs in the system, and then ask you for confirmation to add the competitor.\n\nOnce added the competitor should automatically be added to the %s Entry Sheet" % self.category
		page = Page(add_help=mes,
				title='%s %s Competitor - Tournament Entry' % (self.category, 
				self.type))
	
		## Display
		box = GridLayout(cols=1, padding=50, spacing=50)

		action = BoxLayout(size_hint_y=None, height=150)
		box.add_widget(action)

		add_comp = Button(text='Add Competitor', font_size=20)
		add_comp.bind(on_press=self.command_competitor)
		action.add_widget(add_comp)

		page.add_display(box)

		zo.change_page(page)

	def page_entry_type(self):

		mes = "This is where %s %s competitors are entered into events.\n\nClick the Grade in the panel under the heading (if there are too many, this panel will scroll left to right), this will open the Grade Entry Page.\n\nSelect which events the competitor will enter. If the red 'Unsaved' label appears in the bottom left corner, the current Grade Entry Page changes are unsaved, press the 'Save' button to save them" % (self.category, self.type)

		page = PageEntry(title='%s %s - Event Entry' % (self.category, 
				self.type), add_help=mes, 
				tournament=self.tournament, typename=self.category)
		zo.change_page(page)

class ProgramPublic():

	def __init__(self, *args, **kwargs):

		## Settings
		self.tournament = kwargs.pop('tournament', None)
		self.db = Database(self.tournament)
		details = self.db.cursor.execute("SELECT * FROM details").fetchone()
		self.sport = details[1]
		self.stage = details[2]
		self.type = details[3]

		## Display
		zo.clear_sidemenu()
		zo.change_title(self.tournament)
		self.build_sidebar()
		self.page_main()

		## Engine

	### FUNCTIONS ###

	### SIDEBAR ###

	def command_sidebar(self, heading, subheading):

		if subheading == 'Logout':
			zologin = ProgramLogin()

		elif subheading == 'Main':
			self.page_main()

	def build_sidebar(self):

		if self.stage == 'Competition':

			zo.add_sidemenu(heading='Tournament', function=self.command_sidebar,
							subheadings=['Main', 'Logout'])

		else:
			zo.add_sidemenu(heading='Tournament', function=self.command_sidebar,
							subheadings=['Logout'])

	### PAGES ###

	def page_main(self):

		page = PageMain(tournament=self.tournament)

		if self.stage in ['Competition', 'Complete']:

			scoreboard = ScoreBoard(tournament=self.tournament)
			page.add_display(scoreboard)

		zo.change_page(page)

class ProgramOfficial():
	
	def __init__(self, *args, **kwargs):
	
		## Settings
		self.tournament = kwargs.pop('tournament', None)
		self.role = 'Official'
		self.category = kwargs.pop('category', None) # All, Track, Field
		self.title = '%s - %s Events' % (self.role, self.category)
		self.db = Database(self.tournament)

		details = self.db.cursor.execute("SELECT * FROM details").fetchone()
		self.sport = details[1]
		self.stage = details[2]
		self.type = details[3]

		## Display
		zo.clear_sidemenu()
		zo.change_title(self.tournament)
		self.build_sidebar()
		self.page_main()

		## Engine

	### FUNCTIONS ###

	def command_competitor(self, i):

		if i.text == 'Add Competitor':
			EntryAddRemove(tournament=self.tournament, style='Add',
							typename=self.category)
		if i.text == 'Remove Competitor':
			EntryAddRemove(tournament=self.tournament, style='Remove',
							typename=self.category)

	### SIDEBAR ###

	def command_sidebar(self, heading, subheading):

		if subheading == 'Main':
			self.page_main()

		elif subheading == 'Logout':
			zologin = ProgramLogin()

		elif heading == 'Events':
			self.page_competition_events(subheading)

	def build_sidebar(self):

		zo.add_sidemenu(heading='Tournament', function=self.command_sidebar,
						subheadings=['Main', 'Logout'])

		if self.stage == 'Competition':

			dbe = self.db.cursor.execute('''SELECT title, contest
										FROM template_events''').fetchall()

			table_list = self.db.list_tables()
			event_list = []
			for tup in dbe:
				table = '_'.join(tup[0].split(' '))
				if table in table_list:
					event_list.append(tup[1])

			contest_list = []
			for contest in calculate_contest_list(self.tournament):
				if contest in event_list:

					contest_type = calculate_contest_type(self.tournament, 
										contest)

					if self.category == 'Track':
						if contest_type == 'time':
							contest_list.append(contest)
					elif self.category == 'Field':
						if contest_type == 'distance':
							contest_list.append(contest)
					else:
						contest_list.append(contest)

			zo.add_sidemenu(heading='Events', function=self.command_sidebar,
						subheadings=contest_list)

	### PAGES ###

	def page_main(self):

		page = PageMain(tournament=self.tournament, 
						motto=self.title)

		if self.stage == 'Competition':

			scoreboard = ScoreBoard(tournament=self.tournament)
			page.add_display(scoreboard)

		zo.change_page(page)

	def page_competition_events(self, contest):

		contest_type = calculate_contest_type(self.tournament, contest)

		if contest_type == 'time':

			help_screen = "This page is where the %s event is recorded.\n\nThe competitors are in alphabetical order, their %s initials and colour to their left.\n\nPP stands for Participation Points. These will automatically select if you enter any scores or you can select them yourself.\n\nTime is measured in minutes (m) and seconds (s). If you enter 72 seconds, when you next load this page it will automatically convert that to 1m 12s.\n\nPlacing can be manually assigned or you can click the small 'Calculate' button, this will calculate the placing based on the scores currently entered.\n\nIf you need to add a new competitor, this can be done by clicking the 'Add Entry' button under the %s column on the left\n\nNext to the 'Save' button are the 'Add Round' button which will allow you to make a final/semi-final etc and the 'Complete' button. The 'Complete button will finalise the event and create the Results page. Be aware that the Results will finalise based on the placings of the current page.\n\nDifferent rounds and the results page can be found by clicking the main title and choosing from the dropdown menu." % (contest, self.type, self.type)

		elif contest_type == 'distance':

			if contest in ['High Jump', 'Pole Vault']:

				help_screen = "This page is where the %s event is recorded.\n\nThe competitors are in alphabetical order, their %s initials and colour to their left.\n\nPP stands for Participation Points. These will automatically select if you enter any scores or you can select them yourself.\n\nHeights are measured in metres (m) and are entered at the top of each score column. Pass(P), Miss(X), Made(O) are entered by clicking the blank buttons\n\nPlacing can be manually assigned or you can click the small 'Calculate' button, this will calculate the placing based on the scores currently entered.\n\nIf you need to add a new competitor, this can be done by clicking the 'Add Entry' button under the %s column on the left\n\nNext to the 'Save' button are the 'Add Round' button which will allow you to make a final/semi-final etc and the 'Complete' button. The 'Complete button will finalise the event and create the Results page. Be aware that the Results will finalise based on the placings of the current page.\n\nDifferent rounds and the results page can be found by clicking the main title and choosing from the dropdown menu." % (contest, self.type, self.type)

			else:

				help_screen = "This page is where the %s event is recorded.\n\nThe competitors are in alphabetical order, their %s initials and colour to their left.\n\nPP stands for Participation Points. These will automatically select if you enter any scores or you can select them yourself.\n\nDistances are measured in metres(m) and there are up to 6 attempts which can be recorded. The best score will automatically calculate.\n\nPlacing can be manually assigned or you can click the small 'Calculate' button, this will calculate the placing based on the scores currently entered.\n\nIf you need to add a new competitor, this can be done by clicking the 'Add Entry' button under the %s column on the left\n\nNext to the 'Save' button are the 'Add Round' button which will allow you to make a final/semi-final etc and the 'Complete' button. The 'Complete button will finalise the event and create the Results page. Be aware that the Results will finalise based on the placings of the current page.\n\nDifferent rounds and the results page can be found by clicking the main title and choosing from the dropdown menu." % (contest, self.type, self.type)
		
		page = PageEvent(tournament=self.tournament, contest=contest,
							add_help=help_screen)
		zo.change_page(page)

#-----------------------------------------------------------------------------#

### ENGINE ###

if __name__ == '__main__':

	# Settings
	Clock.max_iteration = 33

	if datetime.now().year == 2018:

		pass

	elif os.path.isfile(os.path.join(os.path.abspath("."), 'main.db')) == False:

		MainApp(version='Main').run()

	else:
		db = Database('main')

		data = db.cursor.execute('''SELECT title, registration
									FROM details''').fetchone()
		title = data[0]
		registration = data[1]

		if champollion(title) == registration:
			MainApp(version='Login').run()

		else:
			MainApp(version='Register').run() #

#-----------------------------------------------------------------------------#

### CREATION ###

created_by = 'Atheros'

#-----------------------------------------------------------------------------#